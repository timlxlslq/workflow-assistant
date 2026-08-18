from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR = Path("/Users/lantian/Library/Mobile Documents/com~apple~CloudDocs/PacificPride/Order/PP0018")
CATALOG_PATH = Path("data/inventory/current-products.xlsx")
OUTPUT_JSON = Path("outputs/pp0018-material-cost/material_data.json")


def text(value) -> str:
    return "" if value is None else str(value).strip()


def parse_quantity(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = text(value).replace("，", ",")
    if not raw:
        return 0.0
    parts = [part.strip() for part in re.split(r"[+＋]", raw)]
    total = 0.0
    for part in parts:
        if not re.fullmatch(r"\d+(?:\.\d+)?", part):
            return 0.0
        total += float(part)
    return total


def normalize_name(value: str) -> str:
    value = re.sub(r"\s+", " ", text(value)).strip()
    value = value.replace("--", "--")
    return value


def material_rows(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    header_row = None
    name_col = qty_col = spec_col = remark_col = None
    active = False
    rows = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = [text(value) for value in row]
        if "Panel板材" in values:
            active = True
        if active and "Accessories配件" in values:
            break
        if active and "Name名字" in values and "QTY数量" in values:
            header_row = row_number
            name_col = values.index("Name名字")
            qty_col = values.index("QTY数量")
            spec_col = values.index("Spec规格") if "Spec规格" in values else None
            remark_col = values.index("Remarks备注") if "Remarks备注" in values else None
            continue
        if active and header_row and row_number > header_row:
            if not row or not isinstance(row[0], (int, float)):
                continue
            name = normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            if not name:
                continue
            qty_raw = row[qty_col] if qty_col is not None and qty_col < len(row) else None
            rows.append({
                "row": row_number,
                "name": name,
                "spec": text(row[spec_col]) if spec_col is not None and spec_col < len(row) else "",
                "quantity": parse_quantity(qty_raw),
                "quantity_raw": text(qty_raw),
                "remark": text(row[remark_col]) if remark_col is not None and remark_col < len(row) else "",
            })
    wb.close()
    return rows


def room_name(path: Path) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["WorkOrderTraveler"]
    value = text(ws["D6"].value)
    wb.close()
    return re.sub(r"\s+", " ", value).strip()


def load_prices() -> dict[str, dict]:
    wb = load_workbook(CATALOG_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = {text(value): index for index, value in enumerate(rows[1]) if text(value)}
    result = {}
    for row in rows[2:]:
        code = text(row[headers["*商品编号"]])
        if not code:
            continue
        result[code] = {
            "code": code,
            "category": text(row[headers["商品类别"]]),
            "name": text(row[headers["商品名称"]]),
            "spec": text(row[headers["规格型号"]]),
            "unit": text(row[headers["计量单位"]]),
            "price": float(row[headers["预计采购价"]] or 0),
        }
    wb.close()
    return result


def mapped_price(material_name: str, raw_remark: str, prices: dict[str, dict]) -> dict:
    name = normalize_name(material_name).replace("\t", "")
    remark = normalize_name(raw_remark)
    canonical = name
    if name.lower() == "edge banding" and remark:
        canonical = f"Edge banding--{remark}"
    # These codes follow the project's existing material matching rules.
    fixed = {
        "18mm--plywood": "M0004",
        "14.5mm--plywood": "M0003",
        "5.4mm--plywood": "M0002",
        "19.1mm--Woodline 4": "M0019",
        "19.1mm--Walnut": "M0159",
        "19.1mm--Cashmere SM": "M0049",
        "19.1mm--Blanco SM": "M0055",
        "19.1mm--Ivory Oak": "M0144",
        "9mm--Walnut": None,
        "Edge banding--Woodline 4": "M0020",
        "Edge banding--Walnut": "M0160",
        "Edge banding--Ivory Oak": "M0145",
        "Edge banding--Cashmere SM": "M0050",
    }
    code = fixed.get(canonical)
    if code and code in prices:
        product = prices[code]
        return {
            "canonical_name": canonical,
            "code": code,
            "product_name": product["name"],
            "product_spec": product["spec"],
            "unit": product["unit"],
            "price": product["price"],
            "price_status": "已知",
            "price_note": "库存商品资料 / 预计采购价；沿用项目现有材料映射规则",
        }
    return {
        "canonical_name": canonical,
        "code": "",
        "product_name": "",
        "product_spec": "",
        "unit": "",
        "price": 0.0,
        "price_status": "未知价格",
        "price_note": "商品库中没有可确认的对应价格",
    }


def main() -> None:
    prices = load_prices()
    files = sorted(SOURCE_DIR.glob("Work Order Traveler(*.xlsx"))
    # Workbook glob is intentionally narrowed below to avoid hidden/system files.
    files = sorted(path for path in SOURCE_DIR.glob("*.xlsx") if path.name.lower().startswith("work order traveler("))
    raw_by_room: dict[str, list[dict]] = defaultdict(list)
    source_counts: dict[str, set[str]] = defaultdict(set)
    room_files: dict[str, list[str]] = defaultdict(list)

    for path in files:
        label = room_name(path)
        short_file = path.stem.removeprefix("Work Order Traveler(").removesuffix(")")
        # The user-defined reporting unit is one Traveler file per room. Keep
        # the file identity in the room key so same-named files remain separate.
        room = f"{label}（{short_file}）"
        picking = material_rows(path, "Pickinglist")
        cutting = material_rows(path, "Cuttinglist")
        # Use Pickinglist quantities when present. For blank material rows, use
        # the corresponding Cuttinglist material row; this covers older and
        # update Traveler files where only one of the two sheets is populated.
        cutting_by_key: dict[tuple[str, str], list[dict]] = defaultdict(list)
        for row in cutting:
            cutting_by_key[(normalize_name(row["name"]).replace("\t", ""), row["spec"])].append(row)
        used_cutting_keys: set[tuple[str, str]] = set()
        chosen: list[dict] = []
        for row in picking:
            key = (normalize_name(row["name"]).replace("\t", ""), row["spec"])
            quantity = row["quantity"]
            source = "Pickinglist"
            if quantity <= 0:
                candidates = cutting_by_key.get(key, [])
                if not candidates and row["name"].startswith("19.1mm--") and row["name"] == "19.1mm--":
                    candidates = [item for item in cutting if item["name"].startswith("19.1mm--") and item["quantity"] > 0]
                if candidates:
                    quantity = sum(item["quantity"] for item in candidates)
                    source = "Cuttinglist（Pickinglist为空）"
                    used_cutting_keys.add(key)
                    if row["name"] == "19.1mm--" and candidates:
                        row = {
                            **row,
                            "name": candidates[0]["name"],
                            "spec": candidates[0]["spec"],
                            "remark": candidates[0]["remark"],
                        }
            if quantity > 0:
                chosen.append({**row, "quantity": quantity, "quantity_source": source})
        for row in cutting:
            key = (normalize_name(row["name"]).replace("\t", ""), row["spec"])
            if key in used_cutting_keys:
                continue
            # Add cutting-only rows, including the update kitchen file's
            # color-specific panel row that is blank in Pickinglist.
            if row["quantity"] > 0 and not any(
                normalize_name(item["name"]).replace("\t", "") == key[0]
                and item["quantity"] > 0
                for item in chosen
            ):
                chosen.append({**row, "quantity_source": "Cuttinglist（Pickinglist缺少）"})
        room_files[room].append(path.name)
        for item in chosen:
            item_room = room
            item["room"] = item_room
            item["source_file"] = path.name
            raw_by_room[room].append(item)
            source_counts[(room, normalize_name(item["name"]).replace("\t", ""), item["spec"])].add(path.name)

    # Infer a generic edge-band color only when a room has exactly one
    # positive colored panel; this is an explicit, reviewable inference.
    room_panel_colors: dict[str, set[str]] = defaultdict(set)
    for room, items in raw_by_room.items():
        for item in items:
            match = re.fullmatch(r"19\.1mm--(.+)", normalize_name(item["name"]).replace("\t", ""))
            if match and match.group(1):
                room_panel_colors[room].add(match.group(1))

    grouped: dict[tuple[str, str, str], dict] = {}
    for room, items in raw_by_room.items():
        for item in items:
            name = normalize_name(item["name"]).replace("\t", "")
            remark = normalize_name(item["remark"])
            inferred = False
            if name.lower() == "edge banding" and remark:
                name = f"Edge banding--{remark}"
            elif name.lower() == "edge banding" and len(room_panel_colors[room]) == 1:
                inferred_color = next(iter(room_panel_colors[room]))
                name = f"Edge banding--{inferred_color}"
                inferred = True
            elif name.lower() == "edge banding":
                name = "Edge banding--未注明颜色"
            key = (room, name, item["spec"])
            entry = grouped.setdefault(key, {
                "room": room,
                "material_name": name,
                "spec": item["spec"],
                "quantity": 0.0,
                "quantity_sources": set(),
                "source_files": set(),
                "inferred_color": False,
                "raw_names": set(),
            })
            entry["quantity"] += item["quantity"]
            entry["quantity_sources"].add(item["quantity_source"])
            entry["source_files"].add(item["source_file"])
            entry["inferred_color"] = entry["inferred_color"] or inferred
            entry["raw_names"].add(item["name"])

    rows = []
    for entry in sorted(grouped.values(), key=lambda value: (value["room"].lower(), value["material_name"].lower(), value["spec"])):
        price_info = mapped_price(entry["material_name"], "", prices)
        entry.update(price_info)
        entry["source_files"] = sorted(entry["source_files"])
        entry["quantity_sources"] = sorted(entry["quantity_sources"])
        entry["raw_names"] = sorted(entry["raw_names"])
        entry["amount"] = round(entry["quantity"] * entry["price"], 4)
        if entry["inferred_color"]:
            entry["price_note"] += "；封边颜色按该房间唯一 Panel 颜色推定"
        rows.append(entry)

    room_summaries = []
    for room in sorted({row["room"] for row in rows}, key=str.lower):
        room_rows = [row for row in rows if row["room"] == room]
        room_summaries.append({
            "room": room,
            "amount": round(sum(row["amount"] for row in room_rows), 4),
            "quantity_rows": len(room_rows),
            "unknown_price_rows": sum(row["price_status"] == "未知价格" for row in room_rows),
            "source_files": room_files[room],
        })

    payload = {
        "order": "PP0018",
        "source_dir": str(SOURCE_DIR),
        "catalog_path": str(CATALOG_PATH.resolve()),
        "catalog_price_field": "预计采购价",
        "method_note": "每个 Traveler 文件作为一个房间单独汇总，共 8 个房间；Pickinglist数量优先，数量为空时回退Cuttinglist；材料单价来自当前库存商品资料预计采购价；未知价格按0计算并标红。",
        "rows": rows,
        "room_summaries": room_summaries,
        "grand_total": round(sum(row["amount"] for row in rows), 4),
        "unknown_materials": sorted({row["material_name"] for row in rows if row["price_status"] == "未知价格"}),
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"files": len(files), "rows": len(rows), "rooms": len(room_summaries), "grand_total": payload["grand_total"], "unknown_materials": payload["unknown_materials"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
