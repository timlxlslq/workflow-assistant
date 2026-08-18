import json
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from traveler_assistant.order_workflow import parse_order_materials, repair_material_color_table


source = Path("/Volumes/server/Optimized Orders/pp0035-2/PP0035-2 materials.xlsx")
target = Path("outputs/019ff173-07cd-75d3-a5e7-5ef8c8ec6992/PP0035-2-material-after.xlsx")
shutil.copy2(source, target)
result = repair_material_color_table(target)
try:
    workbook = load_workbook(target, data_only=False)
    sheet = workbook.active
    _, materials, edges = parse_order_materials("PP0035-2", target)
    payload = {
        "repair": result,
        "color_table": [sheet.cell(16, col).value for col in range(3, 10)],
        "formulas": {
            coordinate: sheet[coordinate].value
            for coordinate in ("D17", "E17", "D18", "E18", "D19", "E19")
        },
        "panels": [
            {"color": item.color, "quantity": item.quantity}
            for item in materials
            if item.kind == "panel"
        ],
        "edges": edges,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
finally:
    workbook.close()
