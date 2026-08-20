import os
import sqlite3
import shutil
import tempfile
import time
import unittest
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from traveler_assistant.core import Config, RuleError, parse_fittings_groups
from traveler_assistant.inventory import InventoryMappings, parse_traveler, set_ignored_mapping
from traveler_assistant.order_workflow import (
    _choose_fittings,
    _fill_usage_list,
    MaterialItem,
    generate_material_from_travelers,
    add_manual_hardware,
    find_existing_traveler,
    generate_material_from_reports,
    generate_database_order_traveler,
    generate_order_traveler,
    list_order_folders,
    parse_order_materials,
    parse_material_room_rows,
    _select_room_materials,
    preview_payload,
    preview_order,
    preview_manual_hardware,
    repair_material_color_table,
    set_ignored,
    update_order_traveler,
    update_related_orders,
)
from unittest.mock import patch
from traveler_assistant.test_data import create_local_test_source


def make_materials(path: Path, order_id: str = "PP9999", fractional: bool = False, edge: float = 12.5):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"] = "Job:", order_id
    headers = [
        "Room/section", "", "3/4 Plywood", "5/8 Plywood", "1/4 Plywood",
        "3/4 Finish Panel", "1/4 Finish Panel", "Edge Banding (m)", "Color",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(2, col).value = value
    ws.append(["Kitchen", "", 2.5 if fractional else 2, 1, 3, 4, 1, edge, "Basalto SM"])
    for col, value in enumerate(["Total Qty:", "", 2.5 if fractional else 2, 1, 3, 4, 1, edge, "Basalto SM"], 1):
        ws.cell(14, col).value = value
    ws["A15"] = "Color Table"
    ws["A16"] = "Color:"
    ws["A17"] = "Sheets (3/4):"
    ws["A18"] = "Sheets (1/4):"
    ws["A19"] = "Edge Banding (m):"
    wb.save(path)


def make_board(path: Path, factory: str, name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Page1"
    ws["A2"], ws["B2"] = "订 单 号", factory
    ws["F2"], ws["G2"] = "订单名称", name
    wb.save(path)


def make_board_material_report(
    path: Path,
    factory: str = "F100",
    name: str = "PP9999-KITCHEN",
    plywood_qty: int = 2,
    panel_qty: int = 3,
    edge_qty: float = 12.5,
):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Page1"
    ws.append(["板材清单"])
    ws.append(["订 单 号", factory, "", "", "", "订单名称", name])
    ws.append([])
    ws.append(["大板统计"])
    ws.append(["序号", "颜    色", "", "", "规 格", "数 量"])
    ws.append([1, "18mm/Finished/", "", "", "2440*1220*18", plywood_qty])
    ws.append([2, "19.1mm/Test Oak/MDF", "", "", "2745*1220*19.1", panel_qty])
    ws.append(["小板统计", "", "", "", "封边统计"])
    ws.append(["厚度", "颜色", "", "件数", "面积合计/平方", "", "封边条/米"])
    ws.append([19.1, "Test Oak", "", 6, 2.4, "", edge_qty])
    wb.save(path)


def make_fittings(path: Path, groups: list[tuple[str, float]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Page1"
    cursor = 1
    for factory, hinge_qty in groups:
        ws.cell(cursor, 1).value = "Order No."
        ws.cell(cursor, 3).value = factory
        header = cursor + 5
        ws.cell(header, 3).value = "Name"
        ws.cell(header, 5).value = "Code"
        ws.cell(header, 6).value = "Size"
        ws.cell(header, 11).value = "Quantity"
        ws.cell(header + 1, 3).value = "TestFullHinge"
        ws.cell(header + 1, 5).value = "71T950A"
        ws.cell(header + 1, 9).value = "Piece"
        ws.cell(header + 1, 11).value = hinge_qty
        cursor = header + 3
    wb.save(path)


def make_template(path: Path):
    project_template = Path(__file__).resolve().parents[1] / "resources/templates/Work Order Traveler.xlsx"
    shutil.copy2(project_template, path)


def make_product_catalog(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品类别", "*商品编号", "商品名称", "规格型号", "状态"])
    rows = [
        ["Hardware", "M2000", "Manual Handle", "Black", "启用"],
        ["Plywood", "M0002", "1/4 Finished,UV1S", "5.2*1220*2440mm", "启用"],
        ["Plywood", "M0003", "5/8 Finished,UV2S", "15*1220*2440mm", "启用"],
        ["Plywood", "M0004", "3/4 Finished,UV2S", "18*1220*2440mm", "启用"],
        ["Panel", "M0065", "Basalto SM", "19.1*1220*2745mm", "启用"],
        ["Panel", "M0066", "Basalto SM", "8*1220*2745mm", "启用"],
        ["Edge band", "M0066E", "Basalto SM Edge Banding", "22mm*1mm*225m", "启用"],
        ["Hardware", "M1001", "Hinge", "", "启用"],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def picking_layout_snapshot(sheet):
    structural_last_column = max(item.max_col for item in sheet.merged_cells.ranges)
    return {
        "dimensions": (sheet.max_row, structural_last_column),
        "merges": sorted(str(item) for item in sheet.merged_cells.ranges),
        "row_heights": [sheet.row_dimensions[row].height for row in range(1, sheet.max_row + 1)],
        "column_widths": [
            sheet.column_dimensions[get_column_letter(column)].width
            for column in range(1, structural_last_column + 1)
        ],
        "styles": [
            [
                (
                    str(sheet.cell(row, column).font),
                    str(sheet.cell(row, column).fill),
                    str(sheet.cell(row, column).border),
                    str(sheet.cell(row, column).alignment),
                    str(sheet.cell(row, column).protection),
                    sheet.cell(row, column).number_format,
                )
                for column in range(1, structural_last_column + 1)
            ]
            for row in range(1, sheet.max_row + 1)
        ],
    }


class OrderWorkflowTests(unittest.TestCase):
    def test_database_order_traveler_uses_sqlite_facts_without_source_material(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                order_root=root / "generated",
                template=root / "Work Order Traveler.xlsx",
            )
            make_template(config.template)
            config.prepare_storage()

            from traveler_assistant.order_index import OrderIndexStore

            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, source_folder, updated_at) values(?,?,?,?)",
                ("PP9999", "owned", "/server/PP9999", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, ownership_status, optimized, updated_at) values(?,?,?,?,?,?)",
                ("F9999", "PP9999", "PP9999-KITCHEN", "已确认", 1, "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("PP9999", "plywood", "", "18", 4, "pcs", "aihouse", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("PP9999", "panel", "Test Oak", "19.1", 3, "pcs", "aihouse", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, edge, source_type, updated_at) values(?,?,?,?,?,?,?,?,?)",
                ("PP9999", "edge", "Test Oak", "", 12.5, "m", "Test Oak", "aihouse", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into hardware_items(order_id, factory_order, scope, product_code, name, spec, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?,?,?)",
                ("PP9999", "F9999", "factory_order", "71T950A", "Hinge", "Full", 6, "pcs/个", "aicnc", "2026-08-15T10:00:00"),
            )
            store.connection.commit()
            store.close()

            output = generate_database_order_traveler(config, "PP9999")
            self.assertTrue(output.is_file())
            self.assertNotIn("server", str(output))
            workbook = load_workbook(output, data_only=False, read_only=True)
            self.assertEqual(workbook.sheetnames[:3], ["WorkOrderTraveler", "Usage List", "Picking List"])
            self.assertEqual(workbook["WorkOrderTraveler"]["B5"].value, "PP9999")
            self.assertEqual(workbook["Usage List"]["A3"].value, "数据库汇总")
            self.assertEqual(workbook["Usage List"]["C3"].value, 4)
            self.assertEqual(workbook["Usage List"]["F3"].value, 3)
            self.assertEqual(workbook["Usage List"]["H3"].value, 12.5)
            self.assertIn("PP9999-KITCHEN", " ".join(str(cell.value or "") for row in workbook["Picking List"].iter_rows() for cell in row))
            self.assertIn("Hinge", " ".join(str(cell.value or "") for row in workbook["Picking List"].iter_rows() for cell in row))
            purchase = workbook["Purchase List"]
            self.assertIsNone(purchase["A4"].value)
            self.assertEqual(purchase["D5"].value, "PP9999-KITCHEN")
            purchase_values = [
                cell.value
                for row in purchase.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            ]
            self.assertIn("18mm--Plywood", purchase_values)
            self.assertIn("19.1mm--Test Oak", purchase_values)
            self.assertIn("Edge banding--Test Oak", purchase_values)
            self.assertIn("Hinge", purchase_values)
            self.assertNotIn("TB18", purchase_values)

    def test_legacy_traveler_gets_usage_list_and_material_from_picking_list(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order_root = root / "orders"
            local_order = order_root / "CS001"
            local_order.mkdir(parents=True)
            traveler = local_order / "Work Order Traveler(CS001).xlsx"
            make_template(traveler)
            workbook = load_workbook(traveler, data_only=False)
            workbook.remove(workbook["Usage List"])
            workbook["Picking List"].title = "Pickinglist"
            work_order = workbook["WorkOrderTraveler"]
            work_order["B5"] = "CS001"
            picking = workbook["Pickinglist"]
            picking["D3"] = "CS001-KITCHEN"
            picking["C6"] = "19.1mm--Test Oak"
            picking["G6"] = 3
            picking["C7"] = "Edge banding--Test Oak"
            picking["G7"] = 12.5
            workbook.save(traveler)

            target = root / "server" / "cs001"
            target.mkdir(parents=True)
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root / "source",
                order_root=order_root,
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )

            result = generate_material_from_travelers(
                config, target, "CS001", confirm_write=True
            )

            self.assertEqual(result["traveler_count"], 1)
            self.assertEqual(result["legacy_traveler_updates"], 1)
            self.assertTrue(result["traveler_backups"])
            material = target / "CS001 materials.xlsx"
            self.assertTrue(material.is_file())
            _, materials, edges = parse_order_materials("CS001", material)
            self.assertIn(
                ("panel", 19.1, "Test Oak", 3.0),
                [(item.kind, item.thickness, item.color, item.quantity) for item in materials],
            )
            self.assertEqual(edges, {"Test Oak": 13})
            updated = load_workbook(traveler, data_only=False, read_only=True)
            self.assertEqual(updated.sheetnames[:3], ["WorkOrderTraveler", "Usage List", "Picking List"])
            self.assertEqual(parse_traveler(traveler).documents["CS001"][0].name, "19.1mm--Test Oak")

    def test_missing_material_can_be_generated_from_report_summary(self):
        with tempfile.TemporaryDirectory() as temp:
            order = Path(temp) / "PP9999"
            make_board_material_report(order / "Kitchen" / "Report" / "pp-板材清单.xlsx")

            created = generate_material_from_reports(order, "PP9999")

            self.assertEqual(created, (order / "PP9999 materials.xlsx").resolve())
            _, materials, edges = parse_order_materials("PP9999", created)
            self.assertEqual(
                {(item.kind, item.thickness, item.color): item.quantity for item in materials if item.quantity},
                {("plywood", 18.0, ""): 2, ("panel", 19.1, "Test Oak"): 3},
            )
            self.assertEqual(edges, {"Test Oak": 13})
            generated = load_workbook(created, data_only=False).active
            self.assertEqual(generated["A14"].value, "Total Qty:")
            self.assertEqual(generated["C16"].value, "Test Oak")
            self.assertEqual(generated["C14"].value, "=SUM(C3:C13)")
            self.assertIn("SUMIF($I$3:$I$13,C$16,$F$3:$F$13)", generated["C17"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$G$3:$G$13)", generated["C18"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$H$3:$H$13)", generated["C19"].value)
            template = load_workbook(
                Path(__file__).resolve().parents[1] / "resources/templates/Order Materials.xlsx",
                data_only=False,
            ).active
            self.assertEqual(sorted(map(str, generated.merged_cells.ranges)), sorted(map(str, template.merged_cells.ranges)))
            self.assertEqual(
                [generated.column_dimensions[get_column_letter(column)].width for column in range(1, 10)],
                [template.column_dimensions[get_column_letter(column)].width for column in range(1, 10)],
            )
            for coordinate in ("A1", "B1", "A2", "C3", "A14", "C16", "A20"):
                self.assertEqual(generated[coordinate].style_id, template[coordinate].style_id)

    def test_generated_material_color_table_aggregates_repeated_report_colors(self):
        with tempfile.TemporaryDirectory() as temp:
            order = Path(temp) / "PP9999"
            make_board_material_report(
                order / "Kitchen" / "Report" / "pp-板材清单.xlsx",
                panel_qty=3,
                edge_qty=12.5,
            )
            make_board_material_report(
                order / "Laundry" / "Report" / "pp-板材清单.xlsx",
                factory="F200",
                name="PP9999-LAUNDRY",
                plywood_qty=1,
                panel_qty=4,
                edge_qty=7.25,
            )

            created = generate_material_from_reports(order, "PP9999")
            generated = load_workbook(created, data_only=False).active

            self.assertEqual(generated["C14"].value, "=SUM(C3:C13)")
            self.assertIn("SUM(F3:F13)", generated["F14"].value)
            self.assertIn("SUM(H3:H13)", generated["H14"].value)
            self.assertEqual(generated["C16"].value, "Test Oak")
            self.assertIn("SUMIF($I$3:$I$13,C$16,$F$3:$F$13)", generated["C17"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$G$3:$G$13)", generated["C18"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$H$3:$H$13)", generated["C19"].value)
            _, materials, edges = parse_order_materials("PP9999", created)
            self.assertEqual(
                next(item.quantity for item in materials if item.kind == "panel" and item.color == "Test Oak"),
                7,
            )
            self.assertEqual(edges, {"Test Oak": 20})

    def test_complex_report_generation_requests_manual_material(self):
        with tempfile.TemporaryDirectory() as temp:
            order = Path(temp) / "PP9999"
            report = order / "Kitchen" / "Report" / "pp-板材清单.xlsx"
            make_board_material_report(report)
            wb = load_workbook(report)
            wb.active["B6"] = "12mm/Finished/"
            wb.save(report)

            with self.assertRaises(RuleError) as raised:
                generate_material_from_reports(order, "PP9999")
            self.assertEqual(raised.exception.code, "material_generation_failed")
            self.assertIn("手动生成 material", str(raised.exception))

    def test_local_test_source_is_repeatable_and_matches_server_layout(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp) / "test-source"
            first = create_local_test_source(root)
            sentinel = root / "keep-my-file.txt"
            sentinel.write_text("keep", encoding="utf-8")
            second = create_local_test_source(root)

            self.assertEqual(first["orders"], ["PP9001", "CS900"])
            self.assertEqual(second["root"], str(root.resolve()))
            self.assertEqual(sentinel.read_text(encoding="utf-8"), "keep")
            owned_root = root / "Optimized Orders"
            cut_root = root / "CUT TO SIZE"
            self.assertEqual(
                [item["order_id"] for item in list_order_folders(Config(source_root=owned_root))],
                ["PP9001"],
            )
            self.assertEqual(
                [item["order_id"] for item in list_order_folders(Config(source_root=cut_root))],
                ["CS900"],
            )
            owned = preview_order(Config(source_root=owned_root, state_dir=root / "state"), owned_root / "PP9001")
            cut = preview_order(Config(source_root=cut_root, state_dir=root / "state"), cut_root / "CS900")
            self.assertEqual([item.factory_order for item in owned.factories], ["F9001", "F9002"])
            self.assertEqual(sum(len(item.fittings) for item in owned.factories), 2)
            self.assertEqual(cut.factories, [])
            self.assertTrue((root / "test-data-manifest.json").is_file())

    def test_usage_list_expands_and_rewrites_summary_formulas(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "materials.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"], sheet["B1"] = "Job:", "PP9999"
            headers = [
                "Room/section", "", "3/4 Plywood", "5/8 Plywood", "1/4 Plywood",
                "3/4 Finish Panel", "1/4 Finish Panel", "Edge Banding (m)", "Color",
            ]
            for col, value in enumerate(headers, 1):
                sheet.cell(2, col).value = value
            for row in range(3, 16):
                sheet.cell(row, 1).value = "Kitchen" if row == 3 else None
                sheet.cell(row, 6).value = 1
                sheet.cell(row, 8).value = 2.5
                sheet.cell(row, 9).value = "Color A" if row % 2 else "Color B"
            sheet["A16"] = "Total Qty:"
            workbook.save(source)

            template = root / "template.xlsx"
            make_template(template)
            traveler = load_workbook(template)
            _fill_usage_list(source, traveler, "PP9999")
            usage = traveler["Usage List"]

            self.assertEqual(usage["A3"].value, "Kitchen")
            self.assertIsNone(usage["A4"].value)
            self.assertEqual(usage["I15"].value, "Color A")
            self.assertEqual(usage["C16"].value, "=SUM(C3:C15)")
            self.assertEqual(usage["F16"].value, "check color table")
            self.assertEqual(usage["C18"].value, "Color A")
            self.assertEqual(usage["D18"].value, "Color B")
            self.assertIn("$F$3:$F$15", usage["C19"].value)
            self.assertIn("$H$3:$H$15", usage["C21"].value)
            formulas = [
                cell.value
                for row in usage.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            self.assertFalse(any("UNIQUE" in formula or "FILTER" in formula for formula in formulas))
            self.assertNotIn("A3:A4", {str(item) for item in usage.merged_cells.ranges})

    def test_usage_list_normalizes_alias_color_before_color_table_formulas(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "materials.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"], sheet["B1"] = "Job:", "PP0035-2"
            headers = [
                "Room/section", "", "3/4 Plywood", "5/8 Plywood", "1/4 Plywood",
                "3/4 Finish Panel", "1/4 Finish Panel", "Edge Banding (m)", "Color",
            ]
            for col, value in enumerate(headers, 1):
                sheet.cell(2, col).value = value
            sheet.cell(3, 1).value = "office"
            sheet.cell(3, 6).value = 1
            sheet.cell(3, 8).value = 36.72
            sheet.cell(3, 9).value = "Khaki"
            sheet.cell(4, 1).value = "Openshelf"
            sheet.cell(4, 6).value = 3
            sheet.cell(4, 8).value = 18.5
            sheet.cell(4, 9).value = "Ivory Oak"
            sheet["A5"] = "Total Qty:"
            workbook.save(source)

            template = root / "template.xlsx"
            make_template(template)
            traveler = load_workbook(template)
            _fill_usage_list(source, traveler, "PP0035-2")
            usage = traveler["Usage List"]

            self.assertEqual(usage["I3"].value, "Penelope FA44")
            self.assertEqual(usage["I4"].value, "Ivory Oak")
            self.assertEqual(usage["C16"].value, "Penelope FA44")
            self.assertEqual(usage["D16"].value, "Ivory Oak")
            self.assertIn("C$16", usage["C17"].value)
            self.assertIn("$I$3:$I$13", usage["C17"].value)

    def test_order_folders_are_sorted_by_modified_time_descending(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            oldest = root / "PP9999"
            newest = root / "PP0001"
            middle = root / "PP5000-2"
            ignored = root / "notes"
            for folder in (oldest, newest, middle, ignored):
                folder.mkdir()

            os.utime(oldest, (100, 100))
            os.utime(middle, (200, 200))
            os.utime(newest, (300, 300))

            orders = list_order_folders(Config(source_root=root))

            self.assertEqual(
                [order["order_id"] for order in orders],
                ["PP0001", "PP5000-2", "PP9999"],
            )

    def test_cut_to_size_requires_materials_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "CS004"
            order.mkdir()
            config = Config(source_root=root, state_dir=root / "state")
            with self.assertRaises(RuleError) as raised:
                preview_order(config, order)
            self.assertEqual(raised.exception.code, "missing_materials")
            self.assertIn("material", str(raised.exception))

    def test_cut_to_size_generates_materials_only_traveler(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "CS004"
            order.mkdir()
            make_materials(order / "material.xlsx", order_id="CS004")
            # Even if an old source folder happens to contain a fitting file,
            # cut-to-size orders must not read or write it.
            make_fittings(order / "Fittingslist-old.xlsx", [("F999", 7)])
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )

            preview = preview_order(config, order)
            self.assertEqual(preview.order_id, "CS004")
            self.assertEqual(preview.factories, [])
            self.assertTrue(any("不读取或写入五金" in item for item in preview.warnings))

            # openpyxl normalizes merged-cell style records on every save, so
            # compare with an otherwise untouched workbook after the same round trip.
            normalized_template = root / "normalized-template.xlsx"
            untouched = load_workbook(template, data_only=False)
            untouched.save(normalized_template)
            template_picking = load_workbook(normalized_template, data_only=False)["Picking List"]
            template_layout = picking_layout_snapshot(template_picking)
            output = generate_order_traveler(config, preview)
            with zipfile.ZipFile(output) as archive:
                usage_xml = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
            self.assertNotIn("UNIQUE", usage_xml)
            self.assertNotIn("FILTER", usage_xml)
            self.assertIn("SUMIF", usage_xml)
            wb = load_workbook(output, data_only=False)
            self.assertEqual(wb["WorkOrderTraveler"]["B5"].value, "CS004")
            self.assertEqual(wb["Usage List"]["B1"].value, "CS004")
            purchase_values = [
                cell.value
                for row in wb["Purchase List"].iter_rows()
                for cell in row
                if cell.value not in (None, "")
            ]
            self.assertIsNone(wb["Purchase List"]["A4"].value)
            self.assertNotIn("TB18", purchase_values)
            self.assertNotIn("Poket Door Pantry Hinge", purchase_values)
            self.assertNotIn("Touch Open", purchase_values)
            picking = wb["Picking List"]
            self.assertEqual(picking["A1"].value, "Picking List领料单")
            self.assertEqual(picking_layout_snapshot(picking), template_layout)
            self.assertIsNone(picking["D3"].value)
            for row in range(1, picking.max_row + 1):
                if isinstance(picking.cell(row, 1).value, (int, float)):
                    self.assertTrue(
                        all(
                            picking.cell(row, column).value is None
                            or not str(picking.cell(row, column).value).strip()
                            for column in range(2, picking.max_column + 1)
                        ),
                        f"Picking List 第 {row} 行仍有业务数据",
                    )
            self.assertNotIn(
                "板材和封边请查看 Usage List",
                [cell.value for row in picking.iter_rows() for cell in row],
            )
            traveler = parse_traveler(output)
            self.assertEqual(set(traveler.documents), {"CS004"})
            self.assertTrue(traveler.documents["CS004"])

    def test_source_component_code_is_not_written_as_sku(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report = order / "Report"
            report.mkdir(parents=True)
            make_materials(order / "material.xlsx")
            make_board(report / "板材清单.xlsx", "F100", "PP9999-KITCHEN")
            make_fittings(report / "Fittingslist.xlsx", [("F100", 4)])
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )
            output = generate_order_traveler(config, preview_order(config, order))
            picking = load_workbook(output, data_only=False)["Picking List"]
            hinge_rows = [
                row for row in range(1, picking.max_row + 1)
                if picking.cell(row, 3).value == "Hinge"
            ]
            self.assertTrue(hinge_rows)
            self.assertEqual(hinge_rows, [6])
            self.assertTrue(all(picking.cell(row, 2).value is None for row in hinge_rows))

    def test_opt_out_hardware_omits_report_fittings_from_traveler(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report = order / "Report"
            report.mkdir(parents=True)
            make_materials(order / "material.xlsx")
            make_board(report / "板材清单.xlsx", "F100", "PP9999-KITCHEN")
            make_fittings(report / "Fittingslist.xlsx", [("F100", 4)])
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )

            preview = preview_order(config, order, include_hardware=False)
            self.assertFalse(preview.include_hardware)
            self.assertEqual(preview.factories[0].fittings, [])
            self.assertTrue(any("不包含五金" in warning for warning in preview.warnings))
            output = generate_order_traveler(config, preview)
            picking = load_workbook(output, data_only=False)["Picking List"]
            self.assertFalse(any(
                picking.cell(row, 3).value == "Hinge"
                for row in range(1, picking.max_row + 1)
            ))

    def test_invalid_empty_dimension_returns_one_business_error(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "source.xlsx"
            broken = root / "Fittingslist-broken.xlsx"
            Workbook().save(source)
            with zipfile.ZipFile(source, "r") as src, zipfile.ZipFile(broken, "w") as dst:
                for item in src.infolist():
                    payload = src.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml":
                        payload = payload.replace(b'<dimension ref="A1:A1"/>', b'<dimension ref="A1:?0"/>')
                    dst.writestr(item, payload)
            with self.assertRaises(RuleError) as raised:
                parse_fittings_groups(broken)
            self.assertEqual(raised.exception.code, "fittings_schema")

    def test_single_color_materials_and_integer_validation(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / "PP9999 materials.xlsx"
            make_materials(path)
            _, materials, edge = parse_order_materials("PP9999", path)
            self.assertIn(("panel", 19.1, "Basalto SM", 4), [(x.kind, x.thickness, x.color, x.quantity) for x in materials])
            self.assertEqual(edge, {"Basalto SM": 12.5})
            make_materials(path, fractional=True)
            with self.assertRaises(RuleError) as raised:
                parse_order_materials("PP9999", path)
            self.assertEqual(raised.exception.code, "fractional_material")

    def test_single_color_materials_keep_edge_when_panel_is_zero(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP9999 materials.xlsx"
            make_materials(path, edge=12.5)
            workbook = load_workbook(path)
            sheet = workbook.active
            for coordinate in ("F3", "G3", "F14", "G14"):
                sheet[coordinate] = 0
            sheet["C16"] = None
            sheet["C17"] = None
            sheet["C18"] = None
            sheet["C19"] = None
            workbook.save(path)

            _, materials, edges = parse_order_materials("PP9999", path)

            self.assertFalse([item for item in materials if item.kind == "panel"])
            self.assertEqual(edges, {"Basalto SM": 12.5})

    def test_repairs_empty_single_color_table_from_detail_rows_without_changing_details(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0072 materials.xlsx"
            make_materials(path, order_id="PP0072", edge=319.64)
            workbook = load_workbook(path, data_only=False)
            sheet = workbook.active
            detail_snapshot = [
                [sheet.cell(row, column).value for column in range(1, 10)]
                for row in range(3, 14)
            ]
            total_snapshot = [sheet.cell(14, column).value for column in range(1, 10)]
            for coordinate in (
                "C16", "D16", "E16", "C17", "D17", "E17",
                "C18", "D18", "E18", "C19", "D19", "E19",
            ):
                sheet[coordinate] = None
            workbook.save(path)

            result = repair_material_color_table(path)
            self.assertTrue(result["corrected"])
            self.assertEqual(result["colors"], ["Basalto SM"])
            self.assertEqual(result["panel_count"], 5)
            self.assertAlmostEqual(result["edge_banding"], 319.64)

            repaired = load_workbook(path, data_only=False).active
            self.assertEqual(repaired["C16"].value, "Basalto SM")
            self.assertIn("SUMIF($I$3:$I$13,C$16,$F$3:$F$13)", repaired["C17"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$G$3:$G$13)", repaired["C18"].value)
            self.assertIn("SUMIF($I$3:$I$13,C$16,$H$3:$H$13)", repaired["C19"].value)
            self.assertEqual(
                [repaired.cell(row, column).value for row in range(3, 14) for column in range(1, 10)],
                [value for row in detail_snapshot for value in row],
            )
            self.assertEqual(
                [repaired.cell(14, column).value for column in range(1, 10)],
                total_snapshot,
            )

            _, materials, edges = parse_order_materials("PP0072", path)
            self.assertEqual(
                [(item.kind, item.color, item.quantity) for item in materials if item.kind == "panel"],
                [("panel", "Basalto SM", 4.0), ("panel", "Basalto SM", 1.0)],
            )
            self.assertEqual(edges, {"Basalto SM": 319.64})

    def test_repairs_missing_color_table_colors_and_rebuilds_all_summary_formulas(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0035-2 materials.xlsx"
            make_materials(path, order_id="PP0035-2")
            workbook = load_workbook(path, data_only=False)
            sheet = workbook.active
            sheet["F3"], sheet["G3"], sheet["H3"], sheet["I3"] = 3, 0, 18.5, "Ivory Oak"
            sheet["F4"], sheet["H4"], sheet["I4"] = 1, 36.72, "Khaki"
            sheet["F5"], sheet["H5"], sheet["I5"] = 5, 87.52, "Frappe 3"
            sheet["F14"], sheet["G14"], sheet["H14"] = 9, 0, 142.74
            sheet["C16"] = "Ivory Oak"
            sheet["D16"], sheet["E16"] = None, None
            workbook.save(path)

            result = repair_material_color_table(path)
            self.assertTrue(result["corrected"])
            self.assertEqual(result["colors"], ["Khaki", "Frappe 3"])
            self.assertEqual(result["panel_count"], 6)
            self.assertAlmostEqual(result["edge_banding"], 124.24)
            self.assertIn("补充 2 种 Panel 颜色", result["message"])
            self.assertIn("6 张 Panel", result["message"])
            self.assertIn("124.24 m", result["message"])

            repaired = load_workbook(path, data_only=False).active
            self.assertEqual(repaired["C16"].value, "Ivory Oak")
            self.assertEqual(repaired["D16"].value, "Khaki")
            self.assertEqual(repaired["E16"].value, "Frappe 3")
            self.assertIn("D$16", repaired["D17"].value)
            self.assertIn("E$16", repaired["E18"].value)
            self.assertIn("$I$3:$I$13", repaired["E19"].value)

            _, materials, edges = parse_order_materials("PP0035-2", path)
            panel_quantities = {
                item.color: item.quantity
                for item in materials
                if item.kind == "panel"
            }
            self.assertEqual(panel_quantities, {"Ivory Oak": 3, "Penelope FA44": 1, "Frappe 3": 5})
            self.assertEqual(edges, {"Ivory Oak": 18.5, "Penelope FA44": 36.72, "Frappe 3": 87.52})

    def test_preview_includes_material_color_table_repair_warning(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "CS004"
            order.mkdir()
            path = order / "CS004 materials.xlsx"
            make_materials(path, order_id="CS004")
            workbook = load_workbook(path)
            sheet = workbook.active
            sheet["F3"], sheet["H3"], sheet["I3"] = 3, 18.5, "Ivory Oak"
            sheet["F4"], sheet["H4"], sheet["I4"] = 1, 36.72, "Khaki"
            sheet["F5"], sheet["H5"], sheet["I5"] = 5, 87.52, "Frappe 3"
            sheet["F14"], sheet["G14"], sheet["H14"] = 9, 1, 142.74
            sheet["D16"], sheet["E16"] = None, None
            workbook.save(path)

            preview = preview_order(Config(source_root=root, state_dir=root / "state"), order)
            self.assertTrue(any("自动修正 Color Table" in warning for warning in preview.warnings))
            self.assertEqual(
                {item.color: item.quantity for item in preview.materials if item.kind == "panel"},
                {"Ivory Oak": 1, "Penelope FA44": 1, "Frappe 3": 5},
            )

    def test_repair_aggregates_repeated_detail_color_rows(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0035-2 materials.xlsx"
            make_materials(path, order_id="PP0035-2")
            workbook = load_workbook(path)
            sheet = workbook.active
            sheet["F3"], sheet["G3"], sheet["H3"], sheet["I3"] = 3, 0, 18.5, "Ivory Oak"
            sheet["F4"], sheet["H4"], sheet["I4"] = 1, 36.72, "Khaki"
            sheet["F5"], sheet["H5"], sheet["I5"] = 2, 41.28, "Khaki"
            sheet["F6"], sheet["H6"], sheet["I6"] = 5, 87.52, "Frappe 3"
            sheet["F14"], sheet["G14"], sheet["H14"] = 11, 0, 184.02
            sheet["C16"], sheet["D16"], sheet["E16"] = "Ivory Oak", None, None
            workbook.save(path)

            result = repair_material_color_table(path)
            self.assertEqual(result["panel_count"], 8)
            self.assertAlmostEqual(result["edge_banding"], 165.52)
            _, materials, edges = parse_order_materials("PP0035-2", path)
            self.assertEqual(
                {item.color: item.quantity for item in materials if item.kind == "panel"},
                {"Ivory Oak": 3, "Penelope FA44": 3, "Frappe 3": 5},
            )
            self.assertEqual(
                edges,
                {"Ivory Oak": 18.5, "Penelope FA44": 78.0, "Frappe 3": 87.52},
            )

    def test_existing_complete_color_table_mismatch_requires_manual_handling(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0035-2 materials.xlsx"
            make_materials(path, order_id="PP0035-2")
            workbook = load_workbook(path)
            sheet = workbook.active
            sheet["C16"] = "Basalto SM"
            # A complete table is still the parsing source, but a mismatch
            # with Total Qty must stop automatic material writing.
            sheet["C17"], sheet["C18"], sheet["C19"] = 2, 0, 9.25
            workbook.save(path)

            result = repair_material_color_table(path)
            self.assertFalse(result["corrected"])
            with self.assertRaises(RuleError) as raised:
                parse_order_materials("PP0035-2", path)
            self.assertEqual(raised.exception.code, "material_summary_mismatch")

    def test_eight_color_table_is_read_without_seven_color_limit_error(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0035 materials.xlsx"
            make_materials(path, order_id="PP0035")
            workbook = load_workbook(path)
            sheet = workbook.active
            colors = [
                "Cashmere SM", "Cinnamon Triba", "Frappe 3", "Ivory Oak",
                "Milk Oak", "Penelope FA44", "Raven Oak", "Woodline 3",
            ]
            for row, color in enumerate(colors, start=3):
                sheet.cell(row, 6).value = 1
                sheet.cell(row, 8).value = 10
                sheet.cell(row, 9).value = color
                sheet.cell(row, 1).value = "Traveler 汇总"
            for column, color in enumerate(colors, start=3):
                sheet.cell(16, column).value = color
                sheet.cell(17, column).value = 1
                sheet.cell(18, column).value = 0
                sheet.cell(19, column).value = 10
            sheet["F14"], sheet["G14"], sheet["H14"] = 8, 0, 80
            workbook.save(path)

            result = repair_material_color_table(path)
            self.assertFalse(result["corrected"])
            _, materials, edges = parse_order_materials("PP0035", path)
            self.assertEqual(
                {item.color for item in materials if item.kind == "panel"},
                set(colors),
            )
            self.assertEqual(set(edges), set(colors))

    def test_integer_display_format_uses_the_total_qty_values_excel_shows(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0068 materials.xlsx"
            make_materials(path)
            wb = load_workbook(path)
            ws = wb.active
            for coordinate, value in {
                "C14": 10.25,
                "D14": 0.75,
                "E14": 3.25,
                "F14": 6.25,
                "H14": 319.64,
            }.items():
                ws[coordinate] = value
                ws[coordinate].number_format = "0;\\-0;;@"
            wb.save(path)
            _, materials, edges = parse_order_materials("PP0068", path)
            quantities = {(item.kind, item.thickness): item.quantity for item in materials}
            self.assertEqual(quantities[("plywood", 18.0)], 10)
            self.assertEqual(quantities[("plywood", 14.5)], 1)
            self.assertEqual(quantities[("plywood", 5.4)], 3)
            self.assertEqual(quantities[("panel", 19.1)], 6)
            self.assertEqual(edges, {"Basalto SM": 320})

    def test_material_detail_quantity_requires_color(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "CS004 materials.xlsx"
            make_materials(path, order_id="CS004")
            workbook = load_workbook(path)
            workbook.active["I3"] = None
            workbook.save(path)

            with self.assertRaises(RuleError) as raised:
                parse_order_materials("CS004", path)

            self.assertEqual(raised.exception.code, "material_color_required")
            self.assertIn("Kitchen", str(raised.exception))

    def test_total_qty_and_color_table_must_match_before_material_write(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "CS004 materials.xlsx"
            make_materials(path, order_id="CS004")
            workbook = load_workbook(path)
            sheet = workbook.active
            sheet["C16"] = "Basalto SM"
            sheet["C17"], sheet["C18"], sheet["C19"] = 4, 1, 12.5
            workbook.save(path)

            _, materials, edges = parse_order_materials("CS004", path)
            self.assertEqual(
                {(item.kind, item.color, item.quantity) for item in materials if item.kind == "panel"},
                {("panel", "Basalto SM", 4.0), ("panel", "Basalto SM", 1.0)},
            )
            self.assertEqual(edges, {"Basalto SM": 12.5})

            workbook = load_workbook(path)
            workbook.active["F14"] = 5
            workbook.save(path)
            with self.assertRaises(RuleError) as raised:
                parse_order_materials("CS004", path)
            self.assertEqual(raised.exception.code, "material_summary_mismatch")

    def test_formula_without_cached_values_uses_display_values_for_totals_and_color_table(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0072 materials.xlsx"
            make_materials(path, order_id="PP0072", fractional=True, edge=319.64)
            wb = load_workbook(path)
            ws = wb.active
            ws["C3"], ws["D3"], ws["E3"] = 21.5, 4.5, 9
            for column in "CDEFGH":
                ws[f"{column}14"] = f"=SUM({column}3:{column}13)"
                ws[f"{column}14"].number_format = r"0;\-0;;@"
            ws["C16"] = "Basalto SM"
            ws["C17"] = "=SUMIF($I$3:$I$13,C$16,$F$3:$F$13)"
            ws["C18"] = "=SUMIF($I$3:$I$13,C$16,$G$3:$G$13)"
            ws["C19"] = "=SUMIF($I$3:$I$13,C$16,$H$3:$H$13)"
            for coordinate in ("C17", "C18", "C19"):
                ws[coordinate].number_format = r"0;\-0;;@"
            wb.save(path)

            _, materials, edges = parse_order_materials("PP0072", path)

            quantities = {(item.kind, item.thickness): item.quantity for item in materials}
            self.assertEqual(quantities[("plywood", 18.0)], 22)
            self.assertEqual(quantities[("plywood", 14.5)], 5)
            self.assertEqual(quantities[("plywood", 5.4)], 9)
            self.assertEqual(quantities[("panel", 19.1)], 4)
            self.assertEqual(quantities[("panel", 8.0)], 1)
            self.assertEqual(edges, {"Basalto SM": 320})

    def test_integer_display_format_is_also_used_for_room_rows(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "PP0068 materials.xlsx"
            make_materials(path)
            wb = load_workbook(path)
            ws = wb.active
            # The source formula result is fractional, but the workbook's
            # integer display format is what the operator sees and what the
            # summary parser already uses.
            ws["F3"] = 6.25
            ws["F3"].number_format = "0;\\-0;;@"
            wb.save(path)
            rows = parse_material_room_rows(path)
            panel = next(item for item in rows[0][1] if item.kind == "panel")
            self.assertEqual(panel.quantity, 6)

    def test_room_section_factory_name_extracts_exact_order_and_rejects_ambiguous_rows(self):
        rows = [
            ("PP0035-OFFICE", [MaterialItem("panel", 19.1, "Frappe 3", 4)], {}),
            ("PP0035-2-MASTER", [MaterialItem("panel", 19.1, "Ivory Oak", 6)], {}),
        ]
        materials, edges, warnings = _select_room_materials(
            rows,
            "PP0035-2",
            {},
            known_order_ids={"PP0035", "PP0035-2"},
        )
        self.assertEqual([(item.color, item.quantity) for item in materials], [("Ivory Oak", 6)])
        self.assertEqual(edges, {})
        self.assertEqual(warnings, [])

        with self.assertRaisesRegex(RuleError, "包含订单号的工厂单名称"):
            _select_room_materials(
                [("MASTER", [MaterialItem("panel", 19.1, "Ivory Oak", 6)], {})],
                "PP0035-2",
                {},
                known_order_ids={"PP0035", "PP0035-2"},
            )

    def test_related_update_reports_the_specific_order_error(self):
        group = {
            "errors": [{
                "order_id": "PP0067",
                "code": "missing_materials",
                "message": "PP0067 根目录找不到文件名包含 material 的 Excel",
            }],
            "orders": [],
        }
        with patch("traveler_assistant.order_workflow.preview_related_orders", return_value=group):
            with self.assertRaises(RuleError) as raised:
                update_related_orders(Config(), Path("/orders/PP0067"))
        self.assertEqual(raised.exception.code, "related_orders_need_review")
        self.assertIn("PP0067", str(raised.exception))
        self.assertIn("找不到文件名包含 material", str(raised.exception))

    def test_related_update_response_keeps_all_orders_and_factories(self):
        group = {
            "errors": [],
            "order_ids": ["PP0035", "PP0035-2"],
            "orders": [
                {"order_id": "PP0035", "existing_traveler": "/tmp/PP0035.xlsx"},
                {"order_id": "PP0035-2", "existing_traveler": "/tmp/PP0035-2.xlsx"},
            ],
            "factories": [
                {"factory_order": "F2608050220", "order_name": "PP0035-OFFICE"},
                {"factory_order": "F2607290217", "order_name": "PP0035-2-OPENSHELF"},
                {"factory_order": "F2608050221", "order_name": "PP0035-2-MASTER"},
            ],
        }
        with patch("traveler_assistant.order_workflow.preview_related_orders", return_value=group), \
                patch("traveler_assistant.order_workflow.preview_order", side_effect=[object(), object()]), \
                patch(
                    "traveler_assistant.order_workflow.update_order_traveler",
                    side_effect=[(Path("/tmp/PP0035.xlsx"), Path("/tmp/PP0035.bak")),
                                 (Path("/tmp/PP0035-2.xlsx"), Path("/tmp/PP0035-2.bak"))],
                ):
            result = update_related_orders(Config(), Path("/orders/PP0035-2"))

        self.assertEqual(result["order_ids"], ["PP0035", "PP0035-2"])
        self.assertEqual(len(result["orders"]), 2)
        self.assertEqual(len(result["factories"]), 3)
        self.assertEqual(len(result["updated_orders"]), 2)

    def test_duplicate_fittings_use_newest_and_tied_conflict_stops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            older = root / "A" / "Fittingslist-old.xlsx"
            newer = root / "B" / "Fittingslist-new.xlsx"
            older.parent.mkdir()
            newer.parent.mkdir()
            make_fittings(older, [("F999", 2)])
            make_fittings(newer, [("F999", 5)])
            now = time.time()
            os.utime(older, (now - 30, now - 30))
            os.utime(newer, (now, now))
            selected, warnings = _choose_fittings(root)
            self.assertEqual(selected["F999"][0].quantity, 5)
            self.assertTrue(any("修改时间最新" in warning for warning in warnings))
            os.utime(older, (now, now))
            with self.assertRaises(RuleError) as raised:
                _choose_fittings(root)
            self.assertEqual(raised.exception.code, "fittings_timestamp_tie")

    def test_empty_malformed_fittings_is_skipped_and_traveler_can_generate(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report = order / "Room" / "Report"
            report.mkdir(parents=True)
            materials_path = order / "PP9999 materials.xlsx"
            make_materials(materials_path)
            # Room details may contain cut/layout fractions; order-level
            # plywood must still come from Total Qty and not these rows.
            workbook = load_workbook(materials_path)
            sheet = workbook.active
            for coordinate, value in {"C3": 2.5, "D3": 0.75, "E3": 3.25, "F3": 6.25}.items():
                sheet[coordinate] = value
            workbook.save(materials_path)
            make_board(report / "板材清单.xlsx", "F100", "PP9999-CLOSET")
            empty = Workbook()
            empty.active["A1"] = "No fittings were used"
            empty.save(report / "Fittingslist-empty.xlsx")
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )

            preview = preview_order(config, order)
            self.assertEqual([factory.factory_order for factory in preview.factories], ["F100"])
            self.assertEqual(preview.factories[0].fittings, [])
            self.assertTrue(any("没有内容" in warning for warning in preview.warnings))

            output = generate_order_traveler(config, preview)
            wb = load_workbook(output, data_only=False)
            picking = wb["Picking List"]
            self.assertIn(
                "PP9999-CLOSET",
                [picking.cell(row, 4).value for row in range(1, picking.max_row + 1)],
            )
            self.assertEqual(
                [
                    picking.cell(row, 7).value
                    for row in range(1, picking.max_row + 1)
                    if isinstance(picking.cell(row, 1).value, int)
                    and 3 <= row <= 9
                ],
                [None, None, None, None],
            )

    def test_global_ignore_and_generate_one_order_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report_a = order / "A" / "Report"
            report_b = order / "B" / "Report"
            report_a.mkdir(parents=True)
            report_b.mkdir(parents=True)
            make_materials(order / "anything materials final.xlsx")
            make_board(report_a / "板材清单-a.xlsx", "F100", "PP9999-KITCHEN")
            make_board(report_b / "板材清单-b.xlsx", "F200", "PP9999-LAUNDRY")
            make_fittings(report_a / "Fittingslist-a.xlsx", [("F100", 4), ("F200", 7)])
            make_fittings(report_b / "Fittingslist-b.xlsx", [("F100", 4), ("F200", 7)])
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )
            preview = preview_order(config, order)
            ignored_name = preview.factories[0].fittings[0].name
            set_ignored(config, ignored_name, True)
            preview = preview_order(config, order)
            self.assertTrue(preview.factories[0].fittings[0].ignored)
            self.assertIsNotNone(InventoryMappings(config.workflow_database).ignored_reason(ignored_name))
            set_ignored_mapping(config, ignored_name, False)
            preview = preview_order(config, order)
            self.assertFalse(preview.factories[0].fittings[0].ignored)
            set_ignored(config, preview.factories[0].fittings[0].name, True)
            preview = preview_order(config, order)
            self.assertTrue(preview.factories[0].fittings[0].ignored)
            output = generate_order_traveler(config, preview)
            wb = load_workbook(output, data_only=False)
            self.assertEqual(wb.sheetnames[:3], ["WorkOrderTraveler", "Usage List", "Picking List"])
            self.assertEqual(wb["WorkOrderTraveler"]["B5"].value, "PP9999")
            self.assertEqual(wb["WorkOrderTraveler"]["D4"].value.date(), date.today())
            picking = wb["Picking List"]
            values = [picking.cell(row, 1).value for row in range(1, picking.max_row + 1)]
            self.assertEqual(values.count("Name/工厂单名称"), 2)
            self.assertEqual(picking["A1"].value, "Picking List领料单")
            self.assertEqual(picking["A1"].alignment.horizontal, "center")
            self.assertEqual(picking["A1"].alignment.vertical, "center")
            self.assertIn("A1:I1", {str(item) for item in picking.merged_cells.ranges})
            name_rows = [
                row for row in range(1, picking.max_row + 1)
                if picking.cell(row, 1).value == "Name/工厂单名称"
            ]
            self.assertNotEqual(
                picking.cell(name_rows[0], 1).fill.fgColor.rgb,
                picking.cell(name_rows[1], 1).fill.fgColor.rgb,
            )
            self.assertTrue(
                any(
                    all(picking.cell(row, col).value is None for col in range(1, 8))
                    for row in range(name_rows[0] + 1, name_rows[1])
                )
            )
            self.assertEqual(find_existing_traveler(config, "PP9999"), output)
            self.assertEqual(preview_payload(config, preview)["existing_traveler"], str(output))

            wb["WorkOrderTraveler"]["B20"] = "人工填写内容"
            wb["WorkOrderTraveler"]["D4"] = "保留原日期"
            picking = wb["Picking List"]
            hardware_title = next(
                row for row in range(1, picking.max_row + 1)
                if picking.cell(row, 1).value == "Hardware Accessory五金功能件"
            )
            manual_row = hardware_title + 2
            picking.cell(manual_row, 2).value = "M2000"
            picking.cell(manual_row, 3).value = "Manual Handle"
            picking.cell(manual_row, 5).value = "Black"
            picking.cell(manual_row, 7).value = 6
            picking.cell(manual_row, 9).value = "现场增加"
            wb.save(output)
            materials_wb = load_workbook(order / "anything materials final.xlsx")
            materials_wb.active["C3"] = 8
            materials_wb.active["C14"] = 8
            materials_wb.save(order / "anything materials final.xlsx")
            refreshed = preview_order(config, order)
            updated, backup = update_order_traveler(config, refreshed)
            self.assertEqual(updated, output)
            self.assertTrue(backup.is_file())
            updated_wb = load_workbook(updated, data_only=False)
            self.assertEqual(updated_wb["WorkOrderTraveler"]["B20"].value, "人工填写内容")
            self.assertEqual(updated_wb["WorkOrderTraveler"]["D4"].value, "保留原日期")
            self.assertEqual(updated_wb["Usage List"]["C3"].value, 8)
            self.assertEqual(updated_wb["Usage List"]["C14"].value, "=SUM(C3:C13)")
            self.assertEqual(updated_wb.sheetnames.count("Usage List"), 1)
            updated_picking = updated_wb["Picking List"]
            manual_rows = [
                row for row in range(1, updated_picking.max_row + 1)
                if updated_picking.cell(row, 3).value == "Manual Handle"
            ]
            self.assertEqual(len(manual_rows), 1)
            self.assertEqual(updated_picking.cell(manual_rows[0], 2).value, "M2000")
            self.assertEqual(updated_picking.cell(manual_rows[0], 5).value, "Black")
            self.assertEqual(updated_picking.cell(manual_rows[0], 7).value, 6)
            self.assertEqual(updated_picking.cell(manual_rows[0], 9).value, "现场增加")

    def test_ignored_hardware_is_not_persisted_when_order_is_rescanned(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report = order / "Kitchen" / "Report"
            report.mkdir(parents=True)
            make_materials(order / "PP9999 materials.xlsx")
            make_board(report / "板材清单.xlsx", "F100", "PP9999-KITCHEN")
            make_fittings(report / "Fittingslist.xlsx", [("F100", 4)])
            config = Config(source_root=root, state_dir=root / "state")
            config.prepare_storage()
            make_product_catalog(config.state_dir / "inventory" / "current-products.xlsx")

            preview = preview_order(config, order)
            ignored = preview.factories[0].fittings[0]
            connection = sqlite3.connect(config.workflow_database)
            try:
                self.assertGreater(connection.execute("select count(*) from hardware_items").fetchone()[0], 0)
            finally:
                connection.close()

            set_ignored(config, ignored.name, True)
            preview_order(config, order)
            connection = sqlite3.connect(config.workflow_database)
            try:
                self.assertEqual(
                    connection.execute(
                        "select count(*) from hardware_items where product_code=?",
                        (ignored.code,),
                    ).fetchone()[0],
                    0,
                )
            finally:
                connection.close()

    def test_add_manual_hardware_previews_backs_up_and_aggregates_same_sku(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            order = root / "PP9999"
            report = order / "Kitchen" / "Report"
            report.mkdir(parents=True)
            make_materials(order / "PP9999 materials.xlsx")
            make_board(report / "板材清单.xlsx", "F100", "PP9999-KITCHEN")
            make_fittings(report / "Fittingslist.xlsx", [("F100", 4)])
            template = root / "template.xlsx"
            make_template(template)
            config = Config(
                source_root=root,
                order_root=root / "orders",
                template=template,
                backup_root=root / "backups",
                state_dir=root / "state",
            )
            make_product_catalog(config.state_dir / "inventory" / "current-products.xlsx")
            traveler = generate_order_traveler(config, preview_order(config, order))

            preview = preview_manual_hardware(
                config, "PP9999", "pp9999-kitchen", "m2000", 3, "现场增加"
            )
            self.assertEqual(preview["factory_name"], "PP9999-KITCHEN")
            self.assertEqual(preview["product_name"], "Manual Handle")
            self.assertEqual(preview["spec"], "Black")
            self.assertEqual(preview["quantity"], 3)
            with self.assertRaises(RuleError) as raised:
                preview_manual_hardware(config, "PP9999", "PP9999-KITCHEN", "M2000", 1.5)
            self.assertEqual(raised.exception.code, "manual_hardware_quantity")

            updated, first_backup, first = add_manual_hardware(
                config, "PP9999", "PP9999-KITCHEN", "M2000", 3, "现场增加"
            )
            self.assertEqual(updated, traveler)
            self.assertTrue(first_backup.is_file())
            self.assertEqual(first["result"], "added")
            _, second_backup, second = add_manual_hardware(
                config, "PP9999", "PP9999-KITCHEN", "M2000", 2, "补充两个"
            )
            self.assertTrue(second_backup.is_file())
            self.assertNotEqual(first_backup, second_backup)
            self.assertEqual(second["result"], "increased")
            self.assertEqual(second["saved_quantity"], 5)

            workbook = load_workbook(traveler, data_only=False)
            picking = workbook["Picking List"]
            rows = [
                row for row in range(1, picking.max_row + 1)
                if picking.cell(row, 2).value == "M2000"
            ]
            self.assertEqual(len(rows), 1)
            row = rows[0]
            self.assertEqual(picking.cell(row, 3).value, "Manual Handle")
            self.assertEqual(picking.cell(row, 5).value, "Black")
            self.assertEqual(picking.cell(row, 7).value, 5)
            self.assertEqual(picking.cell(row, 9).value, "现场增加；补充两个")
            self.assertIn("Hinge", [picking.cell(item, 3).value for item in range(1, picking.max_row + 1)])


if __name__ == "__main__":
    unittest.main()
