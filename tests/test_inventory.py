import json
import os
import sqlite3
import subprocess
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from traveler_assistant.core import Config, RuleError
from traveler_assistant.inventory import (
    InventoryMappings,
    InventoryPreview,
    InventorySyncStore,
    OutboundItem,
    Product,
    ProductDatabase,
    ProductCatalog,
    _catalog_change_summary,
    close_inventory_chrome,
    _keychain_password,
    _jdy_error_detail,
    _find_existing_inventory_page,
    _resolve_jdy_runtime,
    _is_inventory_authenticated_url,
    _is_inventory_domain_url,
    _is_inventory_service_workbench_url,
    _persist_completed_outbound_results,
    build_database_preview,
    build_preview,
    check_stock,
    database_outbound_fingerprint,
    import_catalog,
    list_traveler_names,
    mark_customer_supplied_outbound,
    open_inventory_chrome,
    order_stock_requirements,
    parse_traveler,
    list_inventory_mappings,
    remove_manual_mapping,
    reconcile_folder_status,
    run_jdy,
    resolve_inventory_items,
    outbound_scope_decisions,
    set_ignored_mapping,
    save_manual_mapping,
    set_outbound_scope,
    stock_requirements,
    bootstrap_product_database,
    TravelerItem,
    TravelerData,
    update_catalog_online,
    update_manual_mapping,
)
from traveler_assistant.order_index import (
    OrderIndexStore,
    _refresh_outbound_status,
    assert_factory_orders_outbound_allowed,
    reconcile_outbound_statuses,
)


def make_traveler(path: Path, items):
    workbook = Workbook()
    main = workbook.active
    main.title = "WorkOrderTraveler"
    main.append(["Name/工厂单名称", "PP0099"])
    usage = workbook.create_sheet("Usage List")
    usage.append(["Job:", "PP0099"])
    usage.append([
        "Room/section", "", "3/4 Plywood", "5/8 Plywood", "1/4 Plywood",
        "3/4 Finish Panel", "1/4 Finish Panel", "Edge Banding (m)", "Color",
    ])
    plywood = {"18mm--Plywood": 0, "14.5mm--Plywood": 0, "5.4mm--Plywood": 0}
    colored_materials = {}
    hardware = []
    for name, quantity in items:
        if name in plywood:
            plywood[name] = quantity
        elif "--" in name and (
            name.lower().startswith(("8mm--", "9mm--", "19.1mm--", "edge banding--"))
        ):
            color = name.split("--", 1)[1]
            values = colored_materials.setdefault(color, {"panel": 0, "back": 0, "edge": 0})
            if name.lower().startswith("19.1mm--"):
                values["panel"] = quantity
            elif name.lower().startswith(("8mm--", "9mm--")):
                values["back"] = quantity
            else:
                values["edge"] = quantity
        else:
            hardware.append((name, quantity))
    colors = list(colored_materials) or ["TEST COLOR"]
    for index, color in enumerate(colors):
        values = colored_materials.get(color, {"panel": 0, "back": 0, "edge": 0})
        usage.append([
            "Test" if index == 0 else "", "",
            plywood["18mm--Plywood"] if index == 0 else 0,
            plywood["14.5mm--Plywood"] if index == 0 else 0,
            plywood["5.4mm--Plywood"] if index == 0 else 0,
            values["panel"], values["back"], values["edge"], color,
        ])
    usage.cell(13, 1).value = None
    usage.append(["Total Qty:"])
    usage.append(["Color Table"])
    usage.append(["Color:"])
    usage.append(["Sheets (3/4):"])
    usage.append(["Sheets (1/4):"])
    usage.append(["Edge Banding (m):"])
    picking = workbook.create_sheet("Picking List")
    picking.append(["Picking List领料单"])
    picking.append(["Name/工厂单名称", "PP0099-KITCHEN"])
    picking.append(["Panel板材"])
    picking.append(["No.", "Name名字", "QTY数量"])
    for number, (name, quantity) in enumerate(hardware, 1):
        picking.append([number, name, quantity])
    workbook.save(path)


def make_catalog(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品类别", "*商品编号", "商品名称", "规格型号", "状态", "计量单位"])
    rows = [
        ("Plywood", "M0004", "3/4 Finished UV2S", "18mm", "启用", "张"),
        ("Hardware", "M1001", "Unihopper Hinge", "", "启用", "件"),
        ("Hardware", "M1068", "Push Open A", "", "启用", "件"),
        ("Hardware", "M1069", "Push Open B", "", "启用", "件"),
        ("Edge band", "M0020", "Woodline 4 Edge Banding", "22mm", "启用", "m"),
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def make_priced_catalog(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品类别", "*商品编号", "商品名称", "规格型号", "状态", "计量单位", "预计采购价"])
    sheet.append(["Plywood", "M0004", "3/4 Finished UV2S", "18mm", "启用", "张", 31.76])
    sheet.append(["Panel", "M0005", "No price panel", "19.1mm", "启用", "张", None])
    workbook.save(path)


class InventoryTests(unittest.TestCase):
    def test_customer_supplied_outbound_marks_database_only_and_reopens_on_fact_change(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("CS001", "cutToSize", "now"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F1001", "CS001", "CS001-KITCHEN", 1, "未出库", "now"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("CS001", "panel", "Customer Panel", "19.1", 4, "张", "aihouse", "now"),
            )
            store.commit()
            store.close()

            set_outbound_scope(
                config,
                "CS001",
                "material",
                "customer_supplied",
                reason="客户提供材料，本公司只负责加工",
            )
            fingerprint = database_outbound_fingerprint(config, "CS001", "F1001")
            result = mark_customer_supplied_outbound(config, "CS001", ["F1001"])
            self.assertEqual(result["outbound_mode"], "customer_supplied")
            self.assertEqual(result["outbound_status"], "已出库")

            connection = sqlite3.connect(config.workflow_database)
            row = connection.execute(
                "select outbound_status, outbound_document, outbound_mode, outbound_fingerprint from factory_orders where factory_order='F1001'"
            ).fetchone()
            self.assertEqual(row, ("已出库", "", "customer_supplied", fingerprint))
            self.assertEqual(
                _refresh_outbound_status(
                    config,
                    {"order_id": "CS001", "factory_order": "F1001", "factory_name": "CS001-KITCHEN"},
                    [],
                ),
                ("已出库", ""),
            )
            connection.execute(
                "update material_items set quantity=5 where order_id='CS001'"
            )
            connection.commit()
            connection.close()
            self.assertNotEqual(
                database_outbound_fingerprint(config, "CS001", "F1001"),
                fingerprint,
            )
            self.assertEqual(
                _refresh_outbound_status(
                    config,
                    {"order_id": "CS001", "factory_order": "F1001", "factory_name": "CS001-KITCHEN"},
                    [],
                ),
                ("需要更新", ""),
            )

    def test_hardware_scope_requires_actual_positive_hardware_facts(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("CS004", "cutToSize", "now"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F4004", "CS004", "CS004-KITCHEN", 1, "未出库", "now"),
            )
            store.commit()
            store.close()

            scope = outbound_scope_decisions(config, "CS004")
            self.assertEqual(scope["hardware"], {})
            with self.assertRaisesRegex(RuleError, "没有.*可出库五金数据"):
                set_outbound_scope(config, "CS004", "hardware", "required", factory_order="F4004")

    def test_outbound_scope_read_returns_latest_relevant_saved_decision(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("CS005", "cutToSize", "now"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F5005", "CS005", "CS005-KITCHEN", 1, "未出库", "now"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("CS005", "panel", "Blanco HG", "19.1", 1, "张", "manual", "now"),
            )
            store.connection.execute(
                "insert into hardware_items(order_id, factory_order, scope, product_code, name, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?,?)",
                ("CS005", "F5005", "factory_order", "M1001", "Hinge", 1, "件", "database", "now"),
            )
            store.commit()
            store.close()

            set_outbound_scope(config, "CS005", "material", "customer_supplied", reason="客户提供板材")
            material_scope = outbound_scope_decisions(config, "CS005", ["F5005"])
            self.assertEqual(material_scope["last_decision"]["scope_type"], "material")
            self.assertEqual(material_scope["material"]["requirement"], "customer_supplied")
            self.assertEqual(material_scope["last_decision"]["reason"], "客户提供板材")

            set_outbound_scope(config, "CS005", "hardware", "required", factory_order="F5005")
            hardware_scope = outbound_scope_decisions(config, "CS005", ["F5005"])
            self.assertEqual(hardware_scope["last_decision"]["scope_type"], "hardware")
            self.assertEqual(hardware_scope["last_decision"]["factory_order"], "F5005")

    def test_cut_to_size_customer_supplied_material_stays_fact_but_is_not_outbound(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state", order_root=root / "generated")
            config.prepare_storage()
            (config.state_dir / "inventory").mkdir(parents=True, exist_ok=True)
            make_catalog(config.state_dir / "inventory" / "current-products.xlsx")
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, source_folder, updated_at) values(?,?,?,?)",
                ("CS002", "cutToSize", str(root / "source" / "CS002"), "now"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F2002", "CS002", "CS002-Hardware", 1, "未出库", "now"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("CS002", "plywood", "", "18", 5, "pcs", "aihouse", "now"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, edge, source_type, updated_at) values(?,?,?,?,?,?,?,?,?)",
                ("CS002", "edge", "Woodline 4", "", 100, "m", "Woodline 4", "aihouse", "now"),
            )
            store.connection.execute(
                "insert into hardware_items(order_id, factory_order, scope, product_code, name, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?,?)",
                ("CS002", "F2002", "factory_order", "M1001", "Unihopper Hinge", 2, "件", "aicnc", "now"),
            )
            store.commit()
            store.close()
            set_outbound_scope(config, "CS002", "material", "customer_supplied", reason="客户提供板材和封边")
            with patch("traveler_assistant.inventory.bootstrap_product_database", return_value=config.state_dir / "inventory" / "current-products.xlsx"):
                preview = build_database_preview(config, "CS002", ["F2002"])
            self.assertTrue(preview.ready)
            self.assertEqual({item.product_code for item in preview.outbound_items}, {"M1001"})
            self.assertEqual(preview.outbound_items[0].document_remark, "CS002-Hardware")
            self.assertEqual(preview.scope_decisions[0]["requirement"], "customer_supplied")
            facts = sqlite3.connect(config.workflow_database).execute(
                "select quantity from material_items where order_id='CS002' order by material_type"
            ).fetchall()
            self.assertEqual([row[0] for row in facts], [100.0, 5.0])

    def test_customer_supplied_scope_is_rejected_for_owned_order(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("PP9999", "owned", "now"),
            )
            store.commit()
            store.close()
            with self.assertRaisesRegex(RuleError, "自有订单不支持设置出库范围"):
                set_outbound_scope(config, "PP9999", "material", "customer_supplied", reason="误操作")

    def test_owned_order_rejects_even_normalized_outbound_scope_decision(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("PP9998", "owned", "now"),
            )
            store.connection.commit()
            store.close()
            with self.assertRaisesRegex(RuleError, "自有订单不支持设置出库范围"):
                set_outbound_scope(config, "PP9998", "material", "required")

    def test_remainder_decision_allows_empty_order_without_opening_browser(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, updated_at) values(?,?,?)",
                ("CS002", "cutToSize", "now"),
            )
            store.commit()
            store.close()
            set_outbound_scope(config, "CS002", "material", "remainder", reason="本单只用余料生产")
            preview = build_database_preview(config, "CS002", [])
            self.assertTrue(preview.ready)
            self.assertTrue(preview.no_outbound_required)
            self.assertEqual(preview.outbound_items, [])

    def test_database_order_outbound_preview_maps_without_traveler_file(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            config = Config(state_dir=state, order_root=root / "generated")
            config.prepare_storage()
            (state / "inventory").mkdir(parents=True, exist_ok=True)
            make_catalog(state / "inventory" / "current-products.xlsx")
            (state / "inventory" / "mappings.json").write_text(
                '{"manual": {}, "ignored": {}}', encoding="utf-8"
            )
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, source_folder, updated_at) values(?,?,?,?)",
                ("PP9999", "owned", str(root / "source" / "PP9999"), "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F9999", "PP9999", "PP9999-KITCHEN", 1, "未出库", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("PP9999", "plywood", "", "18", 2, "pcs", "database", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, edge, source_type, updated_at) values(?,?,?,?,?,?,?,?,?)",
                ("PP9999", "edge", "Woodline 4", "", 12.5, "m", "Woodline 4", "database", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into hardware_items(order_id, factory_order, scope, product_code, name, spec, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?,?,?)",
                ("PP9999", "F9999", "factory_order", "M1001", "Unihopper Hinge", "", 3, "件", "database", "2026-08-15T10:00:00"),
            )
            store.connection.commit()
            store.close()

            with patch(
                "traveler_assistant.inventory.bootstrap_product_database",
                return_value=state / "inventory" / "current-products.xlsx",
            ):
                preview = build_database_preview(config, "PP9999", ["F9999"])

            self.assertTrue(preview.ready)
            self.assertEqual(
                {(item.document_remark, item.product_code) for item in preview.outbound_items},
                {("PP9999", "M0004"), ("PP9999", "M0020"), ("PP9999-KITCHEN", "M1001")},
            )
            self.assertEqual(preview.traveler.path, config.workflow_database.resolve())
            self.assertFalse((config.order_root / "PP9999" / "Work Order Traveler(PP9999).xlsx").exists())

    def test_order_context_source_no_longer_requires_traveler_gate(self):
        root = Path(__file__).resolve().parents[1]
        dashboard = (root / "macos" / "OrderDashboardView.swift").read_text(encoding="utf-8")
        swift = (root / "macos" / "TravelerAssistant.swift").read_text(encoding="utf-8")
        self.assertNotIn("model.orderExistingTravelerPath.isEmpty", dashboard)
        self.assertIn('"order-preview", "--order-id"', swift)
        self.assertIn('var arguments = ["outbound"]', swift)
        self.assertIn('arguments += ["--order-id", orderID]', swift)
        self.assertIn('"get-outbound-scope", "--order-id"', swift)
        self.assertIn("loadOutboundScope", swift)
        self.assertIn('finishInventoryStep(\n                named: "预检订单出库数据"', swift)
        self.assertNotIn("case .inventory", swift)
        self.assertNotIn("selection = .inventory", swift)
        self.assertNotIn("onChange(of: model.inventoryMappingRequestPath", swift)
        self.assertIn("showInventoryMappingWorkspace", swift)
        self.assertIn("inventoryMappingTargetNames", swift)
        self.assertIn("PendingInventoryMappingWorkspace", swift)
        self.assertIn("PendingInventoryIgnoreSheet", swift)
        self.assertIn('Button("忽略")', swift)
        self.assertIn('Button("加入全局忽略")', swift)
        self.assertIn('Button("处理映射")', swift)
        self.assertIn("inventoryMappingSourceFolderPath", swift)
        self.assertIn("refreshDashboardOrdersAfterInventoryMapping", swift)
        self.assertIn('runOrder(["list-index"], failureStatus: "订单列表刷新失败"', swift)
        self.assertIn("InventoryView(", dashboard)
        detail_card = dashboard.split("struct OrderDashboardDetailCard", 1)[1].split(
            "struct OutboundScopeSheet", 1
        )[0]
        self.assertIn('if orderType != "owned"', detail_card)
        self.assertIn('Button("设置出库范围")', detail_card)
        detail = dashboard.split("struct OrderDashboardDetailPage", 1)[1].split(
            "struct OrderDashboardDetailCard", 1
        )[0]
        self.assertNotIn('Text("订单详情")', detail)
        self.assertIn("VStack(alignment: .leading, spacing: 14)", detail)
        self.assertIn("min(max(0, geometry.size.width - AppLayout.contentPadding * 2), 1120)", detail)
        self.assertIn("ScrollView(.vertical)", detail)
        self.assertEqual(detail.count("ScrollView(.vertical)"), 1)
        self.assertIn("ScrollView(.vertical)", detail)
        self.assertIn("boardAndEdgeSection", detail)
        self.assertIn("hardwareSection", detail)
        self.assertIn("HStack(alignment: .center", detail)
        self.assertIn('Text("订单 (\\(order.orderId))")', detail)
        self.assertNotIn("order.orderType", detail)
        self.assertIn('Button("生成 Traveler")', detail)
        self.assertIn('Button("关闭")', detail)
        self.assertIn(".font(.title2.weight(.semibold))", detail)
        self.assertIn("orderDetailCardMinHeight", detail)
        self.assertIn("value: row.quantity.formatted()", detail)
        self.assertNotIn("model.loadOrderDetailFromDatabase(order)", detail)
        self.assertNotIn("if model.selectedOrderId.caseInsensitiveCompare(order.orderId)", detail)
        self.assertIn('Text("订单材料")', swift)
        self.assertIn("orderDetailGridColumnCount", detail)
        self.assertIn("inventoryIgnoredMappings", swift)
        self.assertIn("inventoryManualMappings", swift)
        self.assertIn("五金全局忽略列表", swift)
        self.assertIn("struct InventoryManualMappingsSheet", swift)
        self.assertIn("saveInventoryIgnoredMapping", swift)
        self.assertIn("updateInventoryIgnoredMapping", swift)
        self.assertIn("removeInventoryIgnoredMapping", swift)
        self.assertIn("saveSettingsManualMapping", swift)
        self.assertIn("updateSettingsManualMapping", swift)
        self.assertIn("removeSettingsManualMapping", swift)
        settings = swift.split("struct SettingsView", 1)[1].split(
            "struct OperationLogViewerView", 1
        )[0]
        self.assertNotIn("Traveler 材料名称", settings)
        self.assertNotIn("设置材料映射", settings)
        self.assertIn("库存商品资料与全局忽略", settings)
        self.assertIn('Button("查看") { showIgnoredHardwareList = true }', settings)
        self.assertIn("struct InventoryIgnoredMappingsSheet", swift)

    def test_database_outbound_status_changes_when_persisted_order_data_changes(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            config = Config(state_dir=state, order_root=root / "generated")
            config.prepare_storage()
            (state / "inventory").mkdir(parents=True, exist_ok=True)
            make_catalog(state / "inventory" / "current-products.xlsx")
            (state / "inventory" / "mappings.json").write_text(
                '{"manual": {}, "ignored": {}}', encoding="utf-8"
            )
            store = OrderIndexStore(config.workflow_database)
            store.connection.execute(
                "insert into orders(order_id, order_type, source_folder, updated_at) values(?,?,?,?)",
                ("PP9999", "owned", str(root / "source" / "PP9999"), "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into factory_orders(factory_order, order_id, factory_name, optimized, outbound_status, updated_at) values(?,?,?,?,?,?)",
                ("F9999", "PP9999", "PP9999-KITCHEN", 1, "已出库", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("PP9999", "plywood", "", "18", 2, "pcs", "database", "2026-08-15T10:00:00"),
            )
            store.connection.execute(
                "insert into hardware_items(order_id, factory_order, scope, product_code, name, spec, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?,?,?)",
                ("PP9999", "F9999", "factory_order", "M1001", "Unihopper Hinge", "", 3, "件", "database", "2026-08-15T10:00:00"),
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.inventory.bootstrap_product_database",
                return_value=state / "inventory" / "current-products.xlsx",
            ):
                preview = build_database_preview(config, "PP9999", ["F9999"])

            factory = {
                "order_id": "PP9999",
                "factory_order": "F9999",
                "factory_name": "PP9999-KITCHEN",
            }
            record = {
                "order_id": "PP9999",
                "remark": "PP9999-KITCHEN",
                "status": "已出库",
                "document_number": "QTCK-1",
                "traveler_path": str(config.workflow_database),
                "raw_fingerprint": InventorySyncStore.raw_document_fingerprint(
                    preview.traveler.documents["PP9999-KITCHEN"]
                ),
            }
            self.assertEqual(_refresh_outbound_status(config, factory, [record]), ("已出库", "QTCK-1"))

            connection = sqlite3.connect(config.workflow_database)
            connection.execute(
                "update hardware_items set quantity=4 where order_id=? and factory_order=?",
                ("PP9999", "F9999"),
            )
            connection.commit()
            connection.close()
            self.assertEqual(_refresh_outbound_status(config, factory, [record]), ("需要更新", "QTCK-1"))

    def test_inventory_page_detection_accepts_tenant_workbench_subdomain(self):
        url = "https://vip2-hz.jdy.com/default-new.jsp?dbid=7937191793066#/beginner-guide"

        self.assertTrue(_is_inventory_domain_url(url))
        self.assertTrue(_is_inventory_authenticated_url(url))
        with patch(
            "traveler_assistant.inventory._inventory_cdp_pages",
            return_value=[{"type": "page", "url": url}],
        ):
            self.assertEqual(_find_existing_inventory_page("http://127.0.0.1:9222"), {"type": "page", "url": url})

    def test_inventory_service_workbench_is_an_entry_page(self):
        service_url = "https://service.jdy.com/workbench/web/index.html"

        self.assertTrue(_is_inventory_authenticated_url(service_url))
        self.assertTrue(_is_inventory_service_workbench_url(service_url))

    def test_inventory_page_detection_still_rejects_login_and_non_inventory_urls(self):
        self.assertFalse(_is_inventory_authenticated_url("https://www.jdy.com/login/"))
        self.assertFalse(_is_inventory_authenticated_url("https://www.jdy.com/global/?logout=true"))
        self.assertFalse(_is_inventory_domain_url("https://example.com/default-new.jsp"))

    def test_jdy_cdp_cleanup_uses_playwright_browser_close(self):
        source = (Path(__file__).resolve().parents[1] / "tools" / "jdy_inventory.mjs").read_text(encoding="utf-8")

        self.assertIn("await browser.close()", source)
        self.assertIn("await remoteBrowser.close()", source)
        self.assertIn('request.action === "closeChrome"', source)
        self.assertIn('session.send("Browser.close")', source)
        self.assertNotIn("browser.disconnect()", source)
        self.assertNotIn("remoteBrowser.disconnect()", source)

    def test_outbound_navigation_reuses_existing_list_and_opens_visible_menu_item(self):
        source = (Path(__file__).resolve().parents[1] / "tools" / "jdy_inventory.mjs").read_text(encoding="utf-8")

        self.assertIn("const currentOtherOutboundListFrame", source)
        self.assertIn("const isOtherOutboundListFrame", source)
        self.assertIn("const isOtherOutboundListURL", source)
        self.assertIn("action=initOiList", source)
        self.assertIn("const currentOtherOutboundFormFrame", source)
        self.assertIn("const isServiceWorkbenchURL", source)
        self.assertIn("const ensureInventoryBusinessWorkbench", source)
        self.assertIn("const waitForInventoryActionShell", source)
        self.assertIn("点击服务工作台“进入使用”", source)
        self.assertIn("const waitForVisibleLeftNavigationItem", source)
        self.assertIn("const waitForOtherOutboundListFrame", source)
        self.assertIn("const waitForOtherOutboundFormFrame", source)
        self.assertIn("const clickOtherOutboundHistory", source)
        self.assertIn("const clickOtherOutboundHistoryWithRetry", source)
        self.assertIn("其他出库单中的历史单据", source)
        self.assertIn("准备点击“历史单据”进入记录列表", source)
        self.assertIn("const assertOutboundFormMatchesRequest", source)
        self.assertIn("const outboundMaterialRows", source)
        self.assertIn('td[aria-describedby="grid_invNumber"]', source)
        self.assertIn("snapshot.productCode === item.productCode", source)
        self.assertNotIn("const meaningfulCells = cells.filter", source)
        self.assertIn("await waitForOtherOutboundListFrame(page)", source)
        self.assertNotIn('log("已复用当前页面中的空白其他出库单表单")', source)
        self.assertIn(".quick-datepicker-start:visible", source)
        self.assertIn("Prefer the tab that already contains the outbound list", source)
        self.assertIn('log("已复用当前页面中的其他出库单记录列表")', source)
        self.assertIn('const openOtherOutboundMenuItem', source)
        self.assertIn('visibleTextLocatorsAcrossFrames(page, label)', source)
        self.assertIn("warehouse.hover({ force: true })", source)
        self.assertIn("本机已有出库单", source)
        self.assertIn("已停止，不新建重复出库单", source)

    def test_outbound_timeout_persists_completed_documents_before_retry(self):
        source = (Path(__file__).resolve().parents[1] / "traveler_assistant" / "inventory.py").read_text(encoding="utf-8")
        self.assertIn("_persist_completed_outbound_results", source)
        self.assertIn("后续单据结果需要先查询库存历史再重试", source)

    def test_jdy_error_detail_explains_reused_page_menu_timeout(self):
        detail = _jdy_error_detail(
            "库存系统自动操作失败：左侧菜单中等待可见“仓库”超过 30 秒；当前页面："
            "https://vip2-hz.jdy.com/default-new.jsp"
        )
        self.assertIn("左侧“仓库”菜单", detail)

    def test_outbound_preview_hides_bottom_write_note(self):
        source = (Path(__file__).resolve().parents[1] / "macos" / "TravelerAssistant.swift").read_text(encoding="utf-8")
        self.assertNotIn("真实写入会创建库存出库单，并保存本机同步记录。", source)

    def test_outbound_preview_rows_are_full_row_clickable_and_scroll_independently(self):
        root = Path(__file__).resolve().parents[1]
        source = (root / "macos" / "TravelerAssistant.swift").read_text(encoding="utf-8")
        dashboard_source = (root / "macos" / "OrderDashboardView.swift").read_text(encoding="utf-8")
        self.assertIn("private func previewRowContent(_ row: InventoryPreviewRow)", source)
        self.assertNotIn("selectedPreviewRows", source)
        self.assertNotIn("togglePreviewRow(row.id)", source)
        self.assertNotIn("点击记录行即可整行选中；按住 Command 可多选", source)
        self.assertIn(".frame(maxWidth: .infinity)\n        .frame(height: AppLayout.operationLogHeight)", source)
        self.assertIn("HStack(spacing: 8) {\n                    Button(\"取消\")", source)
        self.assertIn("ScrollView(.vertical)", source)
        self.assertIn("inventoryWriteBlocked", source)
        self.assertIn("inventoryWriteCompleted", source)
        self.assertIn('text: model.inventoryWriteCompleted', source)
        self.assertIn('active: confirmRealSave || model.inventoryWriteCompleted', source)
        self.assertIn('self.inventoryWriteCompleted = true', source)
        self.assertIn('finishRunningInventoryStep(', source)
        self.assertIn('.contentShape(Rectangle())', source)
        self.assertNotIn('Text(model.inventorySuccessMessage.isEmpty ? model.inventoryStatus : model.inventorySuccessMessage)', source)
        self.assertNotIn('· 已选 \\(selectedPreviewRows.count)', source)
        self.assertNotIn('已选工厂单：', source)
        self.assertNotIn('Text("\\(model.inventoryPreviewRows.count) 行")', source)
        self.assertIn('if model.inventoryChromeStatus.hasPrefix("❌")', source)
        self.assertNotIn('Text("已选 \\(selectedFactoryIDs.count) 个工厂单")', dashboard_source)

    def test_order_dashboard_outbound_refresh_and_panel_color_layout_contract(self):
        root = Path(__file__).resolve().parents[1]
        source = (root / "macos" / "TravelerAssistant.swift").read_text(encoding="utf-8")
        dashboard_source = (root / "macos" / "OrderDashboardView.swift").read_text(encoding="utf-8")
        self.assertIn("func refreshDashboardOrdersAfterOutbound()", source)
        self.assertIn('runOrder(["list-index"], failureStatus: "订单列表刷新失败"', source)
        self.assertIn("pendingDashboardOutboundRefresh", source)
        self.assertIn("HStack(alignment: .center, spacing: AppLayout.actionSpacing)", source)
        self.assertIn(".frame(maxWidth: .infinity, alignment: .center)", source)
        self.assertIn("model.refreshDashboardOrdersAfterOutbound()", dashboard_source)
        self.assertIn('.fixedSize(horizontal: true, vertical: false)', dashboard_source)
        self.assertIn('.lineLimit(2)', dashboard_source)
        self.assertIn('.multilineTextAlignment(.leading)', dashboard_source)

    def test_inventory_chrome_success_result_has_no_hidden_login_prompt(self):
        source = (Path(__file__).resolve().parents[1] / "traveler_assistant" / "inventory.py").read_text(encoding="utf-8")
        swift_source = (Path(__file__).resolve().parents[1] / "macos" / "TravelerAssistant.swift").read_text(encoding="utf-8")
        self.assertNotIn("库存专用 Chrome 已打开，请手工登录并完成验证码后再操作", source)
        self.assertNotIn('object["message"] as? String', swift_source)

    def test_order_preview_materials_can_be_mapped_for_stock_check(self):
        with tempfile.TemporaryDirectory() as directory:
            state = Path(directory) / "state"
            catalog = state / "inventory" / "current-products.xlsx"
            mappings = state / "inventory" / "mappings.json"
            catalog.parent.mkdir(parents=True)
            make_catalog(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = SimpleNamespace(
                order_id="PP0068",
                materials=[SimpleNamespace(kind="plywood", thickness=18.0, color="", quantity=2)],
                edge_banding={"Woodline 4": 12.5},
            )
            with patch("traveler_assistant.order_workflow.preview_order", return_value=preview):
                order_id, rows = order_stock_requirements(Config(state_dir=state), Path(directory) / "PP0068")
            self.assertEqual(order_id, "PP0068")
            self.assertEqual([row["productCode"] for row in rows], ["M0004", "M0020"])
            self.assertEqual([row["requiredQuantity"] for row in rows], [2, 13])
            self.assertEqual([row["unit"] for row in rows], ["张", "m"])

    def test_jdy_runtime_uses_portable_overrides_and_rejects_missing_dependencies(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            node = root / "runtime" / "node"
            node.parent.mkdir(parents=True)
            node.touch()
            node.chmod(0o755)
            modules = root / "runtime" / "node_modules"
            (modules / "playwright").mkdir(parents=True)
            (modules / "playwright-core").mkdir()
            with patch.dict(
                "os.environ",
                {"TRAVELER_NODE": str(node), "TRAVELER_NODE_MODULES": str(modules)},
            ):
                self.assertEqual(_resolve_jdy_runtime(root), (node, modules))

            with patch.dict(
                "os.environ",
                {
                    "TRAVELER_NODE": str(root / "missing-node"),
                    "TRAVELER_NODE_MODULES": str(modules),
                },
            ):
                with self.assertRaises(RuleError) as raised:
                    _resolve_jdy_runtime(root)
                self.assertEqual(raised.exception.code, "inventory_runtime")

    def test_stock_requirements_default_to_materials_and_can_include_hardware(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("18mm--Plywood", 2), ("Hinge", 3)])
            make_catalog(catalog)
            mappings.write_text('{"manual": {"Hinge": "M1001"}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)

            materials = stock_requirements(preview)
            all_items = stock_requirements(preview, include_hardware=True)

            self.assertEqual([item["productCode"] for item in materials], ["M0004"])
            self.assertEqual([item["productCode"] for item in all_items], ["M0004", "M1001"])

    def test_stock_check_compares_required_and_available_quantities(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            state = root / "state"
            catalog = state / "inventory" / "current-products.xlsx"
            mappings = state / "inventory" / "mappings.json"
            catalog.parent.mkdir(parents=True)
            make_traveler(traveler, [("18mm--Plywood", 2), ("Hinge", 3)])
            make_catalog(catalog)
            mappings.write_text('{"manual": {"Hinge": "M1001"}, "ignored": {}}', encoding="utf-8")
            config = Config(state_dir=state)
            browser_result = {
                "ok": True,
                "results": [{
                    "productCode": "M0004",
                    "productName": "3/4 Finished UV2S",
                    "availableQuantity": 1,
                }],
            }
            with patch("traveler_assistant.inventory.run_jdy", return_value=browser_result) as run:
                result = check_stock(config, traveler)

            run.assert_called_once()
            self.assertTrue(result["hasShortage"])
            self.assertEqual(result["rows"][0]["shortageQuantity"], 1)
            self.assertFalse(result["rows"][0]["sufficient"])

    def test_cut_to_size_folder_status_can_be_reconciled(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            folder = root / "orders" / "CS004"
            folder.mkdir(parents=True)
            (folder / "Work Order Traveler(CS004).xlsx").touch()
            config = Config(
                order_root=root / "orders",
                backup_root=root / "backups",
                state_dir=root / "state",
            )
            traveler = SimpleNamespace(order_name="CS004")
            with patch("traveler_assistant.inventory.parse_traveler", return_value=traveler), \
                 patch(
                     "traveler_assistant.inventory.run_jdy",
                     return_value={"matches": ["CS004 QTCK000123"]},
                 ):
                result = reconcile_folder_status(config, "cs004")
            self.assertEqual(result["found"][0]["document_number"], "QTCK000123")

    def test_online_catalog_update_exports_validates_and_installs(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")

            def fake_export(_config, action, **kwargs):
                self.assertEqual(action, "exportProducts")
                make_catalog(kwargs["download_path"])
                return {"ok": True}

            with patch("traveler_assistant.inventory.run_jdy", side_effect=fake_export):
                result = update_catalog_online(config)

            installed = root / "state" / "inventory" / "current-products.xlsx"
            database = root / "state" / "workflow.sqlite3"
            self.assertTrue(result["ok"])
            self.assertEqual(result["count"], len(ProductCatalog(installed).products))
            loaded = ProductDatabase(database)
            self.assertEqual(result["count"], loaded.count())
            self.assertEqual(loaded.require_code("M0004").name, "3/4 Finished UV2S")
            loaded.close()
            self.assertFalse(any(installed.parent.glob(".products-download-*.xlsx")))

    def test_catalog_import_persists_cost_price_and_keeps_missing_price_null(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            source = root / "priced-products.xlsx"
            make_priced_catalog(source)

            parsed = {product.code: product for product in ProductCatalog(source).products}
            self.assertAlmostEqual(parsed["M0004"].cost_price, 31.76)
            self.assertIsNone(parsed["M0005"].cost_price)

            import_catalog(Config(state_dir=state), source)
            database = state / "workflow.sqlite3"
            connection = sqlite3.connect(database)
            columns = {row[1] for row in connection.execute("pragma table_info(products)")}
            connection.close()
            self.assertIn("cost_price", columns)
            self.assertFalse((state / "inventory" / "inventory.sqlite3").exists())
            loaded = ProductDatabase(database)
            self.assertAlmostEqual(loaded.require_code("M0004").cost_price, 31.76)
            self.assertIsNone(loaded.require_code("M0005").cost_price)
            loaded.close()

    def test_catalog_change_summary_reports_added_updated_and_removed_products(self):
        previous = [
            Product("A", "M001", "Old name", "", "启用"),
            Product("B", "M002", "Removed", "", "启用"),
        ]
        current = [
            Product("A", "M001", "New name", "", "启用"),
            Product("C", "M003", "Added", "", "启用"),
        ]

        self.assertEqual(
            _catalog_change_summary(previous, current),
            {"added_count": 1, "updated_count": 1, "removed_count": 1},
        )

    def test_catalog_change_summary_detects_cost_price_change(self):
        previous = [Product("Plywood", "M001", "Panel", "18mm", "启用", cost_price=10.0)]
        current = [Product("Plywood", "M001", "Panel", "18mm", "启用", cost_price=12.5)]

        self.assertEqual(
            _catalog_change_summary(previous, current),
            {"added_count": 0, "updated_count": 1, "removed_count": 0},
        )

    def test_close_inventory_chrome_uses_dedicated_cdp_action(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            browser_result = SimpleNamespace(
                returncode=0,
                stdout='{"ok": true, "closed": true}',
                stderr="",
            )
            with patch("traveler_assistant.inventory._resolve_jdy_runtime", return_value=(Path("/node"), Path("/modules"))), \
                 patch("traveler_assistant.inventory.subprocess.run", return_value=browser_result) as run:
                result = close_inventory_chrome(config)

            request = json.loads(run.call_args.kwargs["input"])
            self.assertTrue(result["closed"])
            self.assertEqual(request["action"], "closeChrome")
            self.assertEqual(request["cdpEndpoint"], "http://127.0.0.1:9222")
            self.assertEqual(run.call_args.kwargs["timeout"], 5)

    def test_lazy_traveler_listing_reports_existing_catalog_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            source = root / "products.xlsx"
            make_catalog(source)
            config = Config(state_dir=state, order_root=root / "orders")
            import_catalog(config, source)

            result = list_traveler_names(config)

            installed = state / "inventory" / "current-products.xlsx"
            self.assertEqual(result["catalog"]["count"], len(ProductCatalog(installed).products))
            self.assertFalse(result["catalog"]["stale"])

    def test_catalog_refresh_keeps_only_latest_xlsx_backup(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            inventory_dir = state / "inventory"
            inventory_dir.mkdir(parents=True)
            old_backup = inventory_dir / "current-products-2026-08-10.xlsx"
            old_backup.write_bytes(b"old")
            source = root / "export.xlsx"
            make_catalog(source)

            result = import_catalog(Config(state_dir=state), source)

            self.assertEqual(result["count"], 5)
            self.assertTrue((inventory_dir / "current-products.xlsx").is_file())
            self.assertFalse(old_backup.exists())
            self.assertTrue((state / "workflow.sqlite3").is_file())
            self.assertFalse((inventory_dir / "inventory.sqlite3").exists())

    def test_runtime_product_search_uses_database_after_xlsx_is_removed(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state = root / "state"
            catalog = state / "inventory" / "current-products.xlsx"
            catalog.parent.mkdir(parents=True)
            make_catalog(catalog)
            database = bootstrap_product_database(Config(state_dir=state))
            catalog.unlink()

            loaded = ProductDatabase(database)
            self.assertEqual([item.code for item in loaded.find(contains="Unihopper")], ["M1001"])
            loaded.close()

    def test_browser_error_preserves_specific_reason(self):
        stderr = (
            '{"event":"progress","message":"正在保存"}\n'
            "库存系统自动操作失败：保存需要人工确认：库存不足\n"
        )
        self.assertEqual(_jdy_error_detail(stderr), "保存需要人工确认：库存不足")

    def test_multiline_browser_error_keeps_first_specific_reason(self):
        stderr = (
            "库存系统自动操作失败：locator.hover: Timeout 30000ms exceeded.\n"
            "Call log:\n"
            "；当前页面：https://example.invalid/\n"
        )
        self.assertEqual(_jdy_error_detail(stderr), "locator.hover: Timeout 30000ms exceeded.")

    def test_login_failure_hides_full_page_dump_and_explains_retry(self):
        stderr = (
            "库存系统自动操作失败：库存系统登录未成功，请检查账号密码、验证码或登录限制后重试"
            "；当前页面：https://www.jdy.com/login/；页面提示：很长的整页内容\n"
        )
        detail = _jdy_error_detail(stderr)
        self.assertIn("库存系统登录未成功", detail)
        self.assertIn("再次查询", detail)
        self.assertNotIn("当前页面", detail)
        self.assertNotIn("整页内容", detail)

    def test_security_challenge_error_explains_visible_login_recovery(self):
        stderr = (
            "库存系统自动操作失败：库存系统要求完成验证码或安全验证；请先使用可见浏览器完成登录验证，再重试\n"
        )
        self.assertEqual(
            _jdy_error_detail(stderr),
            "库存系统要求完成验证码或安全验证；请先使用可见浏览器完成登录验证，再重试",
        )

    def test_profile_lock_failure_is_reported_as_retryable(self):
        stderr = (
            "库存系统自动操作失败：browserType.launchPersistentContext: "
            "Failed to create a ProcessSingleton for your profile directory.\n"
        )
        self.assertEqual(
            _jdy_error_detail(stderr),
            "上一次库存查询浏览器尚未完全退出，请稍候后再次点击查询",
        )

    def test_browser_runtime_noise_does_not_hide_specific_export_timeout(self):
        stderr = (
            "库存系统自动操作失败：点击导出后 60 秒内没有检测到文件下载事件\n"
            "Node.js v24.19.0\n"
            "    at processTicksAndRejections (node:internal/process/task_queues:95:5)\n"
        )
        self.assertEqual(
            _jdy_error_detail(stderr),
            "点击导出后 60 秒内没有检测到文件下载事件",
        )

    def test_run_jdy_passes_attachable_chrome_endpoint(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            browser_result = SimpleNamespace(
                returncode=0,
                stdout='{"ok": true, "url": "https://service.jdy.com/workbench"}',
                stderr="",
            )
            with patch.dict(os.environ, {"TRAVELER_CHROME_CDP_ENDPOINT": "http://127.0.0.1:9333"}), \
                 patch("traveler_assistant.inventory._local_setting", return_value="18108100188"), \
                 patch("traveler_assistant.inventory._keychain_password", return_value="secret"), \
                 patch("traveler_assistant.inventory._resolve_jdy_runtime", return_value=(Path("/node"), Path("/modules"))), \
                 patch("traveler_assistant.inventory.subprocess.run", return_value=browser_result) as run:
                result = run_jdy(config, "preflight")

            request = json.loads(run.call_args.kwargs["input"])
            self.assertTrue(result["ok"])
            self.assertEqual(request["cdpEndpoint"], "http://127.0.0.1:9333")
            self.assertFalse(request["keepBrowserOpen"])
            self.assertEqual(run.call_args.kwargs["timeout"], 90)

    def test_run_jdy_reuses_existing_inventory_page_without_reading_keychain(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            browser_result = SimpleNamespace(
                returncode=0,
                stdout='{"ok": true, "url": "https://www.jdy.com/workbench/web/index.html"}',
                stderr="",
            )
            with patch("traveler_assistant.inventory._local_setting", return_value="18108100188"), \
                 patch(
                     "traveler_assistant.inventory._find_existing_inventory_page",
                     return_value={"url": "https://www.jdy.com/workbench/web/index.html"},
                 ), \
                 patch(
                     "traveler_assistant.inventory._keychain_password",
                     side_effect=AssertionError("已登录页面不应读取钥匙串"),
                 ), \
                 patch("traveler_assistant.inventory._resolve_jdy_runtime", return_value=(Path("/node"), Path("/modules"))), \
                 patch("traveler_assistant.inventory.subprocess.run", return_value=browser_result) as run:
                result = run_jdy(config, "preflight")

            request = json.loads(run.call_args.kwargs["input"])
            self.assertTrue(result["ok"])
            self.assertNotIn("password", request)
            self.assertEqual(request["cdpEndpoint"], "http://127.0.0.1:9222")

    def test_open_inventory_chrome_launches_dedicated_profile_and_debug_port(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            executable = Path(directory) / "Google Chrome"
            executable.touch()
            with patch("traveler_assistant.inventory._inventory_cdp_pages", return_value=[]), \
                 patch("traveler_assistant.inventory._inventory_chrome_executable", return_value=executable), \
                 patch.dict(os.environ, {"TRAVELER_CHROME_CDP_ENDPOINT": "http://127.0.0.1:9333"}), \
                 patch("traveler_assistant.inventory.subprocess.Popen") as popen:
                result = open_inventory_chrome(config)

            self.assertTrue(result["ok"])
            self.assertTrue(result["launched"])
            arguments = popen.call_args.args[0]
            self.assertIn("--remote-debugging-port=9333", arguments)
            self.assertIn("--user-data-dir=" + result["profileDir"], arguments)
            self.assertEqual(arguments[-1], "https://www.jdy.com/login/")

    def test_open_inventory_chrome_does_not_launch_second_browser_on_login_page(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            with patch(
                "traveler_assistant.inventory._inventory_cdp_pages",
                return_value=[{"type": "page", "url": "https://www.jdy.com/login/"}],
            ), patch("traveler_assistant.inventory.subprocess.Popen") as popen:
                result = open_inventory_chrome(config)

            self.assertTrue(result["ok"])
            self.assertTrue(result["waitingForLogin"])
            popen.assert_not_called()

    def test_outbound_stops_before_browser_when_material_mapping_is_missing(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state", order_root=root / "orders")
            traveler = root / "orders" / "Work Order Traveler(PP0099).xlsx"
            traveler.parent.mkdir(parents=True)
            make_traveler(traveler, [("Unmapped Hardware", 2)])
            with patch("traveler_assistant.inventory._local_setting", return_value="18108100188"), \
                 patch("traveler_assistant.inventory._keychain_password", return_value="secret"), \
                 patch("traveler_assistant.inventory.subprocess.run") as run:
                with self.assertRaisesRegex(RuleError, "未映射材料：Unmapped Hardware") as raised:
                    run_jdy(config, "outbound", traveler, confirm_save=True)

            self.assertEqual(raised.exception.code, "inventory_mapping_required")
            run.assert_not_called()

    def test_database_shipped_factory_is_hard_blocked(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state", order_root=root / "orders")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_order("CS005", validation_status="正常")
            store.upsert_factory(
                "F2608120222",
                order_id="CS005",
                factory_name="CS005-KITCHEN",
                sales_order_name="CS005",
                name_source="AIMES",
                ownership_status="已确认",
                optimized=True,
                outbound_status="已出库",
                outbound_document="QTCK20260815001",
            )
            store.commit()
            store.close()

            with self.assertRaisesRegex(RuleError, "已出库") as raised:
                assert_factory_orders_outbound_allowed(config, "CS005", ["F2608120222"])
            self.assertEqual(raised.exception.code, "inventory_already_outbound")

    def test_database_shipped_factory_with_changed_data_can_be_updated(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state", order_root=root / "orders")
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("CS005", validation_status="正常")
            store.upsert_factory(
                "F2608120222",
                order_id="CS005",
                factory_name="CS005-KITCHEN",
                sales_order_name="CS005",
                name_source="AIMES",
                ownership_status="已确认",
                optimized=True,
                outbound_status="已出库",
                outbound_document="QTCK20260815001",
            )
            store.commit()
            store.close()

            with patch("traveler_assistant.order_index.reconcile_outbound_statuses", return_value=0):
                assert_factory_orders_outbound_allowed(
                    config,
                    "CS005",
                    ["F2608120222"],
                    changed_factory_orders=["F2608120222"],
                )

    def test_run_jdy_converts_browser_timeout_to_actionable_rule_error(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Config(state_dir=Path(directory) / "state")
            timeout = subprocess.TimeoutExpired(["node", "jdy_inventory.mjs"], 90, stderr="页面未响应")
            with patch("traveler_assistant.inventory._local_setting", return_value="18108100188"), \
                 patch("traveler_assistant.inventory._keychain_password", return_value="secret"), \
                 patch("traveler_assistant.inventory._resolve_jdy_runtime", return_value=(Path("/node"), Path("/modules"))), \
                 patch("traveler_assistant.inventory.subprocess.run", side_effect=timeout):
                with self.assertRaises(RuleError) as raised:
                    run_jdy(config, "preflight")

            self.assertEqual(raised.exception.code, "jdy_timeout")
            self.assertIn("超过 90 秒", str(raised.exception))

    def test_keychain_timeout_is_reported_as_credentials_error(self):
        with tempfile.TemporaryDirectory() as directory:
            helper = Path(directory) / "keychain-read"
            helper.write_text("placeholder")
            config = Config(state_dir=Path(directory) / "state")
            with patch("traveler_assistant.inventory.subprocess.run", side_effect=subprocess.TimeoutExpired([str(helper), "account"], 10)), \
                 patch("traveler_assistant.inventory.Path.is_file", return_value=True):
                with self.assertRaises(RuleError) as raised:
                    # The patched helper path is resolved through the module's
                    # normal location; only the subprocess call is relevant.
                    _keychain_password("account")

            self.assertEqual(raised.exception.code, "jdy_credentials")
            self.assertIn("超过 10 秒", str(raised.exception))

    def test_parse_dynamic_regions_and_zero(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            make_traveler(path, [("18mm--Plywood", 2), ("Hinge", 0)])
            traveler = parse_traveler(path)
            self.assertEqual(traveler.order_name, "PP0099")
            self.assertEqual(traveler.items[0].quantity, 2)
            self.assertEqual(traveler.zero_items[0].quantity, 0)

    def test_fixed_mapping_and_push_open_expansion(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("18mm--Plywood", 2), ("Push Open", 3), ("Adjustable shelf holder", 4)])
            make_catalog(catalog)
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["商品类别", "*商品编号", "商品名称", "规格型号", "状态"])
            for product in ProductCatalog(catalog).products:
                sheet.append([product.category, product.code, product.name, product.spec, product.status])
            sheet.append(["Hardware", "M1013", "Adjustable Shelf Holder", "", "启用"])
            workbook.save(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)
            self.assertTrue(preview.ready)
            self.assertEqual([item.product_code for item in preview.outbound_items], ["M0004", "M1068", "M1069", "M1013"])
            self.assertEqual([item.quantity for item in preview.outbound_items], [2, 3, 3, 4])

    def test_factory_selection_keeps_order_materials_and_selected_hardware_only(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("18mm--Plywood", 2), ("Hinge", 3)])
            make_catalog(catalog)
            mappings.write_text('{"manual": {"Hinge": "M1001"}, "ignored": {}}', encoding="utf-8")

            preview = build_preview(
                traveler,
                catalog,
                mappings,
                selected_document_remarks=["PP0099-KITCHEN"],
            )

            self.assertTrue(preview.ready)
            self.assertEqual(
                [(item.document_remark, item.product_code) for item in preview.outbound_items],
                [("PP0099", "M0004"), ("PP0099-KITCHEN", "M1001")],
            )
            self.assertEqual(
                set(document["remark"] for document in preview.document_payloads()),
                {"PP0099", "PP0099-KITCHEN"},
            )

    def test_zero_quantity_items_are_not_outbound_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("18mm--Plywood", 2), ("5.4mm--Plywood", 0)])
            make_catalog(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")

            preview = build_preview(traveler, catalog, mappings)

            self.assertEqual([item.traveler_name for item in preview.outbound_items], ["18mm--Plywood"])
            self.assertIn("5.4mm--Plywood", [item["name"] for item in preview.payload()["zero_items"]])

    def test_unique_exact_inventory_name_maps_bls36(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-PANTRY).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("BLS36", 2)])
            make_catalog(catalog)
            workbook = load_workbook(catalog)
            workbook.active.append(["Hardware", "M1093", "BLS36", "TR-270B", "启用"])
            workbook.save(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)
            self.assertTrue(preview.ready)
            self.assertEqual(preview.outbound_items[0].product_code, "M1093")
            self.assertEqual(preview.outbound_items[0].match_source, "库存商品名称精确匹配")

    def test_ignored_material_requires_reason_and_is_visible(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("Old Brand", 1)])
            make_catalog(catalog)
            mappings.write_text(json.dumps({"manual": {}, "ignored": {"Old Brand": "不再采购"}}), encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)
            self.assertTrue(preview.ready)
            self.assertEqual(preview.ignored_items[0]["reason"], "不再采购")

    def test_edge_quantity_is_rounded_half_up_for_inventory(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("Edge banding--Woodline 4", 37.75)])
            make_catalog(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)
            self.assertEqual(preview.outbound_items[0].quantity, 38)
            self.assertIn("37.75m→38m", preview.outbound_items[0].match_source)

    def test_manual_edge_mapping_also_rounds_quantity_for_inventory(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("Edge banding--Woodline 4", 94.25)])
            make_catalog(catalog)
            mappings.write_text(
                json.dumps({"manual": {"Edge banding--Woodline 4": "M0020"}, "ignored": {}}),
                encoding="utf-8",
            )

            preview = build_preview(traveler, catalog, mappings)

            self.assertTrue(preview.ready)
            self.assertEqual(preview.outbound_items[0].product_code, "M0020")
            self.assertEqual(preview.outbound_items[0].quantity, 94)
            self.assertIn("人工指定", preview.outbound_items[0].match_source)
            self.assertIn("94.25m→94m", preview.outbound_items[0].match_source)

    def test_edge_prefers_matching_color_abs_banding_with_24mm_suffix(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            traveler = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_traveler(traveler, [("Edge banding--Ivory Oak", 20.43)])
            make_catalog(catalog)
            workbook = load_workbook(catalog)
            sheet = workbook.active
            sheet.append(["Edge band", "M0145", "Ivory Oak ABS banding 24mm", "1mm*24mm*50M", "启用"])
            sheet.append(["Edge band", "M0146", "Ivory Oak ABS banding 48mm", "1mm*48mm*50M", "启用"])
            sheet.append(["Edge band", "M0147", "Ivory Oak Wood veneer banding 24mm", "0.6mm*24mm*50M", "启用"])
            workbook.save(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(traveler, catalog, mappings)
            self.assertTrue(preview.ready)
            self.assertEqual(preview.outbound_items[0].product_code, "M0145")
            self.assertEqual(preview.outbound_items[0].quantity, 20)
            self.assertIn("ABS banding + 24mm", preview.outbound_items[0].match_source)

    def test_back_panel_8mm_and_9mm_are_bidirectional_aliases(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalog = root / "products.xlsx"
            mappings = root / "mappings.json"
            make_catalog(catalog)
            workbook = load_workbook(catalog)
            workbook.active.append(["Panel", "M1134", "Rosales 3", "8*1220*2745mm", "启用"])
            workbook.save(catalog)
            mappings.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            for thickness in ("8", "9"):
                traveler = root / f"Work Order Traveler(PP0099-{thickness}MM).xlsx"
                make_traveler(traveler, [(f"{thickness}mm--Rosales 3", 2)])
                preview = build_preview(traveler, catalog, mappings)
                self.assertTrue(preview.ready)
                self.assertEqual(preview.outbound_items[0].product_code, "M1134")

    def test_ignored_mapping_can_be_saved_and_removed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "mappings.json"
            mappings = InventoryMappings(path)
            mappings.save_ignored("Special Material", "不再采购")
            self.assertEqual(
                InventoryMappings(path).ignored_reason("special material"),
                "不再采购",
            )
            InventoryMappings(path).remove_ignored("SPECIAL MATERIAL")
            self.assertIsNone(InventoryMappings(path).ignored_reason("Special Material"))

    def test_source_codes_are_not_treated_as_inventory_skus(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            make_catalog(config.state_dir / "inventory" / "current-products.xlsx")
            set_ignored_mapping(config, "LED", True, "用户确认全局忽略")

            result = resolve_inventory_items(
                config,
                [
                    (TravelerItem(1, "五金", "LED", 2, "F0099"), "WJ-CBD"),
                    (TravelerItem(2, "五金", "未建档五金", 1, "F0099"), "WJ-UNKNOWN"),
                ],
            )

            self.assertEqual([item["name"] for item in result["ignored"]], ["LED"])
            self.assertEqual([item["name"] for item in result["missing"]], ["未建档五金"])
            self.assertEqual(result["missing"][0]["source_code"], "WJ-UNKNOWN")
            self.assertEqual(result["outbound"], [])

    def test_ignoring_hardware_removes_existing_database_facts(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            connection = sqlite3.connect(config.workflow_database)
            connection.executemany(
                """insert into hardware_items(
                    order_id, factory_order, product_code, name, quantity, source_type, updated_at
                ) values(?,?,?,?,?,?,?)""",
                [
                    ("PP0099", "F0099", "WJ-CBT", "Adjustable shelf holder", 2, "aicnc", "now"),
                    ("PP0099", "F0099", "71T950A", "TestFullHinge", 4, "aicnc", "now"),
                    ("PP0099", "F0099", "WJ-CBD", "LED", 5, "aicnc", "now"),
                ],
            )
            connection.commit()
            connection.close()

            result = set_ignored_mapping(config, "Shelf Holder", True, "不再采购")
            self.assertEqual(result["removed_database_rows"], 1)
            connection = sqlite3.connect(config.workflow_database)
            try:
                self.assertEqual(connection.execute("select count(*) from hardware_items").fetchone()[0], 2)
            finally:
                connection.close()

            result = set_ignored_mapping(config, "Hinge", True, "不再采购")
            self.assertEqual(result["removed_database_rows"], 1)
            connection = sqlite3.connect(config.workflow_database)
            try:
                self.assertEqual(connection.execute("select count(*) from hardware_items").fetchone()[0], 1)
            finally:
                connection.close()

            result = set_ignored_mapping(config, "LED", True, "不再采购")
            self.assertEqual(result["removed_database_rows"], 1)
            connection = sqlite3.connect(config.workflow_database)
            try:
                self.assertEqual(connection.execute("select count(*) from hardware_items").fetchone()[0], 0)
            finally:
                connection.close()

    def test_manual_mapping_can_be_saved_and_replaces_ignore(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "mappings.json"
            mappings = InventoryMappings(path)
            mappings.save_ignored("Special Material", "旧品牌")
            InventoryMappings(path).save_manual("Special Material", "m1093")
            reloaded = InventoryMappings(path)
            self.assertEqual(reloaded.manual_code("special material"), "M1093")
            self.assertIsNone(reloaded.ignored_reason("Special Material"))

    def test_manual_mapping_can_be_viewed_updated_and_removed(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            make_catalog(config.state_dir / "inventory" / "current-products.xlsx")

            save_manual_mapping(config, "Hinge", "M1001")
            self.assertEqual(list_inventory_mappings(config)["manual"]["Hinge"], "M1001")
            update_manual_mapping(config, "Hinge", "Cabinet Hinge", "M1001")
            self.assertEqual(list_inventory_mappings(config)["manual"], {"Cabinet Hinge": "M1001"})
            remove_manual_mapping(config, "Cabinet Hinge")
            self.assertEqual(list_inventory_mappings(config)["manual"], {})

    def test_sync_status_changes_with_traveler_fingerprint(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            path = root / "Work Order Traveler(PP0099-KITCHEN).xlsx"
            catalog_path = root / "products.xlsx"
            mapping_path = root / "mappings.json"
            make_traveler(path, [("18mm--Plywood", 2)])
            make_catalog(catalog_path)
            mapping_path.write_text('{"manual": {}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(path, catalog_path, mapping_path)
            store = InventorySyncStore(root / "sync.json", root / "backups")
            store.save_success(preview, [{
                "remark": "PP0099",
                "saved": True,
                "documentNumber": "CK-001",
            }])
            self.assertEqual(store.status_for(parse_traveler(path))[0], "已出库")
            make_traveler(path, [("18mm--Plywood", 3)])
            reloaded = InventorySyncStore(root / "sync.json", root / "backups")
            self.assertEqual(reloaded.status_for(parse_traveler(path))[0], "需要更新")

    def test_order_material_outbound_links_only_selected_split_factories(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("CS004", validation_status="正常")
            for factory_order, factory_name in (
                ("F-KITCHEN", "CS004-KITCHEN"),
                ("F-VANITY", "CS004-vanity"),
            ):
                store.upsert_factory(
                    factory_order,
                    order_id="CS004",
                    factory_name=factory_name,
                    sales_order_name="CS004",
                    name_source="AIMES",
                    ownership_status="已确认",
                    optimized=True,
                    outbound_status="未出库",
                )
            store.connection.execute(
                "insert into material_items(order_id, material_type, color, thickness, quantity, unit, source_type, updated_at) values(?,?,?,?,?,?,?,?)",
                ("CS004", "plywood", "", "18", 2, "张", "database", "now"),
            )
            store.commit()
            store.close()

            source_item = TravelerItem(1, "板材与封边", "18mm--Plywood", 2, "CS004")
            outbound_item = OutboundItem(
                traveler_name="18mm--Plywood",
                product_code="M0001",
                product_name="18mm Plywood",
                quantity=2,
                section="板材与封边",
                match_source="test",
                document_remark="CS004",
                unit="张",
            )

            def preview(factory_order):
                traveler = TravelerData(
                    path=config.workflow_database.resolve(),
                    pp_folder="CS004",
                    order_id="CS004",
                    order_name="CS004",
                    items=[source_item],
                    zero_items=[],
                    documents={"CS004": [source_item]},
                    modified_at="2026-08-20T00:00:00",
                    fingerprint="test-fingerprint",
                )
                return InventoryPreview(
                    traveler=traveler,
                    outbound_items=[outbound_item],
                    selected_factory_orders=(factory_order,),
                    source_type="database",
                )

            sync = InventorySyncStore(
                config.state_dir / "inventory-outbound-records.json",
                config.backup_root,
            )
            sync.save_success(preview("F-KITCHEN"), [{
                "remark": "CS004",
                "saved": True,
                "documentNumber": "QTCK-001",
            }])
            links = sqlite3.connect(config.workflow_database).execute(
                "select factory_order from outbound_document_factories "
                "where document_number='QTCK-001' order by factory_order"
            ).fetchall()
            self.assertEqual(links, [("F-KITCHEN",)])

            # The second factory reuses the unchanged order-level material
            # document; it must append its explicit identity without creating
            # another inventory document.
            sync = InventorySyncStore(
                config.state_dir / "inventory-outbound-records.json",
                config.backup_root,
            )
            sync.save_success(preview("F-VANITY"), [{
                "remark": "CS004",
                "unchanged": True,
                "saved": True,
                "documentNumber": "QTCK-001",
            }])
            connection = sqlite3.connect(config.workflow_database)
            links = connection.execute(
                "select factory_order from outbound_document_factories "
                "where document_number='QTCK-001' order by factory_order"
            ).fetchall()
            self.assertEqual(links, [("F-KITCHEN",), ("F-VANITY",)])
            self.assertEqual(reconcile_outbound_statuses(config), 2)
            statuses = connection.execute(
                "select factory_order, outbound_status, outbound_document "
                "from factory_orders where order_id='CS004' order by factory_order"
            ).fetchall()
            connection.close()
            self.assertEqual(
                statuses,
                [
                    ("F-KITCHEN", "已出库", "QTCK-001"),
                    ("F-VANITY", "已出库", "QTCK-001"),
                ],
            )

            connection = sqlite3.connect(config.workflow_database)
            connection.execute("update material_items set quantity=3 where order_id='CS004'")
            connection.commit()
            connection.close()
            self.assertEqual(reconcile_outbound_statuses(config), 2)
            connection = sqlite3.connect(config.workflow_database)
            changed_statuses = connection.execute(
                "select factory_order, outbound_status, outbound_document "
                "from factory_orders where order_id='CS004' order by factory_order"
            ).fetchall()
            connection.close()
            self.assertEqual(
                changed_statuses,
                [
                    ("F-KITCHEN", "需要更新", "QTCK-001"),
                    ("F-VANITY", "需要更新", "QTCK-001"),
                ],
            )

    def test_previous_hardware_block_becoming_empty_requires_manual_void(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            path = root / "Work Order Traveler(PP0099).xlsx"
            catalog_path = root / "products.xlsx"
            mapping_path = root / "mappings.json"
            make_traveler(path, [("Hinge", 2)])
            make_catalog(catalog_path)
            mapping_path.write_text('{"manual": {"Hinge": "M1001"}, "ignored": {}}', encoding="utf-8")
            preview = build_preview(path, catalog_path, mapping_path)
            store = InventorySyncStore(root / "sync.json", root / "backups")
            plans = store.prepare_documents(preview)
            results = [
                {
                    "remark": plan["remark"],
                    "saved": True,
                    "documentNumber": f"CK-{index:03d}",
                }
                for index, plan in enumerate(plans, 1)
            ]
            store.save_success(preview, results)
            make_traveler(path, [])
            empty_preview = build_preview(path, catalog_path, mapping_path)
            with self.assertRaisesRegex(RuleError, "人工删除或作废"):
                InventorySyncStore(root / "sync.json", root / "backups").prepare_documents(empty_preview)

    def test_same_hardware_name_is_aggregated_within_factory(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "Work Order Traveler(PP0099).xlsx"
            make_traveler(path, [("Hinge", 2), ("Hinge", 3)])
            traveler = parse_traveler(path)
            hardware = traveler.documents["PP0099-KITCHEN"]
            self.assertEqual(len(hardware), 1)
            self.assertEqual(hardware[0].quantity, 5)


if __name__ == "__main__":
    unittest.main()
