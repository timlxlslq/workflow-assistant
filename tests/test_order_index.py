import os
import sqlite3
import tempfile
import unittest
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from traveler_assistant.core import Config, RuleError
from traveler_assistant.database import ensure_schema
from traveler_assistant.order_index import (
    OrderIndexStore,
    _aimes_row_issue,
    _business_aimes_message,
    _business_report_message,
    _business_validation_message,
    _clear_stale_mapping_validation_status,
    _complete_aimes_stage_durations,
    _server_change_message,
    _summarize_server_read_items,
    _is_mixed_order_folder,
    _effective_factory_candidate,
    _exact_resolve_unowned_factories,
    _factory_order_before_initial_date,
    _merge_candidate,
    _load_outbound_records,
    _merge_database_factory_candidates,
    _orders_requiring_server_scan,
    _partition_aimes_rows,
    _record_generated_material_baseline,
    _report_files,
    _replace_server_material_facts,
    reconcile_outbound_statuses,
    _server_snapshot,
    _server_folder_rename_pairs,
    _server_folders_for_sync,
    _server_data_change_message,
    _source_path_in_dashboard_scope,
    _valid_aimes_order_id,
    _visible_aimes_row,
    assign_aimes_factory_order,
    ignore_server_folder,
    scan_server_changes,
    ignore_aimes_factories,
    list_order_index,
    save_order_annotations,
    process_server_folder,
    process_server_changes,
    preview_server_changes,
    allocate_server_material,
    confirm_server_material_allocations,
    confirm_server_preview,
    record_temporary_outbound,
    restore_aimes_factories,
    restore_aimes_order_assignment,
    sync_aimes_index,
    sync_order_index,
)
from traveler_assistant.order_workflow import MaterialItem
from traveler_assistant.inventory import InventoryMappings


class OrderIndexTests(unittest.TestCase):
    def test_server_preview_requires_factory_confirmation_before_production_write(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            config.prepare_storage()
            folder = config.source_root / "PP9999"
            report = folder / "Report"
            report.mkdir(parents=True)
            from tests.test_order_workflow import make_board_material_report, make_fittings, make_materials
            make_materials(folder / "PP9999 materials.xlsx")
            make_board_material_report(report / "pp-板材清单.xlsx", factory="F100", name="PP9999-KITCHEN")
            make_fittings(report / "Fittingslist.xlsx", [("F100", 2)])

            with patch(
                "traveler_assistant.order_index.sync_order_index",
                wraps=sync_order_index,
            ) as sync, patch(
                "traveler_assistant.inventory.resolve_inventory_items",
                return_value={"missing": []},
            ):
                preview = preview_server_changes(config, [folder])
            sync.assert_called_once()
            self.assertFalse(sync.call_args.kwargs["full_refresh"])
            self.assertFalse(sync.call_args.kwargs["validate_selected_orders"])
            self.assertFalse(sync.call_args.kwargs["refresh_outbound_statuses"])
            self.assertFalse(sync.call_args.kwargs["reconcile_outbound"])
            payload = preview["server_write_preview"]
            self.assertEqual(payload["orders"][0]["order_id"], "PP9999")
            self.assertEqual(payload["orders"][0]["factories"][0]["factory_order"], "F100")
            self.assertIn("materials", payload)
            self.assertEqual(
                [item["material_type"] for item in payload["orders"][0]["materials"]],
                ["plywood", "plywood", "plywood", "panel", "panel", "edge"],
            )
            store = OrderIndexStore(config.workflow_database)
            self.assertEqual(store.connection.execute("select count(*) from orders").fetchone()[0], 0)
            store.close()

            with self.assertRaises(RuleError):
                confirm_server_preview(config, payload["token"], "PP9999", "F100")
            confirm_server_preview(
                config,
                payload["token"],
                "PP9999",
                "F100",
                confirm_write=True,
            )
            store = OrderIndexStore(config.workflow_database)
            self.assertEqual(
                store.connection.execute("select order_id from orders").fetchall(),
                [("PP9999",)],
            )
            self.assertEqual(
                store.connection.execute("select factory_order, order_id from factory_orders").fetchall(),
                [("F100", "PP9999")],
            )
            material_rows = store.connection.execute(
                "select order_id, material_type, quantity from material_items"
            ).fetchall()
            self.assertEqual(
                sorted(material_rows),
                sorted([
                    ("PP9999", "edge", 12.5),
                    ("PP9999", "panel", 4.0),
                    ("PP9999", "panel", 1.0),
                    ("PP9999", "plywood", 2.0),
                    ("PP9999", "plywood", 1.0),
                    ("PP9999", "plywood", 3.0),
                ]),
            )
            store.close()

    def test_server_scan_blocks_material_preview_until_source_file_is_fixed(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            config.prepare_storage()
            folder = config.source_root / "PP9999"
            folder.mkdir(parents=True)
            material_path = folder / "PP9999 materials.xlsx"
            material_path.write_bytes(b"not an Excel workbook")

            first_scan = scan_server_changes(config)
            self.assertTrue(any(
                issue["kind"] == "material_validation"
                for issue in first_scan["current_issues"]
            ))
            with self.assertRaises(RuleError) as raised:
                preview_server_changes(config, [folder])
            self.assertEqual(raised.exception.code, "material_validation")

            from tests.test_order_workflow import make_materials
            make_materials(material_path)
            with patch(
                "traveler_assistant.inventory.resolve_inventory_items",
                return_value={"missing": []},
            ):
                second_scan = scan_server_changes(config)
                self.assertFalse(any(
                    issue["kind"] == "material_validation"
                    for issue in second_scan["current_issues"]
                ))
                preview = preview_server_changes(config, [folder])

            self.assertEqual(
                preview["server_write_preview"]["orders"][0]["order_id"],
                "PP9999",
            )

    def test_server_preview_validates_material_before_room_allocation(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            config.prepare_storage()
            folder = config.source_root / "CS004"
            folder.mkdir(parents=True)
            material_path = folder / "CS004 material.xlsx"
            from tests.test_order_workflow import make_materials
            make_materials(material_path, order_id="CS004")
            workbook = load_workbook(material_path)
            workbook.active["I3"] = None
            workbook.save(material_path)

            with patch(
                "traveler_assistant.inventory.resolve_inventory_items",
                return_value={"missing": []},
            ):
                with self.assertRaises(RuleError) as raised:
                    preview_server_changes(config, [folder])

            self.assertEqual(raised.exception.code, "material_validation")
            self.assertIn("Color 为空", str(raised.exception))

    def test_server_material_allocation_splits_one_source_row_between_orders(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            config.prepare_storage()
            token = "a" * 32
            preview_path = config.state_dir / "server-previews" / token / "workflow.sqlite3"
            ensure_schema(preview_path)
            preview = OrderIndexStore(preview_path)
            preview.upsert_order("PP9999", source_folder="/server/mixed")
            preview.upsert_order("PP8888", source_folder="/server/mixed")
            preview.upsert_factory(
                "F100", order_id="PP9999", factory_name="PP9999 KITCHEN",
                sales_order_name="PP9999", name_source="AIMES",
            )
            preview.upsert_factory(
                "F200", order_id="PP8888", factory_name="PP8888 KITCHEN",
                sales_order_name="PP8888", name_source="AIMES",
            )
            preview.connection.execute(
                """
                insert into material_items(
                    order_id, material_type, color, thickness, quantity, unit,
                    edge, source_type, source_path, source_fingerprint, updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    "PP9999", "panel", "Frappe 3", "19.1", 13, "pcs", "",
                    "aihouse", "/server/mixed/material.xlsx", "fingerprint", "now",
                ),
            )
            preview.connection.execute(
                "insert into server_material_preview_scopes(source_folder) values(?)",
                ("/server/mixed",),
            )
            preview.commit()
            material_id = preview.connection.execute(
                "select id from material_items"
            ).fetchone()[0]
            preview.close()

            first = allocate_server_material(config, token, material_id, "PP9999", 8)
            second = allocate_server_material(config, token, material_id, "PP8888", 5)
            self.assertEqual(first["material"]["remaining_quantity"], 5)
            self.assertEqual(second["material"]["remaining_quantity"], 0)

            confirmed = confirm_server_material_allocations(
                config, token, confirm_write=True
            )
            self.assertEqual(confirmed["orders"], ["PP8888", "PP9999"])
            store = OrderIndexStore(config.workflow_database)
            rows = store.connection.execute(
                """
                select order_id, quantity
                from material_items
                order by order_id
                """
            ).fetchall()
            allocations = store.connection.execute(
                """
                select order_id, allocated_quantity
                from server_material_allocations
                order by order_id
                """
            ).fetchall()
            store.close()

        self.assertEqual(rows, [("PP8888", 5.0), ("PP9999", 8.0)])
        self.assertEqual(allocations, [("PP8888", 5.0), ("PP9999", 8.0)])

    def test_server_material_allocation_ignores_stale_sqlite_row_ids(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            config.prepare_storage()
            token = "b" * 32
            preview_path = config.state_dir / "server-previews" / token / "workflow.sqlite3"
            ensure_schema(preview_path)
            preview = OrderIndexStore(preview_path)
            preview.upsert_order("PP9999", source_folder="/server/one")
            source_path = "/server/one/material.xlsx"
            preview.connection.execute(
                """
                insert into material_items(
                    order_id, material_type, color, thickness, quantity, unit,
                    edge, source_type, source_path, source_fingerprint, updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?)
                """,
                ("PP9999", "panel", "Ivory Oak", "19.1", 3, "pcs", "",
                 "aihouse", source_path, "fingerprint", "now"),
            )
            ivory_id = preview.connection.execute(
                "select id from material_items where color='Ivory Oak'"
            ).fetchone()[0]
            preview.connection.execute(
                """
                insert into material_items(
                    order_id, material_type, color, thickness, quantity, unit,
                    edge, source_type, source_path, source_fingerprint, updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?)
                """,
                ("PP9999", "plywood", "", "5.4", 6, "pcs", "",
                 "aihouse", source_path, "fingerprint", "now"),
            )
            # This row has the old panel id but the fields of the plywood
            # fact.  The former implementation counted it against Ivory Oak.
            preview.connection.execute(
                """
                insert into server_material_allocations(
                    source_material_id, source_path, source_material_key,
                    material_type, color, thickness, unit, edge,
                    source_quantity, order_id, allocated_quantity,
                    source_fingerprint, created_at, updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (ivory_id, source_path, str(ivory_id), "plywood", "", "5.4", "pcs", "",
                 6, "PP9999", 6, "fingerprint", "now", "now"),
            )
            preview.connection.execute(
                "insert into server_material_preview_scopes(source_folder) values(?)",
                ("/server/one",),
            )
            preview.commit()
            preview.close()

            confirmed = confirm_server_material_allocations(
                config, token, confirm_write=True
            )
            self.assertEqual(confirmed["orders"], ["PP9999"])
            store = OrderIndexStore(config.workflow_database)
            rows = store.connection.execute(
                """
                select material_type, color, quantity
                from material_items
                where order_id='PP9999'
                order by material_type, color
                """
            ).fetchall()
            store.close()

        self.assertEqual(rows, [("panel", "Ivory Oak", 3.0), ("plywood", "", 6.0)])

    def test_resolved_mapping_clears_stale_order_validation_error(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order(
                "PP9999",
                order_type="temporary",
                validation_status="数据异常",
                source_folder=str(root / "source" / "PP9999"),
            )
            store.connection.execute(
                "update orders set validation_message = ? where order_id = ?",
                ("订单存在未完成商品 SKU 处理：LED。", "PP9999"),
            )
            cleared = _clear_stale_mapping_validation_status(
                store,
                {"PP9999"},
                set(),
                "2026-08-19T10:00:00",
            )
            row = store.connection.execute(
                "select validation_status, validation_message from orders where order_id = ?",
                ("PP9999",),
            ).fetchone()
            store.close()

        self.assertEqual(cleared, 1)
        self.assertEqual(row, ("正常", ""))

    def test_unresolved_mapping_keeps_order_validation_error(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order(
                "PP9999",
                order_type="temporary",
                validation_status="数据异常",
                source_folder=str(root / "source" / "PP9999"),
            )
            store.connection.execute(
                "update orders set validation_message = ? where order_id = ?",
                ("订单存在未完成商品 SKU 处理：LED。", "PP9999"),
            )
            cleared = _clear_stale_mapping_validation_status(
                store,
                {"PP9999"},
                {"hardware_mapping:PP9999:/server/Fittingslist.xlsx"},
                "2026-08-19T10:00:00",
            )
            row = store.connection.execute(
                "select validation_status, validation_message from orders where order_id = ?",
                ("PP9999",),
            ).fetchone()
            store.close()

        self.assertEqual(cleared, 0)
        self.assertEqual(row, ("数据异常", "订单存在未完成商品 SKU 处理：LED。"))

    def test_aimes_stage_durations_exclude_aggregate_and_account_for_backend_overhead(self):
        stages = _complete_aimes_stage_durations([
            {"stage": "login", "label": "登录 AIMES", "duration_seconds": 1.2},
            {"stage": "attempt", "label": "获取 AIMES 数据成功，总计用时", "duration_seconds": 9.9},
        ], 2.0)

        self.assertEqual(
            [item["label"] for item in stages],
            ["登录 AIMES", "后台准备与收尾"],
        )
        self.assertAlmostEqual(
            sum(float(item["duration_seconds"]) for item in stages),
            2.0,
            places=6,
        )

    def test_order_annotations_store_note_and_nonconsecutive_installation_days(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state")
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("PP9999", source_folder="/server/PP9999")
            store.upsert_factory(
                "F999",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                name_source="AIMES",
                ownership_status="已确认",
            )
            store.commit()
            store.close()

            saved = save_order_annotations(
                config,
                "PP9999",
                user_note="客户要求安装前确认台面颜色",
                planned_days=[
                    {"date": "2026-07-08", "installer": "安装组 A"},
                    {"date": "2026-07-11", "installer": "安装组 B"},
                ],
                actual_days=[
                    {"date": "2026-07-08", "installer": "安装组 A"},
                    {"date": "2026-07-11", "installer": "安装组 A"},
                ],
            )

            self.assertTrue(saved["saved"])
            row = saved["order"]
            self.assertEqual(row["user_note"], "客户要求安装前确认台面颜色")
            self.assertEqual(row["installation"]["planned"]["start_date"], "2026-07-08")
            self.assertEqual(row["installation"]["planned"]["end_date"], "2026-07-11")
            self.assertEqual(row["installation"]["planned"]["day_count"], 2)
            self.assertEqual(row["installation"]["actual"]["days"][1]["installer"], "安装组 A")

            reopened = OrderIndexStore(config.workflow_database)
            reopened.upsert_order("PP9999", server_seen="2026-08-18T12:00:00")
            persisted = next(item for item in reopened.summaries() if item["order_id"] == "PP9999")
            reopened.close()
            self.assertEqual(persisted["user_note"], "客户要求安装前确认台面颜色")
            self.assertEqual(persisted["installation"]["actual"]["day_count"], 2)

    def test_order_annotations_allow_missing_installer_but_reject_duplicate_dates(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state")
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("PP9999")
            store.commit()
            saved = store.save_order_annotations(
                "PP9999",
                user_note="",
                planned_days=[{"date": "2026-07-08", "installer": ""}],
                actual_days=[],
            )
            self.assertTrue(saved["saved"])
            self.assertEqual(
                saved["order"]["installation"]["planned"]["days"],
                [{"date": "2026-07-08", "installer": ""}],
            )
            with self.assertRaisesRegex(ValueError, "日期重复"):
                store.save_order_annotations(
                    "PP9999",
                    user_note="",
                    planned_days=[
                        {"date": "2026-07-08", "installer": "安装组"},
                        {"date": "2026-07-08", "installer": "安装组"},
                    ],
                    actual_days=[],
                )
            store.close()

    def test_server_folder_rename_requires_unique_identical_report_signature(self):
        old = "/server/PP0099"
        new = "/server/PP0099-renamed"
        previous = {
            old: {"source_folder": old, "kind": "folder", "order_id": "PP0099", "modified_at": 10, "size": 1},
            f"{old}/material.xlsx": {"source_folder": old, "kind": "material", "order_id": "PP0099", "modified_at": 20, "size": 200},
        }
        current = {
            new: {"source_folder": new, "kind": "folder", "order_id": "PP0099", "modified_at": 99, "size": 1},
            f"{new}/material.xlsx": {"source_folder": new, "kind": "material", "order_id": "PP0099", "modified_at": 20, "size": 200},
        }
        self.assertEqual(_server_folder_rename_pairs(previous, current), [(old, new)])
        current[f"{new}/extra.xlsx"] = {"source_folder": new, "kind": "material", "order_id": "PP0099", "modified_at": 20, "size": 200}
        self.assertEqual(_server_folder_rename_pairs(previous, current), [])
    def test_server_material_replacement_collapses_old_source_path_rows(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            old_path = root / "Downloads" / "PP0065 materials.xlsx"
            current_path = root / "server" / "PP0065 materials.xlsx"
            current_path.parent.mkdir(parents=True)
            current_path.write_bytes(b"same workbook content")
            store.connection.execute(
                """insert into material_items(
                    order_id, material_type, color, thickness, quantity, unit,
                    source_type, source_path, updated_at
                ) values(?,?,?,?,?,?,?,?,?)""",
                ("PP0065", "panel", "Rosales 3", "19.1", 17, "pcs",
                 "aihouse", str(old_path), "before"),
            )
            _replace_server_material_facts(
                store,
                "PP0065",
                current_path,
                [MaterialItem("panel", 19.1, "Rosales 3", 17)],
                {"Rosales 3": 239.56},
                InventoryMappings(config.workflow_database),
                "after",
            )
            rows = store.connection.execute(
                "select source_path, source_fingerprint, count(*) from material_items where order_id='PP0065' group by source_path, source_fingerprint"
            ).fetchall()
            store.close()

            self.assertEqual(len(rows), 1)
            self.assertEqual({row[0] for row in rows}, {str(current_path)})
            self.assertEqual(rows[0][2], 2)
            self.assertTrue(rows[0][1])

    def test_prepared_sync_can_resolve_material_mappings_before_fittings_import_path(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            material = config.source_root / "PP9999" / "PP9999 materials.xlsx"
            material.parent.mkdir(parents=True)
            material.write_bytes(b"placeholder")
            config.prepare_storage()

            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[]), \
                 patch(
                     "traveler_assistant.order_workflow.parse_order_materials",
                     return_value=(
                         "Sheet1",
                         [MaterialItem("panel", 19.1, "Test Oak", 1)],
                         {"Test Oak": 1},
                     ),
                 ), \
                 patch(
                     "traveler_assistant.inventory.resolve_inventory_items",
                     return_value={"missing": []},
                 ):
                result = sync_order_index(config)

            self.assertEqual(result["orders"], [])

    def test_server_read_trace_is_grouped_by_folder_and_file_kind(self):
        items = [
            ("/Volumes/server/CUT TO SIZE/cs001", "/Volumes/server/CUT TO SIZE/cs001", "folder"),
            ("/Volumes/server/CUT TO SIZE/cs001/CS001 materials.xlsx", "/Volumes/server/CUT TO SIZE/cs001", "material"),
            ("/Volumes/server/CUT TO SIZE/cs001/Report/pp-板材清单-new.xlsx", "/Volumes/server/CUT TO SIZE/cs001", "board"),
            ("/Volumes/server/CUT TO SIZE/cs001/Report/Fittingslist.xlsx", "/Volumes/server/CUT TO SIZE/cs001", "fittings"),
            ("/Volumes/server/Optimized Orders/PP0035", "/Volumes/server/Optimized Orders/PP0035", "folder"),
            ("/Volumes/server/Optimized Orders/PP0035/PP0035 materials.xlsx", "/Volumes/server/Optimized Orders/PP0035", "material"),
        ]

        summary = _summarize_server_read_items(items)

        self.assertIn("读取 2 个文件夹及 4 个相关文件", summary)
        self.assertIn("cs001 文件夹（下属：1 个 material 文件、1 个板材清单、1 个五金清单）", summary)
        self.assertIn("PP0035 文件夹（下属：1 个 material 文件）", summary)
        self.assertNotIn("/Volumes/server/CUT TO SIZE", summary)

    def test_server_read_trace_limits_folder_examples(self):
        items = [
            (f"/server/CS{i:03d}", f"/server/CS{i:03d}", "folder")
            for i in range(1, 8)
        ]

        summary = _summarize_server_read_items(items)

        self.assertIn("读取 7 个文件夹及 0 个相关文件", summary)
        self.assertIn("另有 1 个文件夹已省略", summary)
        self.assertNotIn("CS007 文件夹", summary)

    def test_server_scan_covers_owned_and_cut_to_size_roots(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "server" / "Optimized Orders")
            owned = config.source_root / "PP9999"
            cut_to_size = config.source_root.parent / "CUT TO SIZE" / "CS999"
            owned.mkdir(parents=True)
            cut_to_size.mkdir(parents=True)
            (owned / "PP9999 materials.xlsx").write_bytes(b"material")
            (cut_to_size / "CS999 materials.xlsx").write_bytes(b"material")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100", order_id="PP9999", factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999", split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.upsert_aimes_factory(
                "F101", order_id="CS999", factory_name="CS999-KITCHEN",
                sales_order_name="CS999", split_time="2026-08-10T08:31:00",
                seen_at="2026-08-10T08:31:00",
            )
            store.commit()
            root_path, snapshot = _server_snapshot(config, store)
            store.close()

            self.assertEqual(root_path, config.source_root)
            folder_paths = {
                path for path, item in snapshot.items() if item["kind"] == "folder"
            }
            self.assertIn(str(owned), folder_paths)
            self.assertIn(str(cut_to_size), folder_paths)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            _, folders = _server_folders_for_sync(config, None, store=store)
            store.close()
            self.assertEqual({str(item) for item in folders}, {str(owned), str(cut_to_size)})

    def test_successful_cut_to_size_preview_marks_indexed_factory_optimized(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "server" / "Optimized Orders")
            folder = config.source_root.parent / "CUT TO SIZE" / "CS005"
            report = folder / "Report" / "板材清单.xlsx"
            report.parent.mkdir(parents=True)
            report.write_bytes(b"placeholder")
            aimes_row = {
                "factory_order": "F100",
                "factory_name": "CS005-KITCHEN",
                "sales_order_name": "CS005",
                "split_time": "2026-08-10T08:30:00",
            }
            preview = SimpleNamespace(factories=[])
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[aimes_row]), \
                 patch("traveler_assistant.order_workflow.parse_board_identity", return_value=("F100", "CS005-KITCHEN")), \
                 patch("traveler_assistant.order_workflow.preview_order", return_value=preview):
                result = sync_order_index(config)

            self.assertEqual(result["orders"][0]["order_id"], "CS005")
            self.assertEqual(result["orders"][0]["optimized_count"], 1)
            self.assertEqual(result["orders"][0]["stage"], "已优化")

    def test_cut_to_size_optimization_artifact_marks_status_when_material_is_absent(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "server" / "Optimized Orders")
            folder = config.source_root.parent / "CUT TO SIZE" / "CS005"
            report = folder / "Report" / "pp-板材清单.xlsx"
            artifact = folder / "New Nesting" / "Optimize file" / "Optimize file.xml"
            report.parent.mkdir(parents=True)
            artifact.parent.mkdir(parents=True)
            report.write_bytes(b"placeholder")
            artifact.write_text("<optimized />", encoding="utf-8")
            aimes_row = {
                "factory_order": "F100",
                "factory_name": "CS005-KITCHEN",
                "sales_order_name": "CS005",
                "split_time": "2026-08-10T08:30:00",
            }
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[aimes_row]), \
                 patch("traveler_assistant.order_workflow.parse_board_identity", return_value=("F100", "CS005-KITCHEN")), \
                 patch("traveler_assistant.order_workflow.preview_order", side_effect=RuleError("missing_material", "缺少 material")):
                result = sync_order_index(config)

            order = result["orders"][0]
            self.assertEqual(order["stage"], "已优化")
            self.assertEqual(order["optimized_count"], 1)
            self.assertEqual(order["validation_status"], "待校验")
            self.assertEqual(order["material_status"], "待校验")

    def test_exact_standard_order_folder_wins_over_mixed_factory_report_folder(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "server" / "Optimized Orders")
            exact_folder = config.source_root / "PP0035"
            mixed_folder = config.source_root / "PP0035-2"
            (mixed_folder / "Report").mkdir(parents=True)
            (mixed_folder / "Report" / "板材清单.xlsx").write_bytes(b"board")
            exact_folder.mkdir(parents=True)
            rows = [
                {
                    "factory_order": "F100",
                    "factory_name": "PP0035-OFFICE",
                    "sales_order_name": "PP0035",
                    "split_time": "2026-08-10T08:30:00",
                },
                {
                    "factory_order": "F101",
                    "factory_name": "PP0035-2-MASTER",
                    "sales_order_name": "PP0035-2",
                    "split_time": "2026-08-10T08:31:00",
                },
            ]
            preview = SimpleNamespace(factories=[])
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=rows), \
                 patch("traveler_assistant.order_workflow.parse_board_identity", return_value=("F100", "PP0035-OFFICE")), \
                 patch("traveler_assistant.order_workflow.preview_order", return_value=preview):
                sync_order_index(config)

            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            source_folders = {
                row[0]: row[1]
                for row in store.connection.execute(
                    "select order_id, source_folder from orders where order_id in ('PP0035', 'PP0035-2')"
                ).fetchall()
            }
            store.close()
            self.assertEqual(source_folders["PP0035"], str(exact_folder))
            self.assertEqual(source_folders["PP0035-2"], str(mixed_folder))

    def test_cut_to_size_fittings_are_not_persisted_as_hardware(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "server" / "Optimized Orders")
            folder = config.source_root.parent / "CUT TO SIZE" / "CS999"
            folder.mkdir(parents=True)
            (folder / "CS999 materials.xlsx").write_bytes(b"material")
            (folder / "Fittingslist.xlsx").write_bytes(b"fittings")
            config.prepare_storage()
            fitting = SimpleNamespace(code="M0001", name="Test hardware", size="", unit="Piece", quantity=1)
            preview = SimpleNamespace(factories=[])
            row = {
                "factory_order": "F999",
                "factory_name": "CS999-KITCHEN",
                "sales_order_name": "CS999",
                "split_time": "2026-08-10T08:30:00",
            }
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[row]), \
                 patch("traveler_assistant.order_workflow.parse_order_materials", return_value=("Sheet1", [], {})), \
                 patch("traveler_assistant.order_workflow.parse_fittings_groups", return_value=[("F999", [fitting])]), \
                 patch("traveler_assistant.order_workflow.preview_order", return_value=preview):
                sync_order_index(config)

            connection = __import__("sqlite3").connect(config.workflow_database)
            count = connection.execute(
                "select count(*) from hardware_items where order_id='CS999'"
            ).fetchone()[0]
            connection.close()
            self.assertEqual(count, 0)

    def test_incremental_sync_reuses_unchanged_server_report_and_rechecks_changes(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            folder = config.source_root / "PP9999"
            report = folder / "Report" / "板材清单.xlsx"
            report.parent.mkdir(parents=True)
            report.write_bytes(b"placeholder")
            aimes_row = {
                "factory_order": "F100",
                "factory_name": "PP9999-KITCHEN",
                "sales_order_name": "PP9999",
                "split_time": "2026-08-10T08:30:00",
            }
            preview = SimpleNamespace(factories=[])
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[aimes_row]), \
                 patch("traveler_assistant.order_workflow.parse_board_identity", return_value=("F100", "PP9999-KITCHEN")) as parse, \
                 patch("traveler_assistant.order_workflow.preview_order", return_value=preview) as preview_order:
                first = sync_order_index(config)
                first_parse_count = parse.call_count
                second = sync_order_index(config)

                self.assertGreaterEqual(first_parse_count, 1)
                self.assertEqual(parse.call_count, first_parse_count)
                self.assertEqual(preview_order.call_count, 1)
                self.assertEqual(first["index_stats"]["parsed_report_count"], 1)
                self.assertEqual(second["index_stats"]["reused_report_count"], 1)
                self.assertEqual(second["index_stats"]["parsed_report_count"], 0)
                self.assertEqual(second["index_stats"]["validated_order_count"], 0)
                self.assertEqual(second["orders"][0]["optimized_count"], 1)

                report.write_bytes(b"changed")
                third = sync_order_index(config)

            self.assertGreater(parse.call_count, first_parse_count)
            self.assertEqual(preview_order.call_count, 2)
            self.assertEqual(third["index_stats"]["parsed_report_count"], 1)
            self.assertEqual(third["index_stats"]["validated_order_count"], 1)

    def test_server_snapshot_uses_bounded_read_only_workers_and_preserves_records(self):
        class TrackingExecutor:
            def __init__(self, max_workers):
                self.max_workers = max_workers

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc_value, traceback):
                return False

            def map(self, function, folders):
                return [function(folder) for folder in folders]

        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            for name in ("PP0035", "PP0046"):
                folder = config.source_root / name
                folder.mkdir(parents=True)
                (folder / f"{name} materials.xlsx").write_bytes(b"material")

            executors = []

            def make_executor(max_workers):
                executor = TrackingExecutor(max_workers)
                executors.append(executor)
                return executor

            with patch(
                "traveler_assistant.order_index._orders_requiring_server_scan",
                return_value={"PP0035", "PP0046"},
            ), patch(
                "traveler_assistant.order_index.ThreadPoolExecutor",
                side_effect=make_executor,
            ):
                store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
                root_path, snapshot = _server_snapshot(config, store)
                store.close()

            self.assertEqual(root_path, config.source_root)
            self.assertEqual(len(executors), 1)
            self.assertEqual(executors[0].max_workers, 2)
            self.assertEqual(
                sorted(Path(path).name for path, item in snapshot.items() if item["kind"] == "folder"),
                ["PP0035", "PP0046"],
            )
            self.assertEqual(
                sorted(Path(path).name for path, item in snapshot.items() if item["kind"] != "folder"),
                ["PP0035 materials.xlsx", "PP0046 materials.xlsx"],
            )

    def test_sync_index_reuses_scan_snapshot_and_reports_phase_durations(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            folder = config.source_root / "PP9999"
            report = folder / "Report" / "板材清单.xlsx"
            report.parent.mkdir(parents=True)
            report.write_bytes(b"placeholder")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.commit()
            store.close()

            scan = scan_server_changes(config)
            snapshot_path = Path(scan["server"]["snapshot_path"])
            preview = SimpleNamespace(factories=[])
            with patch("traveler_assistant.order_index.load_aimes_order_cache", return_value=[]), \
                 patch("traveler_assistant.order_index._server_folders_for_sync", side_effect=AssertionError("snapshot should avoid folder discovery")), \
                 patch("traveler_assistant.order_workflow.parse_board_identity", return_value=("F100", "PP9999-KITCHEN")), \
                 patch("traveler_assistant.order_workflow.preview_order", return_value=preview):
                result = sync_order_index(config, server_snapshot_path=snapshot_path)

            self.assertTrue(result["index_stats"]["server_snapshot_reused"])
            self.assertEqual(result["index_stats"]["server_snapshot_entry_count"], 2)
            self.assertIn("server_metadata_and_report_sync", result["index_stats"]["phase_durations"])
            self.assertIn("索引阶段耗时：", result["operation_trace"]["sync"][-1])

    def test_prebaseline_temporary_folder_is_excluded_and_stale_pending_cleared(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                server_scan_baseline_at="2099-01-01T00:00:00-08:00",
            )
            folder = config.source_root / "pp001"
            folder.mkdir(parents=True)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_source_file(
                folder,
                source_folder=folder,
                kind="folder",
                changed_at="2026-08-12T09:00:00",
            )
            store.upsert_active_issue(
                issue_key=f"temporary_processing:{folder}",
                kind="temporary_processing",
                path=str(folder),
                message="旧的临时文件夹问题",
            )
            store.commit()
            store.close()

            result = scan_server_changes(config)

            self.assertEqual(result["server"]["changes"], [])
            self.assertEqual(result["current_issues"], [])
            reopened = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            self.assertEqual(
                reopened.connection.execute("select count(*) from source_files").fetchone()[0],
                0,
            )
            self.assertEqual(reopened.active_issues(), [])
            reopened.close()

    def test_named_mixed_folder_is_not_marked_as_temporary(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "CS003 PP0047"
            folder.mkdir(parents=True)

            self.assertTrue(_is_mixed_order_folder(folder))
            changes = scan_server_changes(config)["server"]["changes"]
            folder_change = next(item for item in changes if item["path"] == str(folder))
            self.assertFalse(folder_change["manual_only"])
            self.assertTrue(folder_change["mixed_order"])
            self.assertEqual(folder_change["order_id"], "CS003、PP0047")
            self.assertIn("混单文件夹", folder_change["message"])

    def test_fully_shipped_mixed_folder_is_skipped_until_aimes_reopens_an_order(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "CS003 PP0047"
            folder.mkdir(parents=True)
            store = OrderIndexStore(config.workflow_database)
            store.upsert_aimes_factory(
                "F100", order_id="CS003", factory_name="CS003 KITCHEN",
                sales_order_name="CS003", split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.upsert_aimes_factory(
                "F101", order_id="PP0047", factory_name="PP0047 KITCHEN",
                sales_order_name="PP0047", split_time="2026-08-10T08:31:00",
                seen_at="2026-08-10T08:31:00",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[
                    {"order_id": "CS003", "remark": "CS003 KITCHEN", "status": "已出库", "document_number": "OUT-CS003"},
                    {"order_id": "PP0047", "remark": "PP0047 KITCHEN", "status": "已出库", "document_number": "OUT-PP0047"},
                ],
            ):
                skipped = scan_server_changes(config)["server"]
                self.assertFalse(skipped["changed"])
                self.assertEqual(skipped["scan_stats"]["order_folder_count"], 0)

                reopened = OrderIndexStore(config.workflow_database)
                reopened.upsert_aimes_factory(
                    "F102", order_id="PP0047", factory_name="PP0047 VANITY",
                    sales_order_name="PP0047", split_time="2026-08-17T08:31:00",
                    seen_at="2026-08-17T08:31:00",
                )
                reopened.commit()
                reopened.close()
                reopened_scan = scan_server_changes(config)["server"]

        self.assertTrue(reopened_scan["changed"])
        self.assertEqual(reopened_scan["scan_stats"]["order_folder_count"], 1)
        self.assertEqual(reopened_scan["changes"][0]["path"], str(folder))

    def test_reportless_mixed_folder_requires_review_and_can_be_ignored_for_one_month(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "CS003 PP0047"
            folder.mkdir(parents=True)

            first = scan_server_changes(config)
            self.assertTrue(any(
                item["change_type"] == "missing_report" and item["path"] == str(folder)
                for item in first["server"]["changes"]
            ))
            issue = next(item for item in first["current_issues"] if item["kind"] == "server_missing_report")
            self.assertIn("缺少可识别的报表", issue["message"])

            ignored = ignore_server_folder(config, folder)
            self.assertEqual(ignored["pending_server_changes"], [])
            self.assertEqual(ignored["current_issues"], [])

            unchanged = scan_server_changes(config)
            self.assertEqual(unchanged["server"]["changes"], [])
            self.assertEqual(unchanged["current_issues"], [])

            folder_stat = folder.stat()
            os.utime(folder, ns=(folder_stat.st_atime_ns, folder_stat.st_mtime_ns + 1_000_000_000))
            changed = scan_server_changes(config)
            self.assertEqual(changed["server"]["changes"], [])
            self.assertEqual(changed["current_issues"], [])

            report = folder / "Report" / "pp-板材清单.xlsx"
            from tests.test_order_workflow import make_board_material_report
            make_board_material_report(report, factory="F100", name="PP9999-KITCHEN")
            report_changed = scan_server_changes(config)
            self.assertTrue(any(item["path"] == str(report) for item in report_changed["server"]["changes"]))
            self.assertFalse(any(item["kind"] == "server_missing_report" for item in report_changed["current_issues"]))

    def test_reportless_mixed_folder_becomes_permanently_ignored_after_watch_window(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "CS003 PP0047"
            folder.mkdir(parents=True)
            scan_server_changes(config)
            ignore_server_folder(config, folder)

            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.connection.execute(
                "update ignored_server_folders set watch_until = ? where path = ?",
                ("2000-01-01T00:00:00", str(folder)),
            )
            store.commit()
            store.close()

            expired = scan_server_changes(config)
            self.assertEqual(expired["server"]["changes"], [])
            self.assertEqual(expired["current_issues"], [])
            reopened = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            permanent = reopened.connection.execute(
                "select permanent from ignored_server_folders where path = ?", (str(folder),)
            ).fetchone()[0]
            reopened.close()
            self.assertEqual(permanent, 1)

    def test_failed_temporary_processing_remains_in_pending_server_changes(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "temporary-production"
            folder.mkdir(parents=True)
            (folder / "material.xlsx").write_bytes(b"not-a-workbook")

            result = process_server_changes(config)

            self.assertTrue(result["temporary_processing"]["failed"])
            pending_paths = {item["path"] for item in result["pending_server_changes"]}
            self.assertIn(str(folder), pending_paths)
            self.assertIn(str(folder / "material.xlsx"), pending_paths)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            self.assertTrue(any(item["kind"] == "temporary_processing" for item in store.active_issues()))
            self.assertEqual(
                store.connection.execute("select count(*) from source_files").fetchone()[0],
                2,
            )
            store.close()

            repeated = scan_server_changes(config)["server"]
            self.assertTrue(repeated["changed"])
            self.assertTrue(all(
                item["change_type"] == "processing_failed"
                for item in repeated["changes"]
            ))
            self.assertFalse(any(
                "首次发现" in item["message"] or "新增" in item["message"]
                for item in repeated["changes"]
            ))

    def test_temporary_fittings_report_is_deferred_until_user_approves_processing(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(state_dir=root / "state", source_root=root / "source")
            folder = config.source_root / "temporary-production"
            fittings = folder / "Report" / "Fittingslist.xlsx"
            fittings.parent.mkdir(parents=True)
            fittings.write_bytes(b"not-a-workbook")

            result = sync_order_index(config)

            self.assertFalse(any(
                issue["kind"] == "report_error" and issue["path"] == str(fittings)
                for issue in result["current_issues"]
            ))

    def test_manual_temporary_outbound_records_server_baseline_case_insensitively(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "inserthood cabinet OLD CNC"
            report = folder / "Report"
            report.mkdir(parents=True)
            (folder / "INSERTHOOD CABINET OLD CNC materials.xlsx").write_bytes(b"material")
            (report / "Fittingslist.xlsx").write_bytes(b"fittings")
            (report / "pp-板材清单-new.xlsx").write_bytes(b"board")
            traveler = config.order_root / "INSERTHOOD CABINET OLD CNC" / (
                "Work Order Traveler(INSERTHOOD CABINET OLD CNC).xlsx"
            )

            record_temporary_outbound(
                config,
                traveler,
                {"saved": True, "results": [{"documentNumber": "QTCK-001"}]},
            )

            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            rows = store.connection.execute(
                "select source_folder, outbound_status, outbound_document from temporary_orders"
            ).fetchall()
            source_paths = {
                row[0]
                for row in store.connection.execute("select path from source_files").fetchall()
            }
            store.close()
            self.assertEqual(rows, [(str(folder), "已出库", "QTCK-001")])
            self.assertEqual(len(source_paths), 4)
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

    def test_shipped_temporary_folder_is_skipped_without_report_rescan(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "temporary-production"
            report = folder / "Report" / "pp-板材清单.xlsx"
            report.parent.mkdir(parents=True)
            report.write_bytes(b"board")
            traveler = config.order_root / "TEMPORARY-PRODUCTION" / "Work Order Traveler(TEMPORARY-PRODUCTION).xlsx"
            record_temporary_outbound(
                config,
                traveler,
                {"saved": True, "results": [{"documentNumber": "OUT-001"}]},
            )

            with patch("traveler_assistant.order_index._report_files", wraps=_report_files) as report_files:
                result = scan_server_changes(config)

            self.assertFalse(result["server"]["changed"])
            self.assertFalse(any(
                call.args and str(folder) == str(call.args[0])
                for call in report_files.call_args_list
            ))

    def test_old_temporary_folder_is_filtered_before_report_rescan(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                server_scan_baseline_at="2099-01-01T00:00:00-08:00",
            )
            folder = config.source_root / "old-temporary"
            report = folder / "Report" / "pp-板材清单.xlsx"
            report.parent.mkdir(parents=True)
            report.write_bytes(b"board")

            with patch("traveler_assistant.order_index._report_files", wraps=_report_files) as report_files:
                result = scan_server_changes(config)

            self.assertFalse(result["server"]["changed"])
            self.assertFalse(any(
                call.args and str(folder) == str(call.args[0])
                for call in report_files.call_args_list
            ))

    def test_temporary_processing_generates_material_traveler_and_outbounds(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "temporary-production"
            report = folder / "Report" / "pp-板材清单.xlsx"
            fittings = folder / "Report" / "Fittingslist.xlsx"
            from tests.test_order_workflow import make_board_material_report, make_fittings
            make_board_material_report(report, factory="F100", name="PP9999-KITCHEN")
            make_fittings(fittings, [("F100", 2)])

            with patch(
                "traveler_assistant.inventory.run_jdy",
                return_value={"saved": True, "results": [{"documentNumber": "OUT-1"}]},
            ) as outbound:
                result = process_server_changes(config)

            self.assertEqual(result["temporary_processing"]["failed"], [])
            self.assertEqual(result["temporary_processing"]["succeeded"][0]["order_ids"], ["PP9999"])
            self.assertTrue((folder / "PP9999 materials.xlsx").is_file())
            self.assertTrue((root / "orders" / "PP9999" / "Work Order Traveler(PP9999).xlsx").is_file())
            outbound.assert_called_once()
            self.assertEqual(result["pending_server_changes"], [])

    def test_temporary_processing_can_skip_hardware_in_traveler_and_outbound(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "temporary-production"
            report = folder / "Report" / "pp-板材清单.xlsx"
            fittings = folder / "Report" / "Fittingslist.xlsx"
            from tests.test_order_workflow import make_board_material_report, make_fittings
            make_board_material_report(report, factory="F100", name="PP9999-KITCHEN")
            make_fittings(fittings, [("F100", 2)])

            with patch(
                "traveler_assistant.inventory.run_jdy",
                return_value={"saved": True, "results": [{"documentNumber": "OUT-2"}]},
            ) as outbound:
                result = process_server_changes(config, include_hardware=False)

            self.assertEqual(result["temporary_processing"]["failed"], [])
            traveler = root / "orders" / "PP9999" / "Work Order Traveler(PP9999).xlsx"
            from openpyxl import load_workbook
            picking = load_workbook(traveler, data_only=False)["Picking List"]
            self.assertFalse(any(
                picking.cell(row, 3).value == "Hinge"
                for row in range(1, picking.max_row + 1)
            ))
            outbound.assert_called_once()

    def test_temporary_folder_without_aimes_identity_uses_folder_name_everywhere(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "B12"
            report = folder / "Report" / "pp-板材清单.xlsx"
            fittings = folder / "Report" / "Fittingslist.xlsx"
            from tests.test_order_workflow import make_board_material_report, make_fittings
            make_board_material_report(report, factory="", name="")
            make_fittings(fittings, [("", 2)])

            with patch(
                "traveler_assistant.inventory.run_jdy",
                return_value={"saved": True, "results": [{"documentNumber": "OUT-B12"}]},
            ) as outbound:
                result = process_server_changes(config)

            self.assertEqual(result["temporary_processing"]["failed"], [])
            self.assertEqual(result["temporary_processing"]["succeeded"][0]["order_ids"], ["B12"])
            traveler = root / "orders" / "B12" / "Work Order Traveler(B12).xlsx"
            self.assertTrue(traveler.is_file())
            workbook = __import__("openpyxl").load_workbook(traveler, data_only=False)
            self.assertEqual(workbook["WorkOrderTraveler"]["B5"].value, "B12")
            self.assertEqual(workbook["Usage List"]["B1"].value, "B12")
            picking = workbook["Picking List"]
            self.assertTrue(any(
                picking.cell(row, 4).value == "B12"
                for row in range(1, picking.max_row + 1)
            ))
            outbound.assert_called_once()
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            record = store.temporary_order(str(folder))
            self.assertEqual(record["folder_name"], "B12")
            self.assertEqual(record["processing_status"], "Traveler 已生成")
            self.assertEqual(record["outbound_status"], "已出库")
            self.assertEqual(record["outbound_document"], "OUT-B12")
            store.close()

    def test_temporary_outbound_is_not_repeated_when_folder_content_is_unchanged(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "B12"
            report = folder / "Report" / "pp-板材清单.xlsx"
            from tests.test_order_workflow import make_board_material_report
            make_board_material_report(report, factory="", name="")

            with patch(
                "traveler_assistant.inventory.run_jdy",
                return_value={"saved": True, "results": [{"documentNumber": "OUT-B12"}]},
            ) as outbound:
                first = process_server_changes(config)
                second = process_server_changes(config)

            self.assertEqual(first["temporary_processing"]["failed"], [])
            self.assertEqual(second["temporary_processing"]["failed"], [])
            self.assertEqual(second["temporary_processing"]["succeeded"], [])
            outbound.assert_called_once()

    def test_failed_outbound_reuses_unchanged_generated_traveler_on_retry(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "B12"
            report = folder / "Report" / "pp-板材清单.xlsx"
            from tests.test_order_workflow import make_board_material_report
            make_board_material_report(report, factory="", name="")

            from traveler_assistant.order_workflow import update_order_traveler as real_update
            with patch(
                "traveler_assistant.inventory.run_jdy",
                side_effect=[
                    {"saved": False, "results": []},
                    {"saved": True, "results": [{"documentNumber": "OUT-B12-RETRY"}]},
                ],
            ) as outbound, patch(
                "traveler_assistant.order_workflow.update_order_traveler",
                wraps=real_update,
            ) as updater:
                first = process_server_changes(config)
                second = process_server_changes(config)

            self.assertTrue(first["temporary_processing"]["failed"])
            self.assertEqual(second["temporary_processing"]["failed"], [])
            processed = second["temporary_processing"]["succeeded"][0]["processed"]
            self.assertEqual(processed[0]["traveler_action"], "reused")
            updater.assert_not_called()
            self.assertEqual(outbound.call_count, 2)

    def test_temporary_folder_uses_unique_aimes_review_match_when_available(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "source",
                order_root=root / "orders",
            )
            folder = config.source_root / "B12"
            report = folder / "Report" / "pp-板材清单.xlsx"
            fittings = folder / "Report" / "Fittingslist.xlsx"
            from tests.test_order_workflow import make_board_material_report, make_fittings
            make_board_material_report(report, factory="F100", name="B12")
            make_fittings(fittings, [("F100", 2)])
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.replace_aimes_review_rows([{
                "ignore_key": "factory:F100",
                "factory_order": "F100",
                "factory_name": "B12",
                "sales_order_name": "临时 B12",
                "reason": "销售单名称不符合标准规则",
            }])
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.inventory.run_jdy",
                return_value={"saved": True, "results": [{"documentNumber": "OUT-B12-AIMES"}]},
            ) as outbound:
                result = process_server_changes(config)

            self.assertEqual(result["temporary_processing"]["failed"], [])
            self.assertEqual(result["temporary_processing"]["succeeded"][0]["order_ids"], ["B12"])
            workbook = __import__("openpyxl").load_workbook(
                root / "orders" / "B12" / "Work Order Traveler(B12).xlsx",
                data_only=False,
            )
            self.assertTrue(any(
                workbook["Picking List"].cell(row, 4).value == "B12"
                for row in range(1, workbook["Picking List"].max_row + 1)
            ))
            outbound.assert_called_once()
    def test_factory_order_initial_date_cutoff_uses_embedded_date(self):
        self.assertTrue(_factory_order_before_initial_date("F2605260119", "", "2026-07-22"))
        self.assertFalse(_factory_order_before_initial_date("F2608010001", "", "2026-07-22"))
        self.assertTrue(_factory_order_before_initial_date("F100", "2026-07-21 10:00:00", "2026-07-22"))
        self.assertFalse(_factory_order_before_initial_date("F100", "", "2026-07-22"))

    def test_initial_date_removes_stale_ownership_issue_and_does_not_recreate_it(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "missing-source",
                initial_date="2026-07-22",
            )
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_factory("F2605260119", ownership_status="待确认")
            store.upsert_active_issue(
                issue_key="factory_ownership:F2605260119",
                kind="factory_ownership",
                factory_order="F2605260119",
                message="历史工厂单不应继续提醒",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index.load_aimes_order_cache",
                return_value=[{
                    "factory_order": "F2605260119",
                    "factory_name": "PP0037 KITCHEN",
                    "sales_order_name": "PP0037",
                    "split_time": "",
                }],
            ):
                result = sync_order_index(config)

            reopened = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            self.assertFalse(any(
                item["issue_key"] == "factory_ownership:F2605260119"
                for item in result["current_issues"]
            ))
            self.assertFalse(any(
                item["issue_key"] == "factory_ownership:F2605260119"
                for item in reopened.active_issues()
            ))
            reopened.close()

    def test_server_change_message_identifies_order_factory_and_data(self):
        self.assertEqual(
            _server_data_change_message("modified", ["PP0035-2"], "F20050502", "五金信息"),
            "修改订单 PP0035-2（工厂单 F20050502）的五金信息",
        )

    def test_server_change_message_explains_action_and_path(self):
        item = {"kind": "folder", "order_id": "PP0035-2", "manual_only": False}
        self.assertEqual(
            _server_change_message("modified", item, "/Volumes/server/Optimized Orders/pp0035-2"),
            "Server 订单文件夹修改：pp0035-2（订单 PP0035-2）（路径：/Volumes/server/Optimized Orders/pp0035-2）",
        )
        self.assertEqual(
            _server_data_change_message(
                "added", ["PP0035-2"], "F20050502", "五金信息", "/Volumes/server/Optimized Orders/pp0035-2/Fittingslist.xlsx"
            ),
            "新增订单 PP0035-2（工厂单 F20050502）的五金信息（来源：/Volumes/server/Optimized Orders/pp0035-2/Fittingslist.xlsx）",
        )

    def test_invalid_and_test_aimes_rows_are_warnings_only(self):
        invalid = _aimes_row_issue({
            "factory_order": "F200",
            "factory_name": "KITCHEN",
            "sales_order_name": "ORDER-200",
            "split_time": "2026-08-10 08:30:00",
        })
        test_row = _aimes_row_issue({
            "factory_order": "F201",
            "factory_name": "Kitchen Test",
            "sales_order_name": "PP0035",
            "split_time": "2026-08-10 08:30:00",
        })

        self.assertIn("销售单名称不是", invalid["reason"])
        self.assertIn("包含 test", test_row["reason"])
        self.assertEqual(test_row["ignore_key"], "factory:F201")
        visible, warnings = _partition_aimes_rows(
            [
                {
                    "factory_order": "F2606300182",
                    "factory_name": "CS001-Unnamed",
                    "sales_order_name": "SHERRY 001",
                    "split_time": "2026-06-30 10:00:00",
                }
            ],
            set(),
            {"factory:F2606300182": "CS001"},
        )
        self.assertEqual(visible[0]["sales_order_name"], "CS001")
        self.assertEqual(visible[0]["factory_order"], "F2606300182")
        self.assertEqual(warnings, [])

        unassigned_visible, unassigned_warnings = _partition_aimes_rows(
            [{
                "factory_order": "F2606300182",
                "factory_name": "CS001-Unnamed",
                "sales_order_name": "SHERRY 001",
                "split_time": "2026-06-30 10:00:00",
            }],
            set(),
            {},
        )
        self.assertEqual(unassigned_visible, [])
        self.assertEqual(unassigned_warnings[0]["suggested_order_id"], "CS001")

    def test_fittings_factory_order_uses_order_folder_hint(self):
        candidates = {}
        _merge_candidate(
            candidates,
            "F2606060141",
            source="server",
            order_id="PP0037",
            folder="/Volumes/server/Optimized Orders/pp0037",
        )

        result = _effective_factory_candidate("F2606060141", candidates["F2606060141"])

        self.assertEqual(result["order_id"], "PP0037")
        self.assertEqual(result["ownership_status"], "已确认")

    def test_unowned_factory_uses_exact_aimes_name_to_derive_order(self):
        candidates = {}
        _merge_candidate(
            candidates,
            "F2606060141",
            source="server",
            folder="/tmp/pp0037",
        )
        config = Config()

        with patch(
            "traveler_assistant.core.lookup_aimes_names",
            return_value={"F2606060141": "PP0037"},
        ):
            error = _exact_resolve_unowned_factories(config, candidates)

        result = _effective_factory_candidate("F2606060141", candidates["F2606060141"])
        self.assertEqual(error, "")
        self.assertEqual(result["order_id"], "PP0037")
        self.assertEqual(result["name_source"], "AIMES精确查询")

    def test_existing_database_factory_skips_exact_aimes_lookup(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.upsert_factory(
                "F2606170156",
                order_id="PP0035",
                factory_name="PP0035-ROOM 5",
                name_source="server_report",
                ownership_status="已确认",
            )
            store.commit()

            candidates = {}
            _merge_candidate(
                candidates,
                "F2606170156",
                source="server",
                folder=str(Path(temp) / "PP0035"),
            )
            _merge_database_factory_candidates(store, candidates)
            with patch(
                "traveler_assistant.core.lookup_aimes_names",
                side_effect=AssertionError("已有本地工厂单身份时不应精确查询 AIMES"),
            ):
                error = _exact_resolve_unowned_factories(config, candidates)

            result = _effective_factory_candidate(
                "F2606170156", candidates["F2606170156"]
            )
            store.close()

        self.assertEqual(error, "")
        self.assertEqual(result["order_id"], "PP0035")
        self.assertEqual(result["factory_name"], "PP0035-ROOM 5")
        self.assertEqual(result["name_source"], "server_report")

    def test_active_issue_is_persisted_and_resolved(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "order-index.sqlite3"
            store = OrderIndexStore(path)
            store.upsert_active_issue(
                issue_key="factory_ownership:F100",
                kind="factory_ownership",
                factory_order="F100",
                path="/tmp/pp0037",
                message="工厂单 F100 的订单归属无法唯一确认",
            )
            store.commit()
            self.assertEqual(store.active_issues()[0]["issue_key"], "factory_ownership:F100")
            store.resolve_active_issue("factory_ownership:F100")
            store.commit()
            self.assertEqual(store.active_issues(), [])
            store.close()

    def test_deleted_aimes_factory_is_audit_only_and_not_in_summaries(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "order-index.sqlite3"
            store = OrderIndexStore(path)
            store.upsert_order("PP9999", order_type="owned")
            store.upsert_aimes_factory(
                "F100", order_id="PP9999", factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999", split_time="2026-08-16T10:00:00",
                seen_at="2026-08-16T10:00:00",
            )
            store.upsert_aimes_factory(
                "F101", order_id="PP9999", factory_name="PP9999-BATH",
                sales_order_name="PP9999", split_time="2026-08-16T10:01:00",
                seen_at="2026-08-16T10:01:00",
            )
            store.mark_aimes_deleted(["F100"], verified_at="2026-08-16T11:00:00")
            store.commit()
            summary = next(item for item in store.summaries() if item["order_id"] == "PP9999")
            status = store.connection.execute(
                "select aimes_status, aimes_deleted_at from factory_orders where factory_order='F100'"
            ).fetchone()
            store.close()
        self.assertEqual([item["factory_order"] for item in summary["factories"]], ["F101"])
        self.assertEqual(tuple(status), ("deleted", "2026-08-16T11:00:00"))

    def test_invalid_aimes_rows_are_transient_warnings_only(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            source_rows = [{
                "factory_order": "F200",
                "factory_name": "TEST ROOM",
                "sales_order_name": "BAD-ORDER",
                "split_time": "2026-08-10 09:30:00",
            }]
            with patch(
                "traveler_assistant.core.refresh_aimes_recent_orders",
                return_value=source_rows,
            ), patch(
                "traveler_assistant.order_index.load_aimes_order_cache",
                return_value=[],
            ):
                result = sync_aimes_index(config, force=True)

            self.assertEqual(result["aimes"]["issue_count"], 0)
            self.assertEqual(result["aimes"]["warning_count"], 1)
            self.assertEqual(result["aimes_issues"], [])
            self.assertEqual(result["aimes_warnings"][0]["factory_order"], "F200")
            store = OrderIndexStore(config.workflow_database)
            self.assertEqual(store.connection.execute("select count(*) from aimes_review_rows").fetchone()[0], 0)
            self.assertEqual(store.connection.execute("select count(*) from factory_orders").fetchone()[0], 0)
            store.close()

    def test_exactly_verified_aimes_factory_is_persisted_as_aimes_identity(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("PP0035", validation_status="正常")
            store.upsert_factory(
                "F099",
                order_id="PP0035",
                factory_name="PP0035-ROOM 5",
                name_source="server_report",
                ownership_status="已确认",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.core.refresh_aimes_recent_orders_and_verify",
                return_value=(
                    [{
                        "factory_order": "F100",
                        "factory_name": "PP0035-KITCHEN",
                        "sales_order_name": "PP0035",
                        "split_time": "2026-08-19 10:00:00",
                    }],
                    {
                        "rows": [{
                            "factory_order": "F099",
                            "factory_name": "PP0035-ROOM 5",
                            "sales_order_name": "PP0035",
                            "split_time": "",
                        }],
                        "missing": [],
                    },
                ),
            ):
                result = sync_aimes_index(config, force=True)

            self.assertTrue(result["aimes"]["succeeded"])
            store = OrderIndexStore(config.workflow_database)
            row = store.connection.execute(
                "select order_id, factory_name, sales_order_name, name_source, last_aimes_seen from factory_orders where factory_order='F099'"
            ).fetchone()
            store.close()

        self.assertEqual(tuple(row[:4]), ("PP0035", "PP0035-ROOM 5", "PP0035", "AIMES"))
        self.assertTrue(row[4])

    def test_business_errors_are_actionable_and_hide_technical_details(self):
        validation = _business_validation_message(
            RuntimeError('Traceback: sqlite3.OperationalError: database is locked')
        )
        aimes = _business_aimes_message(RuntimeError("HTTP status code 500"))
        report = _business_report_message("fittings", Path("FittingslistPC123.xlsx"))

        self.assertIn("请检查", validation)
        self.assertNotIn("Traceback", validation)
        self.assertNotIn("sqlite3", validation)
        self.assertIn("重新获取", aimes)
        self.assertNotIn("500", aimes)
        self.assertIn("FittingslistPC123.xlsx", report)
        self.assertIn("重新扫描 Server", report)

    def test_old_status_is_migrated_and_validation_reason_is_persisted(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "order-index.sqlite3"
            store = OrderIndexStore(path)
            store.upsert_order("PP9999", validation_status="数据异常", stage="待拆单")
            store.upsert_factory(
                "F100",
                order_id="PP9999",
                factory_name="KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                name_source="AIMES",
                ownership_status="已确认",
            )
            store.connection.execute(
                "update orders set validation_message = ? where order_id = ?",
                ("未找到 material 文件。请补充后重新扫描 Server。", "PP9999"),
            )
            store.commit()
            store.close()

            reopened = OrderIndexStore(path)
            row = reopened.summaries()[0]
            stored_stage = reopened.connection.execute(
                "select stage from orders where order_id = ?", ("PP9999",)
            ).fetchone()[0]
            version = reopened.connection.execute("pragma user_version").fetchone()[0]
            reopened.close()

        self.assertEqual(row["stage"], "数据异常")
        self.assertEqual(stored_stage, "已设计")
        self.assertEqual(row["validation_message"], "未找到 material 文件。请补充后重新扫描 Server。")
        self.assertEqual(version, 7)

    def test_schema_migration_resolves_legacy_warning_from_unique_order_folder(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source_folder = root / "pp0037"
            source_folder.mkdir()
            path = root / "order-index.sqlite3"
            store = OrderIndexStore(path)
            store.upsert_factory(
                "F2606060141",
                source_folder=str(source_folder),
                ownership_status="待确认",
            )
            store.connection.execute("pragma user_version = 6")
            store.commit()
            store.close()

            reopened = OrderIndexStore(path)
            factory = reopened.connection.execute(
                "select order_id, ownership_status, name_source from factory_orders where factory_order = ?",
                ("F2606060141",),
            ).fetchone()
            self.assertEqual(factory, ("PP0037", "已确认", "server_folder"))
            self.assertEqual(reopened.active_issues(), [])
            self.assertIn(
                "自动确认工厂单 F2606060141 归属订单 PP0037",
                reopened.latest_changes()[0]["message"],
            )
            reopened.close()

    def test_aimes_owner_wins_over_stale_server_owner(self):
        candidate = {
            "names": {
                "server": {"PP0035-OFFICE"},
                "aimes": {"OFFICE"},
            },
            "orders": {"PP0035", "PP0035-2"},
            "sales_orders": {"PP0035-2"},
            "split_times": {"2026-08-10T08:30:00"},
            "folders": {"/Volumes/server/Optimized Orders/pp0035-2"},
            "has_hardware": False,
            "optimized": False,
        }

        result = _effective_factory_candidate("F2608050220", candidate)

        self.assertEqual(result["order_id"], "PP0035-2")
        self.assertEqual(result["factory_name"], "OFFICE")
        self.assertEqual(result["sales_order_name"], "PP0035-2")
        self.assertEqual(result["split_time"], "2026-08-10T08:30:00")
        self.assertEqual(result["name_source"], "AIMES")
        self.assertEqual(result["ownership_status"], "已确认")

    def test_aimes_order_validation_and_test_filter(self):
        self.assertEqual(_valid_aimes_order_id("PP0035"), "PP0035")
        self.assertEqual(_valid_aimes_order_id("PP0035-2"), "PP0035-2")
        self.assertEqual(_valid_aimes_order_id("CS123"), "CS123")
        self.assertEqual(_valid_aimes_order_id("PP0034-2"), "")
        self.assertEqual(_valid_aimes_order_id("PP035"), "")
        self.assertEqual(_valid_aimes_order_id("PP0035-A"), "")
        self.assertIsNone(_visible_aimes_row({
            "factory_order": "F100",
            "factory_name": "Kitchen TEST",
            "sales_order_name": "PP0035",
            "split_time": "2026-08-10 08:30:00",
        }))
        self.assertIsNone(_visible_aimes_row({
            "factory_order": "F101",
            "factory_name": "Kitchen",
            "sales_order_name": "PP0035-test",
            "split_time": "2026-08-10 08:30:00",
        }))

    def test_old_pp_server_paths_are_outside_dashboard_scope(self):
        root = Path("/Volumes/server/Optimized Orders")

        self.assertFalse(_source_path_in_dashboard_scope(root, str(root / "PP0034")))
        self.assertFalse(_source_path_in_dashboard_scope(root, str(root / "PP0034-2" / "report.xlsx")))
        self.assertTrue(_source_path_in_dashboard_scope(root, str(root / "PP0035")))
        self.assertTrue(_source_path_in_dashboard_scope(root, str(root / "PP0035-2" / "report.xlsx")))
        self.assertTrue(_source_path_in_dashboard_scope(root, str(root / "CS123")))

    def test_summary_aggregates_factory_status(self):
        with tempfile.TemporaryDirectory() as temp:
            store = OrderIndexStore(Path(temp) / "order-index.sqlite3")
            store.upsert_order("PP9999", validation_status="正常")
            store.upsert_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-01T10:00:00",
                name_source="AIMES",
                ownership_status="已确认",
                optimized=True,
                outbound_status="已出库",
            )
            store.upsert_factory(
                "F200",
                order_id="PP9999",
                factory_name="PP9999-LAUNDRY",
                sales_order_name="PP9999",
                split_time="2026-08-02T10:00:00",
                name_source="AIMES",
                ownership_status="已确认",
                optimized=True,
                outbound_status="未出库",
            )
            store.commit()

            row = store.summaries()[0]
            store.close()

        self.assertEqual(row["factory_count"], 2)
        self.assertEqual(row["optimized_count"], 2)
        self.assertEqual(row["shipped_count"], 1)
        self.assertEqual(row["stage"], "部分出货")
        self.assertEqual(row["optimization_progress"], "2 / 2")
        self.assertEqual(row["outbound_progress"], "1 / 2")
        self.assertEqual(row["latest_split_time"], "2026-08-02T10:00:00")

    def test_standard_outbound_status_reconciles_and_survives_reopen(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_order("CS005", validation_status="正常")
            store.upsert_factory(
                "F2608120222",
                order_id="CS005",
                factory_name="CS005-KITCHEN",
                sales_order_name="CS005",
                name_source="AIMES",
                ownership_status="已确认",
                optimized=True,
                outbound_status="未出库",
            )
            store.commit()
            store.close()

            records = [{
                "order_id": "CS005",
                "remark": "CS005",
                "status": "已出库",
                "document_number": "QTCK20260815001",
            }]
            with patch("traveler_assistant.order_index._load_outbound_records", return_value=records):
                store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
                self.assertEqual(reconcile_outbound_statuses(config, store), 1)
                store.close()
                listed = list_order_index(config)

            self.assertEqual(listed["orders"][0]["stage"], "已出货")
            reopened = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            row = reopened.connection.execute(
                "select outbound_status, outbound_document from factory_orders where factory_order = ?",
                ("F2608120222",),
            ).fetchone()
            reopened.close()
            self.assertEqual(row, ("已出库", "QTCK20260815001"))

    def test_fully_shipped_order_is_completed_even_if_optimization_evidence_is_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_order("PP9998", validation_status="正常")
            for factory_order, factory_name in (("F100", "PP9998-KITCHEN"), ("F101", "PP9998-LAUNDRY")):
                store.upsert_factory(
                    factory_order,
                    order_id="PP9998",
                    factory_name=factory_name,
                    sales_order_name="PP9998",
                    name_source="AIMES",
                    ownership_status="已确认",
                    optimized=False,
                    outbound_status="已出库",
                )
            store.commit()

            row = store.summaries()[0]
            store.close()

        self.assertEqual(row["optimized_count"], 0)
        self.assertEqual(row["shipped_count"], 2)
        self.assertEqual(row["stage"], "已出货")

    def test_grouped_outbound_document_reconciles_all_factory_orders_after_reindex(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            config.prepare_storage()
            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("PP9999", validation_status="正常")
            factories = [
                ("F100", "PP9999-MASTER"),
                ("F101", "PP9999-1ST"),
                ("F102", "PP9999-2ND"),
            ]
            for factory_order, factory_name in factories:
                store.upsert_factory(
                    factory_order,
                    order_id="PP9999",
                    factory_name=factory_name,
                    sales_order_name="PP9999",
                    name_source="AIMES",
                    ownership_status="已确认",
                    optimized=True,
                    outbound_status="未出库",
                )
            store.commit()
            store.close()

            connection = sqlite3.connect(config.workflow_database)
            connection.execute(
                """
                insert into outbound_documents(
                    document_number, document_type, order_id, factory_order,
                    status, source, issued_at, updated_at
                ) values(?,?,?,?,?,?,?,?)
                """,
                (
                    "QTCK-GROUPED",
                    "库存出库单",
                    "PP9999",
                    "F100,F101,F102",
                    "已出库",
                    "user-confirmed-grouped",
                    "2026-08-19",
                    "2026-08-19T12:00:00-07:00",
                ),
            )
            connection.commit()
            connection.close()
            ensure_schema(config.workflow_database)

            store = OrderIndexStore(config.workflow_database)
            records = _load_outbound_records(config)
            self.assertEqual(
                sorted(record["factory_order"] for record in records),
                ["F100", "F101", "F102"],
            )
            self.assertEqual(reconcile_outbound_statuses(config, store), 3)
            store.close()

            listed = list_order_index(config)

        order = listed["orders"][0]
        self.assertEqual(order["stage"], "已出货")
        self.assertEqual(order["outbound_progress"], "3 / 3")
        self.assertEqual(
            {
                (item["factory_order"], item["outbound_status"], item["outbound_document"])
                for item in order["factories"]
            },
            {
                ("F100", "已出库", "QTCK-GROUPED"),
                ("F101", "已出库", "QTCK-GROUPED"),
                ("F102", "已出库", "QTCK-GROUPED"),
            },
        )

    def test_order_level_outbound_record_is_not_broadcast_to_split_factories(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_order("CS005", validation_status="正常")
            for factory_order, factory_name in (("F100", "CS005-KITCHEN"), ("F101", "CS005-LAUNDRY")):
                store.upsert_factory(
                    factory_order,
                    order_id="CS005",
                    factory_name=factory_name,
                    sales_order_name="CS005",
                    name_source="AIMES",
                    ownership_status="已确认",
                    optimized=True,
                    outbound_status="未出库",
                )
            store.commit()
            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "CS005",
                    "remark": "CS005",
                    "status": "已出库",
                    "document_number": "QTCK-ONE",
                }],
            ):
                self.assertEqual(reconcile_outbound_statuses(config, store), 0)
            statuses = store.connection.execute(
                "select outbound_status from factory_orders order by factory_order"
            ).fetchall()
            store.close()
            self.assertEqual(statuses, [("未出库",), ("未出库",)])

    def test_partial_factory_upsert_preserves_persisted_business_statuses(self):
        with tempfile.TemporaryDirectory() as temp:
            store = OrderIndexStore(Path(temp) / "order-index.sqlite3")
            store.upsert_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                report_state="已发现",
                ownership_status="已确认",
                has_hardware=True,
                optimized=True,
                outbound_status="已出库",
                outbound_document="QTCK-ONE",
            )
            store.upsert_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                name_source="AIMES",
            )
            row = store.connection.execute(
                """
                select report_state, ownership_status, has_hardware, optimized,
                       outbound_status, outbound_document
                from factory_orders where factory_order = 'F100'
                """
            ).fetchone()
            store.close()
            self.assertEqual(row, ("已发现", "已确认", 1, 1, "已出库", "QTCK-ONE"))

    def test_fully_shipped_aimes_order_is_not_a_server_scan_candidate(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(state_dir=Path(temp) / "state")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999 KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "PP9999",
                    "remark": "PP9999 KITCHEN",
                    "status": "已出库",
                    "document_number": "QTCK001",
                }],
            ):
                candidates = _orders_requiring_server_scan(config, store)
            store.close()

        self.assertNotIn("PP9999", candidates)

    def test_scan_does_not_parse_material_source_as_traveler_during_outbound_reconcile(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            config.prepare_storage()
            folder = config.source_root / "PP9999"
            folder.mkdir(parents=True)
            material_path = folder / "PP0035-2 materials.xlsx"
            from openpyxl import Workbook

            workbook = Workbook()
            workbook.active.title = "Sheet1"
            workbook.save(material_path)

            store = OrderIndexStore(config.workflow_database)
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999-KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "PP9999",
                    "remark": "PP9999-KITCHEN",
                    "status": "已出库",
                    "document_number": "QTCK001",
                    "raw_fingerprint": "stored-material-fingerprint",
                    "traveler_path": str(material_path),
                }],
            ):
                store = OrderIndexStore(config.workflow_database)
                candidates = _orders_requiring_server_scan(config, store)
                store.close()
                result = scan_server_changes(config)

            self.assertNotIn("PP9999", candidates)
            self.assertFalse(result["server"]["changed"])
            self.assertEqual(result["server"]["changes"], [])

    def test_fully_shipped_folder_resolves_stale_material_validation_issue(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            config.prepare_storage()
            folder = config.source_root / "PP9999"
            folder.mkdir(parents=True)
            material_path = folder / "PP9999 materials.xlsx"
            material_path.write_bytes(b"invalid source retained for history")

            store = OrderIndexStore(config.workflow_database)
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999 KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.upsert_active_issue(
                issue_key=f"material_validation:PP9999:{material_path}",
                kind="material_validation",
                order_id="PP9999",
                path=str(material_path),
                message="历史材料校验失败",
                seen_at="2026-08-10T09:00:00",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "PP9999",
                    "remark": "PP9999 KITCHEN",
                    "status": "已出库",
                    "document_number": "QTCK001",
                }],
            ):
                result = scan_server_changes(config)

            self.assertFalse(any(
                issue["kind"] == "material_validation"
                for issue in result["current_issues"]
            ))
            reopened = OrderIndexStore(config.workflow_database)
            issue = reopened.connection.execute(
                "select status, resolved_at from active_issues where issue_key = ?",
                (f"material_validation:PP9999:{material_path}",),
            ).fetchone()
            reopened.close()

        self.assertEqual(issue[0], "resolved")
        self.assertTrue(issue[1])

    def test_fully_shipped_folder_resolves_stale_hardware_selection_issue(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            config.prepare_storage()
            folder = config.source_root / "PP9999"
            folder.mkdir(parents=True)

            store = OrderIndexStore(config.workflow_database)
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999 KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            issue_key = f"hardware_selection:PP9999:{folder}"
            store.upsert_active_issue(
                issue_key=issue_key,
                kind="hardware_selection",
                order_id="PP9999",
                path=str(folder),
                message="五金清单缺少 Order No. 区块",
                seen_at="2026-08-10T09:00:00",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "PP9999",
                    "remark": "PP9999 KITCHEN",
                    "status": "已出库",
                    "document_number": "QTCK001",
                }],
            ):
                result = scan_server_changes(config)

            self.assertFalse(any(
                issue["kind"] == "hardware_selection"
                for issue in result["current_issues"]
            ))
            reopened = OrderIndexStore(config.workflow_database)
            issue = reopened.connection.execute(
                "select status, resolved_at from active_issues where issue_key = ?",
                (issue_key,),
            ).fetchone()
            reopened.close()

        self.assertEqual(issue[0], "resolved")
        self.assertTrue(issue[1])

    def test_automatic_server_snapshot_skips_shipped_order_until_aimes_adds_factory(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            shipped_folder = config.source_root / "PP9999"
            active_folder = config.source_root / "PP8888"
            unindexed_folder = config.source_root / "PP7777"
            shipped_folder.mkdir(parents=True)
            active_folder.mkdir(parents=True)
            unindexed_folder.mkdir(parents=True)
            (shipped_folder / "PP9999 materials.xlsx").write_bytes(b"shipped")
            (active_folder / "PP8888 materials.xlsx").write_bytes(b"active")
            (unindexed_folder / "PP7777 materials.xlsx").write_bytes(b"unindexed")
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100",
                order_id="PP9999",
                factory_name="PP9999 KITCHEN",
                sales_order_name="PP9999",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.upsert_aimes_factory(
                "F200",
                order_id="PP8888",
                factory_name="PP8888 KITCHEN",
                sales_order_name="PP8888",
                split_time="2026-08-10T08:31:00",
                seen_at="2026-08-10T08:31:00",
            )
            store.commit()
            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[{
                    "order_id": "PP9999",
                    "remark": "PP9999 KITCHEN",
                    "status": "已出库",
                    "document_number": "QTCK001",
                }],
            ):
                _, snapshot = _server_snapshot(config, store)
                self.assertNotIn(str(shipped_folder), snapshot)
                self.assertIn(str(active_folder), snapshot)
                self.assertIn(str(unindexed_folder), snapshot)

                # A newly persisted AIMES factory order reopens the order for
                # the next automatic scan.
                store.upsert_aimes_factory(
                    "F101",
                    order_id="PP9999",
                    factory_name="PP9999 CLOSET",
                    sales_order_name="PP9999",
                    split_time="2026-08-17T08:30:00",
                    seen_at="2026-08-17T08:30:00",
                )
                store.commit()
                _, reopened_snapshot = _server_snapshot(config, store)
            store.close()

        self.assertIn(str(shipped_folder), reopened_snapshot)

    def test_new_current_aimes_factory_reopens_server_scan_candidate(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(
                state_dir=Path(temp) / "state",
                source_root=Path(temp) / "source",
            )
            config.source_root.mkdir()
            (config.source_root / "PP9999").mkdir()
            row = {
                "factory_order": "F100",
                "factory_name": "PP9999 KITCHEN",
                "sales_order_name": "PP9999",
                "split_time": "2026-08-10 08:30:00",
            }
            with patch(
                "traveler_assistant.core.refresh_aimes_recent_orders",
                return_value=[row],
            ), patch(
                "traveler_assistant.order_index.load_aimes_order_cache",
                return_value=[],
            ):
                result = sync_order_index(config, aimes_if_needed=True)

        self.assertEqual(result["sync"]["server_folder_count"], 1)
        self.assertEqual(result["orders"][0]["order_id"], "PP9999")

    def test_skipped_standard_order_does_not_resolve_its_old_issue(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config(
                state_dir=Path(temp) / "state",
                source_root=Path(temp) / "source",
            )
            config.source_root.mkdir()
            folder = config.source_root / "PP9999"
            folder.mkdir()
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_active_issue(
                issue_key="factory_ownership:F100",
                kind="factory_ownership",
                order_id="PP9999",
                factory_order="F100",
                path=str(folder),
                message="需要人工确认",
            )
            store.commit()
            store.close()

            result = sync_order_index(config)

            self.assertTrue(any(item["issue_key"] == "factory_ownership:F100" for item in result["current_issues"]))

    def test_orders_sort_by_latest_factory_split_time(self):
        with tempfile.TemporaryDirectory() as temp:
            store = OrderIndexStore(Path(temp) / "order-index.sqlite3")
            for order_id in ("PP0034", "PP0035", "PP0036"):
                store.upsert_order(order_id, validation_status="正常")
            store.upsert_factory("F100", order_id="PP0035", factory_name="Kitchen", sales_order_name="PP0035", split_time="2026-08-01T10:00:00", name_source="AIMES", ownership_status="已确认")
            store.upsert_factory("F101", order_id="PP0035", factory_name="Master", sales_order_name="PP0035", split_time="2026-08-03T10:00:00", name_source="AIMES", ownership_status="已确认")
            store.upsert_factory("F200", order_id="PP0036", factory_name="Office", sales_order_name="PP0036", split_time="2026-08-02T10:00:00", name_source="AIMES", ownership_status="已确认")
            store.upsert_factory("F099", order_id="PP0034", factory_name="Old", sales_order_name="PP0034", split_time="2026-08-04T10:00:00", name_source="AIMES", ownership_status="已确认")
            store.commit()
            rows = store.summaries()
            store.close()

        self.assertEqual([row["order_id"] for row in rows], ["PP0035", "PP0036"])
        self.assertEqual([row["factory_order"] for row in rows[0]["factories"]], ["F101", "F100"])

    def test_summary_includes_confirmed_server_report_factory_assigned_to_normal_order(self):
        with tempfile.TemporaryDirectory() as temp:
            store = OrderIndexStore(Path(temp) / "order-index.sqlite3")
            store.upsert_order("PP0035", validation_status="正常")
            store.upsert_factory(
                "F100",
                order_id="PP0035",
                factory_name="PP0035-ROOM 5",
                name_source="server_report",
                ownership_status="已确认",
            )
            store.commit()
            rows = store.summaries()
            store.close()

        summary = next(row for row in rows if row["order_id"] == "PP0035")
        self.assertEqual([item["factory_order"] for item in summary["factories"]], ["F100"])
        self.assertEqual(summary["factory_count"], 1)

    def test_aimes_if_needed_runs_once_per_day_after_success(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = Path(temp) / "source"
            config.source_root.mkdir()

            cached_row = {
                "factory_order": "F100",
                "factory_name": "KITCHEN",
                "sales_order_name": "PP9999",
                "split_time": "2026-08-10 08:30:00",
            }
            with patch(
                "traveler_assistant.core.refresh_aimes_recent_names",
                return_value={"F100": "KITCHEN"},
            ), patch(
                "traveler_assistant.core.refresh_aimes_recent_orders",
                return_value=[cached_row],
            ) as refresh, patch(
                "traveler_assistant.order_index.load_aimes_order_cache",
                side_effect=[[], [cached_row]],
            ):
                first = sync_order_index(config, aimes_if_needed=True)
                second = sync_order_index(config, aimes_if_needed=True)

            self.assertEqual(first["sync"]["aimes_count"], 1)
            self.assertEqual(second["sync"]["aimes_count"], 1)
            self.assertFalse(second["sync"]["aimes_attempted"])
            self.assertEqual(refresh.call_count, 1)
            self.assertEqual(refresh.call_args.args[1], 50)
            self.assertEqual(second["orders"][0]["order_id"], "PP9999")

    def test_aimes_only_sync_reports_change_then_skips_after_daily_success(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            row = {
                "factory_order": "F100",
                "factory_name": "KITCHEN",
                "sales_order_name": "PP0035-2",
                "split_time": "2026-08-10 08:30:00",
            }
            with patch(
                "traveler_assistant.core.refresh_aimes_recent_orders",
                return_value=[row],
            ) as refresh, patch(
                "traveler_assistant.order_index.load_aimes_order_cache",
                side_effect=[[], [row]],
            ):
                first = sync_aimes_index(config, if_needed=True)
                second = sync_aimes_index(config, if_needed=True)

            self.assertTrue(first["aimes"]["attempted"])
            self.assertTrue(first["aimes"]["succeeded"])
            self.assertTrue(first["aimes"]["changed"])
            self.assertEqual(first["orders"][0]["order_id"], "PP0035-2")
            self.assertFalse(second["aimes"]["attempted"])
            self.assertTrue(second["aimes"]["skipped_today"])
            self.assertEqual(refresh.call_count, 1)
            self.assertEqual(refresh.call_args.args[1], 50)

    def test_server_scan_is_non_mutating_until_full_processing(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = Path(temp) / "source"
            folder = config.source_root / "PP0035-2"
            folder.mkdir(parents=True)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100",
                order_id="PP0035-2",
                factory_name="PP0035-2 KITCHEN",
                sales_order_name="PP0035-2",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.commit()
            store.close()

            first = scan_server_changes(config)["server"]
            repeated = scan_server_changes(config)["server"]
            self.assertTrue(first["changed"])
            self.assertEqual(first["changes"], repeated["changes"])
            self.assertEqual(first["scan_stats"]["order_folder_count"], 1)
            self.assertEqual(first["scan_stats"]["related_excel_count"], 0)
            self.assertEqual(first["scan_stats"]["quick_checked_file_count"], 0)
            self.assertEqual(first["scan_stats"]["reused_folder_count"], 0)
            self.assertEqual(first["scan_stats"]["deep_scanned_folder_count"], 1)
            self.assertEqual(first["scan_stats"]["added_count"], 1)
            self.assertEqual(first["scan_stats"]["modified_count"], 0)
            self.assertEqual(first["scan_stats"]["deleted_count"], 0)
            self.assertTrue(first["changes"][0]["event_time"])
            trace = scan_server_changes(config)["operation_trace"]["server"]
            self.assertIn(
                "快速检查 0 个相关 Excel 文件，复用 0 个订单文件夹，深度扫描 1 个订单文件夹，总用时",
                trace[0],
            )
            self.assertIn("扫描范围：订单文件夹 1 个，相关 Excel 文件 0 个", trace[1])

            sync_order_index(config)
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

            folder_stat = folder.stat()
            os.utime(
                folder,
                ns=(folder_stat.st_atime_ns, folder_stat.st_mtime_ns + 1_000_000_000),
            )
            folder_only_change = scan_server_changes(config)["server"]
            self.assertFalse(folder_only_change["changed"])

            store = OrderIndexStore(config.workflow_database)
            folder_changes_before = store.connection.execute(
                "select count(*) from sync_changes where kind = 'folder_changed'"
            ).fetchone()[0]
            store.close()
            sync_order_index(config)
            store = OrderIndexStore(config.workflow_database)
            folder_changes_after = store.connection.execute(
                "select count(*) from sync_changes where kind = 'folder_changed'"
            ).fetchone()[0]
            store.close()
            self.assertEqual(folder_changes_after, folder_changes_before)

            report = folder / "material.xlsx"
            report.write_bytes(b"placeholder")
            changed = scan_server_changes(config)["server"]
            self.assertTrue(changed["changed"])
            self.assertEqual(changed["scan_stats"]["related_excel_count"], 1)
            self.assertEqual(changed["scan_stats"]["quick_checked_file_count"], 1)
            self.assertEqual(changed["scan_stats"]["reused_folder_count"], 0)
            self.assertEqual(changed["scan_stats"]["deep_scanned_folder_count"], 1)
            self.assertEqual(changed["scan_stats"]["added_count"], 1)
            # Creating a child report may also update the parent folder mtime;
            # only the recognized report is a business change.
            self.assertEqual(changed["scan_stats"]["modified_count"], 0)
            self.assertEqual(changed["scan_stats"]["deleted_count"], 0)
            self.assertTrue(any(item["path"] == str(report) for item in changed["changes"]))
            report_change = next(item for item in changed["changes"] if item["path"] == str(report))
            self.assertTrue(report_change["event_time"])
            self.assertEqual(changed["changes"], scan_server_changes(config)["server"]["changes"])

            sync_order_index(config)
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

            report.unlink()
            removed = scan_server_changes(config)["server"]
            self.assertEqual(removed["scan_stats"]["related_excel_count"], 0)
            self.assertEqual(removed["scan_stats"]["added_count"], 0)
            self.assertEqual(
                removed["scan_stats"]["modified_count"],
                sum(item["change_type"] == "modified" for item in removed["changes"]),
            )
            self.assertEqual(removed["scan_stats"]["deleted_count"], 1)
            self.assertTrue(any(
                item["change_type"] == "removed" and item["path"] == str(report)
                for item in removed["changes"]
            ))
            sync_order_index(config)
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

    def test_server_scan_baseline_covers_both_server_roots(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "Optimized Orders",
            )
            folder = root / "CUT TO SIZE" / "CS003"
            folder.mkdir(parents=True)
            store = OrderIndexStore(config.workflow_database)
            store.upsert_aimes_factory(
                "F100",
                order_id="CS003",
                factory_name="CS003 KITCHEN",
                sales_order_name="CS003",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.commit()
            store.close()

            with patch(
                "traveler_assistant.order_index._load_outbound_records",
                return_value=[],
            ):
                first = scan_server_changes(config)["server"]
                self.assertTrue(any(item["path"] == str(folder) for item in first["changes"]))

                sync_order_index(config)
                repeated = scan_server_changes(config)["server"]

        self.assertFalse(repeated["changed"])
        self.assertEqual(repeated["changes"], [])

    def test_app_generated_material_is_baselined_but_later_edit_is_detected(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            config = Config(
                state_dir=root / "state",
                source_root=root / "server" / "Optimized Orders",
            )
            folder = config.source_root.parent / "CUT TO SIZE" / "CS005"
            folder.mkdir(parents=True)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            store.upsert_aimes_factory(
                "F100",
                order_id="CS005",
                factory_name="CS005-KITCHEN",
                sales_order_name="CS005",
                split_time="2026-08-10T08:30:00",
                seen_at="2026-08-10T08:30:00",
            )
            store.upsert_source_file(
                folder,
                source_folder=folder,
                kind="folder",
                order_id="CS005",
                changed_at="2026-08-14T10:00:00",
            )
            store.commit()
            store.close()

            materials = folder / "CS005 materials.xlsx"
            materials.write_bytes(b"generated material")
            before_registration = scan_server_changes(config)["server"]
            self.assertTrue(before_registration["changed"])
            self.assertTrue(any(item["path"] == str(materials) for item in before_registration["changes"]))

            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            _record_generated_material_baseline(store, folder, materials, order_id="CS005")
            store.commit()
            store.close()
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

            materials.write_bytes(b"external edit")
            after_edit = scan_server_changes(config)["server"]
            self.assertTrue(after_edit["changed"])
            self.assertTrue(any(
                item["path"] == str(materials) and item["change_type"] == "modified"
                for item in after_edit["changes"]
            ))

    def test_selected_server_folder_reuses_index_processing_for_one_folder(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = (Path(temp) / "source").resolve()
            folder = config.source_root / "PP0035-2"
            folder.mkdir(parents=True)

            store = OrderIndexStore(config.workflow_database)
            store.upsert_order("PP0035-2", order_type="temporary")
            store.commit()
            store.close()

            result = process_server_folder(config, folder)

            self.assertEqual(result["sync"]["server_folder_count"], 1)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            order_type = store.connection.execute(
                "select order_type from orders where order_id = ?", ("PP0035-2",)
            ).fetchone()[0]
            source_paths = {
                row[0] for row in store.connection.execute("select path from source_files").fetchall()
            }
            store.close()
            self.assertEqual(order_type, "owned")
            self.assertIn(str(folder.resolve()), source_paths)
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

    def test_selected_non_order_folder_is_rejected_without_processing_children(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = (Path(temp) / "source").resolve()
            selected = config.source_root / "temporary-production"
            child_order = selected / "PP0035-2"
            child_order.mkdir(parents=True)
            (child_order / "material.xlsx").write_bytes(b"placeholder")

            with self.assertRaisesRegex(RuleError, "选错了文件夹"):
                process_server_folder(config, selected)

            self.assertFalse((config.state_dir / "order-index.sqlite3").exists())

    def test_selected_non_order_folder_with_recognized_report_is_temporary_order(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = (Path(temp) / "source").resolve()
            selected = config.source_root / "temporary-production"
            selected.mkdir(parents=True)
            (selected / "material.xlsx").write_bytes(b"placeholder")

            result = process_server_folder(config, selected)

            self.assertEqual(result["sync"]["server_folder_count"], 1)
            store = OrderIndexStore(config.state_dir / "order-index.sqlite3")
            source_paths = {
                row[0] for row in store.connection.execute("select path from source_files").fetchall()
            }
            store.close()
            self.assertTrue(
                str(selected / "material.xlsx") in source_paths
            )
            self.assertFalse(scan_server_changes(config)["server"]["changed"])

    def test_scan_reports_unprocessed_non_order_folder_as_manual_only(self):
        with tempfile.TemporaryDirectory() as temp:
            config = Config()
            config.state_dir = Path(temp) / "state"
            config.source_root = (Path(temp) / "source").resolve()
            selected = config.source_root / "temporary-production"
            selected.mkdir(parents=True)
            (selected / "material.xlsx").write_bytes(b"placeholder")

            server = scan_server_changes(config)["server"]
            manual_changes = [item for item in server["changes"] if item.get("manual_only")]

            self.assertTrue(manual_changes)
            self.assertTrue(any("临时订单文件夹" in item["message"] for item in manual_changes))


if __name__ == "__main__":
    unittest.main()
