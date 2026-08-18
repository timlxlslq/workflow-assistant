import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from traveler_assistant.core import Config
from traveler_assistant.costing import calculate_order_cost, export_order_cost
from traveler_assistant.database import ensure_schema
from traveler_assistant.inventory import Product, _replace_product_database


class CostingTests(unittest.TestCase):
    def _config(self, root: Path) -> Config:
        config = Config(
            state_dir=root / "state",
            order_root=root / "orders",
            backup_root=root / "backups",
        )
        config.prepare_storage()
        _replace_product_database(config.workflow_database, [
            Product("Panel", "M0004", "Plywood", "18mm", "启用", unit="pcs", cost_price=10.0),
            Product("Panel", "M1010", "Ivory Oak", "19.1mm", "启用", unit="pcs", cost_price=25.0),
            Product("Edge band", "M2010", "Ivory Oak Edge Banding", "", "启用", unit="m", cost_price=2.0),
        ])
        return config

    def test_order_total_uses_order_material_summary_and_raw_quantities(self):
        with tempfile.TemporaryDirectory() as temporary:
            config = self._config(Path(temporary))
            connection = sqlite3.connect(config.workflow_database)
            connection.execute(
                """insert into material_items(
                    order_id,material_type,color,thickness,quantity,unit,
                    source_type,source_path,source_fingerprint,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?)""",
                ("PP9999", "plywood", "", "18", 2.5, "pcs", "aihouse", "", "", "now"),
            )
            connection.execute(
                """insert into material_items(
                    order_id,material_type,color,thickness,quantity,unit,
                    source_type,source_path,source_fingerprint,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?)""",
                ("PP9999", "panel", "Ivory Oak", "19.1", 1, "pcs", "aihouse", "", "", "now"),
            )
            connection.commit()
            connection.close()

            report = calculate_order_cost(config, "PP9999")
            self.assertEqual(report["status"], "已完成")
            self.assertAlmostEqual(report["total_cost"], 50.0)
            self.assertEqual(report["factory_totals"][0]["factory_order"], "材料汇总")
            self.assertAlmostEqual(report["factory_totals"][0]["total"], 50.0)
            self.assertTrue(all(row["factory_order"] == "材料汇总" for row in report["lines"]))

            exported = export_order_cost(config, "PP9999")
            self.assertTrue(Path(exported["export_path"]).is_file())
            workbook = load_workbook(exported["export_path"], data_only=False, read_only=True)
            self.assertEqual(workbook.sheetnames, ["成本汇总", "价格与来源", "成本明细", "缺失项目"])
            summary_values = [cell for row in workbook["成本汇总"].iter_rows(values_only=True) for cell in row]
            self.assertIn("材料汇总", summary_values)
            self.assertNotIn("待分配", summary_values)
            workbook.close()

    def test_missing_cost_price_is_not_treated_as_zero(self):
        with tempfile.TemporaryDirectory() as temporary:
            config = self._config(Path(temporary))
            connection = sqlite3.connect(config.workflow_database)
            connection.execute(
                """insert into material_items(
                    order_id,material_type,color,thickness,quantity,unit,
                    source_type,source_path,source_fingerprint,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?)""",
                ("PP9998", "panel", "Unknown", "19.1", 1, "pcs", "aihouse", "", "", "now"),
            )
            connection.commit()
            connection.close()
            report = calculate_order_cost(config, "PP9998")
            self.assertIsNone(report["total_cost"])
            self.assertEqual(report["status"], "待补充")
            self.assertTrue(report["missing_items"])


if __name__ == "__main__":
    unittest.main()
