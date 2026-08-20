from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .operation_log import log_progress_payload
from .database import (
    database_path,
    ensure_schema,
    migrate_inventory_mapping_file,
    migrate_legacy_databases,
    read_cache,
    write_cache,
)


FACTORY_RE = re.compile(r"^F\d+$", re.IGNORECASE)
AIMES_BULK_FETCH_LIMIT = 50


class RuleError(RuntimeError):
    def __init__(self, code: str, message: str, **context):
        super().__init__(message)
        self.code = code
        self.context = context


def progress(message: str, **details) -> None:
    payload = {"event": "progress", "message": message, **details}
    log_progress_payload(payload)
    print(json.dumps(payload, ensure_ascii=False), file=sys.stderr, flush=True)


@dataclass
class Config:
    source_root: Path = Path("/Volumes/server/Optimized Orders")
    order_root: Path = Path.home() / "Documents/pp-flowhub/runtime/travelers"
    template: Path = Path(__file__).resolve().parent.parent / "resources/templates/Work Order Traveler.xlsx"
    # Traveler file backups remain configurable. Database backups use the
    # local project state directory so they are independent from that setting.
    backup_root: Path = Path("/Volumes/server/g/pp-flowhub/database-backups")
    state_dir: Path = Path.home() / "Documents/pp-flowhub/data"
    initial_date: str = "2026-07-22"
    server_scan_baseline_folder: str = "CS003 PP0047"
    server_scan_baseline_at: str = "2026-07-25T11:00:25-07:00"
    aimes_username: str = ""
    aimes_keychain_service: str = "com.pacificpride.ppflowhub.aimes"
    aimes_retry_delays: tuple[float, float] = (2.0, 4.0)
    operation_log_enabled: bool = True
    storage_prepared: bool = False

    @property
    def operation_log_file(self) -> Path:
        return self.state_dir / "operation-log.jsonl"

    @property
    def workflow_database(self) -> Path:
        central = database_path(self.state_dir)
        if self.storage_prepared:
            return central
        # Test fixtures and users interrupted during a pre-cutover launch may
        # still have only the old index.  Keep that file readable until the
        # explicit storage preparation merges it into workflow.sqlite3.
        legacy = self.state_dir / "order-index.sqlite3"
        return legacy if legacy.exists() else self.state_dir / "order-index.sqlite3"

    @property
    def database_backup_root(self) -> Path:
        return self.state_dir / "database-backups"

    def prepare_storage(self) -> None:
        self.storage_prepared = True
        ensure_schema(database_path(self.state_dir))
        migrate_legacy_databases(self.state_dir)
        migrate_inventory_mapping_file(self.state_dir)

    @property
    def settings_file(self) -> Path:
        return self.state_dir / "settings.json"

    def load_settings(self, source_profile: str | None = None) -> None:
        if not self.settings_file.is_file():
            return
        try:
            values = json.loads(self.settings_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise RuleError("settings_invalid", f"设置文件无法读取：{self.settings_file}") from exc
        if source_profile == "server":
            path_keys = {
                "source_root": ("server_source_root", "source_root"),
                "order_root": ("production_order_root", "order_root"),
                "backup_root": ("production_backup_root", "backup_root"),
            }
        elif source_profile == "local":
            local_root = values.get("local_test_root")
            if local_root:
                root = Path(str(local_root)).expanduser()
                self.source_root = root / "Optimized Orders"
                self.order_root = root / "Generated Travelers"
                self.backup_root = root / "Backups"
            path_keys = {}
        else:
            path_keys = {
                "source_root": ("source_root",),
                "order_root": ("order_root",),
                "backup_root": ("backup_root",),
            }
        for attribute, keys in path_keys.items():
            for key in keys:
                if key in values:
                    setattr(self, attribute, Path(str(values[key])).expanduser())
                    break
        # iCloud Traveler backup paths were part of the old design. They are
        # intentionally ignored so a stale settings file cannot reintroduce
        # cloud persistence after the database cutover.
        if "icloud" in str(self.backup_root).lower() or "mobile documents" in str(self.backup_root).lower():
            self.backup_root = Path("/Volumes/server/g/pp-flowhub/database-backups")
        if "initial_date" in values:
            self.initial_date = str(values["initial_date"])
        if "server_scan_baseline_at" in values:
            self.server_scan_baseline_at = str(values["server_scan_baseline_at"])
        if "aimes_username" in values:
            self.aimes_username = str(values["aimes_username"])
        if isinstance(values.get("operation_log_enabled"), bool):
            self.operation_log_enabled = values["operation_log_enabled"]
        try:
            datetime.strptime(self.initial_date, "%Y-%m-%d")
        except ValueError as exc:
            raise RuleError("settings_invalid", f"初始扫描日期格式错误：{self.initial_date}") from exc
        try:
            datetime.fromisoformat(self.server_scan_baseline_at)
        except ValueError as exc:
            raise RuleError("settings_invalid", f"Server 扫描基准时间格式错误：{self.server_scan_baseline_at}") from exc

    @property
    def factory_names_file(self) -> Path:
        return self.state_dir / "factory-names.json"

    @property
    def aimes_orders_file(self) -> Path:
        return self.state_dir / "aimes-orders.json"

    @property
    def material_assignments_file(self) -> Path:
        return self.state_dir / "material-assignments.json"

    @property
    def node_path(self) -> str:
        configured = os.environ.get("TRAVELER_NODE", "").strip()
        if configured:
            return configured
        bundled = Path(__file__).resolve().parent.parent / "bin" / "node"
        if bundled.is_file():
            return str(bundled)
        workspace_runtime = self.playwright_node_modules.resolve().parent / "bin" / "node"
        return str(workspace_runtime) if workspace_runtime.is_file() else "node"

    @property
    def playwright_node_modules(self) -> Path:
        return Path(os.environ.get("TRAVELER_NODE_MODULES", Path(__file__).resolve().parent.parent / "node_modules"))


def load_factory_name_cache(config: Config) -> dict[str, str]:
    if not config.storage_prepared:
        try:
            values = json.loads(config.factory_names_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        return {str(key).upper(): str(value).strip() for key, value in values.items() if str(value).strip()}
    values = read_cache(config.workflow_database, "factory_names", config.factory_names_file, {})
    if not isinstance(values, dict):
        return {}
    return {str(key).upper(): str(value).strip() for key, value in values.items() if str(value).strip()}


def save_factory_name_cache(config: Config, values: dict[str, str]) -> None:
    if not config.storage_prepared:
        config.state_dir.mkdir(parents=True, exist_ok=True)
        temporary = config.factory_names_file.with_suffix(".tmp")
        temporary.write_text(json.dumps(dict(sorted(values.items())), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary.replace(config.factory_names_file)
        return
    write_cache(config.workflow_database, "factory_names", dict(sorted(values.items())))


def load_aimes_order_cache(config: Config) -> list[dict[str, str]]:
    if not config.storage_prepared:
        try:
            values = json.loads(config.aimes_orders_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return []
    else:
        values = read_cache(config.workflow_database, "aimes_orders", config.aimes_orders_file, [])
    if not isinstance(values, list):
        return []
    return [
        {
            "factory_order": str(row.get("factory_order", "")).upper().strip(),
            "factory_name": str(row.get("factory_name", "")).strip(),
            "sales_order_name": str(row.get("sales_order_name", "")).upper().strip(),
            "split_time": str(row.get("split_time", "")).strip(),
        }
        for row in values
        if isinstance(row, dict) and str(row.get("factory_order", "")).strip()
    ]


def save_aimes_order_cache(config: Config, values: list[dict[str, str]]) -> None:
    if not config.storage_prepared:
        config.state_dir.mkdir(parents=True, exist_ok=True)
        temporary = config.aimes_orders_file.with_suffix(".tmp")
        temporary.write_text(json.dumps(values, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary.replace(config.aimes_orders_file)
        return
    write_cache(config.workflow_database, "aimes_orders", values)


def load_material_assignments(config: Config) -> dict[str, str]:
    if not config.storage_prepared:
        try:
            values = json.loads(config.material_assignments_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
    else:
        values = read_cache(config.workflow_database, "material_assignments", config.material_assignments_file, {})
    return {str(key): str(value) for key, value in values.items() if str(value).strip()}


def save_material_assignment(config: Config, key: str, path: str) -> None:
    values = load_material_assignments(config)
    values[str(key)] = str(path)
    if not config.storage_prepared:
        config.state_dir.mkdir(parents=True, exist_ok=True)
        temporary = config.material_assignments_file.with_suffix(".tmp")
        temporary.write_text(json.dumps(dict(sorted(values.items())), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary.replace(config.material_assignments_file)
    else:
        write_cache(config.workflow_database, "material_assignments", dict(sorted(values.items())))


def _run_aimes_lookup(
    config: Config,
    factory_orders: list[str],
    *,
    recent_limit: int = 0,
    include_order_metadata: bool = False,
    verify_factory_orders: bool = False,
):
    missing = [order.upper() for order in factory_orders if order]
    if not missing and not recent_limit:
        return {"rows": []} if include_order_metadata else {}
    if not config.aimes_username.strip():
        raise RuleError("aimes_credentials", "缺少 AIMES 用户名，无法查询工厂单名称", factory_orders=missing)
    helper = Path(__file__).resolve().parent.parent / "tools/aimes_lookup.mjs"
    if not helper.is_file():
        raise RuleError("aimes_unavailable", "AIMES 查询程序不存在", factory_orders=missing)
    try:
        password = subprocess.check_output([
            "/usr/bin/security", "find-generic-password", "-w", "-a", config.aimes_username,
            "-s", config.aimes_keychain_service,
        ], text=True, stderr=subprocess.DEVNULL).strip()
    except (subprocess.CalledProcessError, FileNotFoundError) as exc:
        raise RuleError("aimes_credentials", "无法从 macOS 钥匙串读取 AIMES 密码", factory_orders=missing) from exc
    payload = json.dumps({
        "username": config.aimes_username,
        "password": password,
        "factoryOrders": missing,
        "recentLimit": recent_limit,
        "includeOrderMetadata": include_order_metadata,
        "verifyFactoryOrders": verify_factory_orders,
    })
    env = os.environ.copy()
    env["NODE_PATH"] = str(config.playwright_node_modules)
    last_error = ""
    operation_timings: list[dict[str, object]] = []

    def consume_progress(stderr: str, attempt_timings: list[dict[str, object]]) -> None:
        for line in stderr.splitlines():
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(event, dict) or event.get("event") != "progress":
                continue
            log_progress_payload(event)
            stage = str(event.get("stage", "")).strip()
            duration = event.get("duration_seconds")
            if stage and isinstance(duration, (int, float)):
                attempt_timings.append({
                    "stage": stage,
                    "label": str(event.get("stage_label", stage)),
                    "duration_seconds": round(float(duration), 2),
                })

    for attempt in range(3):
        attempt_timings: list[dict[str, object]] = []
        progress(f"正在获取 AIMES 数据（第 {attempt + 1}/3 次）", factory_orders=missing)
        try:
            completed = subprocess.run(
                [config.node_path, str(helper)], input=payload, text=True,
                capture_output=True, timeout=120, env=env, check=True,
            )
            consume_progress(completed.stderr, attempt_timings)
            result = json.loads(completed.stdout)
            if not isinstance(result, dict):
                raise json.JSONDecodeError("AIMES 返回结果不是对象", completed.stdout, 0)
            node_timings = result.pop("timings", [])
            if not isinstance(node_timings, list):
                node_timings = []
            normalized_timings = [
                {
                    "stage": str(item.get("stage", "")).strip(),
                    "label": str(item.get("label", item.get("stage", ""))).strip(),
                    "duration_seconds": round(float(item.get("duration_seconds", 0)), 2),
                }
                for item in node_timings
                if isinstance(item, dict) and str(item.get("label", item.get("stage", ""))).strip()
            ]
            # The attempt duration is an envelope around the helper process,
            # not a child stage. Keep only non-overlapping stages here; the
            # caller owns the authoritative end-to-end duration.
            result["_aimes_timings"] = operation_timings + (normalized_timings or attempt_timings)
            result["_aimes_retry_count"] = attempt
            return result
        except subprocess.CalledProcessError as exc:
            consume_progress(exc.stderr or "", attempt_timings)
            last_error = (exc.stderr or "AIMES 查询失败").strip()[-1000:]
            operation_timings.extend(attempt_timings)
            progress(
                f"AIMES 第 {attempt + 1}/3 次尝试失败，准备重试",
                retry_attempt=attempt + 1,
            )
            if "账号或密码错误" in last_error:
                raise RuleError("aimes_credentials", "AIMES 账号或密码错误", factory_orders=missing) from exc
        except (subprocess.TimeoutExpired, OSError, json.JSONDecodeError) as exc:
            last_error = str(exc)
            operation_timings.extend(attempt_timings)
            progress(
                f"AIMES 第 {attempt + 1}/3 次尝试失败，准备重试",
                retry_attempt=attempt + 1,
            )
        if attempt < 2:
            retry_started = time.perf_counter()
            progress(f"等待 {config.aimes_retry_delays[attempt]:.1f} 秒后重试 AIMES")
            time.sleep(config.aimes_retry_delays[attempt])
            operation_timings.append({
                "stage": "retry_wait",
                "label": f"重试等待（第 {attempt + 1} 次后）",
                "duration_seconds": round(time.perf_counter() - retry_started, 2),
            })
    raise RuleError(
        "aimes_unavailable",
        f"AIMES 查询失败：{last_error}",
        factory_orders=missing,
        aimes_timings=operation_timings,
        aimes_retry_count=2,
    )


def lookup_aimes_names(config: Config, factory_orders: list[str], recent_limit: int = 0) -> dict[str, str]:
    missing = [order.upper() for order in factory_orders if order]
    result = _run_aimes_lookup(config, missing, recent_limit=recent_limit)
    if not isinstance(result, dict):
        raise RuleError("aimes_unavailable", "AIMES 返回的工厂单名称格式无效", factory_orders=missing)
    values = {
        str(factory).upper(): str(name).strip()
        for factory, name in result.items()
        if str(factory).strip() and str(name).strip()
    }
    if missing and not all(values.get(factory) for factory in missing):
        raise RuleError("aimes_unavailable", "AIMES 返回结果缺少工厂单名称", factory_orders=missing)
    return values


def lookup_aimes_recent_orders(
    config: Config,
    limit: int = AIMES_BULK_FETCH_LIMIT,
    *,
    include_trace: bool = False,
) -> list[dict[str, str]] | tuple[list[dict[str, str]], list[dict[str, object]]]:
    result = _run_aimes_lookup(config, [], recent_limit=limit, include_order_metadata=True)
    rows = result.get("rows") if isinstance(result, dict) else None
    if not isinstance(rows, list):
        raise RuleError("aimes_unavailable", "AIMES 返回的订单明细格式无效")
    normalized = [
        {
            "factory_order": str(row.get("factory_order", "")).upper().strip(),
            "factory_name": str(row.get("factory_name", "")).strip(),
            "sales_order_name": str(row.get("sales_order_name", "")).upper().strip(),
            "split_time": str(row.get("split_time", "")).strip(),
        }
        for row in rows
        if isinstance(row, dict)
    ]
    if include_trace:
        timings = result.get("_aimes_timings", [])
        return normalized, _non_aggregate_aimes_timings(timings)
    return normalized


def _non_aggregate_aimes_timings(values: object) -> list[dict[str, object]]:
    """Return only flat AIMES stages, excluding an old aggregate timing item."""
    if not isinstance(values, list):
        return []
    return [
        item for item in values
        if isinstance(item, dict)
        and str(item.get("stage", "")).strip() not in {"attempt", "total"}
        and "总计用时" not in str(item.get("label", ""))
    ]


def verify_aimes_factory_orders(config: Config, factory_orders: list[str]) -> dict[str, list[dict[str, str]] | list[str]]:
    """Check exact AIMES existence without treating a missing row as a transport error."""
    requested = [str(order).upper().strip() for order in factory_orders if str(order).strip()]
    result = _run_aimes_lookup(config, requested, verify_factory_orders=True)
    rows = result.get("rows") if isinstance(result, dict) else None
    missing = result.get("missing") if isinstance(result, dict) else None
    if not isinstance(rows, list) or not isinstance(missing, list):
        raise RuleError("aimes_unavailable", "AIMES 返回的精确核验格式无效", factory_orders=requested)
    normalized_rows = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        normalized_rows.append({
            "factory_order": str(row.get("factory_order", "")).upper().strip(),
            "factory_name": str(row.get("factory_name", "")).strip(),
            "sales_order_name": str(row.get("sales_order_name", "")).upper().strip(),
            "split_time": str(row.get("split_time", "")).strip(),
        })
    return {
        "rows": normalized_rows,
        "missing": [str(value).upper().strip() for value in missing if str(value).strip()],
    }


def refresh_aimes_recent_orders_and_verify(
    config: Config,
    limit: int,
    factory_orders: list[str],
    *,
    timing_sink: list[dict[str, object]] | None = None,
) -> tuple[list[dict[str, str]], dict[str, list[dict[str, str]] | list[str]]]:
    """Fetch the recent page and exact-check stale local factories in one session."""
    result = _run_aimes_lookup(
        config,
        [str(order).upper().strip() for order in factory_orders if str(order).strip()],
        recent_limit=limit,
        include_order_metadata=True,
        verify_factory_orders=True,
    )
    recent_rows = result.get("rows") if isinstance(result, dict) else None
    verified_rows = result.get("verify_rows") if isinstance(result, dict) else None
    missing = result.get("missing") if isinstance(result, dict) else None
    if not isinstance(recent_rows, list) or not isinstance(verified_rows, list) or not isinstance(missing, list):
        raise RuleError("aimes_unavailable", "AIMES 返回的获取与精确核验格式无效")

    def normalize(rows: list[object]) -> list[dict[str, str]]:
        return [
            {
                "factory_order": str(row.get("factory_order", "")).upper().strip(),
                "factory_name": str(row.get("factory_name", "")).strip(),
                "sales_order_name": str(row.get("sales_order_name", "")).upper().strip(),
                "split_time": str(row.get("split_time", "")).strip(),
            }
            for row in rows
            if isinstance(row, dict)
        ]

    if timing_sink is not None:
        timing_sink.extend(_non_aggregate_aimes_timings(result.get("_aimes_timings", [])))
    return normalize(recent_rows), {
        "rows": normalize(verified_rows),
        "missing": [str(value).upper().strip() for value in missing if str(value).strip()],
    }


def refresh_aimes_recent_orders(
    config: Config,
    limit: int = AIMES_BULK_FETCH_LIMIT,
    *,
    persist: bool = True,
    timing_sink: list[dict[str, object]] | None = None,
) -> list[dict[str, str]]:
    """Fetch structured AIMES rows, optionally persisting the raw result.

    The order-index workflow passes ``persist=False`` so it can validate the
    rows first. Invalid sales-order names are warnings only and must never be
    written to the business mapping cache.
    """
    traced = lookup_aimes_recent_orders(config, limit, include_trace=True)
    rows, timings = traced
    if timing_sink is not None:
        timing_sink.extend(timings)
    if persist and rows:
        save_aimes_order_cache(config, rows)
        names = {
            row["factory_order"]: row["factory_name"]
            for row in rows
            if row["factory_order"] and row["factory_name"]
        }
        cache = load_factory_name_cache(config)
        cache.update(names)
        save_factory_name_cache(config, cache)
    return rows


def refresh_aimes_recent_names(config: Config, limit: int = AIMES_BULK_FETCH_LIMIT) -> dict[str, str]:
    """Fetch and cache the newest AIMES factory-order rows."""
    rows = refresh_aimes_recent_orders(config, limit, persist=False)
    return {
        row["factory_order"]: row["factory_name"]
        for row in rows
        if row["factory_order"] and row["factory_name"]
    }


@dataclass
class FittingItem:
    name: str
    code: str
    size: str
    unit: str
    quantity: float


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise RuleError("invalid_number", f"无法识别数值：{value}") from exc


def _normalize_name(value: str) -> str:
    return re.sub(r"[\s_-]+", "", value).upper()


def parse_fittings_groups(
    path: Path,
    *,
    allow_missing_factory: bool = False,
    fallback_factory: str = "",
) -> list[tuple[str, list[FittingItem]]]:
    invalid_dimension = False
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    if re.search(br'<dimension[^>]+ref="[^"]*\?[^"]*"', archive.read(name)[:1024]):
                        invalid_dimension = True
                        break
    except (OSError, zipfile.BadZipFile):
        pass

    try:
        workbook = load_workbook(path, data_only=True, read_only=not invalid_dimension)
    except ValueError:
        try:
            workbook = load_workbook(path, data_only=True, read_only=False)
        except Exception as exc:
            raise RuleError("fittings_file_invalid", f"五金清单文件损坏或无法读取：{path.name}") from exc
    if workbook.sheetnames != ["Page1"]:
        raise RuleError("fittings_schema", f"五金清单工作表结构变化：{workbook.sheetnames}", path=str(path))
    sheet = workbook["Page1"]
    starts = [
        row for row in range(1, sheet.max_row + 1)
        if _text(sheet.cell(row, 1).value) == "Order No."
    ]
    if not starts:
        raise RuleError("fittings_schema", "五金清单缺少 Order No. 区块")

    groups = []
    for index, start in enumerate(starts):
        factory = _text(sheet.cell(start, 3).value).upper()
        if not FACTORY_RE.fullmatch(factory):
            if allow_missing_factory and fallback_factory:
                factory = fallback_factory.strip()
            else:
                raise RuleError("fittings_identity", f"五金清单工厂单号异常：{factory}")
        header = start + 5
        expected = {3: "Name", 5: "Code", 6: "Size", 11: "Quantity"}
        if any(_text(sheet.cell(header, col).value) != label for col, label in expected.items()):
            raise RuleError("fittings_schema", f"五金清单第 {start} 行开始的区块字段发生变化")
        end = starts[index + 1] if index + 1 < len(starts) else sheet.max_row + 1
        items = []
        for row in range(header + 1, end):
            name = _text(sheet.cell(row, 3).value)
            if not name or name == "Total":
                continue
            items.append(FittingItem(
                name=name,
                code=_text(sheet.cell(row, 5).value),
                size=_text(sheet.cell(row, 6).value),
                unit=_text(sheet.cell(row, 9).value),
                quantity=_number(sheet.cell(row, 11).value),
            ))
        groups.append((factory, items))
    return groups
