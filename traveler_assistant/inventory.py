from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
from urllib.parse import urlsplit
from urllib.request import urlopen
from collections import Counter
from dataclasses import asdict, dataclass, field, replace
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .core import Config, RuleError, _normalize_name, _text, progress
from .operation_log import configure_operation_log, log_database_statement, log_progress_payload
from .database import ensure_schema


TRAVELER_RE = re.compile(r"^Work Order Traveler\(.+\)\.xlsx$", re.IGNORECASE)
PP_FOLDER_RE = re.compile(r"^(?:PP\d{4}(?:-\d+)?|CS\d{3})$", re.IGNORECASE)
TB_RE = re.compile(r"^TB(\d+)$", re.IGNORECASE)
PANEL_RE = re.compile(r"^(\d+(?:\.\d+)?)mm--(.+)$", re.IGNORECASE)
EDGE_RE = re.compile(r"^Edge banding-+(.+)$", re.IGNORECASE)
REQUIRED_PRODUCT_HEADERS = {"商品类别", "*商品编号", "商品名称", "规格型号", "状态"}
INVENTORY_DATABASE_VERSION = 1
INVENTORY_CDP_DEFAULT_ENDPOINT = "http://127.0.0.1:9222"
INVENTORY_LOGIN_URL = "https://www.jdy.com/login/"
# 登录后会根据租户、区域或产品版本跳转到不同的 jdy.com 子域名，
# 例如 vip2-hz.jdy.com/default-new.jsp，而不是固定的 /workbench/ 路径。
INVENTORY_WORKBENCH_PREFIXES = (
    "https://service.jdy.com/workbench/",
    "http://service.jdy.com/workbench/",
    "https://www.jdy.com/workbench/",
    "http://www.jdy.com/workbench/",
)
INVENTORY_DOMAIN_SUFFIX = ".jdy.com"


@dataclass
class TravelerItem:
    row: int
    section: str
    name: str
    quantity: float
    document_remark: str = ""


@dataclass
class TravelerData:
    path: Path
    pp_folder: str
    order_id: str
    order_name: str
    items: list[TravelerItem]
    zero_items: list[TravelerItem]
    documents: dict[str, list[TravelerItem]]
    modified_at: str
    fingerprint: str

    def content_snapshot(self) -> dict:
        return {
            "order_id": self.order_id,
            "documents": {
                remark: [asdict(item) for item in items]
                for remark, items in sorted(self.documents.items())
            },
            "items": [asdict(item) for item in self.items],
            "zero_items": [asdict(item) for item in self.zero_items],
        }


@dataclass
class Product:
    category: str
    code: str
    name: str
    spec: str
    status: str
    remark: str = ""
    unit: str = ""
    cost_price: float | None = None


@dataclass
class OutboundItem:
    traveler_name: str
    product_code: str
    product_name: str
    quantity: float
    section: str
    match_source: str
    document_remark: str = ""
    unit: str = ""


@dataclass
class InventoryPreview:
    traveler: TravelerData
    outbound_items: list[OutboundItem]
    ignored_items: list[dict] = field(default_factory=list)
    missing_items: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    selected_document_remarks: tuple[str, ...] = ()
    source_type: str = "traveler"
    scope_decisions: list[dict] = field(default_factory=list)
    excluded_items: list[dict] = field(default_factory=list)
    no_outbound_required: bool = False

    @property
    def ready(self) -> bool:
        return not self.missing_items

    def payload(self) -> dict:
        documents = self.document_payloads()
        selected = self._selected_document_set()
        return {
            "source_type": self.source_type,
            "traveler": {
                "path": str(self.traveler.path),
                "pp_folder": self.traveler.pp_folder,
                "order_id": self.traveler.order_id,
                "order_name": self.traveler.order_name,
                "modified_at": self.traveler.modified_at,
                "fingerprint": self.traveler.fingerprint,
            },
            "outbound_items": [asdict(item) for item in self.outbound_items],
            # Keep zero quantities in the machine-readable result for
            # diagnostics, but the Swift preview intentionally does not
            # render them as outbound rows.
            "zero_items": [
                asdict(item) for item in self.traveler.zero_items
                if not selected or _normalize_name(item.document_remark) in selected
            ],
            "ignored_items": self.ignored_items,
            "missing_items": self.missing_items,
            "scope_decisions": self.scope_decisions,
            "excluded_items": self.excluded_items,
            "no_outbound_required": self.no_outbound_required,
            "warnings": self.warnings,
            "documents": documents,
            "ready": self.ready,
        }

    def _selected_document_set(self) -> set[str]:
        selected = {
            _normalize_name(remark)
            for remark in self.selected_document_remarks
            if _normalize_name(remark)
        }
        if selected:
            selected.add(_normalize_name(self.traveler.order_id))
        return selected

    def _document_is_selected(self, remark: str) -> bool:
        selected = self._selected_document_set()
        return not selected or _normalize_name(remark) in selected

    def document_payloads(self) -> list[dict]:
        mapped: dict[str, list[OutboundItem]] = {}
        for item in self.outbound_items:
            mapped.setdefault(item.document_remark, []).append(item)
        ignored_by_remark: dict[str, list[dict]] = {}
        for item in self.ignored_items:
            ignored_by_remark.setdefault(str(item.get("document_remark", "")), []).append(item)
        return [
            {
                "remark": remark,
                "kind": "materials" if remark == self.traveler.order_id else "hardware",
                "items": [asdict(item) for item in mapped.get(remark, [])],
                "ignored_items": ignored_by_remark.get(remark, []),
            }
            for remark in self.traveler.documents
            if self._document_is_selected(remark)
        ]


def stock_requirements(preview: InventoryPreview, include_hardware: bool = False) -> list[dict]:
    """Return one current-stock requirement per SKU.

    Stock checks are read-only and, by default, cover the order-level material
    document only.  Hardware can be included explicitly and is then aggregated
    across factory orders because the balance page reports stock by SKU.
    """
    if not preview.ready:
        raise RuleError(
            "inventory_precheck",
            "Traveler 存在未映射材料，不能查询完整库存",
            missing_items=preview.missing_items,
        )
    selected = [
        item for item in preview.outbound_items
        if include_hardware or item.section == "板材与封边"
    ]
    return _group_stock_requirements(selected)


def _group_stock_requirements(items: Iterable[OutboundItem]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for item in items:
        row = grouped.setdefault(item.product_code, {
            "productCode": item.product_code,
            "productName": item.product_name,
            "unit": item.unit,
            "requiredQuantity": 0.0,
            "travelerNames": [],
        })
        row["requiredQuantity"] += float(item.quantity)
        if item.traveler_name not in row["travelerNames"]:
            row["travelerNames"].append(item.traveler_name)
    return [grouped[code] for code in sorted(grouped)]


def database_stock_requirements(config: Config, order_id: str) -> tuple[str, list[dict]]:
    """Build order-level stock requirements from SQLite facts only."""
    normalized_order_id = order_id.strip().upper()
    detail, _ = _database_factory_rows(config, normalized_order_id)
    if not detail.get("materials"):
        raise RuleError(
            "inventory_database_missing",
            f"数据库中没有订单 {normalized_order_id} 的材料明细",
        )
    _, documents, _, _ = database_document_items(config, normalized_order_id)
    mappings = InventoryMappings(config.workflow_database)
    outbound: list[OutboundItem] = []
    missing: list[dict] = []
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        for item in documents.get(normalized_order_id, []):
            try:
                outbound.extend(match_item(catalog, mappings, item))
            except RuleError as exc:
                missing.append({"name": item.name, "message": str(exc), **exc.context})
    if missing:
        raise RuleError(
            "inventory_precheck",
            "当前订单存在无法映射到库存商品的板材或封边，不能查询完整库存",
            missing_items=missing,
        )
    return normalized_order_id, _group_stock_requirements(outbound)


def order_stock_requirements(config: Config, order_folder: Path) -> tuple[str, list[dict]]:
    """Map the current source-order material preview to inventory SKUs."""
    database_order_id = order_folder.name.strip().upper()
    if database_order_id:
        try:
            return database_stock_requirements(config, database_order_id)
        except RuleError as exc:
            if exc.code != "inventory_database_missing":
                raise
        except (OSError, sqlite3.Error):
            pass

    from .order_workflow import preview_order

    preview = preview_order(config, order_folder)
    mappings = InventoryMappings(config.workflow_database)
    outbound: list[OutboundItem] = []
    missing: list[dict] = []
    plywood_names = {18.0: "18mm--Plywood", 14.5: "14.5mm--Plywood", 5.4: "5.4mm--Plywood"}
    source_items: list[TravelerItem] = []
    for index, material in enumerate(preview.materials, start=1):
        if material.quantity <= 0:
            continue
        if material.kind == "plywood":
            name = next(
                (label for thickness, label in plywood_names.items() if abs(material.thickness - thickness) < 0.01),
                f"{material.thickness:g}mm--Plywood",
            )
        else:
            name = f"{material.thickness:g}mm--{material.color}"
        source_items.append(TravelerItem(index, "板材与封边", name, material.quantity, preview.order_id))
    for color, quantity in preview.edge_banding.items():
        if quantity > 0:
            source_items.append(TravelerItem(
                len(source_items) + 1,
                "板材与封边",
                f"Edge banding--{color}",
                quantity,
                preview.order_id,
            ))
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        for item in source_items:
            try:
                outbound.extend(match_item(catalog, mappings, item))
            except RuleError as exc:
                missing.append({"name": item.name, "message": str(exc), **exc.context})
    if missing:
        raise RuleError(
            "inventory_precheck",
            "当前订单存在无法映射到库存商品的板材或封边，不能查询完整库存",
            missing_items=missing,
        )
    return preview.order_id, _group_stock_requirements(outbound)


def _database_factory_rows(config: Config, order_id: str) -> tuple[dict, dict[str, dict]]:
    from .order_details import order_detail

    normalized_order_id = order_id.strip().upper()
    if not normalized_order_id:
        raise RuleError("inventory_argument", "数据库出库需要订单号")
    detail = order_detail(config, normalized_order_id)
    factory_rows = {
        str(row.get("factory_order", "")).strip().upper(): row
        for row in detail.get("factory_orders", [])
        if str(row.get("factory_order", "")).strip()
    }
    return detail, factory_rows


def database_document_items(
    config: Config,
    order_id: str,
    selected_factory_orders: Iterable[str] | None = None,
) -> tuple[str, dict[str, list[TravelerItem]], list[TravelerItem], dict[str, dict]]:
    """Build outbound source documents directly from persisted order facts.

    The returned documents intentionally use the same normalized names as a
    Traveler, but no workbook is created or read.  The order-level material
    document is keyed by order id; hardware documents are keyed by the factory
    name so existing outbound records remain addressable.
    """
    detail, factory_rows = _database_factory_rows(config, order_id)
    normalized_order_id = order_id.strip().upper()
    requested = {
        str(value).strip().upper()
        for value in (selected_factory_orders or [])
        if str(value).strip()
    }
    unknown = sorted(requested - set(factory_rows))
    if unknown:
        raise RuleError(
            "inventory_factory_unknown",
            "数据库中找不到所选工厂单，已停止出库：" + "、".join(unknown),
            factory_orders=unknown,
        )

    order_type = str(detail.get("order", {}).get("order_type", "")).strip()
    scope = outbound_scope_decisions(config, normalized_order_id, requested)
    material_requirement = scope["material"]["requirement"]
    materials = list(detail.get("materials", []))
    documents: dict[str, list[TravelerItem]] = {normalized_order_id: []}
    zero_items: list[TravelerItem] = []
    row_number = 1
    if material_requirement != "customer_supplied" and material_requirement not in {"remainder", "not_required"}:
        material_rows = materials
    else:
        material_rows = []
    for row in material_rows:
        kind = str(row.get("material_type", "")).strip().casefold()
        color = str(row.get("color", "")).strip()
        thickness = float(str(row.get("thickness", "0") or "0"))
        quantity = float(row.get("quantity", 0) or 0)
        if kind == "plywood":
            name = f"{thickness:g}mm--Plywood"
        elif kind in {"panel", "back"}:
            name = f"{thickness:g}mm--{color}"
        elif kind == "edge":
            name = f"Edge banding--{color}"
        else:
            raise RuleError("inventory_material", f"数据库中的材料类型暂不支持：{kind or '空白'}")
        item = TravelerItem(row_number, "板材与封边", name, quantity, normalized_order_id)
        row_number += 1
        if quantity <= 0:
            zero_items.append(item)
        else:
            documents[normalized_order_id].append(item)

    for row in detail.get("hardware", []):
        factory_order = str(row.get("factory_order", "")).strip().upper()
        if requested and factory_order not in requested:
            continue
        hardware_requirement = scope["hardware"].get(factory_order, {"requirement": "required"})["requirement"]
        if hardware_requirement in {"remainder", "not_required"}:
            continue
        factory = factory_rows.get(factory_order, {})
        remark = str(factory.get("factory_name", "")).strip() or factory_order
        documents.setdefault(remark, [])
        name = str(row.get("name", "")).strip() or str(row.get("product_code", "")).strip()
        quantity = float(row.get("quantity", 0) or 0)
        item = TravelerItem(row_number, "五金", name, quantity, remark)
        row_number += 1
        if quantity <= 0:
            zero_items.append(item)
        else:
            documents[remark].append(item)

    if not any(documents.values()):
        explicit_no_outbound = material_requirement in {"customer_supplied", "remainder", "not_required"}
        if requested:
            explicit_no_outbound = explicit_no_outbound or all(
                scope["hardware"].get(factory_order, {"requirement": "required"})["requirement"]
                in {"remainder", "not_required"}
                for factory_order in requested
            )
        if not explicit_no_outbound:
            raise RuleError("inventory_empty", f"数据库中没有订单 {normalized_order_id} 可出库的材料或五金")
    return normalized_order_id, documents, zero_items, factory_rows


def outbound_scope_decisions(
    config: Config,
    order_id: str,
    selected_factory_orders: Iterable[str] | None = None,
) -> dict:
    """Return explicit outbound-scope decisions without changing source facts.

    A missing decision is conservative: existing facts are treated as required
    for outbound, while an order with no facts remains unresolved and cannot be
    silently classified as remnant production.
    """
    detail, factory_rows = _database_factory_rows(config, order_id)
    normalized = order_id.strip().upper()
    requested = {
        str(value).strip().upper()
        for value in (selected_factory_orders or [])
        if str(value).strip()
    }
    rows = sqlite3.connect(config.workflow_database)
    rows.row_factory = sqlite3.Row
    try:
        decisions = rows.execute(
            "select scope_type, factory_order, requirement, reason, source_fingerprint, updated_at "
            "from outbound_scope_decisions where order_id=?",
            (normalized,),
        ).fetchall()
    finally:
        rows.close()
    by_key = {(row["scope_type"], row["factory_order"]): dict(row) for row in decisions}
    material_exists = bool(detail.get("materials"))
    material = by_key.get(("material", "")) or {
        "scope_type": "material",
        "factory_order": "",
        "requirement": "required" if material_exists else "pending",
        "reason": "",
        "source_fingerprint": "",
        "updated_at": "",
    }
    factory_ids = requested or set(factory_rows)
    hardware = {}
    for factory_order in sorted(factory_ids):
        if factory_order not in factory_rows:
            continue
        hardware[factory_order] = by_key.get(("hardware", factory_order)) or {
            "scope_type": "hardware",
            "factory_order": factory_order,
            "requirement": "required",
            "reason": "",
            "source_fingerprint": "",
            "updated_at": "",
        }
    return {
        "order_id": normalized,
        "order_type": str(detail.get("order", {}).get("order_type", "")),
        "material": material,
        "hardware": hardware,
        "decisions": [material, *hardware.values()],
    }


def set_outbound_scope(
    config: Config,
    order_id: str,
    scope_type: str,
    requirement: str,
    *,
    factory_order: str = "",
    reason: str = "",
) -> dict:
    normalized = order_id.strip().upper()
    scope_type = scope_type.strip().lower()
    requirement = requirement.strip().lower()
    factory_order = factory_order.strip().upper()
    allowed = {"required", "customer_supplied", "remainder", "not_required"}
    if scope_type not in {"material", "hardware"} or requirement not in allowed:
        raise RuleError("inventory_scope", "出库范围类型或决定无效")
    detail, factory_rows = _database_factory_rows(config, normalized)
    order_type = str(detail.get("order", {}).get("order_type", "")).strip()
    if scope_type == "material" and factory_order:
        raise RuleError("inventory_scope", "订单材料出库范围不能填写工厂单")
    if scope_type == "hardware":
        if not factory_order or factory_order not in factory_rows:
            raise RuleError("inventory_scope", f"数据库中找不到订单 {normalized} 的工厂单：{factory_order or '空白'}")
        if requirement == "customer_supplied":
            raise RuleError("inventory_scope", "客户提供材料只适用于来料加工订单的板材和封边，不适用于五金")
    if scope_type == "material" and requirement == "customer_supplied" and order_type != "cutToSize":
        raise RuleError("inventory_scope", "只有来料加工订单支持“客户提供，不出库”")
    if requirement in {"customer_supplied", "remainder", "not_required"} and not reason.strip():
        raise RuleError("inventory_scope_reason", "不出库决定必须填写原因")
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    connection = sqlite3.connect(config.workflow_database)
    try:
        connection.execute(
            """insert into outbound_scope_decisions(
                order_id, scope_type, factory_order, requirement, reason,
                source_fingerprint, created_at, updated_at
            ) values(?,?,?,?,?,?,?,?)
            on conflict(order_id, scope_type, factory_order) do update set
                requirement=excluded.requirement, reason=excluded.reason,
                source_fingerprint=excluded.source_fingerprint, updated_at=excluded.updated_at""",
            (normalized, scope_type, factory_order, requirement, reason.strip(), "", now, now),
        )
        connection.commit()
    finally:
        connection.close()
    return outbound_scope_decisions(config, normalized, [factory_order] if factory_order else None)


def build_database_preview(
    config: Config,
    order_id: str,
    selected_factory_orders: Iterable[str] | None = None,
) -> InventoryPreview:
    """Map persisted SQLite order facts without generating a Traveler file."""
    normalized_order_id, documents, zero_items, _ = database_document_items(
        config, order_id, selected_factory_orders
    )
    scope = outbound_scope_decisions(config, normalized_order_id, selected_factory_orders)
    detail, factory_rows = _database_factory_rows(config, normalized_order_id)
    excluded_items: list[dict] = []
    material_requirement = scope["material"]["requirement"]
    if material_requirement in {"customer_supplied", "remainder", "not_required"}:
        label = {
            "customer_supplied": "客户提供",
            "remainder": "余料生产",
            "not_required": "不需要出库",
        }[material_requirement]
        for row in detail.get("materials", []):
            kind = str(row.get("material_type", "")).strip().casefold()
            color = str(row.get("color", "")).strip()
            thickness = float(str(row.get("thickness", "0") or "0"))
            name = f"{thickness:g}mm--Plywood" if kind == "plywood" else (
                f"{thickness:g}mm--{color}" if kind in {"panel", "back"}
                else f"Edge banding--{color}" if kind == "edge" else kind
            )
            excluded_items.append({
                "name": name,
                "quantity": float(row.get("quantity", 0) or 0),
                "section": "板材与封边",
                "document_remark": normalized_order_id,
                "status": label,
                "reason": scope["material"].get("reason", ""),
            })
    for row in detail.get("hardware", []):
        factory_order = str(row.get("factory_order", "")).strip().upper()
        if selected_factory_orders and factory_order not in {str(value).strip().upper() for value in selected_factory_orders}:
            continue
        decision = scope["hardware"].get(factory_order, {})
        requirement = decision.get("requirement")
        if requirement not in {"remainder", "not_required"}:
            continue
        excluded_items.append({
            "name": str(row.get("name", "")).strip() or str(row.get("product_code", "")).strip(),
            "quantity": float(row.get("quantity", 0) or 0),
            "section": "五金",
            "document_remark": str(factory_rows.get(factory_order, {}).get("factory_name", "")).strip() or factory_order,
            "status": "余料生产" if requirement == "remainder" else "不需要出库",
            "reason": decision.get("reason", ""),
        })
    mappings = InventoryMappings(config.workflow_database)
    selected_remarks = tuple(
        remark for remark in documents
        if remark != normalized_order_id
    )
    outbound: list[OutboundItem] = []
    ignored: list[dict] = []
    missing: list[dict] = []
    catalog_path = bootstrap_product_database(config)
    catalog_context = (
        ProductDatabase(catalog_path)
        if catalog_path.suffix.lower() in {".sqlite", ".sqlite3", ".db"}
        else None
    )
    try:
        catalog = catalog_context or ProductCatalog(catalog_path)
        for remark, items in documents.items():
            for item in items:
                reason = mappings.ignored_reason(item.name)
                if reason is not None:
                    ignored.append({**asdict(item), "reason": reason})
                    continue
                try:
                    outbound.extend(match_item(catalog, mappings, item))
                except RuleError as exc:
                    missing.append({**asdict(item), "code": exc.code, "message": str(exc), **exc.context})
    finally:
        if catalog_context is not None:
            catalog_context.close()

    duplicate_keys = [
        key for key, count in Counter(
            (item.document_remark, item.product_code) for item in outbound
        ).items()
        if count > 1 and _normalize_name(key[1]) not in {"M1068", "M1069"}
    ]
    if duplicate_keys:
        missing.append({
            "code": "duplicate_product_code",
            "message": "多个数据库订单项目映射到同一商品编号",
            "items": [
                asdict(item) for item in outbound
                if (item.document_remark, item.product_code) in duplicate_keys
            ],
        })

    snapshot = {
        "order_id": normalized_order_id,
        "documents": {
            remark: [asdict(item) for item in items]
            for remark, items in sorted(documents.items())
        },
        "zero_items": [asdict(item) for item in zero_items],
    }
    traveler = TravelerData(
        path=config.workflow_database.resolve(),
        pp_folder=normalized_order_id,
        order_id=normalized_order_id,
        order_name=normalized_order_id,
        items=[item for items in documents.values() for item in items],
        zero_items=zero_items,
        documents=documents,
        modified_at=datetime.now().isoformat(timespec="seconds"),
        fingerprint=_fingerprint(snapshot),
    )
    return InventoryPreview(
        traveler,
        outbound,
        ignored,
        missing,
        selected_document_remarks=selected_remarks,
        source_type="database",
        scope_decisions=scope["decisions"],
        excluded_items=excluded_items,
        no_outbound_required=not outbound and not missing and bool(
            excluded_items or scope["material"]["requirement"] in {"customer_supplied", "remainder", "not_required"}
        ),
    )


def changed_factory_orders_for_documents(
    config: Config,
    order_id: str,
    selected_factory_orders: Iterable[str],
    documents: Iterable[dict],
) -> set[str]:
    """Translate changed outbound document remarks back to factory orders."""
    _, factory_rows = _database_factory_rows(config, order_id)
    changed_remarks = {
        _normalize_name(document.get("remark", ""))
        for document in documents
        if document.get("changed")
    }
    order_key = _normalize_name(order_id)
    material_changed = order_key in changed_remarks
    changed: set[str] = set()
    for value in selected_factory_orders:
        factory_order = str(value).strip().upper()
        factory_name = str(factory_rows.get(factory_order, {}).get("factory_name", "")).strip()
        if material_changed or _normalize_name(factory_order) in changed_remarks or _normalize_name(factory_name) in changed_remarks:
            changed.add(factory_order)
    return changed


def _fingerprint(payload: dict) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def _find_header(row: Iterable, label: str) -> int | None:
    matches = [index for index, value in enumerate(row, 1) if _text(value) == label]
    if len(matches) > 1:
        raise RuleError("traveler_schema", f"Picking List 同一行出现多个“{label}”表头")
    return matches[0] if matches else None


def _order_name_candidates(ws) -> list[str]:
    candidates = []
    for row in ws.iter_rows():
        for index, cell in enumerate(row):
            if _text(cell.value) == "Name/工厂单名称":
                for neighbor in row[index + 1:]:
                    value = _text(neighbor.value)
                    if value:
                        candidates.append(value)
                        break
    return candidates


def _normalized_label(value) -> str:
    return re.sub(r"\s+", "", _text(value)).lower()


def _find_label_row(ws, label: str) -> int:
    wanted = _normalized_label(label)
    for row in range(1, ws.max_row + 1):
        if any(_normalized_label(ws.cell(row, col).value) == wanted for col in range(1, ws.max_column + 1)):
            return row
    raise RuleError("traveler_usage_schema", f"Usage List 缺少“{label}”")


def _displayed_integer(cell, label: str) -> float:
    value = cell.value
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
        raise RuleError("traveler_quantity", f"{label} 不是有效数字：{value}")
    number = float(value)
    if number < 0:
        raise RuleError("traveler_quantity", f"{label} 不能为负数：{number:g}")
    if abs(number - round(number)) <= 1e-9:
        return float(round(number))
    first_format = _text(cell.number_format).split(";", 1)[0]
    if first_format.lower() != "general" and "." not in first_format and re.search(r"[0#?]", first_format):
        return float(Decimal(str(number)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    raise RuleError("traveler_quantity", f"{label} 数量为 {number:g}，板材数量必须为整数")


def _nonnegative_number(cell, label: str) -> float:
    value = cell.value
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
        raise RuleError("traveler_quantity", f"{label} 不是有效数字：{value}")
    number = float(value)
    if number < 0:
        raise RuleError("traveler_quantity", f"{label} 不能为负数：{number:g}")
    return number


def _canonical_usage_color(value: str) -> str:
    color = _text(value)
    if re.fullmatch(r"khaki(?:\s*\(7x9\))?", color, re.IGNORECASE):
        return "Penelope FA44"
    return color


def _usage_list_items(ws, order_id: str) -> tuple[list[TravelerItem], list[TravelerItem]]:
    total_row = _find_label_row(ws, "Total Qty:")
    header_map = {
        _normalized_label(ws.cell(2, col).value): col
        for col in range(1, ws.max_column + 1)
        if _text(ws.cell(2, col).value)
    }
    required_headers = (
        "3/4 Plywood",
        "5/8 Plywood",
        "1/4 Plywood",
        "3/4 Finish Panel",
        "1/4 Finish Panel",
        "Edge Banding (m)",
        "Color",
    )
    missing = [label for label in required_headers if _normalized_label(label) not in header_map]
    if missing:
        raise RuleError("traveler_usage_schema", f"Usage List 缺少列：{', '.join(missing)}")

    totals: dict[str, tuple[int, float, bool]] = {}

    def accumulate(row: int, name: str, value, integer: bool) -> None:
        if value in (None, ""):
            number = 0.0
        elif isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            raise RuleError("traveler_quantity", f"Usage List 第 {row} 行 {name} 不是有效数字：{value}")
        else:
            number = float(value)
        if number < 0:
            raise RuleError("traveler_quantity", f"Usage List 第 {row} 行 {name} 不能为负数：{number:g}")
        first_row, current, current_integer = totals.get(name, (row, 0.0, integer))
        totals[name] = (first_row, current + number, current_integer)

    plywood_columns = (
        ("3/4 Plywood", "18mm--Plywood"),
        ("5/8 Plywood", "14.5mm--Plywood"),
        ("1/4 Plywood", "5.4mm--Plywood"),
    )
    panel_columns = (
        ("3/4 Finish Panel", "19.1mm"),
        ("1/4 Finish Panel", "8mm"),
    )
    color_col = header_map[_normalized_label("Color")]
    edge_col = header_map[_normalized_label("Edge Banding (m)")]
    for row in range(3, total_row):
        for label, name in plywood_columns:
            accumulate(row, name, ws.cell(row, header_map[_normalized_label(label)]).value, True)
        color = _canonical_usage_color(ws.cell(row, color_col).value)
        has_colored_material = any(
            ws.cell(row, header_map[_normalized_label(label)]).value not in (None, "", 0, 0.0)
            for label, _ in panel_columns
        ) or ws.cell(row, edge_col).value not in (None, "", 0, 0.0)
        if has_colored_material and not color:
            raise RuleError("traveler_usage_schema", f"Usage List 第 {row} 行有板材或封边数量但缺少颜色")
        if not color:
            continue
        for label, thickness in panel_columns:
            accumulate(
                row,
                f"{thickness}--{color}",
                ws.cell(row, header_map[_normalized_label(label)]).value,
                True,
            )
        accumulate(row, f"Edge banding--{color}", ws.cell(row, edge_col).value, False)

    positive: list[TravelerItem] = []
    zero: list[TravelerItem] = []
    for name, (row, quantity, integer) in totals.items():
        if integer:
            quantity = float(Decimal(str(quantity)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        item = TravelerItem(row, "板材与封边", name, quantity, order_id)
        (positive if quantity > 0 else zero).append(item)
    return positive, zero


def parse_traveler(path: Path) -> TravelerData:
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        raise RuleError("traveler_missing", f"Traveler 文件不存在或格式不是 xlsx：{path}")
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise RuleError("traveler_open", f"无法打开 Traveler：{path.name}") from exc
    required = {"WorkOrderTraveler", "Usage List", "Picking List"}
    if not required.issubset(wb.sheetnames):
        raise RuleError("traveler_schema", f"Traveler 缺少必要工作表：{sorted(required - set(wb.sheetnames))}")
    work_names = _order_name_candidates(wb["WorkOrderTraveler"])
    work_order = work_names[0].upper() if work_names else ""
    usage = wb["Usage List"]
    usage_order = ""
    for row in usage.iter_rows():
        for index, cell in enumerate(row):
            if _normalized_label(cell.value) == "job:":
                usage_order = next((_text(item.value).upper() for item in row[index + 1:] if _text(item.value)), "")
                break
        if usage_order:
            break
    folder_order = next((parent.name.upper() for parent in path.parents if PP_FOLDER_RE.fullmatch(parent.name)), "")
    temporary_folder_name = path.parent.name.strip()
    temporary_order_id = temporary_folder_name.upper() if temporary_folder_name and not folder_order and not PP_FOLDER_RE.fullmatch(temporary_folder_name) else ""
    filename_match = re.fullmatch(r"Work Order Traveler\(([^)]+)\)\.xlsx", path.name, re.IGNORECASE)
    filename_order = filename_match.group(1).strip().upper() if filename_match else ""
    candidates = [value for value in (folder_order, filename_order, usage_order, work_order) if PP_FOLDER_RE.fullmatch(value)]
    temporary_identity = not candidates and bool(temporary_order_id)
    if not candidates:
        if not temporary_identity:
            raise RuleError("traveler_identity", "无法从订单文件夹、文件名、Usage List 或 WorkOrderTraveler 取得订单号")
        order_id = temporary_order_id
    else:
        order_id = candidates[0]
    if any(value != order_id for value in candidates[1:]):
        raise RuleError("traveler_identity", f"Traveler 订单号不一致：{candidates}")

    ws = wb["Picking List"]
    section = ""
    current_factory = ""
    name_col = qty_col = None
    fitting_totals: dict[tuple[str, str], TravelerItem] = {}
    fitting_zero: list[TravelerItem] = []
    documents: dict[str, list[TravelerItem]] = {}
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        for index, value in enumerate(row):
            if _text(value) == "Name/工厂单名称":
                current_factory = next((_text(item) for item in row[index + 1:] if _text(item)), "")
                if current_factory:
                    documents.setdefault(order_id if temporary_identity else current_factory, [])
                section = ""
                name_col = qty_col = None
                break
        first = _text(row[0]) if row else ""
        if first and not isinstance(row[0], (int, float)) and first not in {"No.", "Picking List领料单", "Name/工厂单名称"}:
            if _find_header(row, "Name名字") is None and _find_header(row, "QTY数量") is None:
                section = first
                name_col = qty_col = None
                continue
        found_name = _find_header(row, "Name名字")
        found_qty = _find_header(row, "QTY数量")
        if found_name or found_qty:
            if not found_name or not found_qty:
                raise RuleError("traveler_schema", f"Picking List 第 {row_number} 行名称与数量表头不完整")
            name_col, qty_col = found_name, found_qty
            continue
        if not row or not isinstance(row[0], (int, float)):
            continue
        if not name_col or not qty_col:
            raise RuleError("traveler_schema", f"Picking List 第 {row_number} 行数据没有对应表头")
        name = _text(row[name_col - 1] if name_col <= len(row) else None)
        value = row[qty_col - 1] if qty_col <= len(row) else None
        if not name:
            if value not in (None, "", 0, 0.0):
                raise RuleError("traveler_schema", f"Picking List 第 {row_number} 行有数量但没有名称")
            continue
        if not current_factory:
            raise RuleError("traveler_identity", f"Picking List 第 {row_number} 行五金没有对应工厂单名称")
        if value in (None, "", 0, 0.0):
            fitting_zero.append(TravelerItem(row_number, section, name, 0.0, order_id if temporary_identity else current_factory))
            continue
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise RuleError("traveler_quantity", f"{name} 的数量不是有效数字：{value}", row=row_number)
        if float(value) < 0:
            raise RuleError("traveler_quantity", f"{name} 的数量不能为负数：{value}", row=row_number)
        document_remark = order_id if temporary_identity else current_factory
        key = (document_remark, _normalize_name(name))
        if key in fitting_totals:
            fitting_totals[key].quantity += float(value)
        else:
            fitting_totals[key] = TravelerItem(row_number, section, name, float(value), document_remark)
    material_items, material_zero = _usage_list_items(usage, order_id)
    documents[order_id] = material_items
    for item in fitting_totals.values():
        documents.setdefault(item.document_remark, []).append(item)
    items = material_items + list(fitting_totals.values())
    zero_items = material_zero + fitting_zero
    pp_folder = folder_order or path.parent.name
    snapshot = {
        "order_id": order_id,
        "documents": {
            remark: [asdict(item) for item in values]
            for remark, values in sorted(documents.items())
        },
        "zero_items": [asdict(item) for item in zero_items],
    }
    return TravelerData(
        path=path.resolve(),
        pp_folder=pp_folder,
        order_id=order_id,
        order_name=order_id,
        items=items,
        zero_items=zero_items,
        documents=documents,
        modified_at=datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        fingerprint=_fingerprint(snapshot),
    )


def _catalog_cost_price(value, label: str) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise RuleError("product_catalog_data", f"商品 {label} 的预计采购价不是有效数字：{value}")
    try:
        price = float(value)
    except (TypeError, ValueError) as exc:
        raise RuleError("product_catalog_data", f"商品 {label} 的预计采购价不是有效数字：{value}") from exc
    if not math.isfinite(price) or price < 0:
        raise RuleError("product_catalog_data", f"商品 {label} 的预计采购价不能为负数或非有限数字：{value}")
    return price


class ProductCatalog:
    def __init__(self, path: Path):
        self.path = path
        self.products: list[Product] = []
        self.by_code: dict[str, list[Product]] = {}
        self._load()

    def _load(self) -> None:
        if not self.path.is_file():
            raise RuleError("product_catalog_missing", f"尚未导入库存商品资料：{self.path}")
        try:
            wb = load_workbook(self.path, data_only=True, read_only=True)
        except Exception as exc:
            raise RuleError("product_catalog_open", f"无法打开库存商品资料：{self.path.name}") from exc
        ws = wb[wb.sheetnames[0]]
        header_row = None
        header_map = {}
        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            values = {_text(value): index for index, value in enumerate(row)}
            if REQUIRED_PRODUCT_HEADERS.issubset(values):
                header_row = row_number
                header_map = values
                break
        if not header_row:
            raise RuleError("product_catalog_schema", f"商品资料缺少必要表头：{sorted(REQUIRED_PRODUCT_HEADERS)}")
        optional = {"备注": None, "计量单位": None, "预计采购价": None}
        for label in optional:
            optional[label] = header_map.get(label)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            code = _text(row[header_map["*商品编号"]] if header_map["*商品编号"] < len(row) else None)
            name = _text(row[header_map["商品名称"]] if header_map["商品名称"] < len(row) else None)
            if not code and not name:
                continue
            product = Product(
                category=_text(row[header_map["商品类别"]] if header_map["商品类别"] < len(row) else None),
                code=code,
                name=name,
                spec=_text(row[header_map["规格型号"]] if header_map["规格型号"] < len(row) else None),
                status=_text(row[header_map["状态"]] if header_map["状态"] < len(row) else None),
                remark=_text(row[optional["备注"]] if optional["备注"] is not None and optional["备注"] < len(row) else None),
                unit=_text(row[optional["计量单位"]] if optional["计量单位"] is not None and optional["计量单位"] < len(row) else None),
                cost_price=_catalog_cost_price(
                    row[optional["预计采购价"]]
                    if optional["预计采购价"] is not None and optional["预计采购价"] < len(row)
                    else None,
                    code or name,
                ),
            )
            self.products.append(product)
            self.by_code.setdefault(code.upper(), []).append(product)

    def require_code(self, code: str) -> Product:
        matches = self.by_code.get(code.upper(), [])
        if len(matches) != 1:
            raise RuleError("product_conflict", f"商品编号 {code} 匹配到 {len(matches)} 条记录", product_code=code)
        product = matches[0]
        if not product.code or not product.name:
            raise RuleError("product_invalid", f"当前需要的商品 {code} 缺少编号或名称")
        if product.status and product.status != "启用":
            raise RuleError("product_disabled", f"商品已停用：{code} {product.name}")
        return product

    def find(self, *, category: str | None = None, name: str | None = None, contains: str | None = None,
             spec_thickness: float | None = None) -> list[Product]:
        results = self.products
        if category:
            results = [item for item in results if _normalize_name(item.category) == _normalize_name(category)]
        if name:
            results = [item for item in results if _normalize_name(item.name) == _normalize_name(name)]
        if contains:
            token = _normalize_name(contains)
            results = [item for item in results if token in _normalize_name(item.name) or token in _normalize_name(item.remark)]
        if spec_thickness is not None:
            aliases = {14.5: 15.0, 5.4: 5.2, 8.0: 9.0}
            requested = float(spec_thickness)
            expected_values = (8.0, 9.0) if requested in {8.0, 9.0} else (
                aliases.get(requested, requested),
            )
            results = [
                item for item in results
                if (numbers := re.findall(r"\d+(?:\.\d+)?", item.spec))
                and any(abs(float(numbers[0]) - expected) < 0.6 for expected in expected_values)
            ]
        return results


def _ensure_product_cost_column(connection: sqlite3.Connection) -> None:
    columns = {
        str(row[1])
        for row in connection.execute("pragma table_info(products)").fetchall()
    }
    if "cost_price" not in columns:
        connection.execute("alter table products add column cost_price real")
        connection.commit()


class ProductDatabase:
    """SQLite-backed product catalog used by runtime lookups."""

    def __init__(self, path: Path):
        self.path = path
        if not path.is_file():
            raise RuleError("product_database_missing", f"尚未导入库存商品数据库：{path}")
        try:
            self.connection = sqlite3.connect(path)
            self.connection.set_trace_callback(lambda statement: log_database_statement(path, statement))
            version = int(self.connection.execute("pragma user_version").fetchone()[0])
            # The product catalog lives in the shared workflow database, whose
            # user_version is owned by the workflow schema.
            if version not in (0, INVENTORY_DATABASE_VERSION) and version < 1:
                raise RuleError("product_database_schema", f"库存商品数据库版本不受支持：{version}")
            if self.connection.execute(
                "select 1 from sqlite_master where type = 'table' and name = 'products'"
            ).fetchone() is None:
                raise RuleError("product_database_schema", "库存商品数据库缺少 products 表")
            _ensure_product_cost_column(self.connection)
        except RuleError:
            self.connection.close()
            raise
        except sqlite3.Error as exc:
            if hasattr(self, "connection"):
                self.connection.close()
            raise RuleError("product_database_open", f"无法打开库存商品数据库：{path}") from exc

    def close(self) -> None:
        self.connection.close()

    def __enter__(self) -> "ProductDatabase":
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        self.close()

    @staticmethod
    def _from_row(row: sqlite3.Row | tuple) -> Product:
        return Product(
            category=str(row[0] or ""),
            code=str(row[1] or ""),
            name=str(row[2] or ""),
            spec=str(row[3] or ""),
            status=str(row[4] or ""),
            remark=str(row[5] or ""),
            unit=str(row[6] or ""),
            cost_price=None if row[7] is None else float(row[7]),
        )

    @property
    def products(self) -> list[Product]:
        rows = self.connection.execute(
            """
            select category, code, name, spec, status, remark, unit, cost_price
            from products order by code
            """
        ).fetchall()
        return [self._from_row(row) for row in rows]

    def count(self) -> int:
        return int(self.connection.execute("select count(*) from products").fetchone()[0])

    def require_code(self, code: str) -> Product:
        rows = self.connection.execute(
            """
            select category, code, name, spec, status, remark, unit, cost_price
            from products where normalized_code = ?
            """,
            (_normalize_name(code),),
        ).fetchall()
        if len(rows) != 1:
            raise RuleError("product_conflict", f"商品编号 {code} 匹配到 {len(rows)} 条记录", product_code=code)
        product = self._from_row(rows[0])
        if not product.code or not product.name:
            raise RuleError("product_invalid", f"当前需要的商品 {code} 缺少编号或名称")
        if product.status and product.status != "启用":
            raise RuleError("product_disabled", f"商品已停用：{code} {product.name}")
        return product

    def find(
        self,
        *,
        category: str | None = None,
        name: str | None = None,
        contains: str | None = None,
        spec_thickness: float | None = None,
    ) -> list[Product]:
        clauses = []
        parameters: list[str] = []
        if category:
            clauses.append("normalized_category = ?")
            parameters.append(_normalize_name(category))
        if name:
            clauses.append("normalized_name = ?")
            parameters.append(_normalize_name(name))
        if contains:
            token = _normalize_name(contains)
            clauses.append(
                "(normalized_code like ? or normalized_name like ? or normalized_spec like ? "
                "or normalized_category like ? or normalized_remark like ?)"
            )
            parameters.extend([f"%{token}%"] * 5)
        query = """
            select category, code, name, spec, status, remark, unit, cost_price
            from products
        """
        if clauses:
            query += " where " + " and ".join(clauses)
        query += " order by code"
        results = [
            self._from_row(row)
            for row in self.connection.execute(query, parameters).fetchall()
        ]
        if spec_thickness is None:
            return results
        aliases = {14.5: 15.0, 5.4: 5.2, 8.0: 9.0}
        requested = float(spec_thickness)
        expected_values = (8.0, 9.0) if requested in {8.0, 9.0} else (
            aliases.get(requested, requested),
        )
        return [
            product
            for product in results
            if (numbers := re.findall(r"\d+(?:\.\d+)?", product.spec))
            and any(abs(float(numbers[0]) - expected) < 0.6 for expected in expected_values)
        ]


def _create_product_database(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path)
    connection.set_trace_callback(lambda statement: log_database_statement(path, statement))
    try:
        connection.executescript(
            f"""
            create table if not exists products(
                category text not null default '',
                code text primary key,
                name text not null default '',
                spec text not null default '',
                status text not null default '',
                remark text not null default '',
                unit text not null default '',
                cost_price real,
                normalized_code text not null,
                normalized_name text not null,
                normalized_spec text not null,
                normalized_category text not null,
                normalized_remark text not null
            );
            create index if not exists idx_products_name on products(normalized_name);
            create index if not exists idx_products_category on products(normalized_category);
            """
        )
        _ensure_product_cost_column(connection)
    finally:
        connection.close()


def _replace_product_database(path: Path, products: list[Product]) -> None:
    missing = [product.code or product.name for product in products if not product.code or not product.name]
    duplicate_codes = [
        code for code, count in Counter(_normalize_name(product.code) for product in products).items()
        if code and count > 1
    ]
    if missing:
        raise RuleError("product_catalog_data", "商品资料存在缺少商品编号或商品名称的记录")
    if duplicate_codes:
        raise RuleError(
            "product_catalog_data",
            f"商品资料存在重复商品编号：{', '.join(duplicate_codes[:10])}",
            product_codes=duplicate_codes,
        )
    _create_product_database(path)
    connection = sqlite3.connect(path)
    connection.set_trace_callback(lambda statement: log_database_statement(path, statement))
    try:
        connection.execute("begin immediate")
        connection.execute("delete from products")
        connection.executemany(
            """
            insert into products(
                category, code, name, spec, status, remark, unit, cost_price,
                normalized_code, normalized_name, normalized_spec,
                normalized_category, normalized_remark
            ) values(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            [
                (
                    product.category,
                    product.code,
                    product.name,
                    product.spec,
                    product.status,
                    product.remark,
                    product.unit,
                    product.cost_price,
                    _normalize_name(product.code),
                    _normalize_name(product.name),
                    _normalize_name(product.spec),
                    _normalize_name(product.category),
                    _normalize_name(product.remark),
                )
                for product in products
            ],
        )
        connection.execute(f"pragma user_version = {INVENTORY_DATABASE_VERSION}")
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


class InventoryMappings:
    """Database-backed material rules.

    A temporary JSON compatibility mode remains for isolated legacy fixtures;
    application call sites pass the central workflow SQLite database. Runtime
    data is therefore read from and written to SQLite only.
    """
    def __init__(self, path: Path):
        self.path = path
        self._legacy_path = path if path.suffix.lower() == ".json" else None
        self.database_path = (
            path.parent.parent / "workflow.sqlite3"
            if self._legacy_path is not None
            else path.parent / "workflow.sqlite3"
            if path.name == "order-index.sqlite3"
            else path
        )
        self.manual: dict[str, str] = {}
        self.ignored: dict[str, str] = {}
        if self._legacy_path is not None:
            if path.is_file():
                try:
                    data = json.loads(path.read_text(encoding="utf-8"))
                    self.manual = {_normalize_name(k): str(v) for k, v in data.get("manual", {}).items()}
                    self.ignored = {_normalize_name(k): str(v) for k, v in data.get("ignored", {}).items()}
                except (OSError, json.JSONDecodeError) as exc:
                    raise RuleError("inventory_mapping", f"库存商品映射无法读取：{path}") from exc
            return
        ensure_schema(self.database_path)

    def _rows(self, rule_type: str | None = None) -> list[sqlite3.Row]:
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        try:
            if rule_type:
                return connection.execute(
                    "select source_name, normalized_name, product_code, reason from inventory_resolution_rules "
                    "where rule_type=? order by source_name",
                    (rule_type,),
                ).fetchall()
            return connection.execute(
                "select source_name, normalized_name, product_code, reason from inventory_resolution_rules "
                "order by source_name"
            ).fetchall()
        finally:
            connection.close()

    def entries(self) -> tuple[list[dict], list[dict]]:
        if self._legacy_path is not None:
            return (
                [{"name": name, "product_code": code} for name, code in sorted(self.manual.items())],
                [{"name": name, "reason": reason} for name, reason in sorted(self.ignored.items())],
            )
        rows = self._rows()
        manual = [
            {"name": str(row["source_name"]), "product_code": str(row["product_code"] or "")}
            for row in rows if row["product_code"]
        ]
        ignored = [
            {"name": str(row["source_name"]), "reason": str(row["reason"] or "")}
            for row in rows if not row["product_code"]
        ]
        return manual, ignored

    def ignored_reason(self, name: str) -> str | None:
        normalized = _normalize_name(name)
        if self._legacy_path is not None:
            return self.ignored.get(normalized)
        connection = sqlite3.connect(self.database_path)
        try:
            row = connection.execute(
                "select reason from inventory_resolution_rules where normalized_name=? and rule_type='ignore'",
                (normalized,),
            ).fetchone()
            return str(row[0]) if row else None
        finally:
            connection.close()

    def manual_code(self, name: str) -> str | None:
        normalized = _normalize_name(name)
        if self._legacy_path is not None:
            return self.manual.get(normalized)
        connection = sqlite3.connect(self.database_path)
        try:
            row = connection.execute(
                "select product_code from inventory_resolution_rules where normalized_name=? and rule_type='mapping'",
                (normalized,),
            ).fetchone()
            return str(row[0]) if row else None
        finally:
            connection.close()

    def save_ignored(self, name: str, reason: str) -> None:
        normalized = _normalize_name(name)
        if not normalized:
            raise RuleError("inventory_mapping", "忽略材料名称不能为空")
        value = reason.strip() or "用户在出库预览中选择忽略"
        if self._legacy_path is not None:
            self.ignored[normalized] = value
            self._save()
            return
        self._upsert("ignore", name.strip(), normalized, None, value)

    def save_manual(self, name: str, product_code: str) -> None:
        normalized = _normalize_name(name)
        if not normalized:
            raise RuleError("inventory_mapping", "映射材料名称不能为空")
        code = product_code.strip().upper()
        if not code:
            raise RuleError("inventory_mapping", "映射商品 SKU 不能为空")
        if self._legacy_path is not None:
            self.manual[normalized] = code
            self.ignored.pop(normalized, None)
            self._save()
            return
        self._upsert("mapping", name.strip(), normalized, code, "")

    def remove_ignored(self, name: str) -> None:
        normalized = _normalize_name(name)
        if self._legacy_path is not None:
            self.ignored.pop(normalized, None)
            self._save()
            return
        self._delete(normalized, "ignore")

    def remove_manual(self, name: str) -> None:
        normalized = _normalize_name(name)
        if self._legacy_path is not None:
            self.manual.pop(normalized, None)
            self._save()
            return
        self._delete(normalized, "mapping")

    def _upsert(self, rule_type: str, source_name: str, normalized: str, product_code: str | None, reason: str) -> None:
        connection = sqlite3.connect(self.database_path)
        try:
            connection.execute(
                """insert into inventory_resolution_rules(
                    rule_type, source_name, normalized_name, product_code, reason, created_at, updated_at
                ) values(?,?,?,?,?,?,?)
                on conflict(normalized_name) do update set
                    rule_type=excluded.rule_type,
                    source_name=excluded.source_name,
                    product_code=excluded.product_code,
                    reason=excluded.reason,
                    updated_at=excluded.updated_at""",
                (rule_type, source_name, normalized, product_code, reason, datetime.now().astimezone().isoformat(timespec="seconds"), datetime.now().astimezone().isoformat(timespec="seconds")),
            )
            connection.commit()
        except sqlite3.Error as exc:
            connection.rollback()
            raise RuleError("inventory_mapping", f"库存规则无法保存：{source_name}") from exc
        finally:
            connection.close()

    def _delete(self, normalized: str, rule_type: str) -> None:
        connection = sqlite3.connect(self.database_path)
        try:
            connection.execute(
                "delete from inventory_resolution_rules where normalized_name=? and rule_type=?",
                (normalized, rule_type),
            )
            connection.commit()
        finally:
            connection.close()

    def _save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        payload = {"version": 1, "manual": self.manual, "ignored": self.ignored}
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(self.path)


HARDWARE_DISPLAY_NAMES = {
    "WJCBT": "Shelf Holder",
    "71T950A": "Hinge",
    "HRAIL": "H-Rail",
    "LRAIL": "L-Rail",
}


def ignored_hardware_reason(mappings: InventoryMappings, name: str = "", code: str = "") -> str | None:
    """Match ignored hardware against raw report names, codes, and display aliases."""
    candidates = [name, code]
    display_name = HARDWARE_DISPLAY_NAMES.get(_normalize_name(code))
    if display_name:
        candidates.append(display_name)
    for candidate in candidates:
        reason = mappings.ignored_reason(candidate)
        if reason is not None:
            return reason
    return None


def remove_ignored_hardware_records(config: Config, names: Iterable[str]) -> int:
    """Remove persisted hardware facts that the user has globally ignored."""
    if not config.storage_prepared:
        return 0
    ignored_names = {_normalize_name(name) for name in names if _normalize_name(name)}
    if not ignored_names or not config.workflow_database.is_file():
        return 0
    connection = sqlite3.connect(config.workflow_database)
    try:
        rows = connection.execute(
            "select id, name, product_code from hardware_items"
        ).fetchall()
        ids = [
            row[0]
            for row in rows
            if any(
                _normalize_name(candidate) in ignored_names
                for candidate in (
                    row[1],
                    row[2],
                    HARDWARE_DISPLAY_NAMES.get(_normalize_name(row[2]), ""),
                )
            )
        ]
        if ids:
            connection.executemany("delete from hardware_items where id=?", ((row_id,) for row_id in ids))
            connection.commit()
        return len(ids)
    finally:
        connection.close()


FIXED_CODES = {
    "18MMPLYWOOD": "M0004",
    "14.5MMPLYWOOD": "M0003",
    "5.4MMPLYWOOD": "M0002",
    "ADJUSTABLESHELFHOLDER": "M1013",
    "SHELHOLDER": "M1013",
    "HINGE": "M1001",
    "TESTFULLHINGE": "M1001",
    "HRAIL": "M1002",
    "LRAIL": "M1003",
    "M.C(L)": "M1089",
}


def _single(catalog: ProductCatalog, matches: list[Product], traveler_name: str, source: str) -> tuple[Product, str]:
    if len(matches) != 1:
        raise RuleError(
            "product_match",
            f"{traveler_name} 匹配到 {len(matches)} 个库存商品，需要人工指定",
            traveler_name=traveler_name,
            candidates=[asdict(item) for item in matches],
        )
    return catalog.require_code(matches[0].code), source


def match_item(catalog: ProductCatalog, mappings: InventoryMappings, item: TravelerItem) -> list[OutboundItem]:
    def outbound(product: Product, quantity: float, source: str) -> OutboundItem:
        return OutboundItem(
            item.name, product.code, product.name, quantity, item.section, source,
            item.document_remark, product.unit,
        )

    def edge_outbound(product: Product, source: str) -> OutboundItem:
        rounded = float(Decimal(str(item.quantity)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        return outbound(
            product,
            rounded,
            f"{source}（{item.quantity:g}m→{rounded:g}m，四舍五入取整）",
        )

    ignored = mappings.ignored_reason(item.name)
    if ignored is not None:
        return []
    manual = mappings.manual_code(item.name)
    if manual:
        product = catalog.require_code(manual)
        if _normalize_name(product.category) == _normalize_name("Edge band"):
            return [edge_outbound(product, "人工指定")]
        return [outbound(product, item.quantity, "人工指定")]
    normalized = _normalize_name(item.name)
    if normalized == "PUSHOPEN":
        results = []
        for code in ("M1068", "M1069"):
            product = catalog.require_code(code)
            results.append(outbound(product, item.quantity, "Push Open 1:1拆分"))
        return results
    fixed = FIXED_CODES.get(normalized)
    if fixed:
        product = catalog.require_code(fixed)
        return [outbound(product, item.quantity, "已确认规则")]
    if match := TB_RE.fullmatch(item.name.strip()):
        token = f"TB{match.group(1)}用"
        product, source = _single(catalog, catalog.find(category="Hardware/Trash Can", contains=token), item.name, "TB型号规则")
        return [outbound(product, item.quantity, source)]
    if match := EDGE_RE.fullmatch(item.name.strip()):
        color = match.group(1).strip()
        matches = catalog.find(category="Edge band", name=f"{color} Edge Banding")
        source = "封边颜色精确名称"
        if not matches:
            color_token = _normalize_name(color)
            matches = [
                product for product in catalog.find(category="Edge band")
                if (normalized := _normalize_name(product.name)).startswith(color_token)
                and "ABSBANDING" in normalized
                and normalized.endswith("24MM")
            ]
            source = "封边颜色 + ABS banding + 24mm后缀"
        product, source = _single(catalog, matches, item.name, source)
        return [edge_outbound(product, source)]
    if match := PANEL_RE.fullmatch(item.name.strip()):
        thickness = float(match.group(1))
        color = match.group(2).strip()
        if color.lower() != "plywood":
            product, source = _single(
                catalog,
                catalog.find(category="Panel", name=color, spec_thickness=thickness),
                item.name,
                "Panel颜色和规格",
            )
            return [outbound(product, item.quantity, source)]
    exact_matches = catalog.find(name=item.name)
    if exact_matches:
        product, source = _single(catalog, exact_matches, item.name, "库存商品名称精确匹配")
        return [outbound(product, item.quantity, source)]
    raise RuleError(
        "product_match",
        f"找不到与 {item.name} 名称完全一致的库存商品，需要人工指定",
        traveler_name=item.name,
        candidates=[],
    )


def resolve_inventory_items(
    config: Config,
    items: Iterable[tuple[TravelerItem, str]],
) -> dict:
    """Resolve order materials/hardware before they become active database facts.

    ``source_code`` is retained for hardware because AICNC codes such as
    ``WJ-CBD`` are source identifiers, not necessarily inventory SKUs.  An
    item is accepted only when it is explicitly ignored or resolves to a
    unique enabled catalog product through the normal mapping rules.
    """
    items = list(items)
    if not items:
        return {"outbound": [], "ignored": [], "missing": []}
    mappings = InventoryMappings(config.workflow_database)
    outbound: list[OutboundItem] = []
    ignored: list[dict] = []
    missing: list[dict] = []
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        for item, source_code in items:
            source_code = _text(source_code)
            # Known source codes have a stable display name in the workflow.
            # Use that name only for catalog matching; keep the original
            # source code on the resolution record and in persisted facts.
            match_item_value = item
            display_name = HARDWARE_DISPLAY_NAMES.get(_normalize_name(source_code))
            if item.section == "五金" and display_name and not _normalize_name(item.name) == _normalize_name(display_name):
                match_item_value = replace(item, name=display_name)
            reason = (
                ignored_hardware_reason(mappings, item.name, source_code)
                if item.section == "五金"
                else mappings.ignored_reason(item.name)
            )
            if reason is not None:
                ignored.append({
                    **asdict(item),
                    "source_code": source_code,
                    "reason": reason,
                })
                continue
            try:
                outbound.extend(match_item(catalog, mappings, match_item_value))
            except RuleError as exc:
                missing.append({
                    **asdict(item),
                    "source_code": source_code,
                    "code": exc.code,
                    "message": str(exc),
                    **exc.context,
                })
    return {"outbound": outbound, "ignored": ignored, "missing": missing}


def build_preview(
    path: Path,
    catalog_path: Path,
    mapping_path: Path,
    selected_document_remarks: Iterable[str] | None = None,
) -> InventoryPreview:
    traveler = parse_traveler(path)
    mappings = InventoryMappings(mapping_path)
    requested_remarks = tuple(
        dict.fromkeys(
            str(remark).strip()
            for remark in (selected_document_remarks or [])
            if str(remark).strip()
        )
    )
    selected = {
        _normalize_name(remark)
        for remark in requested_remarks
    }
    # The order-level material document is always part of a factory-order
    # outbound selection. Hardware documents are narrowed to the selected
    # factory-order names. Missing names are tolerated because CS orders can
    # legitimately have no hardware block in the Traveler.
    if selected:
        selected.add(_normalize_name(traveler.order_id))
    outbound = []
    ignored = []
    missing = []
    catalog_context = (
        ProductDatabase(catalog_path)
        if catalog_path.suffix.lower() in {".sqlite", ".sqlite3", ".db"}
        else None
    )
    try:
        catalog = catalog_context or ProductCatalog(catalog_path)
        for item in traveler.items:
            if selected and _normalize_name(item.document_remark) not in selected:
                continue
            reason = mappings.ignored_reason(item.name)
            if reason is not None:
                ignored.append({**asdict(item), "reason": reason})
                continue
            try:
                outbound.extend(match_item(catalog, mappings, item))
            except RuleError as exc:
                missing.append({
                    **asdict(item),
                    "code": exc.code,
                    "message": str(exc),
                    **exc.context,
                })
    finally:
        if catalog_context is not None:
            catalog_context.close()
    duplicate_keys = [
        key for key, count in Counter(
            (item.document_remark, item.product_code) for item in outbound
        ).items()
        if count > 1 and _normalize_name(key[1]) not in {"M1068", "M1069"}
    ]
    if duplicate_keys:
        conflicts = [
            asdict(item) for item in outbound
            if (item.document_remark, item.product_code) in duplicate_keys
        ]
        missing.append({"code": "duplicate_product_code", "message": "多个 Traveler 项目映射到同一商品编号", "items": conflicts})
    return InventoryPreview(
        traveler,
        outbound,
        ignored,
        missing,
        selected_document_remarks=requested_remarks,
    )


def set_ignored_mapping(config: Config, name: str, ignored: bool, reason: str = "") -> dict:
    mappings = InventoryMappings(config.workflow_database)
    if ignored:
        mappings.save_ignored(name, reason)
        removed_database_rows = remove_ignored_hardware_records(config, [name])
    else:
        mappings.remove_ignored(name)
        removed_database_rows = 0
    return {
        "ok": True,
        "traveler_name": name,
        "ignored": ignored,
        "removed_database_rows": removed_database_rows,
    }


def update_ignored_mapping(config: Config, old_name: str, name: str, reason: str = "") -> dict:
    """Rename or edit one global ignore entry and clean matching facts."""
    mappings = InventoryMappings(config.workflow_database)
    old_normalized = _normalize_name(old_name)
    new_normalized = _normalize_name(name)
    if not old_normalized or not new_normalized:
        raise RuleError("inventory_mapping", "忽略项目名称不能为空")
    if old_normalized != new_normalized:
        mappings.remove_ignored(old_name)
    mappings.save_ignored(name, reason)
    removed_database_rows = remove_ignored_hardware_records(config, [name])
    return {
        "ok": True,
        "old_name": old_name,
        "traveler_name": name,
        "ignored": True,
        "removed_database_rows": removed_database_rows,
    }


def list_inventory_mappings(config: Config) -> dict:
    mappings = InventoryMappings(config.workflow_database)
    manual, ignored = mappings.entries()
    return {
        "ok": True,
        "manual": {item["name"]: item["product_code"] for item in manual},
        "ignored": {item["name"]: item["reason"] for item in ignored},
    }


def search_inventory_products(config: Config, query: str) -> dict:
    token = _normalize_name(query)
    if not token:
        raise RuleError("inventory_argument", "请输入商品编号、名称或规格")
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        matches = [
            product for product in catalog.find(contains=token)
            if not product.status or product.status == "启用"
        ]
    matches.sort(key=lambda product: (
        0 if token in {_normalize_name(product.code), _normalize_name(product.name)} else 1,
        product.code,
    ))
    return {"query": query, "products": [asdict(product) for product in matches[:50]]}


def save_manual_mapping(config: Config, name: str, product_code: str) -> dict:
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        product = catalog.require_code(product_code)
    mappings = InventoryMappings(config.workflow_database)
    mappings.save_manual(name, product.code)
    return {
        "ok": True,
        "traveler_name": name,
        "product": asdict(product),
    }


def remove_manual_mapping(config: Config, name: str) -> dict:
    mappings = InventoryMappings(config.workflow_database)
    mappings.remove_manual(name)
    return {"ok": True, "traveler_name": name, "removed": True}


def update_manual_mapping(config: Config, old_name: str, name: str, product_code: str) -> dict:
    mappings = InventoryMappings(config.workflow_database)
    old_normalized = _normalize_name(old_name)
    new_normalized = _normalize_name(name)
    if not old_normalized or not new_normalized:
        raise RuleError("inventory_mapping", "映射材料名称不能为空")
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        product = catalog.require_code(product_code)
    if old_normalized != new_normalized:
        mappings.remove_manual(old_name)
    mappings.save_manual(name, product.code)
    return {
        "ok": True,
        "old_name": old_name,
        "traveler_name": name,
        "product": asdict(product),
    }


class InventorySyncStore:
    def __init__(self, path: Path, backup_root: Path):
        self.path = path
        self.backup_root = backup_root / "Inventory Sync Records"
        self.data = {"version": 2, "records": {}}
        if path.is_file():
            try:
                self.data = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise RuleError("inventory_sync", f"库存同步记录无法读取：{path}") from exc

    @staticmethod
    def key(order_id: str, remark: str) -> str:
        return hashlib.sha256(
            f"{_normalize_name(order_id)}\n{_normalize_name(remark)}".encode()
        ).hexdigest()

    @staticmethod
    def raw_document_fingerprint(items: list[TravelerItem]) -> str:
        return _fingerprint({
            "items": sorted(
                (
                    _normalize_name(item.name),
                    float(item.quantity),
                )
                for item in items
            )
        })

    @staticmethod
    def mapped_document_fingerprint(items: list[OutboundItem]) -> str:
        return _fingerprint({
            "items": sorted(
                (item.product_code.upper(), float(item.quantity))
                for item in items
            )
        })

    def record_for_document(self, order_id: str, remark: str) -> dict | None:
        return self.data.get("records", {}).get(self.key(order_id, remark))

    def records_for_order(self, order_id: str) -> list[dict]:
        normalized = _normalize_name(order_id)
        return [
            record for record in self.data.get("records", {}).values()
            if _normalize_name(str(record.get("order_id", ""))) == normalized
        ]

    def status_for(self, traveler: TravelerData) -> tuple[str, str]:
        records = self.records_for_order(traveler.order_id)
        if not records:
            return "未出库", ""
        current_remarks = set(traveler.documents)
        if any(
            record.get("remark") not in current_remarks
            or not traveler.documents.get(str(record.get("remark", "")), [])
            for record in records
        ):
            return "需要人工处理", ""
        document_numbers = []
        missing = False
        changed = False
        for remark, items in traveler.documents.items():
            if not items:
                continue
            record = self.record_for_document(traveler.order_id, remark)
            if not record:
                missing = True
                continue
            document_numbers.append(str(record.get("document_number", "")))
            if record.get("raw_fingerprint") != self.raw_document_fingerprint(items):
                changed = True
        if changed:
            return "需要更新", "、".join(filter(None, document_numbers))
        if missing:
            return "未出库", "、".join(filter(None, document_numbers))
        return "已出库", "、".join(filter(None, document_numbers))

    def prepare_documents(self, preview: InventoryPreview) -> list[dict]:
        selected = preview._selected_document_set()
        previous = {
            str(record.get("remark", "")): record
            for record in self.records_for_order(preview.traveler.order_id)
            if not selected or _normalize_name(str(record.get("remark", ""))) in selected
        }
        current = {
            remark for remark in preview.traveler.documents
            if not selected or _normalize_name(remark) in selected
        }
        disappeared = sorted(
            remark for remark in previous
            if remark and (remark not in current or not preview.traveler.documents.get(remark))
        )
        if disappeared:
            raise RuleError(
                "inventory_manual_void",
                "以下已出库工厂单在当前 Traveler 中消失或五金为空，请人工删除或作废旧出库单后再处理："
                + "、".join(disappeared),
                remarks=disappeared,
            )
        payloads = []
        for document in preview.document_payloads():
            remark = document["remark"]
            items = [
                OutboundItem(**item) for item in document["items"]
            ]
            if not items:
                continue
            record = self.record_for_document(preview.traveler.order_id, remark)
            mapped_fingerprint = self.mapped_document_fingerprint(items)
            payloads.append({
                "remark": remark,
                "kind": document["kind"],
                "items": [
                    {"productCode": item.product_code, "quantity": item.quantity}
                    for item in items
                ],
                # A changed Traveler source must update the original outbound
                # document even when its mapped product codes happen to stay
                # the same.  The raw fingerprint is the Server/report change
                # signal; the mapped fingerprint protects the actual line data.
                "changed": (
                    not record
                    or record.get("mapped_fingerprint") != mapped_fingerprint
                    or record.get("raw_fingerprint") != self.raw_document_fingerprint(
                        preview.traveler.documents.get(remark, [])
                    )
                ),
                "knownDocumentNumber": str(record.get("document_number", "")) if record else "",
                "mappedFingerprint": mapped_fingerprint,
                "rawFingerprint": self.raw_document_fingerprint(
                    preview.traveler.documents.get(remark, [])
                ),
            })
        return payloads

    def save_success(self, preview: InventoryPreview, results: list[dict]) -> None:
        self._backup_current()
        prepared = {item["remark"]: item for item in self.prepare_documents(preview)}
        for result in results:
            if not result.get("saved") and not result.get("unchanged"):
                continue
            remark = str(result.get("remark", ""))
            plan = prepared.get(remark)
            if not plan:
                continue
            key = self.key(preview.traveler.order_id, remark)
            self.data.setdefault("records", {})[key] = {
                "traveler_path": str(preview.traveler.path),
                "order_id": preview.traveler.order_id,
                "remark": remark,
                "kind": plan["kind"],
                "raw_fingerprint": plan["rawFingerprint"],
                "mapped_fingerprint": plan["mappedFingerprint"],
                "document_number": str(result.get("documentNumber", "")),
                "document_url": str(result.get("url", "")),
                "status": "已出库",
                "items": plan["items"],
                "synced_at": datetime.now().isoformat(timespec="seconds"),
            }
        self.data["version"] = 2
        central = self.path.parent.parent / "workflow.sqlite3"
        if central.is_file():
            connection = sqlite3.connect(central)
            try:
                for result in results:
                    if not result.get("saved") and not result.get("unchanged"):
                        continue
                    remark = str(result.get("remark", ""))
                    document_number = str(result.get("documentNumber", "")).strip()
                    if not document_number:
                        continue
                    factory = connection.execute(
                        "select factory_order from factory_orders where order_id=? and (factory_name=? or factory_order=?) limit 1",
                        (preview.traveler.order_id, remark, remark),
                    ).fetchone()
                    connection.execute(
                        """insert into outbound_documents(
                            document_number,document_type,order_id,factory_order,status,source,issued_at,source_path,updated_at
                        ) values(?,?,?,?,?,?,?,?,?)
                        on conflict(document_number) do update set
                            status=excluded.status, factory_order=excluded.factory_order,
                            issued_at=excluded.issued_at, source_path=excluded.source_path, updated_at=excluded.updated_at""",
                        (document_number, str(result.get("kind", "")), preview.traveler.order_id,
                         factory[0] if factory else remark, "已出库", "金蝶", str(result.get("syncedAt", "")),
                         str(preview.traveler.path), datetime.now().isoformat(timespec="seconds")),
                    )
                connection.commit()
            finally:
                connection.close()
        else:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")

    def _backup_current(self) -> None:
        if not self.path.is_file():
            return
        self.backup_root.mkdir(parents=True, exist_ok=True)
        destination = self.backup_root / f"inventory-sync {datetime.now():%Y-%m-%d %H%M%S.%f}.json"
        shutil.copy2(self.path, destination)
        backups = sorted(self.backup_root.glob("inventory-sync *.json"), key=lambda item: item.stat().st_mtime, reverse=True)
        for old in backups[50:]:
            old.unlink()


def _catalog_path(config: Config) -> Path:
    return config.state_dir / "inventory" / "current-products.xlsx"


def _database_path(config: Config) -> Path:
    """Return the only supported product catalog database."""
    return config.state_dir / "workflow.sqlite3"


def _catalog_info(config: Config) -> dict:
    database = _database_path(config)
    if not database.is_file():
        return {}
    loaded = ProductDatabase(database)
    try:
        backup = _catalog_path(config)
        modified_path = backup if backup.is_file() else database
        modified = datetime.fromtimestamp(modified_path.stat().st_mtime)
        return {
            "path": str(database),
            "backup_path": str(backup),
            "count": loaded.count(),
            "modified_at": modified.isoformat(timespec="seconds"),
            "stale": datetime.now() - modified > timedelta(days=30),
        }
    finally:
        loaded.close()


def _sync_path(config: Config) -> Path:
    return config.state_dir / "inventory-outbound-records.json"


def bootstrap_catalog(config: Config) -> Path:
    destination = _catalog_path(config)
    if destination.is_file():
        return destination
    resource_dir = Path(__file__).resolve().parent.parent / "resources" / "inventory"
    sources = sorted(resource_dir.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not sources:
        return destination
    ProductCatalog(sources[0])
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(sources[0], destination)
    return destination


def bootstrap_product_database(config: Config) -> Path:
    """Return the central runtime catalog database, importing the XLSX if needed."""
    if not config.storage_prepared:
        config.prepare_storage()
    destination = _database_path(config)
    source = bootstrap_catalog(config)
    if destination.is_file():
        try:
            connection = sqlite3.connect(destination)
            has_products = connection.execute(
                "select 1 from sqlite_master where type='table' and name='products'"
            ).fetchone() is not None
            connection.close()
        except sqlite3.Error:
            has_products = False
        if has_products:
            return destination
        if source.is_file():
            import_catalog(config, source)
        return destination
    if not source.is_file():
        return destination
    import_catalog(config, source)
    return destination


def reconcile_folder_status(config: Config, folder: str) -> dict:
    folder = folder.strip().upper()
    if not re.fullmatch(r"(?:PP\d{4}|CS\d{3})", folder):
        raise RuleError("inventory_argument", f"文件夹编号格式无效：{folder}")
    paths = sorted((config.order_root / folder).glob("Work Order Traveler(*).xlsx"))
    travelers = [parse_traveler(path) for path in paths]
    response = run_jdy(config, "findOutbound", order_name=folder)
    rows = [str(row) for row in response.get("matches", [])]
    store = InventorySyncStore(_sync_path(config), config.backup_root)
    found = []
    not_found = []
    needs_review = []
    for traveler in travelers:
        token = _normalize_name(traveler.order_name)
        candidates = [row for row in rows if token and token in _normalize_name(row)]
        document_numbers = sorted(set(
            number
            for row in candidates
            for number in re.findall(r"QTCK\d+", row, flags=re.IGNORECASE)
        ))
        if len(document_numbers) == 1:
            number = document_numbers[0].upper()
            found.append({"order_name": traveler.order_name, "document_number": number})
        elif not document_numbers:
            not_found.append({
                "order_name": traveler.order_name,
                "reason": "库存系统中没有查询到对应出库记录",
            })
        else:
            needs_review.append({
                "order_name": traveler.order_name,
                "reason": "找到多个出库单号，需要人工选择",
                "candidates": document_numbers,
            })
    return {
        "ok": True,
        "folder": folder,
        "found": found,
        "not_found": not_found,
        "needs_review": needs_review,
    }


def list_travelers(config: Config, include_history: bool = False) -> dict:
    database = bootstrap_product_database(config)
    initial = datetime.strptime(config.initial_date, "%Y-%m-%d")
    store = InventorySyncStore(_sync_path(config), config.backup_root)
    entries = []
    errors = []
    for path in sorted(config.order_root.rglob("Work Order Traveler(*).xlsx")):
        if not TRAVELER_RE.fullmatch(path.name):
            continue
        try:
            traveler = parse_traveler(path)
            status, document = store.status_for(traveler)
            modified = datetime.fromisoformat(traveler.modified_at)
            if not include_history and modified < initial and status == "已出库":
                continue
            entries.append({
                "path": str(path),
                "pp_folder": traveler.pp_folder,
                "file_name": path.name,
                "order_name": traveler.order_name,
                "modified_at": traveler.modified_at,
                "fingerprint": traveler.fingerprint,
                "status": status,
                "document_number": document,
            })
        except RuleError as exc:
            errors.append({"path": str(path), "code": exc.code, "message": str(exc), **exc.context})
    return {"travelers": entries, "errors": errors, "catalog": _catalog_info(config)}


def list_traveler_names(config: Config) -> dict:
    store = InventorySyncStore(_sync_path(config), config.backup_root)
    records_by_path = {
        str(Path(record.get("traveler_path", "")).resolve()): record
        for record in store.data.get("records", {}).values()
        if record.get("traveler_path")
    }
    entries = []
    for path in sorted(config.order_root.rglob("Work Order Traveler(*).xlsx")):
        if not TRAVELER_RE.fullmatch(path.name):
            continue
        resolved = str(path.resolve())
        record = records_by_path.get(resolved, {})
        entries.append({
            "path": resolved,
            "pp_folder": path.parent.name,
            "file_name": path.name,
            "order_name": "",
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
            "fingerprint": "",
            "status": record.get("status", "未出库"),
            "document_number": record.get("document_number", ""),
        })
    return {
        "travelers": entries,
        "errors": [],
        "catalog": _catalog_info(config),
        "lazy": True,
    }


def import_catalog(config: Config, source: Path) -> dict:
    if not config.storage_prepared:
        config.prepare_storage()
    catalog = ProductCatalog(source)
    destination = _catalog_path(config)
    destination.parent.mkdir(parents=True, exist_ok=True)
    draft = destination.with_suffix(".tmp.xlsx")
    shutil.copy2(source, draft)
    ProductCatalog(draft)
    _replace_product_database(_database_path(config), catalog.products)
    os.replace(draft, destination)
    for old in destination.parent.glob("current-products*.xlsx"):
        if old != destination and old.is_file():
            old.unlink()
    return {"path": str(destination), "source": str(source), "count": len(catalog.products)}


def _catalog_change_summary(previous: list[Product], current: list[Product]) -> dict[str, int]:
    def signature(product: Product) -> tuple[str, ...]:
        return (
            product.category,
            product.name,
            product.spec,
            product.status,
            product.remark,
            product.unit,
            "" if product.cost_price is None else f"{product.cost_price:.12g}",
        )

    previous_by_code = {
        _normalize_name(product.code): product
        for product in previous if _normalize_name(product.code)
    }
    current_by_code = {
        _normalize_name(product.code): product
        for product in current if _normalize_name(product.code)
    }
    previous_codes = set(previous_by_code)
    current_codes = set(current_by_code)
    return {
        "added_count": len(current_codes - previous_codes),
        "updated_count": sum(
            signature(previous_by_code[code]) != signature(current_by_code[code])
            for code in previous_codes & current_codes
        ),
        "removed_count": len(previous_codes - current_codes),
    }


def _existing_catalog_products(config: Config) -> list[Product]:
    catalog_path = _catalog_path(config)
    if catalog_path.is_file():
        return ProductCatalog(catalog_path).products
    database_path = _database_path(config)
    if database_path.is_file():
        with ProductDatabase(database_path) as database:
            return database.products
    return []


def update_catalog_online(config: Config) -> dict:
    """Export the current product catalog from JDY and install it atomically."""
    inventory_dir = _catalog_path(config).parent
    inventory_dir.mkdir(parents=True, exist_ok=True)
    download = inventory_dir / f".products-download-{os.getpid()}.xlsx"
    try:
        progress("正在从库存系统导出商品资料")
        previous_products = _existing_catalog_products(config)
        run_jdy(config, "exportProducts", download_path=download)
        if not download.is_file():
            raise RuleError("product_catalog_download", "库存系统导出完成，但没有找到下载的商品资料文件")
        current_products = ProductCatalog(download).products
        summary = _catalog_change_summary(previous_products, current_products)
        result = import_catalog(config, download)
        progress(
            f"商品资料已更新，共 {result['count']} 个商品；"
            f"新增 {summary['added_count']} 个，更新 {summary['updated_count']} 个，"
            f"删除 {summary['removed_count']} 个"
        )
        return {"ok": True, **result, **summary}
    finally:
        download.unlink(missing_ok=True)


def _local_setting(config: Config, name: str) -> str:
    settings = config.state_dir / "settings.json"
    if not settings.is_file():
        return ""
    try:
        return str(json.loads(settings.read_text(encoding="utf-8")).get(name, "")).strip()
    except (OSError, json.JSONDecodeError):
        return ""


def _inventory_cdp_endpoint() -> str:
    return os.environ.get(
        "TRAVELER_CHROME_CDP_ENDPOINT", INVENTORY_CDP_DEFAULT_ENDPOINT
    ).strip() or INVENTORY_CDP_DEFAULT_ENDPOINT


def _inventory_cdp_pages(endpoint: str) -> list[dict]:
    """Read the local CDP tab list without touching cookies or credentials."""
    try:
        with urlopen(endpoint.rstrip("/") + "/json/list", timeout=1.0) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (OSError, ValueError, UnicodeDecodeError):
        return []
    return [item for item in payload if isinstance(item, dict)] if isinstance(payload, list) else []


def _find_existing_inventory_page(endpoint: str) -> dict | None:
    pages = [
        page for page in _inventory_cdp_pages(endpoint)
        if page.get("type") == "page"
        and _is_inventory_authenticated_url(str(page.get("url", "")))
    ]
    return next(
        (page for page in pages if not _is_inventory_service_workbench_url(str(page.get("url", "")))),
        None,
    ) or (pages[0] if pages else None)


def _is_inventory_domain_url(url: str) -> bool:
    try:
        parsed = urlsplit(url)
    except ValueError:
        return False
    hostname = (parsed.hostname or "").lower().rstrip(".")
    return (
        parsed.scheme in {"http", "https"}
        and (hostname == "jdy.com" or hostname.endswith(INVENTORY_DOMAIN_SUFFIX))
    )


def _is_inventory_login_or_global_url(url: str) -> bool:
    try:
        parsed = urlsplit(url)
    except ValueError:
        return True
    path = (parsed.path or "/").lower()
    return bool(
        re.search(r"/(?:login|logout)(?:/|$)", path)
        or path.rstrip("/") == "/global"
        or parsed.query.lower().find("logout=true") >= 0
    )


def _is_inventory_authenticated_url(url: str) -> bool:
    """Return whether a CDP page is a logged-in inventory application page."""
    return _is_inventory_domain_url(url) and not _is_inventory_login_or_global_url(url)


def _is_inventory_service_workbench_url(url: str) -> bool:
    try:
        parsed = urlsplit(url)
    except ValueError:
        return False
    return (
        (parsed.hostname or "").lower().rstrip(".") in {"service.jdy.com", "www.jdy.com"}
        and (parsed.path or "/").lower().startswith("/workbench")
    )


def _inventory_chrome_profile(config: Config) -> Path:
    return config.state_dir / "inventory" / "browser-profile-zh"


def _inventory_chrome_executable() -> Path | None:
    configured = os.environ.get("TRAVELER_BROWSER_EXECUTABLE", "").strip()
    candidates = [Path(configured)] if configured else []
    candidates.extend([
        Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        Path("/Applications/Google Chrome Beta.app/Contents/MacOS/Google Chrome Beta"),
    ])
    if executable := shutil.which("google-chrome"):
        candidates.append(Path(executable))
    return next(
        (candidate for candidate in candidates if candidate.is_file() and os.access(candidate, os.X_OK)),
        None,
    )


def open_inventory_chrome(config: Config) -> dict:
    """Launch the dedicated, user-loginable Chrome session used by the App."""
    endpoint = _inventory_cdp_endpoint()
    pages = _inventory_cdp_pages(endpoint)
    existing_candidates = [
        page for page in pages
        if page.get("type") == "page"
        and _is_inventory_authenticated_url(str(page.get("url", "")))
    ]
    existing = next(
        (
            page for page in existing_candidates
            if not _is_inventory_service_workbench_url(str(page.get("url", "")))
        ),
        None,
    ) or (existing_candidates[0] if existing_candidates else None)
    if existing:
        return {
            "ok": True,
            "reused": True,
            "url": str(existing.get("url", "")),
        }
    if any(
        page.get("type") == "page"
        and _is_inventory_domain_url(str(page.get("url", "")))
        for page in pages
    ):
        return {
            "ok": True,
            "waitingForLogin": True,
        }

    executable = _inventory_chrome_executable()
    if executable is None:
        raise RuleError(
            "inventory_runtime",
            "未找到 Google Chrome；请安装 Chrome，或用 TRAVELER_BROWSER_EXECUTABLE 指定路径",
        )
    parsed_endpoint = urlsplit(endpoint)
    port = parsed_endpoint.port or 9222
    profile = _inventory_chrome_profile(config)
    profile.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.Popen(
            [
                str(executable),
                f"--remote-debugging-port={port}",
                f"--user-data-dir={profile}",
                "--no-first-run",
                "--no-default-browser-check",
                INVENTORY_LOGIN_URL,
            ],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except OSError as exc:
        raise RuleError("inventory_runtime", f"无法打开库存专用 Chrome：{exc}") from exc
    return {
        "ok": True,
        "launched": True,
        "cdpEndpoint": endpoint,
        "profileDir": str(profile),
    }


def close_inventory_chrome(config: Config) -> dict:
    """Close only the dedicated Chrome exposed on the inventory CDP endpoint."""
    root = Path(__file__).resolve().parent.parent
    node, node_modules = _resolve_jdy_runtime(root)
    helper = root / "tools" / "jdy_inventory.mjs"
    env = os.environ.copy()
    env["NODE_PATH"] = str(node_modules)
    try:
        result = subprocess.run(
            [str(node), str(helper)],
            input=json.dumps({
                "action": "closeChrome",
                "cdpEndpoint": _inventory_cdp_endpoint(),
            }),
            text=True,
            capture_output=True,
            env=env,
            check=False,
            timeout=5,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuleError("inventory_runtime", "关闭库存专用 Chrome 超过 5 秒未完成") from exc
    if result.returncode != 0:
        raise RuleError("inventory_runtime", "关闭库存专用 Chrome 失败")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuleError("inventory_runtime", "关闭库存专用 Chrome 返回结果无法解析") from exc
    return payload if isinstance(payload, dict) else {"ok": True, "closed": False}


def _keychain_password(account: str) -> str:
    if not account:
        raise RuleError("jdy_credentials", "请先在配置中心填写库存系统用户名")
    helper = Path(__file__).resolve().parent.parent / "bin" / "keychain-read"
    if not helper.is_file():
        raise RuleError("jdy_credentials", "库存系统钥匙串读取器尚未构建")
    try:
        result = subprocess.run(
            [str(helper), account],
            capture_output=True, text=True, check=False,
            timeout=10,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuleError("jdy_credentials", "读取库存系统密码超过 10 秒未完成，请检查钥匙串访问权限") from exc
    if result.returncode != 0 or not result.stdout.strip():
        raise RuleError("jdy_credentials", "库存系统密码尚未保存到 macOS 钥匙串")
    return result.stdout.strip()


def _jdy_error_detail(stderr: str) -> str:
    def concise(detail: str) -> str:
        if "Failed to create a ProcessSingleton" in detail or "profile is already in use" in detail:
            return "上一次库存查询浏览器尚未完全退出，请稍候后再次点击查询"
        detail = detail.split("；当前页面：", 1)[0].strip()
        if "库存系统登录未成功" in detail:
            return "库存系统登录未成功；请检查密码是否已更新，或网站是否要求验证码/扫码登录，然后再次查询"
        if "验证码或安全验证" in detail or "安全验证未完成" in detail:
            return "库存系统要求完成验证码或安全验证；请先使用可见浏览器完成登录验证，再重试"
        if "左侧菜单中等待可见“仓库”" in detail:
            return "库存业务工作台已打开，但左侧“仓库”菜单未在等待时间内加载完成；请保持库存系统页面可见后重试"
        if "左侧菜单中等待可见“商品”" in detail:
            return "库存业务工作台已打开，但左侧“商品”菜单未在等待时间内加载完成；请保持库存系统页面可见后重试"
        if "otheroutbound_menu" in detail.lower() or "仓库菜单已打开，但找不到" in detail:
            return "已找到左侧“仓库”菜单，但“其他出库单”入口未出现；请保持库存系统页面可见后重试"
        if "找不到其他出库单记录列表" in detail or "记录列表控件" in detail:
            return "已进入“其他出库单”，但记录列表控件未加载完成；请保持库存系统页面可见后重试"
        if "内嵌表单" in detail or "编辑表单未加载完成" in detail:
            return "已进入“其他出库单”，但出库表单未加载完成；请保持库存系统页面可见后重试"
        if "保存前发现" in detail or "保存前商品" in detail:
            return detail
        return detail

    lines = [line.strip() for line in stderr.splitlines() if line.strip()]
    for index, text in enumerate(lines):
        prefix = "库存系统自动操作失败："
        if text.startswith(prefix):
            detail = text[len(prefix):].strip()
            if index + 1 < len(lines) and lines[index + 1].startswith("；当前页面"):
                detail += lines[index + 1]
            return concise(detail)
    for text in reversed(lines):
        if (
            text.startswith("Node.js v")
            or text.startswith("at ")
            or text.startswith("triggerUncaughtException")
            or text.startswith("file://")
        ):
            continue
        try:
            event = json.loads(text)
            if isinstance(event, dict) and event.get("event") == "progress":
                continue
        except json.JSONDecodeError:
            pass
        return concise(text)
    return "库存系统浏览器操作失败，未返回具体原因"


def _resolve_jdy_runtime(root: Path) -> tuple[Path, Path]:
    configured_node = os.environ.get("TRAVELER_NODE", "").strip()
    if configured_node:
        node_candidates = [Path(configured_node)]
    else:
        node_candidates = [root / "bin" / "node"]
        if path_node := shutil.which("node"):
            node_candidates.append(Path(path_node))
    node = next(
        (candidate for candidate in node_candidates if candidate.is_file() and os.access(candidate, os.X_OK)),
        None,
    )
    if node is None:
        raise RuleError(
            "inventory_runtime",
            "未找到 Node.js；请重新构建 App，或用 TRAVELER_NODE 指定可执行文件",
        )

    configured_modules = os.environ.get("TRAVELER_NODE_MODULES", "").strip()
    node_modules = Path(configured_modules) if configured_modules else root / "node_modules"
    if not all((node_modules / package).is_dir() for package in ("playwright", "playwright-core")):
        raise RuleError(
            "inventory_runtime",
            "未找到 Playwright；请重新构建 App，或用 TRAVELER_NODE_MODULES 指定依赖目录",
        )
    return node, node_modules


def run_jdy(config: Config, action: str, traveler_path: Path | None = None, confirm_save: bool = False,
            download_path: Path | None = None, order_name: str = "",
            stock_items: list[dict] | None = None,
            selected_document_remarks: Iterable[str] | None = None,
            selected_factory_orders: Iterable[str] | None = None,
            order_id: str = "") -> dict:
    operation_started = time.perf_counter()
    progress(f"库存系统：开始准备 {action} 操作")
    username = _local_setting(config, "jdy_username")
    cdp_endpoint = _inventory_cdp_endpoint()
    existing_inventory_page = _find_existing_inventory_page(cdp_endpoint)
    password = ""
    if existing_inventory_page is not None:
        existing_url = str(existing_inventory_page.get("url", ""))
        if _is_inventory_service_workbench_url(existing_url):
            progress(
                "库存系统：发现库存服务工作台，自动化将先点击“进入使用”进入业务系统"
                f"（{existing_url}）"
            )
        else:
            progress(
                "库存系统：发现已登录的库存专用 Chrome 页面，将直接复用，不重新登录"
                f"（{existing_url}）"
            )
    else:
        credentials_started = time.perf_counter()
        password = _keychain_password(username)
        progress(f"库存系统：未发现已登录页面，账号与钥匙串密码读取完成（用时 {time.perf_counter() - credentials_started:.2f} 秒）")
    preview = None
    if action == "outbound":
        if order_id.strip():
            preview = build_database_preview(config, order_id, selected_factory_orders)
        else:
            if not traveler_path:
                raise RuleError("inventory_argument", "出库必须提供订单号或 Traveler")
            preview = build_preview(
                traveler_path,
                bootstrap_product_database(config),
                config.workflow_database,
                selected_document_remarks=selected_document_remarks,
            )
        if not preview.ready:
            names = sorted({
                str(item.get("name", "")).strip()
                for item in preview.missing_items
                if str(item.get("name", "")).strip()
            })
            missing = "、".join(names) or "材料"
            raise RuleError(
                "inventory_mapping_required",
                f"当前出库数据存在未映射材料：{missing}。请先在出库页面完成材料映射后再出库",
                missing_items=preview.missing_items,
                traveler_path=str(traveler_path) if traveler_path else "",
            )
        store = InventorySyncStore(_sync_path(config), config.backup_root)
        documents = store.prepare_documents(preview)
        if not documents:
            if preview.no_outbound_required:
                return {
                    "ok": True,
                    "saved": False,
                    "no_outbound_required": True,
                    "scope_decisions": preview.scope_decisions,
                    "message": "本订单已明确标记为无需出库，未打开库存系统，也未创建出库单",
                }
            raise RuleError("inventory_empty", "当前订单没有需要出库的板材、封边或五金")
        if selected_factory_orders:
            from .order_index import assert_factory_orders_outbound_allowed

            changed_factory_orders = changed_factory_orders_for_documents(
                config,
                preview.traveler.order_id,
                selected_factory_orders,
                documents,
            )
            assert_factory_orders_outbound_allowed(
                config,
                preview.traveler.order_id,
                selected_factory_orders,
                changed_factory_orders=changed_factory_orders,
            )

    root = Path(__file__).resolve().parent.parent
    runtime_started = time.perf_counter()
    node, node_modules = _resolve_jdy_runtime(root)
    progress(f"库存系统：自动化运行环境准备完成（用时 {time.perf_counter() - runtime_started:.2f} 秒）")
    helper = root / "tools" / "jdy_inventory.mjs"
    request = {
        "action": action,
        "username": username,
        "profileDir": str(config.state_dir / "inventory" / "browser-profile-zh"),
        # Browser automation stays off-screen by default so it cannot steal
        # focus from the user's current app. This override is intentionally
        # process-local and is only for login recovery and diagnostics.
        "headless": os.environ.get("TRAVELER_BROWSER_VISIBLE", "").strip().lower()
        not in {"1", "true", "yes", "on"},
        # Only Chrome instances started with this local CDP endpoint are
        # attachable. A normal user-launched Chrome remains untouched.
        "cdpEndpoint": cdp_endpoint,
        # A one-shot helper must exit after returning its JSON result.  When a
        # controlled browser was already running, the Node helper can still
        # attach to it and disconnect without closing the user's browser.  A
        # newly launched browser is closed after this task so Swift never
        # waits forever for the subprocess to finish.
        "keepBrowserOpen": False,
    }
    if existing_inventory_page is None:
        request["password"] = password
    if order_name:
        request["orderName"] = order_name
        today = datetime.now().date()
        request["queryDateFrom"] = (today - timedelta(days=550)).isoformat()
        request["queryDateTo"] = (today + timedelta(days=45)).isoformat()
    if action == "outbound":
        request.update({
            "orderName": preview.traveler.order_id,
            "documents": documents,
            "confirmSave": confirm_save,
            "queryDateFrom": (datetime.now().date() - timedelta(days=550)).isoformat(),
            "queryDateTo": (datetime.now().date() + timedelta(days=45)).isoformat(),
        })
    elif action == "exportProducts":
        if not download_path:
            raise RuleError("inventory_argument", "更新商品资料缺少下载目标")
        request["downloadPath"] = str(download_path)
    elif action == "stockBalance":
        if not stock_items:
            raise RuleError("inventory_empty", "没有需要查询库存的板材或封边")
        request["items"] = stock_items
    env = os.environ.copy()
    env["NODE_PATH"] = str(node_modules)
    requests = [request]
    if action == "outbound":
        requests = [
            {
                **request,
                "orderName": document["remark"],
                "remark": document["remark"],
                "kind": document["kind"],
                "items": document["items"],
                "changed": document["changed"],
                "knownDocumentNumber": document["knownDocumentNumber"],
            }
            for document in request["documents"]
        ]
    responses = []
    for browser_request in requests:
        browser_started = time.perf_counter()
        progress(f"库存系统：开始浏览器操作（第 {len(responses) + 1}/{len(requests)} 个单据）")
        try:
            result = subprocess.run(
                [str(node), str(helper)],
                input=json.dumps(browser_request, ensure_ascii=False),
                text=True, capture_output=True, env=env, check=False,
                # The browser's individual webpage waits are capped at 60s.
                # Keep the process guard longer so startup plus one 60s page
                # wait can finish and return its specific diagnostic.
                timeout=90,
            )
        except subprocess.TimeoutExpired as exc:
            def timeout_text(value: object) -> str:
                if isinstance(value, bytes):
                    return value.decode(errors="replace")
                return str(value or "")

            detail = _jdy_error_detail(
                "\n".join(
                    part for part in (timeout_text(exc.stderr), timeout_text(exc.stdout)) if part
                )
            )
            raise RuleError(
                "jdy_timeout",
                f"库存系统操作超过 90 秒未完成：{detail}",
            ) from exc
        progress(
            f"库存系统：浏览器操作进程结束（用时 {time.perf_counter() - browser_started:.2f} 秒）"
        )
        if result.stderr:
            for line in result.stderr.splitlines():
                try:
                    event = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if isinstance(event, dict) and event.get("event") == "progress":
                    log_progress_payload(event)
            sys.stderr.write(result.stderr)
        if result.returncode != 0:
            if action == "outbound" and preview is not None and confirm_save and responses:
                completed = [
                    item for item in responses
                    if item.get("saved") or item.get("unchanged")
                ]
                if completed:
                    InventorySyncStore(_sync_path(config), config.backup_root).save_success(
                        preview, completed
                    )
            raise RuleError("jdy_browser", _jdy_error_detail(result.stderr))
        try:
            responses.append(json.loads(result.stdout))
        except json.JSONDecodeError as exc:
            raise RuleError("jdy_browser", "库存系统返回结果无法解析") from exc
    try:
        response = responses[0] if action != "outbound" else {
            "ok": True,
            "saved": confirm_save and all(
                item.get("saved") or item.get("unchanged") for item in responses
            ),
            "results": responses,
            "simulated": not confirm_save,
        }
        if action == "outbound" and response.get("saved"):
            results = response["results"]
            if preview is None or any(
                not str(item.get("documentNumber", "")).strip()
                for item in results
            ):
                raise RuleError("jdy_save_result", "库存单已返回成功，但缺少分单结果或单据编号；结果需要人工核实")
            store = InventorySyncStore(_sync_path(config), config.backup_root)
            store.save_success(preview, results)
            response["syncRecorded"] = True
            try:
                from .order_index import (
                    record_standard_outbound_baseline,
                    record_temporary_outbound,
                    reconcile_outbound_statuses,
                )
                reconcile_outbound_statuses(config)
                record_temporary_outbound(config, preview.traveler.path, response)
                response["serverBaselineRecorded"] = record_standard_outbound_baseline(
                    config, preview.traveler.order_id
                )
            except (OSError, sqlite3.Error):
                # The inventory system has already returned a saved document.
                # Preserve that fact for the UI; the normal Inventory Sync
                # record remains authoritative if the secondary index is
                # temporarily unavailable.
                response["orderIndexReconciled"] = False
                response["temporaryLedgerRecorded"] = False
                response["serverBaselineRecorded"] = False
            else:
                response["orderIndexReconciled"] = True
                response["temporaryLedgerRecorded"] = True
        progress(f"库存系统：{action} 操作结束（总用时 {time.perf_counter() - operation_started:.2f} 秒）")
        return response
    except (TypeError, KeyError) as exc:
        raise RuleError("jdy_browser", "库存系统返回结果结构不完整") from exc


def _check_requirements_stock(config: Config, requirements: list[dict]) -> list[dict]:
    response = run_jdy(config, "stockBalance", stock_items=requirements)
    raw_results = response.get("results")
    if not isinstance(raw_results, list):
        raise RuleError("jdy_browser", "库存系统没有返回库存查询明细")
    by_code = {
        str(item.get("productCode", "")).upper(): item
        for item in raw_results if isinstance(item, dict)
    }
    rows = []
    for requirement in requirements:
        code = requirement["productCode"].upper()
        actual = by_code.get(code)
        if actual is None:
            raise RuleError("jdy_browser", f"库存系统没有返回 {code} 的查询结果")
        actual_name = str(actual.get("productName", "")).strip()
        if actual_name and _normalize_name(actual_name) != _normalize_name(requirement["productName"]):
            raise RuleError(
                "inventory_product_mismatch",
                f"库存商品身份不一致：{code} 本地为 {requirement['productName']}，网页为 {actual_name}",
            )
        try:
            available = float(actual.get("availableQuantity", 0))
        except (TypeError, ValueError) as exc:
            raise RuleError("jdy_browser", f"库存系统返回的 {code} 数量不是数字") from exc
        required = float(requirement["requiredQuantity"])
        rows.append({
            **requirement,
            "availableQuantity": available,
            "shortageQuantity": max(0.0, required - available),
            "sufficient": available >= required,
        })
    return rows


def check_stock(config: Config, traveler_path: Path, include_hardware: bool = False) -> dict:
    preview = build_preview(traveler_path, bootstrap_product_database(config), config.workflow_database)
    requirements = stock_requirements(preview, include_hardware)
    rows = _check_requirements_stock(config, requirements)
    return {
        "ok": True,
        "traveler": str(traveler_path.resolve()),
        "includeHardware": include_hardware,
        "rows": rows,
        "hasShortage": any(not row["sufficient"] for row in rows),
    }


def check_order_stock(config: Config, order_folder: Path) -> dict:
    order_id, requirements = order_stock_requirements(config, order_folder)
    rows = _check_requirements_stock(config, requirements)
    return {
        "ok": True,
        "order_id": order_id,
        "folder": str(order_folder.resolve()),
        "rows": rows,
        "hasShortage": any(not row["sufficient"] for row in rows),
    }


def check_database_stock(config: Config, order_id: str) -> dict:
    """Check order-level stock from persisted facts without a Traveler file."""
    normalized_order_id, requirements = database_stock_requirements(config, order_id)
    rows = _check_requirements_stock(config, requirements)
    return {
        "ok": True,
        "order_id": normalized_order_id,
        "source_type": "database",
        "rows": rows,
        "hasShortage": any(not row["sufficient"] for row in rows),
    }


def inventory_main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="pp-flowhub inventory")
    parser.add_argument("action", choices=(
        "list", "list-names", "preview", "order-preview", "import-products", "preflight", "outbound",
        "find-outbound", "reconcile-folder", "ignore-item", "unignore-item",
        "search-products", "set-mapping", "update-mapping", "remove-mapping", "list-mappings", "update-ignore", "set-outbound-scope", "update-products", "stock-check", "open-chrome", "close-chrome",
    ))
    parser.add_argument("--traveler", type=Path)
    parser.add_argument("--order-id", default="")
    parser.add_argument("--source", type=Path)
    parser.add_argument("--include-history", action="store_true")
    parser.add_argument("--order-root", type=Path)
    parser.add_argument("--state-dir", type=Path)
    parser.add_argument("--confirm-save", action="store_true")
    parser.add_argument("--order-name", default="")
    parser.add_argument("--name", default="")
    parser.add_argument("--ignored", choices=("true", "false"))
    parser.add_argument("--traveler-name", action="append", default=[])
    parser.add_argument("--item-name", action="append", default=[])
    parser.add_argument("--old-name", default="")
    parser.add_argument("--document-remark", action="append", default=[])
    parser.add_argument("--factory-order", action="append", default=[])
    parser.add_argument("--scope-type", choices=("material", "hardware"))
    parser.add_argument("--requirement", choices=("required", "customer_supplied", "remainder", "not_required"))
    parser.add_argument("--reason", default="")
    parser.add_argument("--folder", default="")
    parser.add_argument("--query", default="")
    parser.add_argument("--product-code", default="")
    parser.add_argument("--include-hardware", action="store_true")
    args = parser.parse_args(argv)
    config = Config()
    config.load_settings()
    if args.order_root:
        config.order_root = args.order_root
    if args.state_dir:
        config.state_dir = args.state_dir
    config.prepare_storage()
    logger = configure_operation_log(config)
    logger.event("backend.command.started", "开始库存系统操作", details={"action": args.action})
    try:
        if args.action == "list":
            result = list_travelers(config, args.include_history)
        elif args.action == "list-names":
            result = list_traveler_names(config)
        elif args.action == "preview":
            if not args.traveler:
                raise RuleError("inventory_argument", "preview 必须提供 --traveler")
            result = build_preview(args.traveler, bootstrap_product_database(config), config.workflow_database).payload()
        elif args.action == "order-preview":
            if not args.order_id:
                raise RuleError("inventory_argument", "order-preview 必须提供 --order-id")
            result = build_database_preview(config, args.order_id, args.factory_order).payload()
        elif args.action == "set-outbound-scope":
            if not args.order_id or not args.scope_type or not args.requirement:
                raise RuleError("inventory_argument", "保存出库范围需要订单号、范围类型和决定")
            result = set_outbound_scope(
                config,
                args.order_id,
                args.scope_type,
                args.requirement,
                factory_order=args.factory_order[0] if args.factory_order else "",
                reason=args.reason,
            )
        elif args.action == "import-products":
            if not args.source:
                raise RuleError("inventory_argument", "import-products 必须提供 --source")
            result = import_catalog(config, args.source)
        elif args.action == "update-products":
            result = update_catalog_online(config)
        elif args.action == "open-chrome":
            result = open_inventory_chrome(config)
        elif args.action == "close-chrome":
            result = close_inventory_chrome(config)
        elif args.action == "preflight":
            result = run_jdy(config, "preflight")
        elif args.action == "stock-check":
            if not args.traveler:
                raise RuleError("inventory_argument", "stock-check 必须提供 --traveler")
            result = check_stock(config, args.traveler, args.include_hardware)
        elif args.action == "find-outbound":
            if not args.order_name:
                raise RuleError("inventory_argument", "查询出库单必须提供 --order-name")
            result = run_jdy(config, "findOutbound", order_name=args.order_name)
        elif args.action == "reconcile-folder":
            if not args.folder:
                raise RuleError("inventory_argument", "更新文件夹状态必须提供 --folder")
            result = reconcile_folder_status(config, args.folder)
        elif args.action == "search-products":
            result = search_inventory_products(config, args.query)
        elif args.action == "set-mapping":
            item_names = args.item_name or args.traveler_name
            if not item_names or not args.product_code:
                raise RuleError("inventory_argument", "保存映射必须提供材料名称和商品编号")
            result = save_manual_mapping(config, item_names[0], args.product_code)
        elif args.action == "update-mapping":
            item_names = args.item_name or args.traveler_name
            if not args.old_name or not item_names or not args.product_code:
                raise RuleError("inventory_argument", "update-mapping 需要 --old-name、材料名称和商品编号")
            result = update_manual_mapping(config, args.old_name, item_names[0], args.product_code)
        elif args.action == "remove-mapping":
            item_names = args.item_name or args.traveler_name
            if not item_names:
                raise RuleError("inventory_argument", "删除映射必须提供材料名称")
            result = remove_manual_mapping(config, item_names[0])
        elif args.action == "list-mappings":
            result = list_inventory_mappings(config)
        elif args.action == "update-ignore":
            if not args.old_name or not args.name or args.ignored is None:
                raise RuleError("inventory_argument", "update-ignore 需要 --old-name、--name 和 --ignored")
            if args.ignored != "true":
                raise RuleError("inventory_argument", "update-ignore 只能用于保存忽略项目")
            result = update_ignored_mapping(config, args.old_name, args.name, args.reason)
        elif args.action in {"ignore-item", "unignore-item"}:
            item_names = args.item_name or args.traveler_name
            if not item_names:
                raise RuleError("inventory_argument", "修改忽略状态必须提供材料名称")
            results = [
                set_ignored_mapping(
                    config, name, args.action == "ignore-item", args.reason
                )
                for name in item_names
            ]
            result = {"ok": True, "items": results}
        else:
            result = run_jdy(
                config,
                "outbound",
                args.traveler,
                args.confirm_save,
                selected_document_remarks=args.document_remark,
                selected_factory_orders=args.factory_order,
                order_id=args.order_id,
            )
        logger.event("backend.command.completed", "库存系统操作完成", details={"action": args.action})
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except RuleError as exc:
        logger.event(
            "backend.command.failed",
            "库存系统操作失败",
            details={"action": args.action, "code": exc.code, "error": str(exc)},
        )
        print(json.dumps({"fatal": {"code": exc.code, "message": str(exc), **exc.context}}, ensure_ascii=False, indent=2))
        return 2


if __name__ == "__main__":
    raise SystemExit(inventory_main())
