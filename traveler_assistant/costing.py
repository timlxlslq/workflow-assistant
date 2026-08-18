"""Database-backed order cost calculation and workbook export."""

from __future__ import annotations

import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .core import Config, RuleError
from .database import ensure_schema
from .inventory import (
    HARDWARE_DISPLAY_NAMES,
    InventoryMappings,
    ProductDatabase,
    TravelerItem,
    _normalize_name,
    bootstrap_product_database,
    match_item,
)


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _number(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _material_name(kind: str, thickness: str, color: str) -> str:
    if kind == "edge":
        return f"Edge banding--{color}"
    if kind == "plywood":
        return f"{float(thickness):g}mm--Plywood"
    return f"{float(thickness):g}mm--{color}"


def _resolve_product(catalog, mappings, *, name: str, section: str, code: str = ""):
    """Resolve a cost product without rounding the source quantity."""
    if code.strip():
        try:
            return catalog.require_code(code.strip().upper()), "商品编号"
        except RuleError:
            pass
    resolved_name = name
    if section == "五金" and code.strip():
        resolved_name = HARDWARE_DISPLAY_NAMES.get(_normalize_name(code), name)
    item = TravelerItem(
        row=0,
        section=section,
        name=resolved_name,
        quantity=1,
        document_remark="",
    )
    matched = match_item(catalog, mappings, item)
    if len(matched) != 1:
        raise RuleError("cost_product_match", f"{name} 匹配到多个库存商品，不能安全计算成本")
    return catalog.require_code(matched[0].product_code), matched[0].match_source


def _line(
    *,
    category: str,
    factory_order: str,
    room_name: str,
    name: str,
    spec: str,
    quantity: float,
    unit: str,
    product_code: str,
    cost_price: float | None,
    source: str,
    missing: str = "",
) -> dict:
    amount = None if cost_price is None or missing else quantity * cost_price
    return {
        "category": category,
        "factory_order": factory_order,
        "room_name": room_name,
        "name": name,
        "spec": spec,
        "quantity": quantity,
        "unit": unit,
        "product_code": product_code,
        "cost_price": cost_price,
        "amount": amount,
        "source": source,
        "missing": missing,
    }


def _aggregate_material_rows(rows: list[sqlite3.Row]) -> list[dict]:
    """Aggregate material and edge-banding facts at order scope only."""
    grouped: dict[tuple[str, str, str, str], dict] = {}
    for row in rows:
        key = (
            str(row["material_type"] or ""),
            str(row["color"] or ""),
            str(row["thickness"] or ""),
            str(row["unit"] or ""),
        )
        item = grouped.setdefault(
            key,
            {
                "material_type": key[0],
                "color": key[1],
                "thickness": key[2],
                "unit": key[3],
                "quantity": 0.0,
            },
        )
        item["quantity"] += _number(row["quantity"])
    return sorted(
        grouped.values(),
        key=lambda row: (
            {"panel": 0, "plywood": 0, "edge": 1}.get(row["material_type"], 2),
            row["color"],
            _number(row["thickness"]),
            row["unit"],
        ),
    )


def calculate_order_cost(config: Config, order_id: str) -> dict:
    normalized = order_id.strip().upper()
    if not normalized:
        raise RuleError("cost_order_missing", "计算成本需要订单号")
    ensure_schema(config.workflow_database)
    connection = sqlite3.connect(config.workflow_database)
    connection.row_factory = sqlite3.Row
    try:
        material_rows = connection.execute(
            """
            select material_type, color, thickness, quantity, unit, source_path
            from material_items
            where order_id=?
            order by material_type, color, thickness
            """,
            (normalized,),
        ).fetchall()
        has_factory_table = connection.execute(
            "select 1 from sqlite_master where type='table' and name='factory_orders'"
        ).fetchone() is not None
        hardware_rows = connection.execute(
            """
            select factory_order, product_code, name, spec, quantity, unit, source_path
            from hardware_items
            where order_id=? and active=1
              and exists (
                  select 1 from factory_orders
                  where factory_orders.order_id=hardware_items.order_id
                    and factory_orders.factory_order=hardware_items.factory_order
                    and factory_orders.aimes_status='active'
              )
            order by factory_order, product_code, name
            """,
            (normalized,),
        ).fetchall() if has_factory_table else []
    finally:
        connection.close()

    material_rows = _aggregate_material_rows(material_rows)
    if not material_rows and not hardware_rows:
        raise RuleError("cost_data_missing", f"数据库中没有订单 {normalized} 的材料或五金数据")

    mappings = InventoryMappings(config.workflow_database)
    lines: list[dict] = []
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        for row in material_rows:
            name = _material_name(row["material_type"], row["thickness"], row["color"])
            try:
                product, source = _resolve_product(catalog, mappings, name=name, section="板材与封边")
                missing = "预计采购价缺失" if product.cost_price is None else ""
                product_code = product.code
                cost_price = product.cost_price
                display_name = product.name
                resolved_source = source
            except RuleError as exc:
                product_code = ""
                cost_price = None
                missing = str(exc)
                display_name = name
                resolved_source = "未匹配"
            lines.append(_line(
                category="封边条" if row["material_type"] == "edge" else "板材",
                factory_order="材料汇总",
                room_name="",
                name=display_name,
                spec=str(row["thickness"] or "") + ("mm" if row["thickness"] else ""),
                quantity=_number(row["quantity"]),
                unit=row["unit"] or ("m" if row["material_type"] == "edge" else "pcs"),
                product_code=product_code,
                cost_price=cost_price,
                source=resolved_source,
                missing=missing,
            ))

        for row in hardware_rows:
            try:
                product, source = _resolve_product(
                    catalog,
                    mappings,
                    name=row["name"],
                    section="五金",
                    code=row["product_code"],
                )
                missing = "预计采购价缺失" if product.cost_price is None else ""
                product_code = product.code
                cost_price = product.cost_price
                display_name = product.name
                resolved_source = source
            except RuleError as exc:
                missing = str(exc)
                product_code = row["product_code"] or ""
                cost_price = None
                display_name = row["name"] or product_code
                resolved_source = "未匹配"
            lines.append(_line(
                category="五金",
                factory_order=row["factory_order"] or "五金汇总",
                room_name="",
                name=display_name,
                spec=row["spec"] or "",
                quantity=_number(row["quantity"]),
                unit=row["unit"] or "pcs",
                product_code=product_code,
                cost_price=cost_price,
                source=resolved_source,
                missing=missing,
            ))

    missing_items = [line for line in lines if line["missing"]]
    factory_totals: dict[str, dict] = {}
    for item in lines:
        key = item["factory_order"] or "材料汇总"
        bucket = factory_totals.setdefault(
            key,
            {"factory_order": key, "total": 0.0, "has_missing": False},
        )
        if item["amount"] is not None:
            bucket["total"] += item["amount"]
        if item["missing"]:
            bucket["has_missing"] = True

    total = sum(line["amount"] for line in lines if line["amount"] is not None)
    total_missing = bool(missing_items)
    return {
        "order_id": normalized,
        "status": "待补充" if total_missing else "已完成",
        "total_cost": None if total_missing else total,
        "known_cost": total,
        "missing_items": missing_items,
        "lines": lines,
        "factory_lines": list(lines),
        "factory_totals": sorted(
            factory_totals.values(),
            key=lambda item: (item["factory_order"] != "材料汇总", item["factory_order"]),
        ),
        "generated_at": _now(),
    }


EXCEL_HEADERS = [
    "工厂单/汇总", "类别", "商品名称", "规格", "商品编号", "单位",
    "汇总数量", "单价", "材料成本", "价格状态", "数量来源", "价格说明",
]


def _excel_row(row: dict) -> list:
    return [
        row["factory_order"] or "材料汇总",
        row["category"],
        row["name"],
        row["spec"],
        row["product_code"],
        row["unit"],
        row["quantity"],
        row["cost_price"],
        None,
        "未知价格" if row["missing"] else "已知",
        row["source"],
        row["missing"] or "SQLite 商品资料中的 cost_price",
    ]


def _style_rows(sheet, header_row: int, last_row: int, widths: dict[str, float]) -> None:
    navy = "1F4E78"
    border = Side(style="thin", color="B8C7D9")
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:L{last_row}"
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=border, bottom=border)
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(header_row + 1, last_row + 1):
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
        sheet.row_dimensions[row].height = 32
    sheet.sheet_view.showGridLines = False


def export_order_cost(config: Config, order_id: str) -> dict:
    report = calculate_order_cost(config, order_id)
    output_dir = config.state_dir / "cost-reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "_", report["order_id"])
    destination = output_dir / f"{safe_id}-cost.xlsx"

    navy = "1F4E78"
    pale_blue = "EEF5FB"
    blue = "D9EAF7"
    gray = "F3F4F6"
    red_fill = "FDE2E2"
    red_font = "B91C1C"
    border = Side(style="thin", color="B8C7D9")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "成本汇总"
    summary.merge_cells("A1:L1")
    summary["A1"] = f"{report['order_id']} 成本汇总"
    summary["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    summary.merge_cells("A2:L2")
    summary["A2"] = "材料和五金均来自 SQLite；板材与封边只按订单级 material 汇总，不按工厂单分配；未补充价格不按 0 元计入总成本。"
    summary["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    summary["A2"].font = Font(color="334155", italic=True, size=10)
    summary["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    summary.row_dimensions[2].height = 34

    cards = [
        ("A4:B4", "订单号", "A5:B5", report["order_id"]),
        ("D4:E4", "明细行数", "D5:E5", len(report["lines"])),
        ("G4:H4", "待补充行数", "G5:H5", len(report["missing_items"])),
        ("J4:K4", "总成本", "J5:K5", report["total_cost"] if report["total_cost"] is not None else "待补充"),
    ]
    for label_range, label, value_range, value in cards:
        summary.merge_cells(label_range)
        summary.merge_cells(value_range)
        label_cell = label_range.split(":")[0]
        value_cell = value_range.split(":")[0]
        summary[label_cell] = label
        summary[value_cell] = value
        summary[label_cell].fill = PatternFill("solid", fgColor=blue)
        summary[label_cell].font = Font(bold=True, color=navy)
        summary[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        summary[value_cell].font = Font(bold=True, color=navy, size=14)
        summary[value_cell].alignment = Alignment(horizontal="center", vertical="center")
        for row in summary[value_range]:
            for cell in row:
                cell.border = Border(left=border, right=border, top=border, bottom=border)
    summary["J5"].number_format = '"$"#,##0.00'
    summary.merge_cells("A6:L6")
    summary["A6"] = f"已确认成本：${report['known_cost']:,.2f}；状态：{report['status']}"
    summary["A6"].fill = PatternFill("solid", fgColor=gray)
    summary["A6"].font = Font(color="475569")

    summary.append([])
    summary.append(EXCEL_HEADERS)
    header_row = 8
    row_number = 9
    subtotal_rows: list[int] = []
    grouped: dict[str, list[dict]] = defaultdict(list)
    for line in report["lines"]:
        grouped[line["factory_order"] or "材料汇总"].append(line)
    for group_name in sorted(grouped, key=lambda name: (name != "材料汇总", name)):
        start = row_number
        for line in grouped[group_name]:
            summary.append(_excel_row(line))
            summary.cell(row_number, 9).value = f'=IF(OR(G{row_number}="",H{row_number}=""),"",G{row_number}*H{row_number})'
            if line["missing"]:
                for cell in summary[row_number]:
                    cell.fill = PatternFill("solid", fgColor=red_fill)
                    cell.font = Font(color=red_font)
            row_number += 1
        subtotal = row_number
        summary.merge_cells(start_row=subtotal, start_column=1, end_row=subtotal, end_column=8)
        summary.cell(subtotal, 1).value = f"{group_name} 小计"
        summary.cell(subtotal, 9).value = f"=SUM(I{start}:I{subtotal - 1})"
        summary.merge_cells(start_row=subtotal, start_column=10, end_row=subtotal, end_column=12)
        summary.cell(subtotal, 10).value = "存在待补充项目" if any(item["missing"] for item in grouped[group_name]) else "已确认成本"
        for cell in summary[subtotal]:
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(bold=True, color=navy)
            cell.border = Border(top=border, bottom=border)
        subtotal_rows.append(subtotal)
        row_number += 1

    grand_total = row_number
    summary.merge_cells(start_row=grand_total, start_column=1, end_row=grand_total, end_column=8)
    summary.cell(grand_total, 1).value = f"{report['order_id']} 全部材料和五金总计"
    summary.cell(grand_total, 9).value = f"=SUM({','.join(f'I{row}' for row in subtotal_rows)})"
    summary.merge_cells(start_row=grand_total, start_column=10, end_row=grand_total, end_column=12)
    summary.cell(grand_total, 10).value = "待补充价格不计入总成本" if report["status"] != "已完成" else "全部价格已确认"
    for cell in summary[grand_total]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.border = Border(top=Side(style="medium", color=navy), bottom=Side(style="medium", color=navy))
    summary["J5"].fill = PatternFill("solid", fgColor=red_fill if report["status"] != "已完成" else "FFFFFF")
    _style_rows(
        summary,
        header_row=header_row,
        last_row=grand_total,
        widths={"A": 22, "B": 12, "C": 26, "D": 16, "E": 13, "F": 10, "G": 13, "H": 12, "I": 14, "J": 15, "K": 24, "L": 42},
    )
    for row in range(9, grand_total + 1):
        summary.cell(row, 7).number_format = "#,##0.00"
        summary.cell(row, 8).number_format = '"$"#,##0.00'
        summary.cell(row, 9).number_format = '"$"#,##0.00'

    price_sheet = workbook.create_sheet("价格与来源")
    price_sheet.merge_cells("A1:I1")
    price_sheet["A1"] = f"{report['order_id']} 价格与来源"
    price_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    price_sheet["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    price_sheet["A1"].alignment = Alignment(horizontal="center")
    price_sheet.merge_cells("A2:I2")
    price_sheet["A2"] = "价格字段：SQLite products.cost_price；缺失或未匹配项目保留为空，不按 0 元代替。"
    price_sheet["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    price_sheet["A2"].alignment = Alignment(wrap_text=True)
    price_sheet.append([])
    price_sheet.append(["商品名称", "商品编号", "规格", "单位", "单价", "价格状态", "价格说明", "数量来源", "工厂单/汇总"])
    seen: set[tuple[str, str]] = set()
    price_row = 5
    for line in report["lines"]:
        key = (line["product_code"], line["name"])
        if key in seen:
            continue
        seen.add(key)
        price_sheet.append([
            line["name"], line["product_code"], line["spec"], line["unit"], line["cost_price"],
            "未知价格" if line["missing"] else "已知", line["missing"] or "SQLite products.cost_price",
            line["source"], line["factory_order"] or "材料汇总",
        ])
        if line["missing"]:
            for cell in price_sheet[price_row]:
                cell.fill = PatternFill("solid", fgColor=red_fill)
                cell.font = Font(color=red_font)
        price_row += 1
    for cell in price_sheet[4]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color="FFFFFF")
    price_sheet.freeze_panes = "A5"
    price_sheet.auto_filter.ref = f"A4:I{max(4, price_row - 1)}"
    for col, width in {"A": 30, "B": 14, "C": 18, "D": 10, "E": 12, "F": 12, "G": 42, "H": 18, "I": 20}.items():
        price_sheet.column_dimensions[col].width = width
    for row in range(5, price_row):
        price_sheet.cell(row, 5).number_format = '"$"#,##0.00'
        price_sheet.row_dimensions[row].height = 32
        for cell in price_sheet[row]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
    price_sheet.sheet_view.showGridLines = False

    detail = workbook.create_sheet("成本明细")
    detail.append(EXCEL_HEADERS + ["缺失信息"])
    for line in report["lines"]:
        detail.append(_excel_row(line) + [line["missing"]])
    _style_rows(detail, header_row=1, last_row=max(1, detail.max_row), widths={"A": 22, "B": 12, "C": 28, "D": 16, "E": 13, "F": 10, "G": 13, "H": 12, "I": 14, "J": 15, "K": 24, "L": 42, "M": 42})
    for row in range(2, detail.max_row + 1):
        detail.cell(row, 9).value = f'=IF(OR(G{row}="",H{row}=""),"",G{row}*H{row})'
        detail.cell(row, 7).number_format = "#,##0.00"
        detail.cell(row, 8).number_format = '"$"#,##0.00'
        detail.cell(row, 9).number_format = '"$"#,##0.00'

    missing = workbook.create_sheet("缺失项目")
    missing.append(EXCEL_HEADERS + ["缺失信息"])
    for line in report["missing_items"]:
        missing.append(_excel_row(line) + [line["missing"]])
    _style_rows(missing, header_row=1, last_row=max(1, missing.max_row), widths={"A": 22, "B": 12, "C": 28, "D": 16, "E": 13, "F": 10, "G": 13, "H": 12, "I": 14, "J": 15, "K": 24, "L": 42, "M": 42})

    workbook.save(destination)
    return {**report, "export_path": str(destination)}
