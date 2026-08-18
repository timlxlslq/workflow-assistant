from __future__ import annotations

import json
import os
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .core import progress


def _atomic_save(workbook: Workbook, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=f".{destination.stem}-", suffix=".xlsx", dir=destination.parent, delete=False
    ) as handle:
        draft = Path(handle.name)
    try:
        workbook.save(draft)
        os.replace(draft, destination)
    finally:
        draft.unlink(missing_ok=True)


def _materials(path: Path, order_id: str, color: str, panel_qty: int, edge_qty: float) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"], sheet["B1"] = "Job:", order_id
    headers = [
        "Room/section", "", "3/4 Plywood", "5/8 Plywood", "1/4 Plywood",
        "3/4 Finish Panel", "1/4 Finish Panel", "Edge Banding (m)", "Color",
    ]
    for column, value in enumerate(headers, 1):
        sheet.cell(2, column).value = value
    detail = ["Test Kitchen", "", 2, 1, 1, panel_qty, 1, edge_qty, color]
    for column, value in enumerate(detail, 1):
        sheet.cell(3, column).value = value
    total = ["Total Qty:", "", 2, 1, 1, panel_qty, 1, edge_qty, color]
    for column, value in enumerate(total, 1):
        sheet.cell(14, column).value = value
    sheet["A15"] = "Color Table"
    sheet["A16"] = "Color:"
    sheet["A17"] = "Sheets (3/4):"
    sheet["A18"] = "Sheets (1/4):"
    sheet["A19"] = "Edge Banding (m):"
    _atomic_save(workbook, path)


def _board(path: Path, factory: str, name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Page1"
    sheet["A2"], sheet["B2"] = "订 单 号", factory
    sheet["F2"], sheet["G2"] = "订单名称", name
    _atomic_save(workbook, path)


def _fittings(path: Path, groups: list[tuple[str, int]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Page1"
    cursor = 1
    for factory, quantity in groups:
        sheet.cell(cursor, 1).value = "Order No."
        sheet.cell(cursor, 3).value = factory
        header = cursor + 5
        sheet.cell(header, 3).value = "Name"
        sheet.cell(header, 5).value = "Code"
        sheet.cell(header, 6).value = "Size"
        sheet.cell(header, 11).value = "Quantity"
        sheet.cell(header + 1, 3).value = "Test Full Overlay Hinge"
        sheet.cell(header + 1, 5).value = "71T950A"
        sheet.cell(header + 1, 9).value = "Piece"
        sheet.cell(header + 1, 11).value = quantity
        cursor = header + 3
    _atomic_save(workbook, path)


def create_local_test_source(target_root: Path) -> dict:
    """Create deterministic, non-production order inputs without deleting other files."""
    target_root = target_root.expanduser().resolve()
    owned = target_root / "Optimized Orders" / "PP9001"
    cut_to_size = target_root / "CUT TO SIZE" / "CS900"
    progress("正在生成本机测试订单")

    _materials(owned / "PP9001 materials.xlsx", "PP9001", "Test Oak", 4, 18.5)
    _board(owned / "F9001 板材清单.xlsx", "F9001", "PP9001-TEST-KITCHEN")
    _board(owned / "F9002 板材清单.xlsx", "F9002", "PP9001-TEST-LAUNDRY")
    _fittings(owned / "Fittingslist-test.xlsx", [("F9001", 6), ("F9002", 4)])

    _materials(cut_to_size / "CS900 materials.xlsx", "CS900", "Test Walnut", 3, 12.0)
    manifest = {
        "schema_version": 1,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "description": "工作流程助手自动生成的非生产测试数据",
        "owned_order": "PP9001",
        "cut_to_size_order": "CS900",
    }
    target_root.mkdir(parents=True, exist_ok=True)
    manifest_path = target_root / "test-data-manifest.json"
    temporary = manifest_path.with_suffix(".tmp")
    temporary.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, manifest_path)
    progress("本机测试订单已生成")
    return {
        "created": True,
        "root": str(target_root),
        "owned_root": str(target_root / "Optimized Orders"),
        "cut_to_size_root": str(target_root / "CUT TO SIZE"),
        "orders": ["PP9001", "CS900"],
        "manifest": str(manifest_path),
    }
