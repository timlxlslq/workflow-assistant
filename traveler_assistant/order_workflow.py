from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
import shutil
import tempfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.properties import CalcProperties

from .core import (
    Config, FittingItem, RuleError, load_factory_name_cache, load_material_assignments,
    lookup_aimes_names, parse_fittings_groups,
    save_factory_name_cache, save_material_assignment,
)
from .inventory import (
    InventoryMappings,
    ProductDatabase,
    bootstrap_product_database,
    ignored_hardware_reason,
    parse_traveler,
    resolve_inventory_items,
    set_ignored_mapping,
    TravelerItem,
)
from .fittings import select_latest_fittings
from .operation_log import configure_operation_log
from .database import ensure_schema


ORDER_FOLDER_RE = re.compile(r"^(PP\d{4}(?:-\d+)?|CS\d{3})$", re.IGNORECASE)
FACTORY_RE = re.compile(r"^F\d+$", re.IGNORECASE)
EPSILON = 1e-9
WORK_ORDER_SHEET = "WorkOrderTraveler"
USAGE_LIST_SHEET = "Usage List"
PICKING_LIST_SHEET = "Picking List"
PURCHASE_LIST_SHEET = "Purchase List"
MATERIAL_DETAIL_ROW_CAPACITY = 11
MATERIAL_COLOR_COLUMN_START = 3


@dataclass(frozen=True)
class MaterialItem:
    kind: str
    thickness: float
    color: str
    quantity: float
    room_name: str = ""


@dataclass(frozen=True)
class PreviewFitting:
    key: str
    name: str
    code: str
    size: str
    unit: str
    quantity: float
    ignored: bool


@dataclass
class FactoryPreview:
    factory_order: str
    order_name: str
    fittings: list[PreviewFitting]


@dataclass
class OrderPreview:
    order_id: str
    folder: Path
    materials_path: Path
    materials_sheet_name: str
    materials: list[MaterialItem]
    edge_banding: dict[str, float]
    factories: list[FactoryPreview]
    warnings: list[str]
    material_rooms: list[str] = field(default_factory=list)
    include_hardware: bool = True
    material_room_rows: list[tuple[str, list[MaterialItem], dict[str, float]]] = field(default_factory=list)


ORDER_TOKEN_RE = re.compile(r"(PP\d{4}(?:-\d+)?|CS\d{3})(?=$|[-\s_])", re.IGNORECASE)


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _factory_name_belongs_to_order(order_id: str, factory_name: str) -> bool:
    order = order_id.strip().upper()
    name = factory_name.strip().upper()
    if not order:
        return False
    match = re.match(rf"^{re.escape(order)}(?:-|\s+)(.+)$", name)
    if not match:
        return False
    remainder = match.group(1).strip()
    if not remainder:
        return False
    if re.fullmatch(r"PP\d{4}", order) and re.match(r"^\d+(?:-|\s|$)", remainder):
        return False
    return True


def _order_ids_in_text(value: str) -> set[str]:
    return {match.group(1).upper() for match in ORDER_TOKEN_RE.finditer(_text(value))}


def related_order_ids(folder: Path) -> list[str]:
    """Discover all complete order ids represented by a shared source folder."""
    found = _order_ids_in_text(folder.name)
    # Keep the standard-folder path lightweight: filenames and XML names are
    # enough to detect the PP0035/PP0035-2 shared-folder case, while opening
    # every board workbook here would defeat the incremental index fast path.
    if ORDER_FOLDER_RE.fullmatch(folder.name):
        for path in folder.rglob("*"):
            if path.is_file():
                found.update(_order_ids_in_text(path.name))
        return sorted(found)
    for path in folder.rglob("*.xlsx"):
        found.update(_order_ids_in_text(path.name))
        if "板材清单" in path.name:
            try:
                _, name = parse_board_identity(path)
            except Exception:
                continue
            found.update(_order_ids_in_text(name))
    for path in folder.rglob("*.xml"):
        found.update(_order_ids_in_text(path.name))
    return sorted(found)


def _number(value, field: str) -> float:
    if value in (None, "") or (isinstance(value, str) and not value.strip()):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise RuleError("invalid_number", f"{field} 不是有效数字：{value}") from exc


def _display_number(value, number_format: str, field: str) -> float:
    """Return the numeric value Excel displays for a supported number format.

    ``openpyxl`` exposes the formula result and the number format separately;
    it does not expose Excel's rendered cell value. Material workbooks use
    simple numeric formats, so reproduce their rounding in memory without
    rewriting the source workbook.
    """
    number = _number(value, field)
    if not math.isfinite(number):
        raise RuleError("invalid_number", f"{field} 不是有限数字：{value}")
    first_section = _text(number_format).split(";", 1)[0]
    if not first_section or first_section.casefold() == "general":
        return number
    # Ignore quoted literals and escaped characters when locating the
    # numeric decimal section of a custom Excel format.
    numeric_format = re.sub(r'"[^"]*"|\\.', "", first_section)
    percent = "%" in numeric_format
    if percent:
        number *= 100
    match = re.search(r"\.([0#?]+)", numeric_format)
    decimal_places = len(match.group(1)) if match else 0
    quantizer = Decimal("1") if decimal_places == 0 else Decimal(
        "1." + ("0" * decimal_places)
    )
    return float(Decimal(str(number)).quantize(quantizer, rounding=ROUND_HALF_UP))


def _integer(value, field: str) -> float:
    number = _number(value, field)
    if not math.isfinite(number) or abs(number - round(number)) > EPSILON:
        raise RuleError("fractional_material", f"{field} 数量为 {number:g}，板材数量必须为整数，请人工检查")
    if number < 0:
        raise RuleError("negative_material", f"{field} 数量为负数：{number:g}")
    return float(round(number))


def _integer_cell(cell, field: str) -> float:
    """Read the integer Excel displays, while still rejecting visible decimals."""
    number = _number(cell.value, field)
    if not math.isfinite(number) or number < 0:
        raise RuleError("negative_material", f"{field} 数量无效：{number:g}")
    displayed = _display_number(number, cell.number_format, field)
    if abs(displayed - round(displayed)) <= EPSILON:
        return float(round(displayed))
    raise RuleError("fractional_material", f"{field} 数量为 {number:g}，板材数量必须为整数，请人工检查")


def _fmt(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else f"{value:g}"


def _normalized_label(value) -> str:
    return re.sub(r"\s+", "", _text(value)).lower()


def _canonical_color(value: str) -> str:
    color = _text(value)
    if re.fullmatch(r"khaki(?:\s*\(7x9\))?", color, re.IGNORECASE):
        return "Penelope FA44"
    return color


def resolve_source_root(configured: Path) -> Path:
    if configured.is_dir():
        return configured
    raise RuleError("server_unavailable", f"服务器目录不可访问：{configured}")


def list_order_folders(config: Config) -> list[dict]:
    root = resolve_source_root(config.source_root)
    rows = []
    for folder in root.iterdir():
        if not folder.is_dir() or not ORDER_FOLDER_RE.fullmatch(folder.name):
            continue
        modified_at = folder.stat().st_mtime
        rows.append((modified_at, {
            "order_id": folder.name.upper(),
            "path": str(folder),
            "modified_at": datetime.fromtimestamp(modified_at).isoformat(timespec="seconds"),
        }))
    rows.sort(key=lambda item: (item[0], item[1]["order_id"]), reverse=True)
    return [row for _, row in rows]


def _find_label_row(ws, label: str) -> int:
    wanted = _normalized_label(label)
    for row in range(1, ws.max_row + 1):
        if any(_normalized_label(ws.cell(row, col).value) == wanted for col in range(1, ws.max_column + 1)):
            return row
    raise RuleError("materials_schema", f"materials 文件缺少“{label}”")


def _label_column(ws, row: int, label: str) -> int:
    wanted = _normalized_label(label)
    for col in range(1, ws.max_column + 1):
        if _normalized_label(ws.cell(row, col).value) == wanted:
            return col
    raise RuleError("materials_schema", f"materials 文件缺少“{label}”")


def _copy_column_style(ws, source_col: int, target_col: int) -> None:
    """Extend a material Color Table while preserving the template style."""
    if source_col == target_col:
        return
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row, source_col)
        target = ws.cell(row, target_col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    if source_letter in ws.column_dimensions:
        ws.column_dimensions[target_letter] = copy.copy(ws.column_dimensions[source_letter])


def repair_material_color_table(path: Path) -> dict:
    """Add detail colors missing from Color Table and restore its formulas.

    The material workbook is the operator-facing source file, so this repair
    is intentionally limited to the Color Table cells and their existing
    summary formulas. It does not change the detail rows or any formatting.
    """
    try:
        values_wb = load_workbook(path, data_only=True, read_only=True)
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        raise RuleError("materials_schema", f"无法读取 materials 文件：{path.name}：{exc}") from exc
    try:
        if len(values_wb.sheetnames) != 1 or len(wb.sheetnames) != 1:
            raise RuleError("materials_schema", f"materials 文件必须只有一个工作簿，当前为：{wb.sheetnames}")
        values_ws = values_wb[values_wb.sheetnames[0]]
        ws = wb[wb.sheetnames[0]]
        total_row = _find_label_row(values_ws, "Total Qty:")
        color_title_row = _find_label_row(values_ws, "Color Table")
        color_row = _find_label_row(values_ws, "Color:")
        panel_34_row = _find_label_row(values_ws, "Sheets (3/4):")
        panel_14_row = _find_label_row(values_ws, "Sheets (1/4):")
        edge_row = _find_label_row(values_ws, "Edge Banding (m):")
        headers = {
            _normalized_label(values_ws.cell(2, col).value): col
            for col in range(1, values_ws.max_column + 1)
            if _text(values_ws.cell(2, col).value)
        }
        finish_34 = headers.get(_normalized_label("3/4 Finish Panel"))
        finish_14 = headers.get(_normalized_label("1/4 Finish Panel"))
        edge_col = headers.get(_normalized_label("Edge Banding (m)"))
        color_col = headers.get(_normalized_label("Color"))
        if not all((finish_34, finish_14, edge_col, color_col)):
            raise RuleError("materials_schema", "materials 文件缺少 Panel、封边或 Color 列")

        detail_colors: list[str] = []
        seen_detail = set()
        corrected_panel_qty = 0.0
        corrected_edge_qty = 0.0
        for row in range(3, total_row):
            source_color = _text(values_ws.cell(row, color_col).value)
            if not source_color:
                continue
            qty_34 = _number(values_ws.cell(row, finish_34).value, f"{source_color} Sheets (3/4)")
            qty_14 = _number(values_ws.cell(row, finish_14).value, f"{source_color} Sheets (1/4)")
            edge = _number(values_ws.cell(row, edge_col).value, f"{source_color} Edge Banding")
            if qty_34 <= 0 and qty_14 <= 0 and edge <= 0:
                continue
            key = _canonical_color(source_color).casefold()
            if key in seen_detail:
                continue
            seen_detail.add(key)
            detail_colors.append(source_color)

        table_colors: list[tuple[int, str]] = []
        table_by_key = {}
        existing_color_end = max(9, values_ws.max_column)
        for col in range(MATERIAL_COLOR_COLUMN_START, existing_color_end + 1):
            source_color = _text(values_ws.cell(color_row, col).value)
            if not source_color:
                continue
            table_colors.append((col, source_color))
            table_by_key.setdefault(_canonical_color(source_color).casefold(), source_color)

        missing = [
            color for color in detail_colors
            if _canonical_color(color).casefold() not in table_by_key
        ]
        if len(table_colors) + len(missing) > MATERIAL_DETAIL_ROW_CAPACITY:
            raise RuleError(
                "materials_schema",
                f"materials Color Table 需要 {len(table_colors) + len(missing)} 种颜色，已超过明细行容量 {MATERIAL_DETAIL_ROW_CAPACITY} 行",
            )

        required_color_end = max(
            existing_color_end,
            MATERIAL_COLOR_COLUMN_START - 1 + len(table_colors) + len(missing),
        )
        if required_color_end > existing_color_end:
            source_col = max(MATERIAL_COLOR_COLUMN_START, existing_color_end)
            for col in range(existing_color_end + 1, required_color_end + 1):
                _copy_column_style(ws, source_col, col)
        next_columns = [
            col for col in range(MATERIAL_COLOR_COLUMN_START, required_color_end + 1)
            if not _text(values_ws.cell(color_row, col).value)
        ]
        for color, col in zip(missing, next_columns):
            ws.cell(color_row, col).value = color
            table_colors.append((col, color))
            table_by_key[_canonical_color(color).casefold()] = color

        if not missing:
            return {"corrected": False, "panel_count": 0, "edge_banding": 0.0, "colors": []}

        detail_start = 3
        detail_end = total_row - 1
        for col, _ in table_colors:
            letter = get_column_letter(col)
            # Color Table is a derived summary of the detail rows. Use the
            # same SUMIF contract for one-color and multi-color workbooks so
            # an empty table can be populated without changing the detail
            # rows or the Total Qty row.
            ws.cell(panel_34_row, col).value = (
                f'=SUMIF($I${detail_start}:$I${detail_end},{letter}${color_row},'
                f'$F${detail_start}:$F${detail_end})'
            )
            ws.cell(panel_14_row, col).value = (
                f'=SUMIF($I${detail_start}:$I${detail_end},{letter}${color_row},'
                f'$G${detail_start}:$G${detail_end})'
            )
            ws.cell(edge_row, col).value = (
                f'=SUMIF($I${detail_start}:$I${detail_end},{letter}${color_row},'
                f'$H${detail_start}:$H${detail_end})'
            )

        for row in range(3, total_row):
            source_color = _text(values_ws.cell(row, color_col).value)
            key = _canonical_color(source_color).casefold()
            if key not in {_canonical_color(color).casefold() for color in missing}:
                continue
            corrected_panel_qty += max(0.0, _number(values_ws.cell(row, finish_34).value, "Panel"))
            corrected_panel_qty += max(0.0, _number(values_ws.cell(row, finish_14).value, "Panel"))
            corrected_edge_qty += max(0.0, _number(values_ws.cell(row, edge_col).value, "Edge Banding"))

        wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        source_mode = path.stat().st_mode
        draft_fd, draft_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=path.suffix, dir=path.parent)
        os.close(draft_fd)
        draft = Path(draft_name)
        try:
            wb.save(draft)
            os.chmod(draft, source_mode & 0o777)
            os.replace(draft, path)
        finally:
            if draft.exists():
                draft.unlink()
        return {
            "corrected": True,
            "panel_count": corrected_panel_qty,
            "edge_banding": corrected_edge_qty,
            "colors": missing,
            "message": (
                f"material 已自动修正 Color Table：补充 {len(missing)} 种 Panel 颜色 "
                f"（{'、'.join(missing)}），补齐 {_fmt(corrected_panel_qty)} 张 Panel、"
                f"{corrected_edge_qty:g} m 封边条。"
            ),
        }
    finally:
        values_wb.close()
        wb.close()


def parse_order_materials(order_id: str, path: Path) -> tuple[str, list[MaterialItem], dict[str, float]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if len(wb.sheetnames) != 1:
        raise RuleError("materials_schema", f"materials 文件必须只有一个工作簿，当前为：{wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    total_row = _find_label_row(ws, "Total Qty:")
    color_title_row = _find_label_row(ws, "Color Table")
    header_row = 2
    header_map = {
        _normalized_label(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if _text(ws.cell(header_row, col).value)
    }
    finish_34_col = header_map.get(_normalized_label("3/4 Finish Panel"))
    finish_14_col = header_map.get(_normalized_label("1/4 Finish Panel"))
    edge_col = header_map.get(_normalized_label("Edge Banding (m)"))
    color_col = header_map.get(_normalized_label("Color"))
    required = ["3/4 Plywood", "5/8 Plywood", "1/4 Plywood"]
    if any(_normalized_label(label) not in header_map for label in required):
        raise RuleError("materials_schema", "materials 文件缺少 Plywood 数量列")

    formula_wb = load_workbook(path, data_only=False, read_only=True)
    formula_ws = formula_wb[formula_wb.sheetnames[0]]

    def detail_sum(column: int, source_color: str | None = None) -> float:
        total = 0.0
        source_key = _canonical_color(source_color).casefold() if source_color is not None else None
        for detail_row in range(header_row + 1, total_row):
            detail_color = _canonical_color(_text(ws.cell(detail_row, color_col or 9).value)).casefold()
            if source_key is not None and detail_color != source_key:
                continue
            total += _number(ws.cell(detail_row, column).value, f"materials {ws.cell(2, column).value}")
        return total

    def color_integer_fallback(column: int, source_color: str) -> float:
        # A legacy one-color workbook already has the order-level quantity in
        # Total Qty. Keep that authoritative value when the newly written
        # Color Table formula has no cached Excel result; only calculate from
        # detail rows when the upper total is actually empty.
        if len(colors) == 1 and ws.cell(total_row, column).value not in (None, ""):
            return _integer_cell(ws.cell(total_row, column), f"{source_color} total")
        return detail_sum(column, source_color)

    def summary_number(row: int, column: int, label: str, fallback: float) -> float:
        value = ws.cell(row, column).value
        formula = formula_ws.cell(row, column).value
        cell = ws.cell(row, column)
        if value in (None, "") and (
            formula_ws.cell(row, column).data_type == "f"
            or isinstance(formula, str) and formula.startswith("=")
        ):
            return _display_number(fallback, cell.number_format, label)
        return _display_number(value, cell.number_format, label)

    def summary_integer(row: int, column: int, label: str, fallback: float) -> float:
        value = ws.cell(row, column).value
        formula = formula_ws.cell(row, column).value
        cell = ws.cell(row, column)
        if value in (None, "") and (
            formula_ws.cell(row, column).data_type == "f"
            or isinstance(formula, str) and formula.startswith("=")
        ):
            displayed = _display_number(fallback, cell.number_format, label)
            if not math.isclose(displayed, round(displayed), abs_tol=EPSILON):
                raise RuleError("fractional_material", f"{label} 必须是整数，当前为 {fallback}")
            return float(round(displayed))
        return _integer_cell(cell, label)

    plywood = []
    for label, thickness in zip(required, (18.0, 14.5, 5.4)):
        col = header_map[_normalized_label(label)]
        quantity = summary_integer(total_row, col, label, detail_sum(col))
        plywood.append(MaterialItem("plywood", thickness, "", quantity))

    color_header_row = None
    for row in range(color_title_row + 1, min(ws.max_row, color_title_row + 5) + 1):
        if any(_normalized_label(ws.cell(row, col).value) == "color:" for col in range(1, ws.max_column + 1)):
            color_header_row = row
            break
    if color_header_row is None:
        raise RuleError("materials_schema", "materials Color Table 缺少 Color 行")
    row_34 = _find_label_row(ws, "Sheets (3/4):")
    row_14 = _find_label_row(ws, "Sheets (1/4):")
    row_edge = _find_label_row(ws, "Edge Banding (m):")

    # Panel and edge quantities are read from Color Table, but the detail
    # rows still have to be complete enough to explain that table. Empty
    # template rows are normal; only a row with an actual Panel/edge quantity
    # requires a Color value.
    if finish_34_col and finish_14_col and edge_col and color_col:
        missing_colors = []
        for detail_row in range(header_row + 1, total_row):
            panel_34 = _display_number(
                ws.cell(detail_row, finish_34_col).value,
                ws.cell(detail_row, finish_34_col).number_format,
                f"第 {detail_row} 行 3/4 Finish Panel",
            )
            panel_14 = _display_number(
                ws.cell(detail_row, finish_14_col).value,
                ws.cell(detail_row, finish_14_col).number_format,
                f"第 {detail_row} 行 1/4 Finish Panel",
            )
            edge = _display_number(
                ws.cell(detail_row, edge_col).value,
                ws.cell(detail_row, edge_col).number_format,
                f"第 {detail_row} 行 Edge Banding",
            )
            if panel_34 <= 0 and panel_14 <= 0 and edge <= 0:
                continue
            if not _text(ws.cell(detail_row, color_col).value):
                room = _text(ws.cell(detail_row, 1).value) or f"第 {detail_row} 行"
                missing_colors.append(room)
        if missing_colors:
            raise RuleError(
                "material_color_required",
                "材料明细存在数量但 Color 为空："
                f"{'、'.join(missing_colors)}；请补充颜色后重新扫描 Server",
            )

    colors: list[tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        color = _text(ws.cell(color_header_row, col).value)
        if color and _normalized_label(color) != "color:":
            colors.append((col, color))

    panels: list[MaterialItem] = []
    edges: dict[str, float] = {}
    if colors:
        color_table_34_total = 0.0
        color_table_14_total = 0.0
        color_table_edge_total = 0.0
        for col, source_color in colors:
            color = _canonical_color(source_color)
            qty_34 = summary_integer(
                row_34,
                col,
                f"{source_color} Sheets (3/4)",
                color_integer_fallback(6, source_color),
            )
            qty_14 = summary_integer(
                row_14,
                col,
                f"{source_color} Sheets (1/4)",
                color_integer_fallback(7, source_color),
            )
            edge = summary_number(row_edge, col, f"{source_color} Edge Banding", detail_sum(8, source_color))
            color_table_34_total += qty_34
            color_table_14_total += qty_14
            color_table_edge_total += edge
            if qty_34:
                panels.append(MaterialItem("panel", 19.1, color, qty_34))
            if qty_14:
                panels.append(MaterialItem("panel", 8.0, color, qty_14))
            if qty_34 or qty_14 or edge:
                if (qty_34 or qty_14) and edge <= 0:
                    raise RuleError("missing_edge", f"{source_color} 有 Panel 数量，但封边条为空或为 0")
                if edge > 0:
                    edges[color] = edge

        summary_mismatches = []
        if finish_34_col:
            total_34 = summary_integer(
                total_row,
                finish_34_col,
                "Total Qty 3/4 Finish Panel",
                detail_sum(finish_34_col),
            )
            if not math.isclose(total_34, color_table_34_total, abs_tol=EPSILON):
                summary_mismatches.append(
                    f"3/4 Finish Panel：Total Qty={_fmt(total_34)}，Color Table={_fmt(color_table_34_total)}"
                )
        if finish_14_col:
            total_14 = summary_integer(
                total_row,
                finish_14_col,
                "Total Qty 1/4 Finish Panel",
                detail_sum(finish_14_col),
            )
            if not math.isclose(total_14, color_table_14_total, abs_tol=EPSILON):
                summary_mismatches.append(
                    f"1/4 Finish Panel：Total Qty={_fmt(total_14)}，Color Table={_fmt(color_table_14_total)}"
                )
        if edge_col:
            total_edge = summary_number(
                total_row,
                edge_col,
                "Total Qty Edge Banding",
                detail_sum(edge_col),
            )
            if not math.isclose(total_edge, color_table_edge_total, abs_tol=EPSILON):
                summary_mismatches.append(
                    f"Edge Banding：Total Qty={_fmt(total_edge)}，Color Table={_fmt(color_table_edge_total)}"
                )
        if summary_mismatches:
            raise RuleError(
                "material_summary_mismatch",
                "Total Qty 与 Color Table 合计不一致："
                f"{'；'.join(summary_mismatches)}；请手工检查后再写入订单材料",
            )
    else:
        finish_34 = finish_34_col
        finish_14 = finish_14_col
        if not all((finish_34, finish_14, edge_col, color_col)):
            raise RuleError("materials_schema", "单颜色 materials 缺少 Panel、封边或 Color 列")
        source_color = _text(ws.cell(total_row, color_col).value)
        qty_34 = _integer_cell(ws.cell(total_row, finish_34), "Sheets (3/4)")
        qty_14 = _integer_cell(ws.cell(total_row, finish_14), "Sheets (1/4)")
        edge = summary_number(
            total_row,
            edge_col,
            "Edge Banding",
            detail_sum(edge_col, source_color),
        )
        if qty_34 or qty_14 or edge:
            if not source_color:
                raise RuleError("materials_schema", "单颜色 materials 有 Panel 或封边数量，但 Total Qty 行缺少 Color")
            if (qty_34 or qty_14) and edge <= 0:
                raise RuleError("missing_edge", f"{source_color} 有 Panel 数量，但封边条为空或为 0")
            color = _canonical_color(source_color)
            if qty_34:
                panels.append(MaterialItem("panel", 19.1, color, qty_34))
            if qty_14:
                panels.append(MaterialItem("panel", 8.0, color, qty_14))
            edges[color] = edge
    return wb.sheetnames[0], plywood + panels, edges


def _board_report_materials(
    path: Path,
    order_id: str,
    *,
    allow_unscoped: bool = False,
    fallback_name: str = "",
) -> dict:
    """Read the compact board/edge summary emitted in a Report folder."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        try:
            _, order_name = parse_board_identity(path)
        except RuleError:
            if not allow_unscoped:
                raise
            order_name = fallback_name.strip() or path.parent.name
    except Exception as exc:
        raise RuleError(
            "material_generation_failed",
            f"{path.name} 无法读取，目录或报表结构较复杂，请手动生成 material 文件：{exc}",
        ) from exc
    if not allow_unscoped and not _factory_name_belongs_to_order(order_id, order_name):
        return {}

    large_row = small_row = None
    for row in range(1, ws.max_row + 1):
        labels = {_normalized_label(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if "大板统计" in labels:
            large_row = row
        if "小板统计" in labels:
            small_row = row
            break
    if not large_row or not small_row or small_row <= large_row + 1:
        raise RuleError(
            "material_generation_failed",
            f"{path.name} 缺少可识别的大板/小板统计，请手动生成 material 文件",
        )

    quantity_col = None
    for row in range(large_row + 1, small_row):
        for col in range(1, ws.max_column + 1):
            if _normalized_label(ws.cell(row, col).value) == "数量":
                quantity_col = col
                break
        if quantity_col:
            break
    if not quantity_col:
        raise RuleError("material_generation_failed", f"{path.name} 找不到大板数量列，请手动生成 material 文件")

    plywood = {18.0: 0.0, 14.5: 0.0, 5.4: 0.0}
    panels: dict[tuple[float, str], float] = defaultdict(float)
    descriptor_re = re.compile(r"^([\d.]+)\s*mm/([^/]*)/", re.IGNORECASE)
    supported = {18.0, 14.5, 5.0, 5.4, 19.1, 8.0, 9.0}
    for row in range(large_row + 1, small_row):
        descriptor = ""
        for col in range(1, ws.max_column + 1):
            candidate = _text(ws.cell(row, col).value)
            if descriptor_re.match(candidate):
                descriptor = candidate
                break
        if not descriptor:
            continue
        match = descriptor_re.match(descriptor)
        thickness = float(match.group(1))
        color = _canonical_color(match.group(2))
        quantity = _integer(ws.cell(row, quantity_col).value, f"{path.name} {descriptor}")
        if not quantity:
            continue
        if thickness not in supported:
            raise RuleError(
                "material_generation_failed",
                f"{path.name} 包含暂不支持的 {thickness:g}mm 板材，请手动生成 material 文件",
            )
        if thickness in {18.0, 14.5, 5.0, 5.4} and color.lower() == "finished":
            plywood[5.4 if thickness in {5.0, 5.4} else thickness] += quantity
        elif thickness in {19.1, 8.0, 9.0} and color:
            panels[(8.0 if thickness in {8.0, 9.0} else 19.1, color)] += quantity
        else:
            raise RuleError(
                "material_generation_failed",
                f"{path.name} 的板材“{descriptor}”无法安全归类，请手动生成 material 文件",
            )

    edge_header_row = edge_col = color_col = None
    for row in range(small_row + 1, min(ws.max_row, small_row + 5) + 1):
        for col in range(1, ws.max_column + 1):
            label = _normalized_label(ws.cell(row, col).value)
            if label == "封边条/米":
                edge_header_row, edge_col = row, col
            elif label == "颜色":
                color_col = col
        if edge_col and color_col:
            break
    if not edge_header_row or not edge_col or not color_col:
        raise RuleError("material_generation_failed", f"{path.name} 找不到封边统计列，请手动生成 material 文件")
    edges: dict[str, float] = defaultdict(float)
    for row in range(edge_header_row + 1, ws.max_row + 1):
        color = _canonical_color(_text(ws.cell(row, color_col).value))
        if not color:
            continue
        edge = _number(ws.cell(row, edge_col).value, f"{path.name} {color} 封边")
        if edge < 0:
            raise RuleError("material_generation_failed", f"{path.name} 的 {color} 封边数量为负数，请手动生成 material 文件")
        if edge:
            edges[color] += edge
    if not any(plywood.values()) and not panels:
        raise RuleError("material_generation_failed", f"{path.name} 没有可生成的板材数量，请手动生成 material 文件")

    room = re.sub(rf"^{re.escape(order_id)}(?:-|\s+)", "", order_name, flags=re.IGNORECASE).strip() or order_name
    return {"room": room, "plywood": plywood, "panels": dict(panels), "edges": dict(edges)}


def _legacy_traveler_factory_name(path: Path, workbook) -> str:
    for sheet_name in (WORK_ORDER_SHEET, "Picking List", "Pickinglist"):
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        for row in ws.iter_rows(values_only=True):
            values = [_text(value) for value in row]
            for index, value in enumerate(values):
                if _normalized_label(value) == _normalized_label("Name/工厂单名称"):
                    return next((item for item in values[index + 1:] if item), path.stem)
    return path.stem


def _material_item_from_traveler_name(name: str, quantity: float, room_name: str) -> MaterialItem | None:
    value = _text(name)
    if not value or quantity <= 0:
        return None
    edge_match = re.match(r"^Edge banding-+(.+)$", value, re.IGNORECASE)
    if edge_match:
        return MaterialItem("edge", 0.0, _canonical_color(edge_match.group(1)), quantity, room_name)
    panel_match = re.match(r"^(\d+(?:\.\d+)?)mm--(.*)$", value, re.IGNORECASE)
    if not panel_match:
        return None
    thickness = float(panel_match.group(1))
    color = _canonical_color(panel_match.group(2))
    if math.isclose(thickness, 18.0, abs_tol=EPSILON):
        return MaterialItem("plywood", 18.0, "", quantity, room_name)
    if math.isclose(thickness, 14.5, abs_tol=EPSILON):
        return MaterialItem("plywood", 14.5, "", quantity, room_name)
    if math.isclose(thickness, 5.4, abs_tol=EPSILON):
        return MaterialItem("plywood", 5.4, "", quantity, room_name)
    if math.isclose(thickness, 19.1, abs_tol=EPSILON):
        return MaterialItem("panel", 19.1, color, quantity, room_name)
    if math.isclose(thickness, 8.0, abs_tol=EPSILON) or math.isclose(thickness, 9.0, abs_tol=EPSILON):
        # Legacy Travelers call this finish-panel size 9mm; the current
        # material workbook stores the same category in its 1/4 Finish Panel column.
        return MaterialItem("panel", 8.0, color, quantity, room_name)
    raise RuleError(
        "material_generation_failed",
        f"{room_name} 使用暂不支持的板材厚度 {thickness:g}mm，请人工确认后再生成 material",
    )


def _legacy_picking_material_items(path: Path, workbook, order_id: str) -> list[MaterialItem]:
    sheet_name = next((name for name in ("Picking List", "Pickinglist") if name in workbook.sheetnames), None)
    if not sheet_name:
        raise RuleError("material_generation_failed", f"{path.name} 缺少 Usage List 或 Picking List，无法提取板材数据")
    ws = workbook[sheet_name]
    room_name = _legacy_traveler_factory_name(path, workbook)
    panel_section = False
    header: tuple[int, int] | None = None
    items: list[MaterialItem] = []
    for row in ws.iter_rows(values_only=True):
        values = [_text(value) for value in row]
        normalized = {_normalized_label(value) for value in values if value}
        if any("panel" in value and "板材" in value for value in normalized):
            panel_section = True
            header = None
            continue
        if panel_section and any("fitting" in value or "配件" in value for value in normalized):
            break
        if not panel_section:
            continue
        name_column = next((index for index, value in enumerate(values) if "name" in value.lower() and "名字" in value), None)
        quantity_column = next((index for index, value in enumerate(values) if "qty" in value.lower() and "数量" in value), None)
        if name_column is not None and quantity_column is not None:
            header = (name_column, quantity_column)
            continue
        if header is None:
            continue
        name_column, quantity_column = header
        if name_column >= len(values) or quantity_column >= len(row):
            continue
        name = values[name_column]
        raw_quantity = row[quantity_column]
        if not name or raw_quantity in (None, "", 0, 0.0):
            continue
        if isinstance(raw_quantity, bool) or not isinstance(raw_quantity, (int, float)):
            raise RuleError("material_generation_failed", f"{path.name} 中 {name} 的数量不是有效数字")
        item = _material_item_from_traveler_name(name, float(raw_quantity), room_name)
        if item:
            items.append(item)
    if not items:
        raise RuleError("material_generation_failed", f"{path.name} 没有可用的板材或封边数据")
    return items


def _traveler_material_items(path: Path, order_id: str) -> list[MaterialItem]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if USAGE_LIST_SHEET in workbook.sheetnames:
        traveler = parse_traveler(path)
        source_items = traveler.documents.get(traveler.order_id, [])
        items = []
        for source in source_items:
            item = _material_item_from_traveler_name(source.name, float(source.quantity), traveler.order_name)
            if item:
                items.append(item)
        if items:
            return items
    return _legacy_picking_material_items(path, workbook, order_id)


def _aggregate_traveler_material_details(items: list[MaterialItem]) -> tuple[list[list[object]], list[str]]:
    plywood = defaultdict(float)
    by_color: dict[str, dict[str, float]] = defaultdict(lambda: {"panel_34": 0.0, "panel_14": 0.0, "edge": 0.0})
    for item in items:
        if item.kind == "plywood":
            plywood[item.thickness] += item.quantity
        elif item.kind == "panel":
            by_color[item.color]["panel_34" if math.isclose(item.thickness, 19.1, abs_tol=EPSILON) else "panel_14"] += item.quantity
        elif item.kind == "edge":
            by_color[item.color]["edge"] += item.quantity
    colors = sorted(
        color for color, values in by_color.items()
        if color and any(value > 0 for value in values.values())
    )
    if len(colors) > 12:
        raise RuleError("material_generation_failed", f"Traveler 中有 {len(colors)} 种颜色，当前 material 模板最多支持 12 种")
    details: list[list[object]] = []
    plywood_values = [plywood[18.0], plywood[14.5], plywood[5.4]]
    if any(value > 0 for value in plywood_values) or not colors:
        details.append(["Traveler 汇总", "", *plywood_values, 0, 0, 0, ""])
    for index, color in enumerate(colors):
        values = by_color[color]
        details.append([
            "Traveler 汇总" if not details else "",
            "",
            0,
            0,
            0,
            values["panel_34"],
            values["panel_14"],
            values["edge"],
            color,
        ])
    if not details:
        raise RuleError("material_generation_failed", "Traveler 中没有可生成的板材或封边数量")
    return details, colors


def _write_generated_material_workbook(destination: Path, order_id: str, details: list[list[object]], colors: list[str]) -> Path:
    template = Path(__file__).resolve().parents[1] / "resources/templates/Order Materials.xlsx"
    if not template.is_file():
        raise RuleError("material_generation_failed", "缺少人工 material 模板，请手动生成 material 文件")
    wb = load_workbook(template)
    ws = wb[wb.sheetnames[0]]
    ws["B1"] = order_id
    if len(details) > 11:
        raise RuleError("material_generation_failed", f"需要 {len(details)} 行 material 明细，超出模板容量")
    if len(colors) > MATERIAL_DETAIL_ROW_CAPACITY:
        raise RuleError(
            "material_generation_failed",
            f"需要 {len(colors)} 种颜色，超出 material 明细行容量 {MATERIAL_DETAIL_ROW_CAPACITY} 行",
        )
    required_color_end = max(9, MATERIAL_COLOR_COLUMN_START - 1 + len(colors))
    for col in range(10, required_color_end + 1):
        _copy_column_style(ws, 9, col)
    for row_index, values in enumerate(details, start=3):
        ws.cell(row_index, 1).value = values[0]
        for column in range(3, 10):
            ws.cell(row_index, column).value = values[column - 1]
    total_row = 14
    for col in range(3, 6):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}3:{letter}13)"
    for col in range(6, 9):
        letter = get_column_letter(col)
        last_color_letter = get_column_letter(max(9, 2 + len(colors)))
        ws.cell(total_row, col).value = (
            f'=IF(COUNTA($C$16:${last_color_letter}$16)>1,"check color table",SUM({letter}3:{letter}13))'
        )
    ws.cell(total_row, 9).value = colors[0] if len(colors) == 1 else ""
    for index, color in enumerate(colors, start=3):
        letter = get_column_letter(index)
        ws.cell(16, index).value = color
        ws.cell(17, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$F$3:$F$13))'
        ws.cell(18, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$G$3:$G$13))'
        ws.cell(19, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$H$3:$H$13))'
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    temporary = destination.with_suffix(".tmp.xlsx")
    try:
        wb.save(temporary)
        parse_order_materials(order_id, temporary)
        temporary.replace(destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, RuleError):
            raise RuleError("material_generation_failed", f"自动生成的 material 校验失败：{exc}") from exc
        raise RuleError("material_generation_failed", f"无法写入 material 文件：{exc}") from exc
    return destination


def _usage_rows_for_traveler(items: list[MaterialItem], room_name: str) -> list[list[object]]:
    details, _ = _aggregate_traveler_material_details(items)
    for row in details:
        if row[0] == "Traveler 汇总":
            row[0] = room_name
    return details


def _update_legacy_traveler_usage_list(config: Config, path: Path, order_id: str, items: list[MaterialItem]) -> Path | None:
    workbook = load_workbook(path, data_only=False)
    if USAGE_LIST_SHEET in workbook.sheetnames:
        return None
    legacy_picking = next((name for name in workbook.sheetnames if name.casefold() == "pickinglist"), None)
    if legacy_picking and PICKING_LIST_SHEET not in workbook.sheetnames:
        workbook[legacy_picking].title = PICKING_LIST_SHEET
    picking = workbook[PICKING_LIST_SHEET]
    factory_name = _legacy_traveler_factory_name(path, workbook)
    for row in picking.iter_rows():
        for index, cell in enumerate(row):
            if _normalized_label(cell.value) == _normalized_label("Name/工厂单名称"):
                for target in row[index + 1:]:
                    if not isinstance(target, MergedCell):
                        target.value = factory_name
                        break
                break
    _restore_template_usage_list(config, workbook)
    target = workbook[USAGE_LIST_SHEET]
    details = _usage_rows_for_traveler(items, factory_name)
    _expand_usage_detail_rows(target, len(details))
    total_row = _find_label_row(target, "Total Qty:")
    for row in range(3, total_row):
        for col in range(1, 10):
            cell = target.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for row_index, values in enumerate(details, start=3):
        for col in range(1, 10):
            target_cell = target.cell(row_index, col)
            if not isinstance(target_cell, MergedCell):
                target_cell.value = values[col - 1]
    _write_merged(target, 1, 2, order_id)
    _write_usage_formulas(target)
    backup = _backup_traveler(config, path, order_id, "legacy-usage-list-backup")
    draft = path.with_name(f".{path.stem}.usage-list.tmp.xlsx")
    try:
        workbook.save(draft)
        check = load_workbook(draft, data_only=False, read_only=True)
        required = [WORK_ORDER_SHEET, USAGE_LIST_SHEET, PICKING_LIST_SHEET]
        if check.sheetnames[:3] != required:
            raise RuleError("write_verification", f"{path.name} 更新后工作表顺序不正确：{check.sheetnames[:3]}")
        parse_traveler(draft)
        os.replace(draft, path)
    except Exception:
        draft.unlink(missing_ok=True)
        raise
    return backup


def generate_material_from_travelers(config: Config, folder: Path, order_id: str, *, confirm_write: bool = False) -> dict:
    if not confirm_write:
        raise RuleError("write_confirmation_required", "此操作会更新本机 Traveler 并写入 Server，请使用 --confirm-write")
    folder = folder.expanduser().resolve()
    order_id = order_id.strip().upper()
    if not folder.is_dir() or not order_id:
        raise RuleError("material_generation_failed", "订单目录或订单号无效")
    existing = [
        path for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx"
        and "material" in path.name.casefold() and not path.name.startswith("~$")
    ]
    if existing:
        raise RuleError("material_generation_failed", f"material 文件已存在：{existing[0].name}")
    local_folder = next(
        (path for path in config.order_root.iterdir() if path.is_dir() and path.name.casefold() == order_id.casefold()),
        None,
    )
    if local_folder is None:
        raise RuleError("material_generation_failed", f"本机订单文件夹不存在：{order_id}")
    travelers = sorted(local_folder.glob("Work Order Traveler(*).xlsx"))
    if not travelers:
        raise RuleError("material_generation_failed", f"本机订单文件夹没有 Traveler：{order_id}")
    all_items: list[MaterialItem] = []
    legacy_updates: list[tuple[Path, list[MaterialItem]]] = []
    for path in travelers:
        items = _traveler_material_items(path, order_id)
        all_items.extend(items)
        workbook = load_workbook(path, read_only=True, data_only=False)
        if USAGE_LIST_SHEET not in workbook.sheetnames:
            legacy_updates.append((path, items))
    details, colors = _aggregate_traveler_material_details(all_items)
    temporary_material = Path(tempfile.mkdtemp(prefix="traveler-material-preview-")) / f"{order_id} materials.xlsx"
    try:
        _write_generated_material_workbook(temporary_material, order_id, details, colors)
        backups = []
        for path, items in legacy_updates:
            backup = _update_legacy_traveler_usage_list(config, path, order_id, items)
            if backup:
                backups.append(str(backup))
        destination = folder / f"{order_id} materials.xlsx"
        shutil.copy2(temporary_material, destination)
        parse_order_materials(order_id, destination)
    finally:
        shutil.rmtree(temporary_material.parent, ignore_errors=True)
    return {
        "order_id": order_id,
        "materials_file": str(destination),
        "traveler_count": len(travelers),
        "legacy_traveler_updates": len(legacy_updates),
        "traveler_backups": backups,
        "colors": colors,
    }


def generate_material_from_reports(folder: Path, order_id: str) -> Path:
    """Create a parser-compatible material workbook from Report board summaries."""
    folder = folder.resolve()
    order_id = order_id.strip().upper()
    if not folder.is_dir() or not order_id:
        raise RuleError("material_generation_failed", "订单目录或订单号无效，请手动生成 material 文件")
    destination = folder / f"{order_id} materials.xlsx"
    if destination.exists():
        raise RuleError("material_generation_failed", f"material 文件已存在：{destination.name}")
    reports = sorted(
        path for path in folder.rglob("*.xlsx")
        if not path.name.startswith("~$") and "板材清单" in path.name
    )
    if not reports:
        raise RuleError("material_generation_failed", "Report 中找不到板材清单，请手动生成 material 文件")
    if ORDER_FOLDER_RE.fullmatch(order_id):
        related = set(related_order_ids(folder))
        foreign = sorted(related - {order_id})
        if foreign:
            raise RuleError(
                "mixed_order_material",
                f"订单目录同时包含其他订单数据：{'、'.join(foreign)}；不能自动生成 material，请人工分配",
                order_ids=sorted(related),
            )
    seen_factories: dict[str, Path] = {}
    for path in reports:
        try:
            factory, report_name = parse_board_identity(path)
        except RuleError:
            # Let the normal parser retain its detailed filename/schema error.
            factory = ""
            report_name = ""
        if factory and _factory_name_belongs_to_order(order_id, report_name):
            if factory in seen_factories:
                raise RuleError(
                    "overlapping_material_reports",
                    f"工厂单 {factory} 在多份板材清单中出现（{seen_factories[factory].name}、{path.name}），报表范围可能重复，不能自动相加",
                    factory_order=factory,
                    files=[str(seen_factories[factory]), str(path)],
                )
            seen_factories[factory] = path
    rows = [
        row for path in reports
        if (row := _board_report_materials(
            path,
            order_id,
            allow_unscoped=not ORDER_FOLDER_RE.fullmatch(order_id),
            fallback_name=folder.name,
        ))
    ]
    if not rows:
        raise RuleError("material_generation_failed", f"Report 中没有属于 {order_id} 的板材清单，请手动生成 material 文件")

    colors = sorted({color for row in rows for color in (set(row["edges"]) | {key[1] for key in row["panels"]})})
    if len(colors) > MATERIAL_DETAIL_ROW_CAPACITY:
        raise RuleError(
            "material_generation_failed",
            f"Report 中有 {len(colors)} 种颜色，超出 material 明细行容量 {MATERIAL_DETAIL_ROW_CAPACITY} 行，请手动生成 material 文件",
        )
    details: list[list[object]] = []
    for source in rows:
        row_colors = sorted(set(source["edges"]) | {key[1] for key in source["panels"]}) or [""]
        for index, color in enumerate(row_colors):
            plywood = source["plywood"] if index == 0 else {}
            details.append([
                source["room"] if index == 0 else "", "",
                plywood.get(18.0, 0), plywood.get(14.5, 0), plywood.get(5.4, 0),
                source["panels"].get((19.1, color), 0), source["panels"].get((8.0, color), 0),
                source["edges"].get(color, 0), color,
            ])
    if len(details) > 11:
        raise RuleError(
            "material_generation_failed",
            f"Report 需要 {len(details)} 行明细，超出人工 material 模板的 11 行容量，请手动生成 material 文件",
        )

    template = Path(__file__).resolve().parents[1] / "resources/templates/Order Materials.xlsx"
    if not template.is_file():
        raise RuleError("material_generation_failed", "缺少人工 material 模板，请手动生成 material 文件")
    wb = load_workbook(template)
    ws = wb[wb.sheetnames[0]]
    ws["B1"] = order_id
    for row_index, values in enumerate(details, start=3):
        ws.cell(row_index, 1).value = values[0]
        for column in range(3, 10):
            ws.cell(row_index, column).value = values[column - 1]

    total_row = 14
    for col in range(3, 6):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}3:{letter}13)"
    required_color_end = max(9, MATERIAL_COLOR_COLUMN_START - 1 + len(colors))
    for col in range(10, required_color_end + 1):
        _copy_column_style(ws, 9, col)
    for col in range(6, 9):
        letter = get_column_letter(col)
        last_color_letter = get_column_letter(required_color_end)
        ws.cell(total_row, col).value = (
            f'=IF(COUNTA($C$16:${last_color_letter}$16)>1,"check color table",SUM({letter}3:{letter}13))'
        )
    ws.cell(total_row, 9).value = colors[0] if len(colors) == 1 else ""

    for index, color in enumerate(colors, start=3):
        letter = get_column_letter(index)
        ws.cell(16, index).value = color
        ws.cell(17, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$F$3:$F$13))'
        ws.cell(18, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$G$3:$G$13))'
        ws.cell(19, index).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$H$3:$H$13))'
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    temporary = destination.with_suffix(".tmp.xlsx")
    try:
        wb.save(temporary)
        parse_order_materials(order_id, temporary)
        temporary.replace(destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, RuleError):
            raise RuleError("material_generation_failed", f"自动生成的 material 校验失败，请手动生成：{exc}") from exc
        raise RuleError("material_generation_failed", f"无法写入 material 文件，请手动生成：{exc}") from exc
    return destination


def parse_material_room_rows(path: Path) -> list[tuple[str, list[MaterialItem], dict[str, float]]]:
    """Read room-level material quantities when the material workbook provides them."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = 2
    headers = {
        _normalized_label(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if _text(ws.cell(header_row, col).value)
    }
    required = {
        "3/4plywood": (18.0, "plywood"), "5/8plywood": (14.5, "plywood"),
        "1/4plywood": (5.4, "plywood"), "3/4finishpanel": (19.1, "panel"),
        "1/4finishpanel": (8.0, "panel"), "edgebanding(m)": (0.0, "edge"), "color": (0.0, "color"),
    }
    if not all(label in headers for label in required):
        return []
    total_row = _find_label_row(ws, "Total Qty:")
    rows = []
    for row in range(header_row + 1, total_row):
        room = _text(ws.cell(row, 1).value)
        if _normalized_label(room) in {"room/section", "color table", "traveler汇总", "total qty:"}:
            continue
        color = _canonical_color(_text(ws.cell(row, headers["color"]).value))
        items = []
        for label, (thickness, kind) in required.items():
            if kind not in {"plywood", "panel"}:
                continue
            # Room rows use the same display-oriented integer convention as
            # the summary rows: some source workbooks store a fractional
            # formula result while formatting the cell as an integer. Read
            # the displayed integer here too, otherwise room aggregation can
            # reintroduce values such as 6.25 after the summary parser has
            # already normalized them correctly.
            try:
                quantity = _integer_cell(ws.cell(row, headers[label]), f"{room} {label}")
            except RuleError as exc:
                # Room rows can contain cut/layout fractions while Total Qty
                # remains the authoritative order quantity. Do not turn such
                # a row into a false factory allocation; the caller will use
                # the order totals and show the material as unallocated.
                if exc.code == "fractional_material":
                    return []
                raise
            if quantity < 0:
                raise RuleError("negative_material", f"{room} {label} 数量不能为负数：{quantity:g}")
            if quantity:
                items.append(MaterialItem(kind, thickness, color if kind == "panel" else "", quantity, room))
        edge_cell = ws.cell(row, headers["edgebanding(m)"])
        edge = _display_number(edge_cell.value, edge_cell.number_format, f"{room} Edge Banding")
        if not room and (items or edge):
            room = "__UNASSIGNED__"
        if not room:
            continue
        rows.append((room, items, {color: edge} if color and edge else {}))
    return rows


def _aggregate_material_sources(order_id: str, paths: list[Path]):
    all_materials: list[MaterialItem] = []
    all_edges: dict[str, float] = defaultdict(float)
    all_rooms = []
    signatures = []
    sheet_name = ""
    for path in paths:
        sheet_name, materials, edges = parse_order_materials(order_id, path)
        rooms = parse_material_room_rows(path)
        all_materials.extend(materials)
        for color, quantity in edges.items():
            all_edges[color] += quantity
        all_rooms.extend(rooms)
        signatures.append((path, tuple((room, tuple((i.kind, i.thickness, i.color, i.quantity) for i in items), tuple(sorted(edge.items()))) for room, items, edge in rooms)))
    duplicate_warning = []
    seen = {}
    for path, signature in signatures:
        if signature and signature in seen:
            duplicate_warning.append(f"{path.name} 与 {seen[signature].name} 的房间材料和封边数据一致，可能重复计数")
        elif signature:
            seen[signature] = path
    return sheet_name, all_materials, dict(all_edges), all_rooms, duplicate_warning


def _next_value_on_row(ws, row: int, col: int) -> str:
    for candidate in range(col + 1, ws.max_column + 1):
        value = _text(ws.cell(row, candidate).value)
        if value:
            return value
    return ""


def parse_board_identity(path: Path) -> tuple[str, str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    factory = ""
    name = ""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            label = _normalized_label(ws.cell(row, col).value)
            if label == "订单号":
                factory = _next_value_on_row(ws, row, col).upper()
            elif label == "订单名称":
                name = _next_value_on_row(ws, row, col)
        if factory and name:
            break
    if not FACTORY_RE.fullmatch(factory) or not name:
        raise RuleError("board_identity", f"板材清单无法取得工厂单号或名称：{path}")
    return factory, name


def _has_positive_fitting_quantity(path: Path) -> bool | None:
    """Return None when the workbook cannot be opened, so corruption is never ignored."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        # Empty AICNC reports can carry the same invalid A1:?0 worksheet
        # dimension as usable reports. Normal mode ignores that optional
        # dimension and lets us distinguish an empty report from corruption.
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
        except Exception:
            return None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=11, max_col=11, values_only=True):
            value = row[0]
            if isinstance(value, (int, float)) and math.isfinite(float(value)) and float(value) > 0:
                return True
    return False


def _fittings_report_is_empty(path: Path) -> bool:
    """Recognize an intentionally empty fittings report, not a broken report."""
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return False
    values = [
        _text(value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows(values_only=True)
        for value in row
        if _text(value)
    ]
    if not values:
        return True
    normalized = {re.sub(r"\s+", " ", value).casefold() for value in values}
    return normalized.issubset({
        "no fittings were used",
        "no fittings used",
        "no fitting was used",
        "没有使用五金",
        "没有五金",
    })


def _choose_fittings(
    folder: Path,
    *,
    allow_missing_factory: bool = False,
    fallback_factory: str = "",
) -> tuple[dict[str, list[FittingItem]], list[str]]:
    paths = sorted(
        path for path in folder.rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.lower().startswith("fittingslist")
    )
    selected, warnings, found_files, skipped_empty = select_latest_fittings(
        paths,
        allow_missing_factory=allow_missing_factory,
        fallback_factory=fallback_factory,
        is_empty_report=_fittings_report_is_empty,
    )
    warnings.extend(
        f"{path.name} 没有内容，已按无五金处理；Traveler 仍可生成"
        for path in skipped_empty
    )
    if not selected:
        if not found_files:
            warnings.append("未找到 Fittingslist Excel，本订单将按无五金继续生成 Traveler")
        elif not warnings:
            warnings.append("所有 Fittingslist 均没有有效五金数量，本订单将按无五金继续生成 Traveler")
        return {}, warnings
    return {factory: list(source.items) for factory, source in selected.items()}, warnings


def _ignored_key(name: str, code: str, size: str, unit: str) -> str:
    payload = "\x1f".join((_text(name).lower(), _text(code).upper(), _text(size).lower(), _text(unit).lower()))
    import hashlib
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]


def _normalize_fittings(items: list[FittingItem], mappings: InventoryMappings) -> list[PreviewFitting]:
    direct = {"WJ-CBT": ("Shelf Holder", "pcs/个"), "71T950A": ("Hinge", "pcs/个")}
    rails = {"H-RAIL": ("H-Rail", "set/套"), "L-RAIL": ("L-Rail", "set/套")}
    aggregate: dict[tuple[str, str, str, str], float] = defaultdict(float)
    rail_sides: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for item in items:
        code = _text(item.code).upper()
        if code in direct:
            name, unit = direct[code]
            aggregate[(name, code, "", unit)] += item.quantity
        elif code in rails:
            side = "right" if "right" in item.name.lower() else "left" if "left" in item.name.lower() else "unknown"
            rail_sides[code][side] += item.quantity
        else:
            aggregate[(_text(item.name), code, _text(item.size), _text(item.unit) or "pcs/个")] += item.quantity
    for code, sides in rail_sides.items():
        name, unit = rails[code]
        if set(sides) != {"left", "right"} or abs(sides["left"] - sides["right"]) > EPSILON:
            raise RuleError("rail_mismatch", f"{name} 左右数量不一致：{dict(sides)}")
        aggregate[(name, code, "", unit)] += sides["left"]
    result = []
    for (name, code, size, unit), quantity in sorted(aggregate.items()):
        key = _ignored_key(name, code, size, unit)
        result.append(PreviewFitting(
            key, name, code, size, unit, quantity,
            mappings.ignored_reason(name) is not None,
        ))
    return result


def _factory_names(
    config: Config,
    folder: Path,
    factories: set[str],
    order_id: str,
    *,
    allow_unscoped: bool = False,
    fallback_name: str = "",
    fallback_factory: str = "",
) -> tuple[dict[str, str], list[str]]:
    names: dict[str, str] = {}
    for path in folder.rglob("*.xlsx"):
        if path.name.startswith("~$") or "板材清单" not in path.name:
            continue
        try:
            factory, name = parse_board_identity(path)
        except RuleError:
            if allow_unscoped:
                continue
            raise
        if not allow_unscoped and not _factory_name_belongs_to_order(order_id, name):
            # A shared folder may contain board lists for a sibling order such
            # as PP0035. Ignore those rows here; they must not contaminate the
            # selected split order PP0035-2.
            continue
        previous = names.get(factory)
        if previous and previous != name:
            raise RuleError("factory_name_conflict", f"{factory} 在板材清单中对应多个名称：{previous} / {name}")
        names[factory] = name
    if allow_unscoped and fallback_name.strip():
        # A no-AIMES temporary order intentionally has no F-number.  Seed the
        # folder identity before the AIMES lookup so offline/manual handling
        # never tries to resolve a synthetic folder key as a real factory.
        names.setdefault(fallback_factory.strip(), fallback_name.strip())
    cache = load_factory_name_cache(config)
    for factory in factories - set(names):
        if cache.get(factory):
            names[factory] = cache[factory]
    missing = sorted(factories - set(names))
    if missing:
        fetched = lookup_aimes_names(config, missing)
        names.update(fetched)
        cache.update(fetched)
        save_factory_name_cache(config, cache)
    foreign = [
        factory for factory, name in names.items()
        if not allow_unscoped and not _factory_name_belongs_to_order(order_id, name)
    ]
    for factory in foreign:
        names.pop(factory, None)
    return names, ([f"已忽略 {len(foreign)} 个属于其他订单的工厂单"] if foreign else [])


def _room_matches_order(room: str, order_id: str, names: dict[str, str]) -> bool:
    room_key = re.sub(r"[^a-z0-9]+", "", room.lower())
    room_tokens = [token for token in re.split(r"[^a-z0-9]+", room.lower()) if len(token) >= 4]
    for name in names.values():
        if not _factory_name_belongs_to_order(order_id, name):
            continue
        name_key = re.sub(r"[^a-z0-9]+", "", name.lower())
        if room_key and (room_key in name_key or name_key.endswith(room_key) or any(token in name_key for token in room_tokens)):
            return True
    return False


def _room_order_ids(room: str) -> list[str]:
    return sorted({match.upper() for match in ORDER_TOKEN_RE.findall(_text(room))})


def _room_has_explicit_order_identity(rows) -> bool:
    return any(_room_order_ids(room) for room, _, _ in rows)


def _select_room_materials(
    rows,
    order_id: str,
    names: dict[str, str],
    known_order_ids: set[str] | None = None,
):
    known = {str(value).strip().upper() for value in (known_order_ids or {order_id}) if str(value).strip()}
    known.add(order_id.upper())
    strict = len(known) > 1
    selected = []
    invalid_rooms: list[str] = []
    for row in rows:
        room, items, edges = row
        has_quantity = any(item.quantity > EPSILON for item in items) or any(value > EPSILON for value in edges.values())
        if room == "__UNASSIGNED__":
            if strict and has_quantity:
                invalid_rooms.append("未填写")
            elif has_quantity:
                selected.append(row)
            continue
        room_orders = _room_order_ids(room)
        if len(room_orders) > 1:
            invalid_rooms.append(f"{room}（包含多个订单号）")
        elif room_orders:
            owner = room_orders[0]
            if owner == order_id.upper():
                selected.append(row)
            elif owner not in known:
                invalid_rooms.append(f"{room}（订单号不在当前文件夹）")
            # A row explicitly belonging to a sibling order is valid, but is
            # intentionally omitted from the current order.
        elif strict:
            invalid_rooms.append(f"{room}（缺少订单号）")
        elif _room_matches_order(room, order_id, names):
            selected.append(row)
        elif has_quantity:
            # A single-order legacy workbook may use a plain room name. Keep
            # that compatibility path; shared-order workbooks must use an
            # explicit order-bearing factory name.
            selected.append(row)
    if invalid_rooms:
        examples = "、".join(dict.fromkeys(invalid_rooms[:5]))
        raise RuleError(
            "material_room_owner_required",
            f"{order_id} 的 material 存在无法确定归属的 Room/section：{examples}；"
            "请填写包含订单号的工厂单名称，例如 PP0035-KITCHEN 或 PP0035-2-MASTER",
        )
    if not selected:
        return [], {}, [f"{order_id} 没有在 material 中匹配到明确房间，需人工检查"]
    grouped: dict[tuple[str, float, str], list[tuple[str, float]]] = defaultdict(list)
    for room, items, _ in selected:
        for item in items:
            grouped[(item.kind, item.thickness, item.color)].append((room, item.quantity))
    materials = [
        MaterialItem(kind, thickness, color, sum(quantity for _, quantity in values), ", ".join(sorted({room for room, _ in values if room != "__UNASSIGNED__"})))
        for (kind, thickness, color), values in sorted(
            grouped.items(),
            key=lambda entry: (
                {"plywood": 0, "panel": 1, "back": 2}.get(entry[0][0].casefold(), 9),
                entry[0][2].casefold(),
                entry[0][1],
            ),
        )
    ]
    totals: dict[tuple[str, float, str], float] = defaultdict(float)
    for item in materials:
        totals[(item.kind, item.thickness, item.color)] += item.quantity
    for (kind, thickness, color), quantity in totals.items():
        if abs(quantity - round(quantity)) > EPSILON:
            label = f"{kind} {thickness:g}mm {color}".strip()
            raise RuleError(
                "fractional_material",
                f"{order_id} 汇总后 {label} 数量为 {quantity:g}，板材总数量必须为整数，请人工检查",
            )
    edges: dict[str, float] = defaultdict(float)
    for _, _, values in selected:
        for color, quantity in values.items():
            edges[color] += quantity
    warnings = [f"{order_id} 有未标注房间的 material 行，已计入当前订单，请人工核对"] if any(room == "__UNASSIGNED__" for room, _, _ in selected) else []
    return materials, dict(edges), warnings


def preview_order(
    config: Config,
    folder: Path,
    requested_order_id: str | None = None,
    *,
    include_hardware: bool = True,
    temporary_factory_order: str = "",
    temporary_factory_name: str = "",
) -> OrderPreview:
    if not folder.is_dir():
        raise RuleError("missing_order_folder", f"订单文件夹不存在：{folder}")
    match = ORDER_FOLDER_RE.fullmatch(folder.name)
    mixed_folder = match is None and len(related_order_ids(folder)) >= 2
    temporary_folder = match is None and not mixed_folder
    if not match and not requested_order_id:
        raise RuleError("invalid_order_folder", f"订单文件夹名称不符合 PP####、PP####-# 或 CS###：{folder.name}")
    order_id = (requested_order_id or match.group(1)).upper()
    materials_files = sorted(
        path for path in folder.rglob("*.xlsx")
        if path.is_file() and path.suffix.lower() == ".xlsx"
        and "material" in path.name.lower()
        and not path.name.startswith("~$")
        and not path.name.lower().startswith("panelmaterial")
    )
    exact_materials = [
        path for path in materials_files
        if re.search(rf"(?<![A-Z0-9]){re.escape(order_id)}(?![A-Z0-9]|-\d)", path.stem, re.IGNORECASE)
    ]
    assignments = load_material_assignments(config)
    assignment_key = f"{folder.resolve()}::{order_id}"
    assigned = Path(assignments[assignment_key]) if assignments.get(assignment_key) else None
    if assigned and assigned.is_file() and assigned in materials_files:
        materials_files = [assigned]
        exact_materials = [assigned]
    if exact_materials:
        materials_files = exact_materials
    elif temporary_folder and len(materials_files) > 1:
        materials_files = [max(materials_files, key=lambda path: path.stat().st_mtime)]
    if not materials_files:
        raise RuleError("missing_materials", f"{order_id} 根目录找不到文件名包含 material 的 Excel")
    if len(materials_files) > 1:
        exact = [path for path in materials_files if re.search(rf"(?<![A-Z0-9]){re.escape(order_id)}(?![A-Z0-9]|-\d)", path.stem, re.IGNORECASE)]
        if assigned and assigned in materials_files:
            materials_path = assigned
        elif exact:
            materials_files = exact
            materials_path = exact[0]
            if len(exact) == 1:
                save_material_assignment(config, assignment_key, str(materials_path))
        else:
            raise RuleError(
                "material_assignment_required",
                f"{order_id} 找到多个 material Excel，需要人工分配",
                files=[str(x) for x in materials_files],
                candidates=[str(x) for x in exact],
                assignment_key=assignment_key,
            )
    else:
        materials_path = materials_files[0]
    repair_warnings = []
    for materials_file in materials_files:
        repair = repair_material_color_table(materials_file)
        if repair.get("corrected"):
            # This is an intentional App write to the source workbook.
            from .order_index import OrderIndexStore, _record_generated_material_baseline

            baseline_store = OrderIndexStore(config.workflow_database)
            try:
                _record_generated_material_baseline(
                    baseline_store,
                    folder,
                    materials_file,
                    order_id=order_id,
                )
                baseline_store.commit()
            finally:
                baseline_store.close()
            repair_warnings.append(repair["message"])
    sheet_name, materials, edges, room_rows, duplicate_warnings = _aggregate_material_sources(order_id, materials_files)
    color_table_materials = list(materials)
    color_table_edges = dict(edges)
    warnings = repair_warnings + list(duplicate_warnings)
    selected_materials = list(materials)
    selected_edges = dict(edges)
    selected_room_rows = list(room_rows)
    if order_id.startswith("CS") and not temporary_folder:
        result = OrderPreview(
            order_id, folder, materials_path, sheet_name, selected_materials, selected_edges, [],
            warnings + ["来料加工订单只读取板材和封边数量，不读取或写入五金"],
            include_hardware=include_hardware,
            material_room_rows=selected_room_rows,
        )
        persist_preview(config, result)
        return result
    if include_hardware:
        fittings, fitting_warnings = _choose_fittings(
            folder,
            allow_missing_factory=temporary_folder and not temporary_factory_order,
            fallback_factory=temporary_factory_order or folder.name,
        )
        warnings.extend(fitting_warnings)
    else:
        fittings = {}
        warnings.append("本次选择不包含五金；即使报表中存在五金，也不会写入 Traveler 或出库")
    names, name_warnings = _factory_names(
        config,
        folder,
        set(fittings),
        order_id,
        allow_unscoped=temporary_folder,
        fallback_name=temporary_factory_name or folder.name,
        fallback_factory=temporary_factory_order or folder.name,
    )
    warnings.extend(name_warnings)
    if room_rows and ORDER_FOLDER_RE.fullmatch(order_id):
        room_materials, room_edges, room_warnings = _select_room_materials(
            room_rows,
            order_id,
            names,
            known_order_ids=set(related_order_ids(folder) or [order_id]),
        )
        warnings.extend(room_warnings)
        if room_materials or room_edges:
            selected_room_rows = [
                row for row in room_rows
                if row[0] == "__UNASSIGNED__"
                or (
                    _room_order_ids(row[0]) == [order_id.upper()]
                    or (not _room_order_ids(row[0]) and _room_matches_order(row[0], order_id, names))
                )
            ]
            authoritative = defaultdict(float)
            selected = defaultdict(float)
            for item in materials:
                authoritative[(item.kind, item.thickness, item.color)] += item.quantity
            for item in room_materials:
                selected[(item.kind, item.thickness, item.color)] += item.quantity
            if any(value > authoritative[key] + EPSILON for key, value in selected.items()):
                raise RuleError(
                    "material_room_overflow",
                    f"{order_id} 的房间材料数量超过 material 总量，无法安全分配，请人工检查",
                )
            identified_rooms = {
                room for room, _, _ in selected_room_rows
                if room and room != "__UNASSIGNED__"
            }
            if selected != authoritative and identified_rooms:
                selected_materials = room_materials
                selected_edges = room_edges
                warnings.append(
                    f"material 包含多个房间或订单；当前只采用能匹配 {order_id} 的房间，未匹配的房间数据不纳入本订单汇总"
                )
    fittings = {factory: items for factory, items in fittings.items() if factory in names}
    board_only = sorted(set(names) - set(fittings))
    for factory in board_only:
        fittings[factory] = []
    if board_only:
        warnings.append(
            f"{', '.join(board_only)} 在板材清单中存在但没有有效五金数量，已按无五金工厂单处理"
        )
    if not fittings and not include_hardware:
        fittings = {factory: [] for factory in names}
    if not fittings:
        raise RuleError("missing_factory_orders", "板材清单和五金文件都无法取得工厂单号，不能生成 Traveler")
    mappings = InventoryMappings(config.workflow_database)
    normalized = {
        factory: _normalize_fittings(items, mappings)
        for factory, items in sorted(fittings.items())
    }
    factories = [
        FactoryPreview(factory, names[factory], normalized[factory])
        for factory in sorted(fittings)
    ]
    result = OrderPreview(
        order_id, folder, materials_path, sheet_name, selected_materials, selected_edges, factories, warnings,
        sorted({room for room, _, _ in selected_room_rows if room and room != "__UNASSIGNED__"}),
        include_hardware=include_hardware,
        material_room_rows=selected_room_rows,
    )
    persist_preview(config, result)
    return result


def _traveler_path(config: Config, order_id: str) -> Path:
    return config.order_root / order_id / f"Work Order Traveler({order_id}).xlsx"


def find_existing_traveler(config: Config, order_id: str) -> Path | None:
    expected = _traveler_path(config, order_id)
    return expected if expected.is_file() else None


def preview_payload(config: Config, preview: OrderPreview) -> dict:
    existing = find_existing_traveler(config, preview.order_id)
    return {
        "order_id": preview.order_id,
        "folder": str(preview.folder),
        "materials_file": str(preview.materials_path),
        "materials_sheet": preview.materials_sheet_name,
        "material_rooms": preview.material_rooms,
        "materials": [asdict(item) for item in preview.materials],
        "edge_banding": preview.edge_banding,
        "include_hardware": preview.include_hardware,
        "factories": [
            {
                "factory_order": factory.factory_order,
                "order_name": factory.order_name,
                "fittings": [asdict(item) for item in factory.fittings],
            }
            for factory in preview.factories
        ],
        "warnings": preview.warnings,
        "existing_traveler": str(existing) if existing else "",
    }


def persist_preview(config: Config, preview: OrderPreview) -> None:
    """Persist parsed material/hardware facts; raw reports remain the source evidence."""
    if not config.storage_prepared:
        return
    resolution = resolve_inventory_items(config, _preview_inventory_resolution_items(preview))
    if resolution["missing"]:
        names = "、".join(dict.fromkeys(
            str(item.get("name", "")).strip()
            for item in resolution["missing"]
            if str(item.get("name", "")).strip()
        )) or "材料或五金"
        raise RuleError(
            "order_inventory_mapping_required",
            f"订单存在未完成商品 SKU 处理：{names}。请先设置映射或加入全局忽略清单，再写入数据库。",
            missing_items=resolution["missing"],
        )
    mappings = InventoryMappings(config.workflow_database)
    import sqlite3
    observed = datetime.now().astimezone().isoformat(timespec="seconds")
    connection = sqlite3.connect(config.workflow_database)
    try:
        connection.execute(
            "delete from material_items where order_id=? and source_type in ('aihouse','derived')",
            (preview.order_id.upper(),),
        )
        for item in preview.materials:
            name = _material_inventory_name(item.kind, item.thickness, item.color)
            if mappings.ignored_reason(name) is not None:
                continue
            connection.execute(
                """insert or replace into material_items(
                    order_id,material_type,color,thickness,quantity,unit,edge,
                    source_type,source_path,source_fingerprint,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?)""",
                (preview.order_id.upper(), item.kind, item.color,
                 str(item.thickness), float(item.quantity), "pcs", "", "aihouse",
                 str(preview.materials_path), "", observed),
            )
        for color, quantity in preview.edge_banding.items():
            if mappings.ignored_reason(f"Edge banding--{color}") is not None:
                continue
            connection.execute(
                """insert or replace into material_items(
                    order_id,material_type,color,thickness,quantity,unit,edge,
                    source_type,source_path,source_fingerprint,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?)""",
                (preview.order_id.upper(), "edge", color, "", float(quantity), "m", color,
                 "aihouse", str(preview.materials_path), "", observed),
            )
        connection.execute(
            "delete from hardware_items where order_id=? and source_type='aicnc'",
            (preview.order_id.upper(),),
        )
        for factory in preview.factories:
            for item in factory.fittings:
                if item.ignored or ignored_hardware_reason(mappings, item.name, item.code) is not None:
                    continue
                connection.execute(
                    """insert into hardware_items(
                        order_id,factory_order,scope,product_code,name,spec,quantity,unit,
                        source_type,source_path,remarks,updated_at
                    ) values(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (preview.order_id.upper(), factory.factory_order, "factory_order", item.code,
                     item.name, item.size, float(item.quantity), item.unit, "aicnc",
                     str(preview.folder), "", observed),
                )
        connection.commit()
    finally:
        connection.close()


def _material_inventory_name(kind: str, thickness: float, color: str = "") -> str:
    if kind == "plywood":
        labels = {18.0: "18mm--Plywood", 14.5: "14.5mm--Plywood", 5.4: "5.4mm--Plywood"}
        return next(
            (label for value, label in labels.items() if abs(float(thickness) - value) < 0.01),
            f"{float(thickness):g}mm--Plywood",
        )
    return f"{float(thickness):g}mm--{color}"


def _preview_inventory_resolution_items(
    preview: OrderPreview,
) -> list[tuple[TravelerItem, str]]:
    items: list[tuple[TravelerItem, str]] = []
    row = 1
    for material in preview.materials:
        items.append((
            TravelerItem(
                row=row,
                section="板材与封边",
                name=_material_inventory_name(material.kind, material.thickness, material.color),
                quantity=material.quantity,
                document_remark=preview.order_id,
            ),
            "",
        ))
        row += 1
    for color, quantity in preview.edge_banding.items():
        items.append((
            TravelerItem(row, "板材与封边", f"Edge banding--{color}", quantity, preview.order_id),
            "",
        ))
        row += 1
    for factory in preview.factories:
        for fitting in factory.fittings:
            items.append((
                TravelerItem(row, "五金", fitting.name or fitting.code, fitting.quantity, factory.order_name),
                fitting.code,
            ))
            row += 1
    return items


def preview_related_orders(
    config: Config,
    folder: Path,
    *,
    include_hardware: bool = True,
) -> dict:
    """Preview every order represented in a shared source folder."""
    order_ids = related_order_ids(folder) if ORDER_FOLDER_RE.fullmatch(folder.name) else [folder.name]
    if not order_ids:
        raise RuleError("no_orders_in_folder", f"文件夹中未识别到完整订单号：{folder}")
    previews = []
    errors = []
    for order_id in order_ids:
        try:
            previews.append(
                preview_payload(
                    config,
                    preview_order(config, folder, order_id, include_hardware=include_hardware),
                )
            )
        except RuleError as exc:
            errors.append({"order_id": order_id, "code": exc.code, "message": str(exc), **exc.context})
    merged_materials = [item for order in previews for item in order.get("materials", [])]
    merged_edges: dict[str, float] = defaultdict(float)
    for order in previews:
        for color, quantity in order.get("edge_banding", {}).items():
            merged_edges[color] += float(quantity)
    merged_factories = [factory for order in previews for factory in order.get("factories", [])]
    merged_warnings = [warning for order in previews for warning in order.get("warnings", [])]
    primary = previews[0] if previews else {}
    return {
        **primary,
        "order_id": "、".join(order_ids),
        "folder": str(folder),
        "order_ids": order_ids,
        "orders": previews,
        "materials": merged_materials,
        "edge_banding": dict(merged_edges),
        "factories": merged_factories,
        "warnings": merged_warnings,
        "errors": errors,
    }


def update_related_orders(
    config: Config,
    folder: Path,
    generate: bool = False,
    *,
    include_hardware: bool = True,
) -> dict:
    """Generate/update each complete order found in a shared source folder."""
    group = preview_related_orders(config, folder, include_hardware=include_hardware)
    if group["errors"]:
        details = "；".join(
            f"{error.get('order_id', '未知订单')}：{error.get('message', '未提供具体原因')}"
            for error in group["errors"]
        )
        raise RuleError(
            "related_orders_need_review",
            f"关联订单校验失败：{details}",
            **group,
        )
    saved = []
    for payload in group["orders"]:
        preview = preview_order(
            config,
            folder,
            payload["order_id"],
            include_hardware=include_hardware,
        )
        if generate or not payload.get("existing_traveler"):
            path = generate_order_traveler(config, preview)
            saved.append({"order_id": payload["order_id"], "created": str(path)})
        else:
            path, backup = update_order_traveler(config, preview)
            saved.append({"order_id": payload["order_id"], "updated": str(path), "backup": str(backup)})
    primary = saved[0] if saved else {}
    # Keep the complete related-order payload in the response. The previous
    # response returned only the first alphabetically sorted order, which made
    # the SwiftUI model replace a three-factory preview with one factory after
    # updating Traveler files.
    return {
        **group,
        **primary,
        "orders": group["orders"],
        "factories": group["factories"],
        "updated_orders": saved,
        "order_ids": group["order_ids"],
    }


def set_ignored(config: Config, name: str, ignored: bool) -> None:
    set_ignored_mapping(
        config,
        name,
        ignored,
        "用户在生产文件预览中选择忽略",
    )


def _copy_cell(source, target) -> None:
    if isinstance(source, MergedCell):
        return
    target.value = source.value
    if source.has_style:
        target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.protection = copy.copy(source.protection)


def _copy_worksheet(source, target_wb, title: str, index: int):
    target = target_wb.create_sheet(title, index)
    for row in source.iter_rows():
        for cell in row:
            _copy_cell(cell, target.cell(cell.row, cell.column))
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key] = copy.copy(dimension)
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key] = copy.copy(dimension)
    target.sheet_format = copy.copy(source.sheet_format)
    target.sheet_properties = copy.copy(source.sheet_properties)
    target.page_margins = copy.copy(source.page_margins)
    target.page_setup = copy.copy(source.page_setup)
    target.print_options = copy.copy(source.print_options)
    target.freeze_panes = source.freeze_panes
    return target


def _shift_merges_for_insert(ws, row: int, count: int) -> None:
    affected = [
        tuple(range_boundaries(str(merged)))
        for merged in ws.merged_cells.ranges
        if merged.max_row >= row
    ]
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= row:
            ws.unmerge_cells(str(merged))
    ws.insert_rows(row, count)
    for min_col, min_row, max_col, max_row in affected:
        if min_row >= row:
            min_row += count
            max_row += count
        else:
            max_row += count
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def _expand_usage_detail_rows(ws, required_rows: int) -> None:
    total_row = _find_label_row(ws, "Total Qty:")
    current_rows = total_row - 3
    if required_rows <= current_rows:
        return
    count = required_rows - current_rows
    template_row = total_row - 1
    old_max_row = ws.max_row
    saved_heights = {
        row: ws.row_dimensions[row].height
        for row in range(total_row, old_max_row + 1)
    }
    _shift_merges_for_insert(ws, total_row, count)
    for row, height in saved_heights.items():
        ws.row_dimensions[row + count].height = height
    for destination_row in range(total_row, total_row + count):
        ws.row_dimensions[destination_row].height = ws.row_dimensions[template_row].height
        for col in range(1, ws.max_column + 1):
            source = ws.cell(template_row, col)
            target = ws.cell(destination_row, col)
            if source.has_style:
                target._style = copy.copy(source._style)
            target.number_format = source.number_format
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.protection = copy.copy(source.protection)
            target.value = None
        ws.merge_cells(
            start_row=destination_row,
            start_column=1,
            end_row=destination_row,
            end_column=2,
        )


def _write_usage_formulas(ws) -> None:
    total_row = _find_label_row(ws, "Total Qty:")
    detail_end = total_row - 1
    color_title_row = _find_label_row(ws, "Color Table")
    color_row = _find_label_row(ws, "Color:")
    panel_34_row = _find_label_row(ws, "Sheets (3/4):")
    panel_14_row = _find_label_row(ws, "Sheets (1/4):")
    edge_row = _find_label_row(ws, "Edge Banding (m):")

    for col in range(3, 6):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}3:{letter}{detail_end})"
    colors = []
    seen_colors = set()
    for row in range(3, detail_end + 1):
        color = _canonical_color(_text(ws.cell(row, 9).value))
        key = color.casefold()
        if color and key not in seen_colors:
            colors.append(color)
            seen_colors.add(key)
    if len(colors) > MATERIAL_DETAIL_ROW_CAPACITY:
        raise RuleError(
            "materials_schema",
            f"Usage List 有 {len(colors)} 种颜色，已超过明细行容量 {MATERIAL_DETAIL_ROW_CAPACITY} 行",
        )
    required_color_end = max(9, MATERIAL_COLOR_COLUMN_START - 1 + len(colors))
    for col in range(10, required_color_end + 1):
        _copy_column_style(ws, 9, col)
    for col in range(6, 9):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = (
            f"=SUM({letter}3:{letter}{detail_end})" if len(colors) <= 1 else "check color table"
        )
    ws.cell(total_row, 9).value = colors[0] if len(colors) == 1 else None

    for col in range(MATERIAL_COLOR_COLUMN_START, required_color_end + 1):
        ws.cell(color_row, col).value = None
        ws.cell(panel_34_row, col).value = None
        ws.cell(panel_14_row, col).value = None
        ws.cell(edge_row, col).value = None
    for col, color in enumerate(colors, start=3):
        ws.cell(color_row, col).value = color
    for col in range(3, 3 + len(colors)):
        letter = get_column_letter(col)
        ws.cell(panel_34_row, col).value = (
            f'=SUMIF($I$3:$I${detail_end},{letter}${color_row},$F$3:$F${detail_end})'
        )
        ws.cell(panel_14_row, col).value = (
            f'=SUMIF($I$3:$I${detail_end},{letter}${color_row},$G$3:$G${detail_end})'
        )
        ws.cell(edge_row, col).value = (
            f'=SUMIF($I$3:$I${detail_end},{letter}${color_row},$H$3:$H${detail_end})'
        )
    ws.cell(color_title_row, 1).value = "Color Table"


def _fill_usage_list(source_path: Path, wb, order_id: str, allowed_rooms: set[str] | None = None) -> None:
    if USAGE_LIST_SHEET not in wb.sheetnames:
        raise RuleError("template_schema", f"Traveler 模板缺少 {USAGE_LIST_SHEET}")
    source_wb = load_workbook(source_path, data_only=False, read_only=True)
    if len(source_wb.sheetnames) != 1:
        raise RuleError("materials_schema", f"materials 文件必须只有一个工作簿：{source_wb.sheetnames}")
    source = source_wb[source_wb.sheetnames[0]]
    source_total_row = _find_label_row(source, "Total Qty:")
    detail_rows = source_total_row - 3
    if detail_rows < 1:
        raise RuleError("materials_schema", "materials 文件没有可写入 Usage List 的明细行")

    target = wb[USAGE_LIST_SHEET]
    _expand_usage_detail_rows(target, detail_rows)
    target_total_row = _find_label_row(target, "Total Qty:")
    for row in range(3, target_total_row):
        for col in range(1, 10):
            cell = target.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for source_row in range(3, source_total_row):
        target_row = source_row
        source_room = _text(source.cell(source_row, 1).value)
        if allowed_rooms and source_room and source_room.casefold() not in {room.casefold() for room in allowed_rooms}:
            continue
        for col in range(1, 10):
            target_cell = target.cell(target_row, col)
            if not isinstance(target_cell, MergedCell):
                value = source.cell(source_row, col).value
                # Color Table headers use the canonical business color name.
                # Normalize copied detail colors as well, otherwise a source
                # value such as Khaki is written beside a Penelope FA44
                # header and the generated SUMIF cannot match it.
                target_cell.value = _canonical_color(_text(value)) if col == 9 and value else value
    _write_merged(target, 1, 2, order_id)
    _write_usage_formulas(target)


def _restore_template_picking_list(config: Config, wb) -> None:
    template_wb = load_workbook(config.template, data_only=False)
    if PICKING_LIST_SHEET not in template_wb.sheetnames:
        raise RuleError("template_schema", f"Traveler 模板缺少 {PICKING_LIST_SHEET}")
    index = wb.sheetnames.index(PICKING_LIST_SHEET) if PICKING_LIST_SHEET in wb.sheetnames else min(2, len(wb.sheetnames))
    if PICKING_LIST_SHEET in wb.sheetnames:
        wb.remove(wb[PICKING_LIST_SHEET])
    _copy_worksheet(template_wb[PICKING_LIST_SHEET], wb, PICKING_LIST_SHEET, index)


def _restore_template_usage_list(config: Config, wb) -> None:
    template_wb = load_workbook(config.template, data_only=False)
    if USAGE_LIST_SHEET not in template_wb.sheetnames:
        raise RuleError("template_schema", f"Traveler 模板缺少 {USAGE_LIST_SHEET}")
    index = wb.sheetnames.index(USAGE_LIST_SHEET) if USAGE_LIST_SHEET in wb.sheetnames else min(1, len(wb.sheetnames))
    if USAGE_LIST_SHEET in wb.sheetnames:
        wb.remove(wb[USAGE_LIST_SHEET])
    _copy_worksheet(template_wb[USAGE_LIST_SHEET], wb, USAGE_LIST_SHEET, index)


def _write_merged(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            ws.cell(merged.min_row, merged.min_col).value = value
            return


def _write_initial_traveler_date(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if _normalized_label(cell.value) == _normalized_label("Date/日期："):
                target = ws.cell(cell.row, cell.column + 1)
                if isinstance(target, MergedCell):
                    _write_merged(ws, cell.row, cell.column + 1, date.today())
                    for merged in ws.merged_cells.ranges:
                        if merged.min_row <= cell.row <= merged.max_row and merged.min_col <= cell.column + 1 <= merged.max_col:
                            target = ws.cell(merged.min_row, merged.min_col)
                            break
                else:
                    target.value = date.today()
                target.number_format = "yyyy.m.d"
                return
    raise RuleError("template_schema", "Traveler 模板缺少 Date/日期 字段")


def _snapshot_rows(ws, first: int, last: int) -> dict:
    return {
        "cells": {
            (row, col): copy.copy(ws.cell(row, col))
            for row in range(first, last + 1)
            for col in range(1, ws.max_column + 1)
            if not isinstance(ws.cell(row, col), MergedCell)
        },
        "heights": {row: ws.row_dimensions[row].height for row in range(first, last + 1)},
        "merges": [
            tuple(range_boundaries(str(merged)))
            for merged in ws.merged_cells.ranges
            if merged.min_row >= first and merged.max_row <= last
        ],
        "first": first,
        "last": last,
        "max_col": ws.max_column,
    }


def _paste_snapshot(ws, snapshot: dict, destination_row: int) -> None:
    offset = destination_row - snapshot["first"]
    for (row, col), source in snapshot["cells"].items():
        _copy_cell(source, ws.cell(row + offset, col))
    for row, height in snapshot["heights"].items():
        ws.row_dimensions[row + offset].height = height
    for min_col, min_row, max_col, max_row in snapshot["merges"]:
        ws.merge_cells(
            start_row=min_row + offset, start_column=min_col,
            end_row=max_row + offset, end_column=max_col,
        )


def _style_merged_row(ws, row: int, last_col: int, color: str) -> None:
    merges = [
        str(merged) for merged in ws.merged_cells.ranges
        if merged.min_row == merged.max_row == row
    ]
    for merged in merges:
        ws.unmerge_cells(merged)
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = copy.copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=True,
            italic=cell.font.italic,
            color=cell.font.color,
        )
        cell.alignment = copy.copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal or "center",
            vertical="center",
            wrap_text=cell.alignment.wrap_text,
        )
    for merged in merges:
        ws.merge_cells(merged)


def _style_picking_list_title(ws, last_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 1 <= merged.max_row:
            ws.unmerge_cells(str(merged))
    ws["A1"] = "Picking List领料单"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = copy.copy(ws["A1"].font)
    ws["A1"].font = Font(
        name=ws["A1"].font.name,
        size=max(ws["A1"].font.size or 0, 24),
        bold=True,
        color=ws["A1"].font.color,
    )
    ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 32)


def _clear_picking_list_business_data(ws) -> None:
    """Keep the template intact when there is no Picking List business content."""
    for row in range(1, ws.max_row + 1):
        first = ws.cell(row, 1).value
        if _text(first) == "Name/工厂单名称":
            first_business_column = 2
        elif isinstance(first, (int, float)):
            first_business_column = 2
        else:
            continue
        for column in range(first_business_column, ws.max_column + 1):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _prepare_picking_list(wb, preview: OrderPreview) -> None:
    ws = wb[PICKING_LIST_SHEET]
    if not preview.factories:
        _clear_picking_list_business_data(ws)
        return
    original_merges = [tuple(range_boundaries(str(merged))) for merged in ws.merged_cells.ranges]
    table_last_col = max(
        [max_col for min_col, min_row, max_col, max_row in original_merges if max_row >= 3]
        or [ws.max_column]
    )
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    ws.delete_rows(4, 7)  # Remove the obsolete Panel section; rows now match the approved screenshot.
    for min_col, min_row, max_col, max_row in original_merges:
        if max_row < 3:
            ws.merge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col,
            )
        elif min_row == max_row == 3:
            ws.merge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col,
            )
        elif min_row >= 11:
            ws.merge_cells(
                start_row=min_row - 7, start_column=min_col,
                end_row=max_row - 7, end_column=max_col,
            )
    base = _snapshot_rows(ws, 3, 13)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 3:
            ws.unmerge_cells(str(merged))
    ws.delete_rows(3, max(ws.max_row - 2, 1))

    _style_picking_list_title(ws, table_last_col)
    cursor = 3
    block_colors = ("DDEBF7", "E2F0D9", "FFF2CC", "E4DFEC")
    for factory_index, factory in enumerate(preview.factories):
        if factory_index:
            ws.insert_rows(cursor, 1)
            ws.row_dimensions[cursor].height = 9
            cursor += 1
        included = [item for item in factory.fittings if not item.ignored]
        slots = max(4, len(included))
        block_height = 11 + (slots - 4)
        ws.insert_rows(cursor, block_height)
        _paste_snapshot(ws, base, cursor)
        if slots > 4:
            insertion = cursor + 7
            ws.insert_rows(insertion, slots - 4)
            template_row = cursor + 6
            for row in range(insertion, insertion + slots - 4):
                ws.row_dimensions[row].height = ws.row_dimensions[template_row].height
                for col in range(1, ws.max_column + 1):
                    _copy_cell(ws.cell(template_row, col), ws.cell(row, col))
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        _write_merged(ws, cursor, 4, factory.order_name)
        _style_merged_row(ws, cursor, table_last_col, block_colors[factory_index % len(block_colors)])
        # Find the fitting column header in the pasted template instead of
        # relying on a fixed offset.  Older approved templates keep the
        # ``Fitting配件`` label row; newer compact ones omit it.  In both
        # layouts the first item must be exactly the row after ``No.``.
        header_row = next(
            (
                row
                for row in range(cursor, cursor + block_height)
                if _text(ws.cell(row, 1).value) == "No."
                and _text(ws.cell(row, 3).value) == "Name名字"
            ),
            None,
        )
        if header_row is None:
            raise RuleError("traveler_schema", "Picking List 模板缺少五金配件表头")
        first_item_row = header_row + 1
        for index in range(slots):
            row = first_item_row + index
            item = included[index] if index < len(included) else None
            if not isinstance(ws.cell(row, 1), MergedCell):
                ws.cell(row, 1).value = index + 1
            # The source report's code identifies its own component; it is
            # not an inventory SKU and must never be written into SKU NO.
            if not isinstance(ws.cell(row, 2), MergedCell):
                ws.cell(row, 2).value = None
            _write_merged(ws, row, 3, item.name if item else None)
            _write_merged(ws, row, 5, item.unit if item else None)
            if not isinstance(ws.cell(row, 7), MergedCell):
                ws.cell(row, 7).value = item.quantity if item else None
        cursor += block_height


def _purchase_material_rows(preview: OrderPreview) -> list[tuple[str, str, float]]:
    rows: list[tuple[str, str, float]] = []
    materials = sorted(
        (item for item in preview.materials if item.quantity > EPSILON),
        key=lambda item: (
            {"plywood": 0, "panel": 1, "back": 2}.get(item.kind.casefold(), 9),
            item.color.casefold(),
            item.thickness,
        ),
    )
    for item in materials:
        kind = item.kind.casefold()
        if kind == "plywood":
            name = f"{item.thickness:g}mm--Plywood"
            spec = "2440*1220"
        elif kind in {"panel", "back"}:
            name = f"{item.thickness:g}mm--{item.color}" if item.color else f"{item.thickness:g}mm--"
            spec = "2745*1220"
        else:
            continue
        rows.append((name, spec, float(item.quantity)))
    for color, quantity in sorted(preview.edge_banding.items(), key=lambda item: item[0].casefold()):
        if float(quantity) > EPSILON:
            rows.append((f"Edge banding--{color}" if color else "Edge banding", "M/米", float(quantity)))
    return rows


def _purchase_hardware_rows(preview: OrderPreview) -> list[tuple[str, str, float, str]]:
    rows: list[tuple[str, str, float, str]] = []
    for factory in sorted(preview.factories, key=lambda item: (item.factory_order.casefold(), item.order_name.casefold())):
        for item in sorted(factory.fittings, key=lambda fitting: (fitting.name.casefold(), fitting.code.casefold())):
            if item.ignored or item.quantity <= EPSILON:
                continue
            rows.append((
                item.name or item.code,
                item.unit,
                float(item.quantity),
                factory.order_name or factory.factory_order,
            ))
    return rows


def _insert_purchase_rows(ws, insertion_row: int, count: int, template_row: int) -> None:
    if count <= 0:
        return
    _shift_merges_for_insert(ws, insertion_row, count)
    for offset in range(count):
        row = insertion_row + offset
        for column in range(1, ws.max_column + 1):
            _copy_cell(ws.cell(template_row, column), ws.cell(row, column))
        for start, end in ((3, 4), (5, 6), (7, 8)):
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)


def _prepare_purchase_list(wb, preview: OrderPreview) -> None:
    """Populate the optional Purchase List sheet from the same preview facts.

    The legacy template contains example materials and accessories.  Leaving
    those cells untouched makes a database-generated Traveler look like it has
    hardware that was never present in SQLite, so this sheet is cleared and
    rebuilt whenever the template provides it.
    """
    if PURCHASE_LIST_SHEET not in wb.sheetnames:
        return
    ws = wb[PURCHASE_LIST_SHEET]
    material_rows = _purchase_material_rows(preview)
    hardware_rows = _purchase_hardware_rows(preview)

    ws["A4"] = None
    factory_names = []
    for factory in preview.factories:
        name = _text(factory.order_name) or _text(factory.factory_order)
        if name and name not in factory_names:
            factory_names.append(name)
    _write_merged(ws, 5, 4, "；".join(factory_names))

    panel_header = next((row for row in range(1, ws.max_row + 1) if _text(ws.cell(row, 1).value) == "Panel板材"), None)
    material_header = next((row for row in range((panel_header or 1) + 1, ws.max_row + 1) if _text(ws.cell(row, 1).value) == "No." and _text(ws.cell(row, 3).value) == "Name名字"), None)
    accessories_header = next((row for row in range((material_header or 1) + 1, ws.max_row + 1) if _text(ws.cell(row, 1).value) == "Accessories配件"), None)
    accessory_header = next((row for row in range((accessories_header or 1) + 1, ws.max_row + 1) if _text(ws.cell(row, 1).value) == "No." and _text(ws.cell(row, 3).value) == "Name名字"), None)
    if not material_header or not accessories_header or not accessory_header:
        raise RuleError("template_schema", "Purchase List 模板缺少材料或配件表头")

    material_first = material_header + 1
    material_capacity = accessories_header - material_first
    desired_material_rows = max(2, len(material_rows))
    if desired_material_rows > material_capacity:
        _insert_purchase_rows(ws, accessories_header, desired_material_rows - material_capacity, accessories_header - 1)
        accessories_header += desired_material_rows - material_capacity
        accessory_header += desired_material_rows - material_capacity

    for row in range(material_first, accessories_header):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for index, (name, spec, quantity) in enumerate(material_rows, start=1):
        row = material_first + index - 1
        ws.cell(row, 1).value = index
        _write_merged(ws, row, 3, name)
        _write_merged(ws, row, 5, spec)
        _write_merged(ws, row, 7, quantity)

    accessory_first = accessory_header + 1
    accessory_capacity = ws.max_row - accessory_first + 1
    if len(hardware_rows) > accessory_capacity:
        _insert_purchase_rows(ws, ws.max_row + 1, len(hardware_rows) - accessory_capacity, ws.max_row)

    for row in range(accessory_first, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for index, (name, unit, quantity, factory_name) in enumerate(hardware_rows, start=1):
        row = accessory_first + index - 1
        ws.cell(row, 1).value = index
        _write_merged(ws, row, 3, name)
        _write_merged(ws, row, 5, unit)
        _write_merged(ws, row, 7, quantity)
        ws.cell(row, 9).value = factory_name


def _manual_hardware(ws) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = defaultdict(list)
    factory = ""
    in_manual_section = False
    name_col = spec_col = qty_col = remarks_col = None
    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if _text(values[0]) == "Name/工厂单名称":
            factory = next((_text(value) for value in values[1:] if _text(value)), "")
            in_manual_section = False
            continue
        if _text(values[0]) == "Hardware Accessory五金功能件":
            in_manual_section = True
            name_col = spec_col = qty_col = remarks_col = None
            continue
        if not in_manual_section:
            continue
        labels = {_text(value): index for index, value in enumerate(values, 1) if _text(value)}
        if "Name名字" in labels and "QTY数量" in labels:
            name_col = labels["Name名字"]
            spec_col = labels.get("Spec规格")
            qty_col = labels["QTY数量"]
            remarks_col = labels.get("Remarks备注")
            continue
        if not isinstance(values[0], (int, float)) or not name_col or not qty_col:
            continue
        name = _text(ws.cell(row, name_col).value)
        if not name:
            continue
        result[factory].append({
            "sku": ws.cell(row, 2).value,
            "name": name,
            "spec": ws.cell(row, spec_col).value if spec_col else None,
            "quantity": ws.cell(row, qty_col).value,
            "remarks": ws.cell(row, remarks_col).value if remarks_col else None,
        })
    return dict(result)


def _manual_hardware_block(ws, factory: str) -> tuple[int, list[int], int] | None:
    current_factory = ""
    in_target = False
    header = None
    slots: list[int] = []
    boundary = ws.max_row + 1
    for row in range(1, ws.max_row + 1):
        first = _text(ws.cell(row, 1).value)
        if first == "Name/工厂单名称":
            if in_target and header is not None:
                boundary = row
                break
            current_factory = next(
                (_text(ws.cell(row, col).value) for col in range(2, ws.max_column + 1) if _text(ws.cell(row, col).value)),
                "",
            )
            in_target = current_factory.casefold() == factory.casefold()
            continue
        if in_target and first == "Hardware Accessory五金功能件":
            header = row + 1
            continue
        if in_target and header is not None and row > header and isinstance(ws.cell(row, 1).value, (int, float)):
            slots.append(row)
    if not in_target and header is None:
        return None
    return header, slots, boundary


def _restore_manual_hardware(ws, hardware: dict[str, list[dict]]) -> None:
    for factory, items in hardware.items():
        unique_items = {}
        for item in items:
            unique_items[(_text(item.get("sku")), _text(item.get("name")))] = item
        items = list(unique_items.values())
        block = _manual_hardware_block(ws, factory)
        if block is None:
            # Legacy sheets can lose the exact factory label while being
            # upgraded. Preserve the manual rows in the first available
            # factory block instead of aborting the whole Traveler update.
            for row in range(1, ws.max_row + 1):
                if _text(ws.cell(row, 1).value) == "Name/工厂单名称":
                    candidate = next((_text(ws.cell(row, col).value) for col in range(2, ws.max_column + 1) if _text(ws.cell(row, col).value)), "")
                    if candidate:
                        block = _manual_hardware_block(ws, candidate)
                        if block:
                            break
            if block is None:
                for item in items:
                    for old_row in reversed([row for row in range(1, ws.max_row + 1) if _text(ws.cell(row, 3).value) == _text(item["name"])]):
                        ws.delete_rows(old_row, 1)
                    row = ws.max_row + 1
                    ws.cell(row, 1).value = 1
                    ws.cell(row, 2).value = item["sku"]
                    ws.cell(row, 3).value = item["name"]
                    ws.cell(row, 5).value = item["spec"]
                    ws.cell(row, 7).value = item["quantity"]
                    ws.cell(row, 9).value = item["remarks"]
                continue
        _, slots, boundary = block
        if block[0] is None:
            for item in items:
                for old_row in reversed([row for row in range(1, ws.max_row + 1) if _text(ws.cell(row, 3).value) == _text(item["name"])]):
                    ws.delete_rows(old_row, 1)
                row = ws.max_row + 1
                ws.cell(row, 1).value = 1
                ws.cell(row, 2).value = item["sku"]
                ws.cell(row, 3).value = item["name"]
                ws.cell(row, 5).value = item["spec"]
                ws.cell(row, 7).value = item["quantity"]
                ws.cell(row, 9).value = item["remarks"]
            continue
        if not slots:
            slots = [block[0] + 1]
        if len(items) > len(slots):
            count = len(items) - len(slots)
            source_row = slots[-1]
            _shift_merges_for_insert(ws, boundary, count)
            for row in range(boundary, boundary + count):
                ws.row_dimensions[row].height = ws.row_dimensions[source_row].height
                for col in range(1, ws.max_column + 1):
                    _copy_cell(ws.cell(source_row, col), ws.cell(row, col))
                    ws.cell(row, col).value = None
                for start, end in ((3, 4), (5, 6), (7, 8)):
                    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
                slots.append(row)
        for index, item in enumerate(items):
            row = slots[index]
            ws.cell(row, 1).value = index + 1
            ws.cell(row, 2).value = item["sku"]
            _write_merged(ws, row, 3, item["name"])
            _write_merged(ws, row, 5, item["spec"])
            _write_merged(ws, row, 7, item["quantity"])
            ws.cell(row, 9).value = item["remarks"]
        # Collapse duplicate legacy rows for the same manually entered item.
        for item in items:
            matches = [row for row in range(1, ws.max_row + 1) if _text(ws.cell(row, 3).value) == _text(item["name"])]
            for row in reversed(matches[1:]):
                ws.delete_rows(row, 1)


def _positive_hardware_quantity(value) -> int:
    number = _number(value, "人工五金")
    if not math.isfinite(number) or number <= 0 or abs(number - round(number)) > EPSILON:
        raise RuleError("manual_hardware_quantity", f"人工五金数量必须是正整数：{number:g}")
    return int(round(number))


def _manual_hardware_target(ws, factory_name: str) -> str:
    requested = _text(factory_name)
    matches = []
    for row in range(1, ws.max_row + 1):
        if _text(ws.cell(row, 1).value) != "Name/工厂单名称":
            continue
        current = next(
            (_text(ws.cell(row, col).value) for col in range(2, ws.max_column + 1) if _text(ws.cell(row, col).value)),
            "",
        )
        if current.casefold() == requested.casefold():
            matches.append(current)
    if len(matches) != 1:
        raise RuleError(
            "manual_hardware_factory_missing",
            f"Traveler 中找不到唯一工厂单名称：{factory_name}",
        )
    return matches[0]


def preview_manual_hardware(
    config: Config,
    order_id: str,
    factory_name: str,
    product_code: str,
    quantity,
    remarks: str = "",
) -> dict:
    traveler = find_existing_traveler(config, order_id)
    with ProductDatabase(bootstrap_product_database(config)) as catalog:
        product = catalog.require_code(product_code.strip().upper())
    factory = factory_name.strip()
    if traveler is not None:
        workbook = load_workbook(traveler, data_only=False, read_only=False)
        required = {WORK_ORDER_SHEET, USAGE_LIST_SHEET, PICKING_LIST_SHEET}
        if not required.issubset(workbook.sheetnames):
            raise RuleError("traveler_schema", f"Traveler 缺少必要工作表：{sorted(required - set(workbook.sheetnames))}")
        factory = _manual_hardware_target(workbook[PICKING_LIST_SHEET], factory_name)
    else:
        import sqlite3
        connection = sqlite3.connect(config.workflow_database)
        try:
            row = connection.execute(
                "select factory_name from factory_orders where order_id=? and (factory_name=? or factory_order=?)",
                (order_id.upper(), factory_name, factory_name),
            ).fetchone()
        finally:
            connection.close()
        if row is not None:
            factory = row[0]
        elif not factory:
            raise RuleError("factory_order_missing", f"数据库中找不到工厂单：{factory_name}")
    return {
        "traveler": str(traveler) if traveler else "",
        "order_id": order_id.upper(),
        "factory_name": factory,
        "product_code": product.code,
        "product_name": product.name,
        "spec": product.spec,
        "quantity": _positive_hardware_quantity(quantity),
        "remarks": _text(remarks) or "人工添加",
    }


def _backup_traveler(config: Config, traveler: Path, order_id: str, label: str = "backup") -> Path:
    backup_dir = config.backup_root / order_id.upper()
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d %H%M%S-%f")
    backup = backup_dir / f"{traveler.stem} {label} {stamp}{traveler.suffix}"
    shutil.copy2(traveler, backup)
    return backup


def add_manual_hardware(
    config: Config,
    order_id: str,
    factory_name: str,
    product_code: str,
    quantity,
    remarks: str = "",
) -> tuple[Path, Path, dict]:
    preview = preview_manual_hardware(
        config, order_id, factory_name, product_code, quantity, remarks
    )
    traveler = Path(preview["traveler"])
    if config.storage_prepared:
        import sqlite3
        observed = datetime.now().astimezone().isoformat(timespec="seconds")
        connection = sqlite3.connect(config.workflow_database)
        try:
            mappings = InventoryMappings(config.workflow_database)
            if ignored_hardware_reason(
                mappings,
                preview["product_name"],
                preview["product_code"],
            ) is not None:
                raise RuleError(
                    "hardware_ignored",
                    f"五金已被设置为忽略，不能写入数据库：{preview['product_name']}",
                )
            connection.execute(
                """insert into hardware_items(
                    order_id,factory_order,scope,product_code,name,spec,quantity,unit,
                    source_type,source_path,remarks,updated_at
                ) values(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (order_id.upper(), preview["factory_name"], "factory_order", preview["product_code"],
                 preview["product_name"], preview["spec"], float(preview["quantity"]), "pcs/个",
                 "manual", "", preview["remarks"], observed),
            )
            connection.commit()
        finally:
            connection.close()
    if not traveler.is_file():
        return Path(""), Path(""), {**preview, "result": "added", "saved_quantity": preview["quantity"], "stored_in_database": True}
    with tempfile.TemporaryDirectory(prefix="manual-hardware-") as temporary:
        draft = Path(temporary) / traveler.name
        shutil.copy2(traveler, draft)
        workbook = load_workbook(draft, data_only=False)
        sheet = workbook[PICKING_LIST_SHEET]
        existing = _manual_hardware(sheet)
        items = existing.setdefault(preview["factory_name"], [])
        matching = next(
            (
                item for item in items
                if _text(item["sku"]).casefold() == preview["product_code"].casefold()
                and _text(item["spec"]).casefold() == preview["spec"].casefold()
            ),
            None,
        )
        if matching is None:
            items.append({
                "sku": preview["product_code"],
                "name": preview["product_name"],
                "spec": preview["spec"],
                "quantity": preview["quantity"],
                "remarks": preview["remarks"],
            })
            result_kind = "added"
        else:
            matching["quantity"] = _positive_hardware_quantity(matching["quantity"]) + preview["quantity"]
            if preview["remarks"] not in {_text(matching["remarks"]), "人工添加"}:
                current = _text(matching["remarks"])
                matching["remarks"] = "；".join(filter(None, (current, preview["remarks"])))
            result_kind = "increased"
        _restore_manual_hardware(sheet, existing)
        workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        workbook.save(draft)
        check = load_workbook(draft, data_only=False, read_only=False)
        saved = _manual_hardware(check[PICKING_LIST_SHEET]).get(preview["factory_name"], [])
        verified = next(
            (
                item for item in saved
                if _text(item["sku"]).casefold() == preview["product_code"].casefold()
                and _text(item["spec"]).casefold() == preview["spec"].casefold()
            ),
            None,
        )
        if verified is None:
            raise RuleError("write_verification", "人工五金写入后重新打开未找到目标记录")
        backup = _backup_traveler(config, traveler, order_id, "manual-hardware-backup")
        os.replace(draft, traveler)
    result = {**preview, "result": result_kind, "saved_quantity": verified["quantity"]}
    return traveler, backup, result


def generate_order_traveler(config: Config, preview: OrderPreview) -> Path:
    destination = _traveler_path(config, preview.order_id)
    destination_dir = destination.parent
    if destination.exists():
        raise RuleError("destination_exists", f"目标 Traveler 已存在，拒绝覆盖：{destination}")
    destination_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="order-traveler-") as temporary:
        draft = Path(temporary) / destination.name
        shutil.copy2(config.template, draft)
        wb = load_workbook(draft)
        required = {WORK_ORDER_SHEET, USAGE_LIST_SHEET, PICKING_LIST_SHEET}
        if not required.issubset(wb.sheetnames):
            raise RuleError("template_schema", f"Traveler 模板缺少必要工作表：{sorted(required - set(wb.sheetnames))}")
        wb[WORK_ORDER_SHEET]["B5"] = preview.order_id
        _write_initial_traveler_date(wb[WORK_ORDER_SHEET])
        _fill_usage_list(preview.materials_path, wb, preview.order_id, set(preview.material_rooms))
        _prepare_picking_list(wb, preview)
        _prepare_purchase_list(wb, preview)
        wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        wb.save(draft)
        check = load_workbook(draft, data_only=False, read_only=True)
        if check.sheetnames[:3] != [WORK_ORDER_SHEET, USAGE_LIST_SHEET, PICKING_LIST_SHEET]:
            raise RuleError("write_verification", "Traveler 工作表名称或顺序不正确")
        os.replace(draft, destination)
    return destination


def _database_material_template(config: Config) -> Path:
    configured = config.template.parent / "Order Materials.xlsx"
    if configured.is_file():
        return configured
    bundled = Path(__file__).resolve().parent.parent / "resources/templates/Order Materials.xlsx"
    if bundled.is_file():
        return bundled
    raise RuleError("template_missing", "找不到数据库 Traveler 所需的材料模板")


def _write_database_material_workbook(
    config: Config,
    destination: Path,
    order_id: str,
    material_rows: list[dict],
) -> None:
    """Create an internal material workbook solely to render DB facts into the Traveler template."""
    shutil.copy2(_database_material_template(config), destination)
    workbook = load_workbook(destination, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet["B1"] = order_id

    for row in range(3, 14):
        for column in range(1, 10):
            cell = worksheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None

    grouped: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in material_rows:
        kind = _text(row.get("material_type")).casefold()
        color = _canonical_color(_text(row.get("color")))
        thickness = _number(row.get("thickness"), "数据库材料厚度")
        quantity = _number(row.get("quantity"), "数据库材料数量")
        if quantity == 0:
            continue
        if kind == "plywood":
            key = ("", "plywood")
            column = {18.0: "plywood_18", 14.5: "plywood_145", 5.4: "plywood_54"}.get(round(thickness, 1))
            if column is None:
                raise RuleError("unsupported_material", f"数据库中的 Plywood 厚度暂不支持：{thickness:g}mm")
        elif kind == "panel":
            key = (color, "panel")
            column = "panel_34" if math.isclose(thickness, 19.1, abs_tol=EPSILON) else "panel_14" if thickness <= 9.0 else None
            if column is None:
                raise RuleError("unsupported_material", f"数据库中的 Panel 厚度暂不支持：{thickness:g}mm")
        elif kind == "edge":
            key = (color, "edge")
            column = "edge"
        else:
            raise RuleError("unsupported_material", f"数据库中的材料类型暂不支持：{kind or '空白'}")
        grouped[key][column] += quantity

    plywood = grouped.pop(("", "plywood"), {})
    colors = sorted({color for color, _ in grouped if color}, key=str.casefold)
    rows: list[tuple[str, dict[str, float]]] = []
    if not colors:
        rows.append(("数据库汇总", dict(plywood)))
    else:
        for index, color in enumerate(colors):
            values = dict(grouped.get((color, "panel"), {}))
            values.update(grouped.get((color, "edge"), {}))
            if index == 0:
                values.update(plywood)
            rows.append(("数据库汇总" if index == 0 else "", values | {"color": color}))
    if len(rows) > MATERIAL_DETAIL_ROW_CAPACITY:
        raise RuleError("material_capacity", f"数据库材料颜色数量超过 Traveler 明细行容量：{len(rows)}")

    for row_index, (room, values) in enumerate(rows, start=3):
        worksheet.cell(row_index, 1).value = room
        worksheet.cell(row_index, 3).value = values.get("plywood_18", 0) or None
        worksheet.cell(row_index, 4).value = values.get("plywood_145", 0) or None
        worksheet.cell(row_index, 5).value = values.get("plywood_54", 0) or None
        worksheet.cell(row_index, 6).value = values.get("panel_34", 0) or None
        worksheet.cell(row_index, 7).value = values.get("panel_14", 0) or None
        worksheet.cell(row_index, 8).value = values.get("edge", 0) or None
        worksheet.cell(row_index, 9).value = values.get("color", "")

    for column in range(3, worksheet.max_column + 1):
        worksheet.cell(16, column).value = None
    for column, color in enumerate(colors, start=MATERIAL_COLOR_COLUMN_START):
        if column > worksheet.max_column:
            _copy_column_style(worksheet, worksheet.max_column, column)
        letter = get_column_letter(column)
        worksheet.cell(16, column).value = color
        worksheet.cell(17, column).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$F$3:$F$13))'
        worksheet.cell(18, column).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$G$3:$G$13))'
        worksheet.cell(19, column).value = f'=IF({letter}$16="","",SUMIF($I$3:$I$13,{letter}$16,$H$3:$H$13))'
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    workbook.save(destination)


def generate_database_order_traveler(config: Config, order_id: str) -> Path:
    """Render an on-demand Traveler from persisted SQLite facts, never from source files."""
    from .order_details import order_detail

    normalized_order_id = order_id.strip().upper()
    if not normalized_order_id:
        raise RuleError("invalid_arguments", "生成 Traveler 需要订单号")
    detail = order_detail(config, normalized_order_id)
    material_rows = detail["materials"]
    if not material_rows:
        raise RuleError("order_materials_missing", f"数据库中没有订单 {normalized_order_id} 的材料明细")

    order_scoped = [
        row for row in material_rows
        if _text(row.get("scope")).casefold() == "order" and not _text(row.get("factory_order"))
    ]
    selected_material_rows = order_scoped or material_rows
    material_items: list[MaterialItem] = []
    edge_banding: dict[str, float] = defaultdict(float)
    for row in selected_material_rows:
        kind = _text(row.get("material_type")).casefold()
        thickness = _number(row.get("thickness"), "数据库材料厚度")
        color = _canonical_color(_text(row.get("color")))
        quantity = _number(row.get("quantity"), "数据库材料数量")
        if kind == "edge":
            edge_banding[color] += quantity
        else:
            material_items.append(MaterialItem(kind, thickness, color, quantity))

    hardware_by_factory: dict[str, list[PreviewFitting]] = defaultdict(list)
    for row in detail["hardware"]:
        factory_order = _text(row.get("factory_order")) or "未分配工厂单"
        hardware_by_factory[factory_order].append(PreviewFitting(
            key=f"{factory_order}:{row.get('product_code', '')}:{row.get('name', '')}:{len(hardware_by_factory[factory_order])}",
            name=_text(row.get("name")),
            code=_text(row.get("product_code")),
            size=_text(row.get("spec")),
            unit=_text(row.get("unit")),
            quantity=_number(row.get("quantity"), "数据库五金数量"),
            ignored=False,
        ))
    factory_rows = { _text(row.get("factory_order")): row for row in detail["factory_orders"] if _text(row.get("factory_order")) }
    orphan_hardware = sorted(set(hardware_by_factory) - set(factory_rows))
    if orphan_hardware:
        raise RuleError(
            "factory_hardware_mismatch",
            f"数据库五金归属了当前订单之外的工厂单：{'、'.join(orphan_hardware)}；请先重新获取 AIMES 并扫描订单数据",
            order_id=normalized_order_id,
            factory_orders=orphan_hardware,
        )
    factories = [
        FactoryPreview(
            factory_order=factory_order,
            order_name=_text(row.get("factory_name")) or factory_order,
            fittings=hardware_by_factory.get(factory_order, []),
        )
        for factory_order, row in sorted(factory_rows.items())
    ]

    temporary_root = Path(tempfile.mkdtemp(prefix="pp-flowhub-db-traveler-"))
    material_workbook = temporary_root / f"{normalized_order_id} materials.xlsx"
    try:
        _write_database_material_workbook(config, material_workbook, normalized_order_id, selected_material_rows)
        preview = OrderPreview(
            normalized_order_id,
            Path(_text(detail["order"].get("source_folder")) or normalized_order_id),
            material_workbook,
            "Sheet1",
            material_items,
            dict(edge_banding),
            factories,
            ["材料与五金来自中央 SQLite"],
            include_hardware=True,
        )
        transient_config = copy.copy(config)
        transient_config.order_root = temporary_root
        traveler = generate_order_traveler(transient_config, preview)
        material_workbook.unlink(missing_ok=True)
        return traveler
    except Exception:
        shutil.rmtree(temporary_root, ignore_errors=True)
        raise


def generate_temporary_traveler(config: Config, preview: OrderPreview) -> Path:
    """Render a Traveler into the OS temporary area; never place it in Order."""
    temporary_root = Path(tempfile.mkdtemp(prefix="pp-flowhub-traveler-"))
    transient_config = copy.copy(config)
    transient_config.order_root = temporary_root
    return generate_order_traveler(transient_config, preview)


def update_order_traveler(config: Config, preview: OrderPreview) -> tuple[Path, Path]:
    existing = find_existing_traveler(config, preview.order_id)
    if not existing:
        raise RuleError("traveler_not_found", f"找不到可更新的 Traveler：{_traveler_path(config, preview.order_id)}")
    backup = _backup_traveler(config, existing, preview.order_id)
    with tempfile.TemporaryDirectory(prefix="order-traveler-update-") as temporary:
        draft = Path(temporary) / existing.name
        shutil.copy2(existing, draft)
        wb = load_workbook(draft)
        if WORK_ORDER_SHEET not in wb.sheetnames:
            raise RuleError("traveler_schema", f"现有 Traveler 缺少核心工作表：{WORK_ORDER_SHEET}")
        manual_hardware = {}
        if PICKING_LIST_SHEET in wb.sheetnames:
            manual_hardware = _manual_hardware(wb[PICKING_LIST_SHEET])
        elif "Pickinglist" in wb.sheetnames:
            manual_hardware = _manual_hardware(wb["Pickinglist"])
            wb.remove(wb["Pickinglist"])
        wb[WORK_ORDER_SHEET]["B5"] = preview.order_id
        if USAGE_LIST_SHEET not in wb.sheetnames:
            _restore_template_usage_list(config, wb)
        if PICKING_LIST_SHEET not in wb.sheetnames:
            _restore_template_picking_list(config, wb)
        _fill_usage_list(preview.materials_path, wb, preview.order_id, set(preview.material_rooms))
        _prepare_picking_list(wb, preview)
        _prepare_purchase_list(wb, preview)
        if preview.include_hardware:
            _restore_manual_hardware(wb[PICKING_LIST_SHEET], manual_hardware)
        wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        wb.save(draft)
        check = load_workbook(draft, data_only=False, read_only=True)
        if check.sheetnames[:3] != [WORK_ORDER_SHEET, USAGE_LIST_SHEET, PICKING_LIST_SHEET]:
            raise RuleError("write_verification", "更新后 Traveler 工作表名称或顺序不正确")
        os.replace(draft, existing)
    return existing, backup


def _config_from_args(args) -> Config:
    config = Config()
    if configured_state := os.environ.get("PP_FLOWHUB_STATE_DIR", "").strip():
        config.state_dir = Path(configured_state).expanduser()
    config.load_settings()
    for name in ("source_root", "order_root", "template", "backup_root", "state_dir"):
        value = getattr(args, name, None)
        if value:
            setattr(config, name, value)
    config.prepare_storage()
    return config


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="pp-flowhub order")
    parser.add_argument("command", choices=("list", "list-index", "detail", "cost", "cost-export", "backup-status", "backup-now", "sync-index", "process-server-changes", "process-server-folder", "preview-server-changes", "confirm-server-preview", "confirm-server-material-preview", "sync-aimes", "scan-server", "ignore-server-folder", "ignore-aimes", "restore-aimes-ignore", "assign-aimes-order", "restore-aimes-assignment", "auto-resolve-issue", "resolve-issue", "save-order-annotations", "preview", "preview-related", "refresh-aimes", "stock-check", "set-ignore", "generate", "generate-db", "temporary", "generate-material", "generate-material-from-travelers", "update", "update-related", "add-hardware", "add-factory", "assign-material", "create-test-data"))
    parser.add_argument("--folder", type=Path)
    parser.add_argument("--server-folder", type=Path, action="append", default=[])
    parser.add_argument("--name", action="append", default=[])
    parser.add_argument("--ignored", choices=("true", "false"))
    parser.add_argument("--source-root", type=Path)
    parser.add_argument("--order-root", type=Path)
    parser.add_argument("--template", type=Path)
    parser.add_argument("--backup-root", type=Path)
    parser.add_argument("--state-dir", type=Path)
    parser.add_argument("--target-root", type=Path)
    parser.add_argument("--order-id", default="")
    parser.add_argument("--note", default="")
    parser.add_argument("--planned-installation-days", default="[]")
    parser.add_argument("--actual-installation-days", default="[]")
    parser.add_argument("--factory-name", default="")
    parser.add_argument("--factory-order", default="")
    parser.add_argument("--product-code", default="")
    parser.add_argument("--quantity")
    parser.add_argument("--remarks", default="")
    parser.add_argument("--materials-file")
    parser.add_argument("--confirm-write", action="store_true")
    parser.add_argument("--include-hardware", choices=("true", "false"), default="true")
    parser.add_argument("--process-temporary", action="store_true")
    parser.add_argument("--refresh-aimes", action="store_true")
    parser.add_argument("--aimes-if-needed", action="store_true")
    parser.add_argument("--full-refresh", action="store_true")
    parser.add_argument("--server-snapshot", type=Path)
    parser.add_argument("--preview-token", default="")
    parser.add_argument("--material-id", type=int)
    parser.add_argument("--ignore-key", action="append", default=[])
    parser.add_argument("--issue-key", default="")
    args = parser.parse_args(argv)
    config = _config_from_args(args)
    logger = configure_operation_log(config)
    logger.event("backend.command.started", "开始订单工作流操作", details={"action": args.command})
    try:
        if args.command == "create-test-data":
            if not args.target_root:
                raise RuleError("invalid_arguments", "create-test-data 需要 --target-root")
            from .test_data import create_local_test_source
            result = create_local_test_source(args.target_root)
        elif args.command == "list":
            result = {"orders": list_order_folders(config)}
        elif args.command == "list-index":
            from .order_index import list_order_index
            result = list_order_index(config)
        elif args.command == "detail":
            if not args.order_id:
                raise RuleError("invalid_arguments", "detail 需要 --order-id")
            from .order_details import order_detail
            result = order_detail(config, args.order_id)
        elif args.command in {"cost", "cost-export"}:
            if not args.order_id:
                raise RuleError("invalid_arguments", f"{args.command} 需要 --order-id")
            from .costing import calculate_order_cost, export_order_cost
            result = export_order_cost(config, args.order_id) if args.command == "cost-export" else calculate_order_cost(config, args.order_id)
        elif args.command == "backup-status":
            from .backup import backup_status
            result = backup_status(config)
        elif args.command == "backup-now":
            from .backup import perform_backup
            result = perform_backup(config)
        elif args.command == "sync-index":
            from .order_index import sync_order_index
            result = sync_order_index(
                config,
                refresh_aimes=args.refresh_aimes,
                aimes_if_needed=args.aimes_if_needed,
                full_refresh=args.full_refresh,
                server_snapshot_path=args.server_snapshot,
            )
        elif args.command == "process-server-changes":
            from .order_index import process_server_changes
            result = process_server_changes(
                config,
                args.server_folder or None,
                include_hardware=args.include_hardware == "true",
            )
        elif args.command == "process-server-folder":
            if not args.folder:
                raise RuleError("invalid_arguments", "process-server-folder 需要 --folder")
            from .order_index import process_server_folder
            try:
                result = process_server_folder(
                    config,
                    args.folder,
                    include_hardware=args.include_hardware == "true",
                    process_temporary=args.process_temporary,
                )
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "preview-server-changes":
            if not args.server_folder:
                raise RuleError("invalid_arguments", "preview-server-changes 需要 --server-folder")
            from .order_index import preview_server_changes
            try:
                result = preview_server_changes(
                    config,
                    args.server_folder,
                    include_hardware=args.include_hardware == "true",
                )
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "confirm-server-preview":
            if not args.preview_token or not args.order_id or not args.factory_order:
                raise RuleError("invalid_arguments", "confirm-server-preview 需要预览标识、订单号和工厂单号")
            from .order_index import confirm_server_preview
            try:
                result = confirm_server_preview(
                    config,
                    args.preview_token,
                    args.order_id,
                    args.factory_order,
                    confirm_write=args.confirm_write,
                )
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "confirm-server-material-preview":
            if not args.preview_token:
                raise RuleError("invalid_arguments", "confirm-server-material-preview 需要预览标识")
            from .order_index import confirm_server_material_preview
            try:
                result = confirm_server_material_preview(
                    config,
                    args.preview_token,
                    confirm_write=args.confirm_write,
                )
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "sync-aimes":
            from .order_index import sync_aimes_index
            result = sync_aimes_index(
                config,
                force=args.refresh_aimes,
                if_needed=args.aimes_if_needed,
            )
        elif args.command == "scan-server":
            from .order_index import scan_server_changes
            result = scan_server_changes(config)
        elif args.command == "ignore-server-folder":
            if not args.folder:
                raise RuleError("invalid_arguments", "ignore-server-folder 需要 --folder")
            from .order_index import ignore_server_folder
            try:
                result = ignore_server_folder(config, args.folder)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "ignore-aimes":
            if not args.ignore_key:
                raise RuleError("invalid_arguments", "请先选择需要忽略的 AIMES 工厂单")
            from .order_index import ignore_aimes_factories
            try:
                result = ignore_aimes_factories(config, args.ignore_key)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "restore-aimes-ignore":
            if not args.ignore_key:
                raise RuleError("invalid_arguments", "请先选择需要恢复的 AIMES 工厂单")
            from .order_index import restore_aimes_factories
            result = restore_aimes_factories(config, args.ignore_key)
        elif args.command == "assign-aimes-order":
            if len(args.ignore_key) != 1 or not args.order_id:
                raise RuleError("invalid_arguments", "确认 AIMES 归属需要一条工厂单记录和建议订单号")
            from .order_index import assign_aimes_factory_order
            try:
                result = assign_aimes_factory_order(config, args.ignore_key[0], args.order_id)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "restore-aimes-assignment":
            if len(args.ignore_key) != 1:
                raise RuleError("invalid_arguments", "撤销 AIMES 建议归属需要一条工厂单记录")
            from .order_index import restore_aimes_order_assignment
            try:
                result = restore_aimes_order_assignment(config, args.ignore_key[0])
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "auto-resolve-issue":
            if not args.issue_key:
                raise RuleError("invalid_arguments", "自动处理当前问题需要 --issue-key")
            from .order_index import auto_resolve_current_issue
            try:
                result = auto_resolve_current_issue(config, args.issue_key)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "resolve-issue":
            if not args.issue_key:
                raise RuleError("invalid_arguments", "确认当前问题需要 --issue-key")
            from .order_index import resolve_current_issue
            try:
                result = resolve_current_issue(config, args.issue_key, args.order_id, args.factory_name)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "save-order-annotations":
            if not args.order_id:
                raise RuleError("invalid_arguments", "save-order-annotations 需要 --order-id")
            try:
                planned_days = json.loads(args.planned_installation_days)
                actual_days = json.loads(args.actual_installation_days)
            except json.JSONDecodeError as exc:
                raise RuleError("invalid_arguments", "安装日期明细不是有效 JSON") from exc
            from .order_index import save_order_annotations
            try:
                result = save_order_annotations(
                    config,
                    args.order_id,
                    user_note=args.note,
                    planned_days=planned_days,
                    actual_days=actual_days,
                )
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "add-factory":
            if not args.order_id or not args.factory_order or not args.factory_name:
                raise RuleError("invalid_arguments", "add-factory 需要 --order-id、--factory-order 和 --factory-name")
            from .order_index import add_manual_factory
            try:
                result = add_manual_factory(config, args.order_id, args.factory_order, args.factory_name)
            except ValueError as exc:
                raise RuleError("invalid_arguments", str(exc)) from exc
        elif args.command == "set-ignore":
            if not args.name or args.ignored is None:
                raise RuleError("invalid_arguments", "set-ignore 需要 --name 和 --ignored")
            ignored = args.ignored == "true"
            for name in args.name:
                set_ignored(config, name, ignored)
            result = {
                "saved": True,
                "names": args.name,
                "ignored": ignored,
            }
        elif args.command == "add-hardware":
            if not args.order_id or not args.factory_name or not args.product_code or args.quantity is None:
                raise RuleError(
                    "invalid_arguments",
                    "add-hardware 需要 --order-id、--factory-name、--product-code 和 --quantity",
                )
            result = preview_manual_hardware(
                config,
                args.order_id,
                args.factory_name,
                args.product_code,
                args.quantity,
                args.remarks,
            )
            if args.confirm_write:
                traveler, backup, saved = add_manual_hardware(
                    config,
                    args.order_id,
                    args.factory_name,
                    args.product_code,
                    args.quantity,
                    args.remarks,
                )
                result = {**saved, "updated": str(traveler), "backup": str(backup)}
            else:
                result = {"approval_required": True, "preview": result}
        elif args.command == "assign-material":
            if not args.folder or not args.order_id or not args.materials_file:
                raise RuleError("invalid_arguments", "assign-material 需要 --folder、--order-id 和 --materials-file")
            folder = args.folder.resolve()
            selected = Path(args.materials_file).expanduser().resolve()
            if not selected.is_file() or selected.parent != folder:
                raise RuleError("invalid_arguments", "所选 material 文件必须位于订单文件夹内")
            save_material_assignment(config, f"{folder}::{args.order_id.upper()}", str(selected))
            result = {"saved": True, "order_id": args.order_id.upper(), "materials_file": str(selected)}
        elif args.command == "generate-material":
            if not args.folder or not args.order_id:
                raise RuleError("invalid_arguments", "generate-material 需要 --folder 和 --order-id")
            folder = args.folder.expanduser().resolve()
            created = generate_material_from_reports(folder, args.order_id)
            from .order_index import OrderIndexStore, _record_generated_material_baseline

            baseline_store = OrderIndexStore(config.workflow_database)
            try:
                _record_generated_material_baseline(
                    baseline_store,
                    folder,
                    created,
                    order_id=args.order_id,
                )
                baseline_store.commit()
            finally:
                baseline_store.close()
            result = {"created": True, "order_id": args.order_id.upper(), "materials_file": str(created)}
        elif args.command == "generate-material-from-travelers":
            if not args.folder or not args.order_id:
                raise RuleError("invalid_arguments", "generate-material-from-travelers 需要 --folder 和 --order-id")
            result = generate_material_from_travelers(
                config,
                args.folder,
                args.order_id,
                confirm_write=args.confirm_write,
            )
        elif args.command == "generate-db":
            if not args.order_id:
                raise RuleError("invalid_arguments", "generate-db 需要 --order-id")
            result = {"created": str(generate_database_order_traveler(config, args.order_id))}
        elif args.command == "stock-check":
            if not args.folder:
                raise RuleError("invalid_arguments", "stock-check 需要 --folder")
            from .inventory import check_order_stock
            result = check_order_stock(config, args.folder)
        elif args.command == "preview-related":
            if not args.folder:
                raise RuleError("invalid_arguments", "preview-related 需要 --folder")
            result = preview_related_orders(config, args.folder)
        elif args.command == "refresh-aimes":
            from .order_index import sync_aimes_index
            result = sync_aimes_index(config, force=True)
            result["saved"] = True
        elif args.command == "update-related":
            if not args.folder:
                raise RuleError("invalid_arguments", "update-related 需要 --folder")
            result = update_related_orders(
                config,
                args.folder,
                include_hardware=args.include_hardware == "true",
            )
        else:
            if not args.folder:
                raise RuleError("invalid_arguments", f"{args.command} 需要 --folder")
            preview = preview_order(config, args.folder)
            result = preview_payload(config, preview)
            if args.command == "generate":
                result["created"] = str(generate_order_traveler(config, preview))
                result["existing_traveler"] = result["created"]
            elif args.command == "temporary":
                result["temporary"] = str(generate_temporary_traveler(config, preview))
            elif args.command == "update":
                updated, backup = update_order_traveler(config, preview)
                result["updated"] = str(updated)
                result["backup"] = str(backup)
        logger.event("backend.command.completed", "订单工作流操作完成", details={"action": args.command})
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except RuleError as exc:
        logger.event(
            "backend.command.failed",
            "订单工作流操作失败",
            details={"action": args.command, "code": exc.code, "error": str(exc)},
        )
        print(json.dumps({"fatal": {"code": exc.code, "message": str(exc), **exc.context}}, ensure_ascii=False, indent=2))
        return 2
    except Exception as exc:
        logger.event(
            "backend.command.failed",
            "订单工作流操作发生未预期错误",
            details={"action": args.command, "code": "unexpected_excel_error", "error": str(exc)},
        )
        print(json.dumps({
            "fatal": {
                "code": "unexpected_excel_error",
                "message": f"Excel 文件无法读取，请检查文件是否损坏或仍在编辑：{exc}",
            }
        }, ensure_ascii=False, indent=2))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
